from io import BytesIO
from typing import Iterable, Sequence

from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


def export_excel_response(
    filename: str,
    headers: Sequence[str],
    rows: Iterable[Sequence],
    sheet_title: str = "Report",
):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31] if sheet_title else "Report"

    ws.append(list(headers))

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    rows = list(rows)

    for row in rows:
        ws.append(list(row))

    for col_idx, _ in enumerate(headers, start=1):
        max_length = 0
        column_letter = get_column_letter(col_idx)

        for cell in ws[column_letter]:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_length:
                max_length = len(value)

        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 40)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    safe_name = filename.replace(" ", "_").replace("/", "_").replace("\\", "_")

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{safe_name}.xlsx"'
        },
    )