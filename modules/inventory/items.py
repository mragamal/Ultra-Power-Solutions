import csv
import io
from urllib.parse import quote

from fastapi import APIRouter, Request, Form, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse

from db import get_conn
from layout import render_page
from modules.inventory.core import (
    next_item_code,
    ensure_inventory_tables,
    get_uom_names,
)

try:
    from openpyxl import Workbook, load_workbook
except Exception:
    Workbook = None
    load_workbook = None

router = APIRouter()


def safe(x):
    return "" if x is None else str(x).strip()


def to_int_flag(value, default=1):
    text = safe(value).lower()
    if text in ("1", "true", "yes", "y", "active"):
        return 1
    if text in ("0", "false", "no", "n", "inactive"):
        return 0
    try:
        return 1 if int(value) == 1 else 0
    except Exception:
        return int(default)


def to_float(value, default=0.0):
    try:
        text = safe(value).replace(",", "")
        if text == "":
            return float(default)
        return float(text)
    except Exception:
        return float(default)


def next_item_code_from_conn(conn):
    row = conn.execute(
        """
        SELECT code
        FROM items
        WHERE COALESCE(code, '') <> ''
        ORDER BY id DESC
        LIMIT 1
        """
    ).fetchone()
    last = safe(row["code"]) if row else ""
    if not last:
        return 1
    try:
        return int(last.split("-")[-1]) + 1
    except Exception:
        return 1


def ensure_uom_exists_in_conn(conn, uom_name):
    name = safe(uom_name)
    if not name:
        return

    exists = conn.execute(
        """
        SELECT id
        FROM inventory_uoms
        WHERE UPPER(COALESCE(name, '')) = UPPER(?)
        LIMIT 1
        """,
        (name,),
    ).fetchone()
    if exists:
        return

    code = name.upper().replace(" ", "_")[:20] or "UOM"
    code_row = conn.execute(
        """
        SELECT id
        FROM inventory_uoms
        WHERE UPPER(COALESCE(code, '')) = UPPER(?)
        LIMIT 1
        """,
        (code,),
    ).fetchone()
    if code_row:
        code = f"{code}_{int(code_row['id']) + 1}"

    conn.execute(
        """
        INSERT INTO inventory_uoms (code, name, is_active)
        VALUES (?, ?, 1)
        """,
        (code, name),
    )


def get_header_value(row, aliases):
    for key in aliases:
        if key in row and safe(row.get(key)):
            return safe(row.get(key))
    return ""


def normalize_import_row(row):
    return {
        "code": get_header_value(row, ["code", "item_code", "sku"]),
        "name": get_header_value(row, ["name", "item_name"]),
        "category": get_header_value(row, ["category", "item_category"]),
        "uom": get_header_value(row, ["uom", "unit", "unit_of_measure", "unit of measure"]) or "Unit",
        "item_type": get_header_value(row, ["item_type", "type"]) or "stock_item",
        "standard_cost": get_header_value(row, ["standard_cost", "cost", "unit_cost", "standard cost"]) or "0",
        "sale_price": get_header_value(row, ["sale_price", "price", "selling_price", "sale price"]) or "0",
        "is_active": get_header_value(row, ["is_active", "active", "status"]) or "1",
    }


def parse_csv_rows(file_bytes):
    text = file_bytes.decode("utf-8-sig", errors="ignore")
    reader = csv.DictReader(io.StringIO(text))
    return [dict(r) for r in reader]


def parse_xlsx_rows(file_bytes):
    if load_workbook is None:
        raise Exception("Excel import is not available right now.")

    workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [safe(h).lower() for h in rows[0]]
    result = []
    for data_row in rows[1:]:
        item = {}
        for i, header in enumerate(headers):
            if not header:
                continue
            item[header] = "" if i >= len(data_row) or data_row[i] is None else str(data_row[i])
        result.append(item)
    return result


def uom_options_html(selected_value="Unit"):
    selected_value = safe(selected_value) or "Unit"
    names = get_uom_names(active_only=True)
    if "Unit" not in names:
        names.insert(0, "Unit")
    if selected_value and selected_value not in names:
        names.append(selected_value)

    out = ""
    for name in names:
        selected = "selected" if name == selected_value else ""
        out += f'<option value="{name}" {selected}>{name}</option>'
    return out


ensure_inventory_tables()


@router.get("/ui/inventory/items", response_class=HTMLResponse)
def items_list(request: Request):
    conn = get_conn()
    rows = conn.execute(
        """
        SELECT *
        FROM items
        ORDER BY code, name
        """
    ).fetchall()
    conn.close()

    msg = safe(request.query_params.get("msg"))
    msg_html = f'<div class="msg ok">{msg}</div>' if msg else ""

    body = ""
    for r in rows:
        active = "Yes" if int(r["is_active"] or 0) == 1 else "No"
        body += f"""
        <tr>
            <td>{r['code'] or ''}</td>
            <td>{r['name'] or ''}</td>
            <td>{r['category'] or ''}</td>
            <td>{r['uom'] or ''}</td>
            <td>{r['item_type'] or ''}</td>
            <td style="text-align:right;">{to_float(r['standard_cost']):,.2f}</td>
            <td style="text-align:right;">{to_float(r['sale_price']):,.2f}</td>
            <td>{active}</td>
            <td><a class="btn gray" href="/ui/inventory/items/{r['id']}/edit">Edit</a></td>
        </tr>
        """

    if not body:
        body = "<tr><td colspan='9' style='text-align:center;'>No items found.</td></tr>"

    html = f"""
    <div class="card">
        {msg_html}
        <div style="display:flex;justify-content:space-between;align-items:center;gap:10px;flex-wrap:wrap;">
            <h2>Items</h2>
            <div style="display:flex;gap:8px;flex-wrap:wrap;">
                <a class="btn blue" href="/ui/inventory/items/template.csv">Template CSV</a>
                <a class="btn blue" href="/ui/inventory/items/template.xlsx">Template Excel</a>
                <a class="btn green" href="/ui/inventory/items/new">+ New Item</a>
            </div>
        </div>

        <form method="post" action="/ui/inventory/items/import" enctype="multipart/form-data" style="margin-top:14px;">
            <div class="row">
                <div class="col">
                    <label>Import Items (CSV / XLSX)</label>
                    <input type="file" name="file" accept=".csv,.xlsx" required>
                </div>
                <div class="col" style="display:flex;align-items:end;">
                    <button class="btn green" type="submit">Import</button>
                </div>
            </div>
        </form>
    </div>

    <div class="card">
        <table>
            <tr>
                <th>Code</th>
                <th>Name</th>
                <th>Category</th>
                <th>UOM</th>
                <th>Type</th>
                <th style="text-align:right;">Cost</th>
                <th style="text-align:right;">Sale Price</th>
                <th>Active</th>
                <th>Actions</th>
            </tr>
            {body}
        </table>
    </div>
    """
    return HTMLResponse(render_page("Items", html, current_path=request.url.path))


def item_form_html(action, data=None):
    data = data or {}
    selected_stock = "selected" if data.get("item_type", "stock_item") == "stock_item" else ""
    selected_service = "selected" if data.get("item_type", "") == "service" else ""
    selected_yes = "selected" if str(data.get("is_active", "1")) == "1" else ""
    selected_no = "selected" if str(data.get("is_active", "1")) != "1" else ""

    return f"""
    <div class="card">
        <h2>{'Edit Item' if '/edit' in action else 'New Item'}</h2>
        <form method="post" action="{action}">
            <div class="row">
                <div class="col">
                    <label>Code</label>
                    <input name="code" value="{data.get('code') or next_item_code()}" readonly required>
                </div>
                <div class="col">
                    <label>Name</label>
                    <input name="name" value="{data.get('name') or ''}" required>
                </div>
            </div>

            <div class="row" style="margin-top:14px;">
                <div class="col">
                    <label>Category</label>
                    <input name="category" value="{data.get('category') or ''}">
                </div>
                <div class="col">
                    <label>UOM</label>
                    <select name="uom">
                        {uom_options_html(data.get('uom') or 'Unit')}
                    </select>
                </div>
            </div>

            <div class="row" style="margin-top:14px;">
                <div class="col">
                    <label>Type</label>
                    <select name="item_type">
                        <option value="stock_item" {selected_stock}>Stock Item</option>
                        <option value="service" {selected_service}>Service</option>
                    </select>
                </div>
                <div class="col">
                    <label>Standard Cost</label>
                    <input name="standard_cost" type="number" step="0.01" value="{data.get('standard_cost') or '0'}">
                </div>
            </div>

            <div class="row" style="margin-top:14px;">
                <div class="col">
                    <label>Sale Price</label>
                    <input name="sale_price" type="number" step="0.01" value="{data.get('sale_price') or '0'}">
                </div>
                <div class="col">
                    <label>Active</label>
                    <select name="is_active">
                        <option value="1" {selected_yes}>Yes</option>
                        <option value="0" {selected_no}>No</option>
                    </select>
                </div>
            </div>

            <div style="margin-top:20px;">
                <button class="btn green" type="submit">Save</button>
                <a class="btn gray" href="/ui/inventory/items">Back</a>
            </div>
        </form>
    </div>
    """


@router.get("/ui/inventory/items/new", response_class=HTMLResponse)
def item_new(request: Request):
    return HTMLResponse(render_page("New Item", item_form_html("/ui/inventory/items/new"), current_path=request.url.path))


@router.post("/ui/inventory/items/new")
def item_create(
    code: str = Form(""),
    name: str = Form(...),
    category: str = Form(""),
    uom: str = Form("Unit"),
    item_type: str = Form("stock_item"),
    standard_cost: str = Form("0"),
    sale_price: str = Form("0"),
    is_active: int = Form(1),
):
    conn = get_conn()
    ensure_uom_exists_in_conn(conn, safe(uom) or "Unit")
    conn.execute(
        """
        INSERT INTO items (code, name, category, uom, item_type, standard_cost, sale_price, is_active)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            safe(code) or next_item_code(),
            safe(name),
            safe(category),
            safe(uom) or "Unit",
            safe(item_type) or "stock_item",
            to_float(standard_cost),
            to_float(sale_price),
            int(is_active or 0),
        ),
    )
    conn.commit()
    conn.close()
    return RedirectResponse("/ui/inventory/items", status_code=302)


@router.get("/ui/inventory/items/{item_id}/edit", response_class=HTMLResponse)
def item_edit(request: Request, item_id: int):
    conn = get_conn()
    row = conn.execute("SELECT * FROM items WHERE id = ?", (item_id,)).fetchone()
    conn.close()
    if not row:
        return HTMLResponse("Item not found", status_code=404)
    return HTMLResponse(
        render_page(
            "Edit Item",
            item_form_html(f"/ui/inventory/items/{item_id}/edit", dict(row)),
            current_path=request.url.path,
        )
    )


@router.post("/ui/inventory/items/{item_id}/edit")
def item_update(
    item_id: int,
    code: str = Form(""),
    name: str = Form(...),
    category: str = Form(""),
    uom: str = Form("Unit"),
    item_type: str = Form("stock_item"),
    standard_cost: str = Form("0"),
    sale_price: str = Form("0"),
    is_active: int = Form(1),
):
    conn = get_conn()
    ensure_uom_exists_in_conn(conn, safe(uom) or "Unit")
    conn.execute(
        """
        UPDATE items
        SET code = ?, name = ?, category = ?, uom = ?, item_type = ?, standard_cost = ?, sale_price = ?, is_active = ?
        WHERE id = ?
        """,
        (
            safe(code),
            safe(name),
            safe(category),
            safe(uom) or "Unit",
            safe(item_type) or "stock_item",
            to_float(standard_cost),
            to_float(sale_price),
            int(is_active or 0),
            item_id,
        ),
    )
    conn.commit()
    conn.close()
    return RedirectResponse("/ui/inventory/items", status_code=302)


@router.get("/ui/inventory/items/template.csv")
def items_template_csv():
    headers = ["code", "name", "category", "uom", "item_type", "standard_cost", "sale_price", "is_active"]
    rows = [
        ["ITM-0001", "Printer Paper A4", "Stationery", "Pack", "stock_item", "150.00", "180.00", "1"],
        ["ITM-0002", "Mouse", "IT", "Piece", "stock_item", "250.00", "300.00", "1"],
        ["ITM-0003", "Installation Service", "Services", "Unit", "service", "0.00", "500.00", "1"],
    ]
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)

    data = output.getvalue().encode("utf-8-sig")
    stream = io.BytesIO(data)
    stream.seek(0)
    headers = {"Content-Disposition": 'attachment; filename="items_template.csv"'}
    return StreamingResponse(stream, media_type="text/csv; charset=utf-8", headers=headers)


@router.get("/ui/inventory/items/template.xlsx")
def items_template_xlsx():
    if Workbook is None:
        return RedirectResponse("/ui/inventory/items?msg=" + quote("Excel template is not available. Use CSV template."))

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Items Template"
    sheet.append(["code", "name", "category", "uom", "item_type", "standard_cost", "sale_price", "is_active"])
    sheet.append(["ITM-0001", "Printer Paper A4", "Stationery", "Pack", "stock_item", "150.00", "180.00", "1"])
    sheet.append(["ITM-0002", "Mouse", "IT", "Piece", "stock_item", "250.00", "300.00", "1"])
    sheet.append(["ITM-0003", "Installation Service", "Services", "Unit", "service", "0.00", "500.00", "1"])

    uom_sheet = workbook.create_sheet("UOMs")
    uom_sheet.append(["uom_name"])
    for name in get_uom_names(active_only=True):
        uom_sheet.append([name])

    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)
    headers = {"Content-Disposition": 'attachment; filename="items_template.xlsx"'}
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.post("/ui/inventory/items/import")
async def items_import(file: UploadFile = File(...)):
    filename = safe(file.filename)
    if not filename:
        return RedirectResponse("/ui/inventory/items?msg=" + quote("Please select a file."), status_code=302)

    ext = filename.lower().split(".")[-1] if "." in filename else ""
    file_bytes = await file.read()
    if not file_bytes:
        return RedirectResponse("/ui/inventory/items?msg=" + quote("Selected file is empty."), status_code=302)

    try:
        if ext == "csv":
            raw_rows = parse_csv_rows(file_bytes)
        elif ext == "xlsx":
            raw_rows = parse_xlsx_rows(file_bytes)
        else:
            return RedirectResponse("/ui/inventory/items?msg=" + quote("File type not supported. Use CSV or XLSX."), status_code=302)
    except Exception as e:
        return RedirectResponse("/ui/inventory/items?msg=" + quote(f"Import failed: {safe(e)}"), status_code=302)

    if not raw_rows:
        return RedirectResponse("/ui/inventory/items?msg=" + quote("No rows found in the file."), status_code=302)

    conn = get_conn()
    imported = 0
    updated = 0
    skipped = 0
    next_code_no = next_item_code_from_conn(conn)

    try:
        for raw_row in raw_rows:
            row = normalize_import_row(raw_row)
            name = safe(row["name"])
            if not name:
                skipped += 1
                continue

            code = safe(row["code"])
            if not code:
                code = f"ITM-{next_code_no:04d}"
                next_code_no += 1

            category = safe(row["category"])
            uom = safe(row["uom"]) or "Unit"
            item_type = safe(row["item_type"]) or "stock_item"
            if item_type not in ("stock_item", "service"):
                item_type = "stock_item"
            standard_cost = to_float(row["standard_cost"])
            sale_price = to_float(row["sale_price"])
            is_active = to_int_flag(row["is_active"], default=1)

            ensure_uom_exists_in_conn(conn, uom)

            exists = conn.execute(
                """
                SELECT id
                FROM items
                WHERE UPPER(COALESCE(code, '')) = UPPER(?)
                LIMIT 1
                """,
                (code,),
            ).fetchone()

            if exists:
                conn.execute(
                    """
                    UPDATE items
                    SET name = ?, category = ?, uom = ?, item_type = ?, standard_cost = ?, sale_price = ?, is_active = ?
                    WHERE id = ?
                    """,
                    (name, category, uom, item_type, standard_cost, sale_price, is_active, exists["id"]),
                )
                updated += 1
            else:
                conn.execute(
                    """
                    INSERT INTO items (code, name, category, uom, item_type, standard_cost, sale_price, is_active)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (code, name, category, uom, item_type, standard_cost, sale_price, is_active),
                )
                imported += 1

        conn.commit()
    except Exception as e:
        conn.rollback()
        conn.close()
        return RedirectResponse("/ui/inventory/items?msg=" + quote(f"Import failed: {safe(e)}"), status_code=302)

    conn.close()
    msg = f"Import completed. Added: {imported}, Updated: {updated}, Skipped: {skipped}."
    return RedirectResponse("/ui/inventory/items?msg=" + quote(msg), status_code=302)
