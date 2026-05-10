
from io import BytesIO
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from datetime import datetime, timedelta

import openpyxl
from fastapi import APIRouter, File, Request, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse

from audit import actor_name_from_request, render_audit_log_card, safe_log_action
from auth import can, current_user
from db import get_conn
from layout import render_page
from modules.accounting.allocation_engine import (
    get_allocated_total_for_document,
    refresh_customer_invoice_payment_status,
    refresh_vendor_bill_payment_status,
)

router = APIRouter()


# =========================================================
# HELPERS
# =========================================================
def safe(x):
    return "" if x is None else str(x).strip()


def get_lang(request: Request | None = None):
    try:
        if request and safe(request.query_params.get("lang")).lower() == "ar":
            return "ar"
    except Exception:
        pass
    return "en"


def tr(lang: str, en: str, ar: str):
    return ar if lang == "ar" else en


def with_lang(url: str, lang: str):
    if lang != "ar":
        return url
    return url + ("&" if "?" in url else "?") + "lang=ar"


def accounting_allowed(request: Request, action: str) -> bool:
    return can(request, "accounting", action)


def is_admin_user(request: Request) -> bool:
    user = current_user(request)
    return bool(user and (user.get("role_code") or "").lower() == "admin")


def can_edit_journal_entry(request: Request, entry) -> bool:
    if not entry:
        return False
    status = safe(entry["status"]).lower()
    if status in ("draft", "pending_final_post"):
        return accounting_allowed(request, "edit")
    return False


def permission_denied(lang: str, en: str, ar: str):
    return HTMLResponse(tr(lang, en, ar), status_code=403)


def to_decimal(value, default="0"):
    try:
        text = safe(value).replace(",", "")
        if text in ["", ".", "-", "-."]:
            text = default
        return Decimal(text)
    except (InvalidOperation, ValueError, TypeError):
        return Decimal(default)


def q2(value):
    return to_decimal(value).quantize(Decimal("1.00"), rounding=ROUND_HALF_UP)


def money(value, places=2):
    try:
        d = Decimal(str(value or 0))
    except Exception:
        d = Decimal("0")
    q = Decimal("1." + ("0" * places))
    d = d.quantize(q, rounding=ROUND_HALF_UP)
    return f"{d:,.{places}f}"


def dec2(value):
    try:
        return Decimal(str(value or 0)).quantize(Decimal("1.00"), rounding=ROUND_HALF_UP)
    except Exception:
        return Decimal("0.00")


def row_value(row, key, default=""):
    try:
        if isinstance(row, dict):
            return row.get(key, default)
        return row[key]
    except Exception:
        try:
            return getattr(row, key)
        except Exception:
            return default


def ensure_column(conn, table_name, column_name, alter_sql):
    cols = conn.execute(f"PRAGMA table_info({table_name})").fetchall()
    names = [c["name"] for c in cols]
    if column_name not in names:
        conn.execute(alter_sql)


def table_exists(conn, table_name):
    row = conn.execute("""
        SELECT name
        FROM sqlite_master
        WHERE type='table' AND name=?
    """, (table_name,)).fetchone()
    return bool(row)


def get_columns(conn, table_name):
    if not table_exists(conn, table_name):
        return []
    return [r["name"] for r in conn.execute(f"PRAGMA table_info({table_name})").fetchall()]


def calc_due_date(entry_date: str, payment_term_days) -> str:
    try:
        base = datetime.strptime(safe(entry_date), "%Y-%m-%d")
        days = int(payment_term_days or 0)
        return (base + timedelta(days=days)).strftime("%Y-%m-%d")
    except Exception:
        return safe(entry_date)


def normalize_partner_type(value: str) -> str:
    key = safe(value).strip().lower().replace(" ", "_").replace("-", "_")
    aliases = {
        "customer": "customer",
        "cust": "customer",
        "client": "customer",
        "????": "customer",
        "?????": "customer",
        "vendor": "vendor",
        "supplier": "vendor",
        "????": "vendor",
        "??????": "vendor",
        "employee": "employee",
        "emp": "employee",
        "staff": "employee",
        "????": "employee",
        "??????": "employee",
    }
    return aliases.get(key, "")


def _normalize_lookup_text(value: str) -> str:
    text = safe(value).strip().lower()
    replacements = {
        "?": "?",
        "?": "?",
        "?": "?",
        "?": "?",
        "?": "?",
        "?": "?",
        "?": "?",
    }
    for src, dst in replacements.items():
        text = text.replace(src, dst)
    text = " ".join(text.split())
    return text


def resolve_partner_reference(conn, partner_type: str, partner_code: str = "", partner_name: str = ""):
    ptype = normalize_partner_type(partner_type)
    code = safe(partner_code).strip()
    name = safe(partner_name).strip()
    if not ptype:
        return "", None

    if ptype in ("customer", "vendor"):
        if code:
            row = conn.execute("""
                SELECT id, code, name
                FROM partners
                WHERE LOWER(COALESCE(partner_type, '')) = ?
                  AND LOWER(COALESCE(code, '')) = ?
                LIMIT 1
            """, (ptype, code.lower())).fetchone()
            if row:
                return ptype, row["id"]

        if name:
            row = conn.execute("""
                SELECT id, code, name
                FROM partners
                WHERE LOWER(COALESCE(partner_type, '')) = ?
                  AND LOWER(COALESCE(name, '')) = ?
                LIMIT 1
            """, (ptype, name.lower())).fetchone()
            if row:
                return ptype, row["id"]

            wanted = _normalize_lookup_text(name)
            candidates = conn.execute("""
                SELECT id, code, name
                FROM partners
                WHERE LOWER(COALESCE(partner_type, '')) = ?
                ORDER BY id
            """, (ptype,)).fetchall()
            exact = [r for r in candidates if _normalize_lookup_text(r["name"]) == wanted]
            if len(exact) == 1:
                return ptype, exact[0]["id"]
            partial = [r for r in candidates if wanted and (wanted in _normalize_lookup_text(r["name"]) or _normalize_lookup_text(r["name"]) in wanted)]
            if len(partial) == 1:
                return ptype, partial[0]["id"]

        raise Exception(f"{ptype.title()} not found for code/name: {code or name}")

    if ptype == "employee":
        if code:
            row = conn.execute("""
                SELECT id, code
                FROM employees
                WHERE LOWER(COALESCE(code, '')) = ?
                LIMIT 1
            """, (code.lower(),)).fetchone()
            if row:
                return ptype, row["id"]

        if name:
            row = conn.execute("""
                SELECT id, code
                FROM employees
                WHERE LOWER(COALESCE(name, '')) = ?
                   OR LOWER(COALESCE(employee_name, '')) = ?
                   OR LOWER(COALESCE(full_name, '')) = ?
                LIMIT 1
            """, (name.lower(), name.lower(), name.lower())).fetchone()
            if row:
                return ptype, row["id"]

            wanted = _normalize_lookup_text(name)
            candidates = conn.execute("""
                SELECT id, code,
                       COALESCE(full_name, employee_name, name, '') AS disp_name
                FROM employees
                ORDER BY id
            """).fetchall()
            exact = [r for r in candidates if _normalize_lookup_text(r["disp_name"]) == wanted]
            if len(exact) == 1:
                return ptype, exact[0]["id"]
            partial = [r for r in candidates if wanted and (wanted in _normalize_lookup_text(r["disp_name"]) or _normalize_lookup_text(r["disp_name"]) in wanted)]
            if len(partial) == 1:
                return ptype, partial[0]["id"]

        raise Exception(f"Employee not found for code/name: {code or name}")

    return "", None


def get_partner_display(conn, partner_type: str, partner_id):
    ptype = normalize_partner_type(partner_type)
    pid = int(partner_id or 0) if safe(partner_id) else 0
    if not ptype or pid <= 0:
        return "", ""

    if ptype in ("customer", "vendor"):
        row = conn.execute("""
            SELECT code, name
            FROM partners
            WHERE id = ?
            LIMIT 1
        """, (pid,)).fetchone()
        if row:
            return safe(row["code"]), safe(row["name"])
        return "", ""

    if ptype == "employee":
        row = conn.execute("""
            SELECT code, COALESCE(name, employee_name, full_name, '') AS display_name
            FROM employees
            WHERE id = ?
            LIMIT 1
        """, (pid,)).fetchone()
        if row:
            return safe(row["code"]), safe(row["display_name"])
        return "", ""

    return "", ""


def partner_type_label(value: str, lang: str = "en"):
    ptype = normalize_partner_type(value)
    labels = {
        "customer": tr(lang, "Customer", "عميل"),
        "vendor": tr(lang, "Vendor", "مورد"),
        "employee": tr(lang, "Employee", "موظف"),
    }
    return labels.get(ptype, "")


def get_partner_display_text(conn, partner_type: str, partner_id):
    code, name = get_partner_display(conn, partner_type, partner_id)
    return " - ".join([part for part in [code, name] if part])


# =========================================================
# SETTINGS
# =========================================================
def get_setting_value(key: str, default=None, conn=None):
    if conn:
        row = conn.execute("""
            SELECT value
            FROM accounting_settings
            WHERE key = ?
            LIMIT 1
        """, (key,)).fetchone()
        if row and row["value"] not in [None, ""]:
            return row["value"]
    else:
        _conn = get_conn()
        try:
            row = _conn.execute("""
                SELECT value
                FROM accounting_settings
                WHERE key = ?
                LIMIT 1
            """, (key,)).fetchone()
            if row and row["value"] not in [None, ""]:
                return row["value"]
        except Exception:
            pass
        finally:
            _conn.close()

    fallback = {
        "journal_prefix": "JV",
    }
    return fallback.get(key, default)


def infer_required_partner_type(conn, account_code: str) -> str:
    code = safe(account_code)
    if not code:
        return ""

    customer_control = safe(get_setting_value("customer_control_account", "112100", conn=conn))
    vendor_control = safe(get_setting_value("vendor_control_account", "211100", conn=conn))
    employee_custody = safe(get_setting_value("employee_custody_account", "", conn=conn))
    employee_advance = safe(get_setting_value("employee_advance_account", "", conn=conn))

    if code == customer_control:
        return "customer"
    if code == vendor_control:
        return "vendor"
    if code and code in {employee_custody, employee_advance}:
        return "employee"

    try:
        row = conn.execute("""
            SELECT name, type, level1, level2
            FROM accounts
            WHERE code = ?
            LIMIT 1
        """, (code,)).fetchone()
    except Exception:
        row = None

    if not row:
        return ""

    haystack = " ".join([
        safe(row["name"]).lower(),
        safe(row["type"]).lower(),
        safe(row["level1"]).lower(),
        safe(row["level2"]).lower(),
    ])

    if any(token in haystack for token in ["custody", "advance", "loan", "employee custody", "employee advance", "?????", "???", "?????", "???", "????"]):
        return "employee"
    if any(token in haystack for token in ["receivable", "customers", "customer", "???????", "????", "??????"]):
        return "customer"
    if any(token in haystack for token in ["vendor", "vendors", "supplier", "payable", "????????", "????", "??????"]):
        return "vendor"

    return ""


# =========================================================
# DB SCHEMA
# =========================================================
def ensure_tables():
    conn = get_conn()
    try:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS journal_entries (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                entry_no TEXT,
                entry_date TEXT,
                description TEXT,
                reference TEXT,
                status TEXT DEFAULT 'draft',
                source_type TEXT,
                source_id INTEGER,
                reversed_by_journal_id INTEGER,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
        """)

        conn.execute("""
            CREATE TABLE IF NOT EXISTS journal_lines (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                journal_id INTEGER,
                line_no INTEGER DEFAULT 1,
                line_description TEXT,
                account_code TEXT,
                debit REAL DEFAULT 0,
                credit REAL DEFAULT 0,
                partner_type TEXT,
                partner_id INTEGER,
                cost_center_id INTEGER
            )
        """)

        ensure_column(conn, "journal_entries", "entry_no", "ALTER TABLE journal_entries ADD COLUMN entry_no TEXT")
        ensure_column(conn, "journal_entries", "entry_date", "ALTER TABLE journal_entries ADD COLUMN entry_date TEXT")
        ensure_column(conn, "journal_entries", "description", "ALTER TABLE journal_entries ADD COLUMN description TEXT")
        ensure_column(conn, "journal_entries", "reference", "ALTER TABLE journal_entries ADD COLUMN reference TEXT")
        ensure_column(conn, "journal_entries", "status", "ALTER TABLE journal_entries ADD COLUMN status TEXT DEFAULT 'draft'")
        ensure_column(conn, "journal_entries", "source_type", "ALTER TABLE journal_entries ADD COLUMN source_type TEXT")
        ensure_column(conn, "journal_entries", "source_id", "ALTER TABLE journal_entries ADD COLUMN source_id INTEGER")
        ensure_column(conn, "journal_entries", "reversed_by_journal_id", "ALTER TABLE journal_entries ADD COLUMN reversed_by_journal_id INTEGER")
        ensure_column(conn, "journal_entries", "created_at", "ALTER TABLE journal_entries ADD COLUMN created_at TEXT DEFAULT CURRENT_TIMESTAMP")

        ensure_column(conn, "journal_lines", "journal_id", "ALTER TABLE journal_lines ADD COLUMN journal_id INTEGER")
        ensure_column(conn, "journal_lines", "line_no", "ALTER TABLE journal_lines ADD COLUMN line_no INTEGER DEFAULT 1")
        ensure_column(conn, "journal_lines", "line_description", "ALTER TABLE journal_lines ADD COLUMN line_description TEXT")
        ensure_column(conn, "journal_lines", "account_code", "ALTER TABLE journal_lines ADD COLUMN account_code TEXT")
        ensure_column(conn, "journal_lines", "debit", "ALTER TABLE journal_lines ADD COLUMN debit REAL DEFAULT 0")
        ensure_column(conn, "journal_lines", "credit", "ALTER TABLE journal_lines ADD COLUMN credit REAL DEFAULT 0")
        ensure_column(conn, "journal_lines", "partner_type", "ALTER TABLE journal_lines ADD COLUMN partner_type TEXT")
        ensure_column(conn, "journal_lines", "partner_id", "ALTER TABLE journal_lines ADD COLUMN partner_id INTEGER")
        ensure_column(conn, "journal_lines", "cost_center_id", "ALTER TABLE journal_lines ADD COLUMN cost_center_id INTEGER")

        conn.commit()
    finally:
        conn.close()


ensure_tables()


# =========================================================
# MASTER FUNCTIONS
# =========================================================
def next_entry_no(conn):
    prefix = get_setting_value("journal_prefix", "JV", conn=conn)
    row = conn.execute("""
        SELECT entry_no
        FROM journal_entries
        WHERE COALESCE(entry_no, '') <> ''
        ORDER BY id DESC
        LIMIT 1
    """).fetchone()

    if not row or not row["entry_no"]:
        return f"{prefix}-0000001"

    try:
        last_num = int(str(row["entry_no"]).split("-")[-1])
    except Exception:
        last_num = 0

    return f"{prefix}-{last_num + 1:07d}"


def account_options(selected_code=""):
    conn = get_conn()
    try:
        rows = conn.execute("""
            SELECT code, name
            FROM accounts
            WHERE COALESCE(is_active, 1) = 1
              AND COALESCE(is_group, 0) = 0
              AND COALESCE(allow_posting, 1) = 1
            ORDER BY code, name
        """).fetchall()
    finally:
        conn.close()

    html = "<option value=''>-- Select Account --</option>"
    for row in rows:
        sel = "selected" if safe(selected_code) == safe(row["code"]) else ""
        html += f"<option value='{safe(row['code'])}' {sel}>{safe(row['code'])} - {safe(row['name'])}</option>"
    return html


def validate_account_for_posting(conn, account_code: str):
    row = conn.execute("""
        SELECT code, name,
               COALESCE(is_group, 0) AS is_group,
               COALESCE(allow_posting, 1) AS allow_posting,
               COALESCE(is_active, 1) AS is_active
        FROM accounts
        WHERE code = ?
        LIMIT 1
    """, (account_code,)).fetchone()

    if not row:
        raise Exception(f"Account {account_code} not found")
    if int(row["is_group"] or 0) == 1:
        raise Exception(f"Account {row['code']} - {row['name']} is a group account")
    if int(row["allow_posting"] or 0) == 0:
        raise Exception(f"Account {row['code']} - {row['name']} is not allowed for posting")
    if int(row["is_active"] or 0) == 0:
        raise Exception(f"Account {row['code']} - {row['name']} is inactive")


def get_entry(conn, journal_id: int):
    return conn.execute("""
        SELECT *
        FROM journal_entries
        WHERE id = ?
        LIMIT 1
    """, (journal_id,)).fetchone()


def get_entry_lines(conn, journal_id: int):
    return conn.execute("""
        SELECT jl.*, COALESCE(a.name, '') AS account_name
        FROM journal_lines jl
        LEFT JOIN accounts a ON a.code = jl.account_code
        WHERE jl.journal_id = ?
        ORDER BY line_no, id
    """, (journal_id,)).fetchall()


def get_entry_totals(conn, journal_id: int):
    row = conn.execute("""
        SELECT
            COALESCE(SUM(debit), 0) AS total_debit,
            COALESCE(SUM(credit), 0) AS total_credit
        FROM journal_lines
        WHERE journal_id = ?
    """, (journal_id,)).fetchone()

    total_debit = q2(row["total_debit"] if row else 0)
    total_credit = q2(row["total_credit"] if row else 0)
    balanced = total_debit == total_credit
    return total_debit, total_credit, balanced


def partner_type_options(selected_value="", lang: str = "en"):
    selected = normalize_partner_type(selected_value)
    options = [
        ("", tr(lang, "None", "بدون")),
        ("customer", tr(lang, "Customer", "عميل")),
        ("vendor", tr(lang, "Vendor", "مورد")),
        ("employee", tr(lang, "Employee", "موظف")),
    ]
    html = ""
    for value, label in options:
        sel = "selected" if selected == value else ""
        html += f"<option value='{value}' {sel}>{label}</option>"
    return html


def parse_lines_from_form(form):
    account_codes = form.getlist("account_code")
    line_descriptions = form.getlist("line_description")
    debits = form.getlist("debit")
    credits = form.getlist("credit")
    partner_types = form.getlist("partner_type")
    partner_refs = form.getlist("partner_ref")

    lines = []
    max_len = max(len(account_codes), len(line_descriptions), len(debits), len(credits), len(partner_types), len(partner_refs), 0)
    conn = get_conn()
    try:
        for i in range(max_len):
            account_code = safe(account_codes[i]) if i < len(account_codes) else ""
            line_description = safe(line_descriptions[i]) if i < len(line_descriptions) else ""
            debit = q2(debits[i] if i < len(debits) else "0")
            credit = q2(credits[i] if i < len(credits) else "0")
            partner_type = safe(partner_types[i]) if i < len(partner_types) else ""
            partner_ref = safe(partner_refs[i]) if i < len(partner_refs) else ""

            if account_code == "" and debit == Decimal("0.00") and credit == Decimal("0.00") and line_description == "" and partner_ref == "":
                continue

            required_partner_type = infer_required_partner_type(conn, account_code)
            effective_partner_type = partner_type or required_partner_type
            resolved_type = ""
            resolved_id = None
            if partner_ref:
                resolved_type, resolved_id = resolve_partner_reference(conn, effective_partner_type, partner_ref, partner_ref)

            if required_partner_type:
                if not resolved_id:
                    raise Exception(
                        f"Partner is required for account {account_code}. "
                        f"Use partner_type={required_partner_type} and enter partner code or name."
                    )
                if resolved_type != required_partner_type:
                    raise Exception(
                        f"Account {account_code} requires partner_type={required_partner_type}, "
                        f"but received partner_type={resolved_type or partner_type or '-'}"
                    )

            lines.append({
                "line_no": i + 1,
                "account_code": account_code,
                "line_description": line_description,
                "debit": debit,
                "credit": credit,
                "partner_type": resolved_type,
                "partner_id": resolved_id,
                "partner_ref": partner_ref,
            })
    finally:
        conn.close()

    return lines


def create_journal_draft(conn, entry_date: str, description: str, reference: str, lines: list):
    entry_no = next_entry_no(conn)

    cur = conn.execute("""
        INSERT INTO journal_entries (
            entry_no, entry_date, description, reference, status
        )
        VALUES (?, ?, ?, ?, 'draft')
    """, (
        entry_no,
        safe(entry_date),
        safe(description),
        safe(reference),
    ))
    journal_id = cur.lastrowid

    for line in lines:
        conn.execute("""
            INSERT INTO journal_lines (
                journal_id, line_no, line_description, account_code, debit, credit, partner_type, partner_id
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            journal_id,
            line["line_no"],
            line["line_description"],
            line["account_code"],
            float(line["debit"]),
            float(line["credit"]),
            safe(line.get("partner_type")),
            line.get("partner_id"),
        ))

    return journal_id


def update_journal_draft(conn, journal_id: int, entry_date: str, description: str, reference: str, lines: list):
    entry = get_entry(conn, journal_id)
    if not entry:
        raise Exception("Journal entry not found")
    if safe(entry["status"]).lower() not in ["draft", "pending_final_post"]:
        raise Exception("Only draft or pending-final entries can be edited")

    conn.execute("""
        UPDATE journal_entries
        SET entry_date = ?, description = ?, reference = ?
        WHERE id = ?
    """, (
        safe(entry_date),
        safe(description),
        safe(reference),
        journal_id,
    ))

    conn.execute("DELETE FROM journal_lines WHERE journal_id = ?", (journal_id,))

    for line in lines:
        conn.execute("""
            INSERT INTO journal_lines (
                journal_id, line_no, line_description, account_code, debit, credit, partner_type, partner_id
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            journal_id,
            line["line_no"],
            line["line_description"],
            line["account_code"],
            float(line["debit"]),
            float(line["credit"]),
            safe(line.get("partner_type")),
            line.get("partner_id"),
        ))


def unlink_source_document_after_journal_delete(conn, entry):
    source_type = safe(entry["source_type"]).lower()
    source_id = int(entry["source_id"] or 0) if safe(entry["source_id"]) else 0
    if source_id <= 0:
        return

    if source_type == "cash_voucher" and table_exists(conn, "cash_vouchers"):
        conn.execute(
            "UPDATE cash_vouchers SET journal_id = NULL, status = 'draft' WHERE id = ?",
            (source_id,),
        )
        voucher = conn.execute("SELECT source_type, source_id FROM cash_vouchers WHERE id = ? LIMIT 1", (source_id,)).fetchone()
        if voucher and safe(voucher["source_type"]).lower() == "expense" and safe(voucher["source_id"]):
            conn.execute("UPDATE expenses SET status = 'pending_payment' WHERE id = ?", (int(voucher["source_id"] or 0),))
        return

    table_map = {
        "customer_invoice": ("customer_invoices", "status = 'draft', journal_id = NULL"),
        "customer_bill": ("customer_invoices", "status = 'draft', journal_id = NULL"),
        "vendor_bill": ("vendor_bills", "status = 'draft', journal_id = NULL"),
        "expense": ("expenses", "status = 'draft', journal_id = NULL"),
        "customer_payment": ("customer_payments", "status = 'draft', journal_id = NULL"),
        "vendor_payment": ("vendor_payments", "status = 'draft', journal_id = NULL"),
        "fixed_asset_acquisition": ("fixed_assets", "journal_id = NULL"),
        "fixed_asset_depreciation": ("fixed_asset_depreciations", "journal_id = NULL"),
        "fixed_asset_disposal": ("fixed_asset_disposals", "journal_id = NULL"),
        "petty_cash_custody": ("employee_custodies", "status = 'draft', journal_id = NULL"),
        "petty_cash_return": ("employee_custody_returns", "status = 'draft', journal_id = NULL"),
        "petty_cash_transfer": ("employee_custody_transfers", "status = 'draft', journal_id = NULL"),
    }
    table_info = table_map.get(source_type)
    if not table_info:
        return
    table_name, set_clause = table_info
    if not table_exists(conn, table_name):
        return
    conn.execute(f"UPDATE {table_name} SET {set_clause} WHERE id = ?", (source_id,))


def delete_unfinalized_journal(conn, journal_id: int):
    entry = get_entry(conn, journal_id)
    if not entry:
        raise Exception("Journal entry not found")
    if safe(entry["status"]).lower() not in ("draft", "pending_final_post"):
        raise Exception("Only draft or pending-final entries can be deleted")

    unlink_source_document_after_journal_delete(conn, entry)
    conn.execute("DELETE FROM journal_lines WHERE journal_id = ?", (journal_id,))
    conn.execute("DELETE FROM journal_entries WHERE id = ?", (journal_id,))


def update_journal_posted_admin(conn, journal_id: int, entry_date: str, description: str, reference: str, lines: list):
    entry = get_entry(conn, journal_id)
    if not entry:
        raise Exception("Journal entry not found")
    if safe(entry["status"]).lower() != "posted":
        raise Exception("Only posted entries can be edited from this action")

    for line in lines:
        if not safe(line["account_code"]):
            raise Exception("Account code is required on all lines")
        validate_account_for_posting(conn, safe(line["account_code"]))

    total_debit = sum((q2(line["debit"]) for line in lines), Decimal("0.00"))
    total_credit = sum((q2(line["credit"]) for line in lines), Decimal("0.00"))
    if q2(total_debit) <= Decimal("0.00") and q2(total_credit) <= Decimal("0.00"):
        raise Exception("Journal entry total cannot be zero")
    if q2(total_debit) != q2(total_credit):
        raise Exception(f"Journal not balanced: DR={q2(total_debit)}, CR={q2(total_credit)}")

    conn.execute("""
        UPDATE journal_entries
        SET entry_date = ?, description = ?, reference = ?
        WHERE id = ?
    """, (
        safe(entry_date),
        safe(description),
        safe(reference),
        journal_id,
    ))

    conn.execute("DELETE FROM journal_lines WHERE journal_id = ?", (journal_id,))

    for line in lines:
        conn.execute("""
            INSERT INTO journal_lines (
                journal_id, line_no, line_description, account_code, debit, credit, partner_type, partner_id
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            journal_id,
            line["line_no"],
            line["line_description"],
            line["account_code"],
            float(line["debit"]),
            float(line["credit"]),
            safe(line.get("partner_type")),
            line.get("partner_id"),
        ))


def sync_customer_invoice_from_journal(conn, journal_id: int, invoice_id: int):
    invoice = conn.execute("""
        SELECT *
        FROM customer_invoices
        WHERE id = ?
        LIMIT 1
    """, (invoice_id,)).fetchone()
    if not invoice:
        return

    entry = get_entry(conn, journal_id)
    lines = get_entry_lines(conn, journal_id)
    if not entry or not lines:
        return

    customer = conn.execute("""
        SELECT *
        FROM partners
        WHERE id = ?
        LIMIT 1
    """, (invoice["customer_id"],)).fetchone()

    customer_account = safe(customer["account_code"]) if customer else ""
    if not customer_account:
        customer_account = safe(get_setting_value("customer_control_account", "112100", conn=conn))
    output_vat_account = safe(get_setting_value("output_vat_account", "212100", conn=conn))
    wht_receivable_account = safe(get_setting_value("wht_receivable_account", "114200", conn=conn))

    invoice_lines = []
    subtotal = Decimal("0.00")
    vat_amount = Decimal("0.00")
    wht_amount = Decimal("0.00")
    other_deductions = Decimal("0.00")
    net_amount = Decimal("0.00")
    debit_candidates = []

    for idx, line in enumerate(lines, start=1):
        account_code = safe(line["account_code"])
        debit = dec2(line["debit"])
        credit = dec2(line["credit"])
        line_desc = safe(line["line_description"]) or f"Invoice line {idx}"

        if account_code == output_vat_account and credit > Decimal("0.00"):
            vat_amount += credit
            continue
        if account_code == wht_receivable_account and debit > Decimal("0.00"):
            wht_amount += debit
            continue
        if account_code == customer_account and debit > Decimal("0.00"):
            net_amount += debit
            continue

        if debit > Decimal("0.00"):
            debit_candidates.append(debit)
            other_deductions += debit
            continue

        if credit > Decimal("0.00"):
            subtotal += credit
            invoice_lines.append({
                "line_no": len(invoice_lines) + 1,
                "item_description": line_desc,
                "account_code": account_code,
                "qty": Decimal("1.0000000"),
                "unit_price": credit.quantize(Decimal("1.0000000"), rounding=ROUND_HALF_UP),
                "line_amount": credit.quantize(Decimal("1.0000000"), rounding=ROUND_HALF_UP),
            })

    if not invoice_lines:
        raise Exception("Posted journal cannot sync to customer invoice because no invoice lines were found")

    wht_amount += other_deductions
    total_amount = subtotal + vat_amount
    if net_amount <= Decimal("0.00"):
        fallback_receivable = max(debit_candidates) if debit_candidates else Decimal("0.00")
        if fallback_receivable > Decimal("0.00"):
            net_amount = fallback_receivable
            wht_amount = (total_amount - net_amount).quantize(Decimal("1.00"), rounding=ROUND_HALF_UP)
            if wht_amount < Decimal("0.00"):
                wht_amount = Decimal("0.00")
        else:
            net_amount = total_amount - wht_amount

    vat_rate = Decimal("0.00") if subtotal <= Decimal("0.00") else ((vat_amount / subtotal) * Decimal("100")).quantize(Decimal("1.00"), rounding=ROUND_HALF_UP)
    wht_rate = Decimal("0.00") if subtotal <= Decimal("0.00") else ((wht_amount / subtotal) * Decimal("100")).quantize(Decimal("1.00"), rounding=ROUND_HALF_UP)

    fk_col = "invoice_id" if "invoice_id" in get_columns(conn, "customer_invoice_lines") else "bill_id"
    conn.execute(f"DELETE FROM customer_invoice_lines WHERE {fk_col} = ?", (invoice_id,))

    for line in invoice_lines:
        conn.execute(f"""
            INSERT INTO customer_invoice_lines (
                {fk_col}, line_no, item_description, account_code, qty, unit_price, line_amount
            )
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (
            invoice_id,
            line["line_no"],
            line["item_description"],
            line["account_code"],
            float(line["qty"]),
            float(line["unit_price"]),
            float(line["line_amount"]),
        ))

    new_due_date = calc_due_date(entry["entry_date"], invoice["payment_term_days"] or 0)
    conn.execute("""
        UPDATE customer_invoices
        SET invoice_no = ?,
            invoice_date = ?,
            due_date = ?,
            description = ?,
            subtotal = ?,
            vat_rate = ?,
            vat_amount = ?,
            wht_rate = ?,
            wht_amount = ?,
            total_amount = ?,
            net_amount = ?,
            status = 'posted'
        WHERE id = ?
    """, (
        safe(entry["reference"]) or safe(invoice["invoice_no"]),
        safe(entry["entry_date"]) or safe(invoice["invoice_date"]),
        new_due_date,
        safe(entry["description"]) or safe(invoice["description"]),
        float(subtotal),
        float(vat_rate),
        float(vat_amount),
        float(wht_rate),
        float(wht_amount),
        float(total_amount),
        float(net_amount),
        invoice_id,
    ))

    refresh_customer_invoice_payment_status(conn, invoice_id)


def sync_vendor_bill_from_journal(conn, journal_id: int, bill_id: int):
    bill = conn.execute("""
        SELECT *
        FROM vendor_bills
        WHERE id = ?
        LIMIT 1
    """, (bill_id,)).fetchone()
    if not bill:
        return

    entry = get_entry(conn, journal_id)
    lines = get_entry_lines(conn, journal_id)
    if not entry or not lines:
        return

    vendor = conn.execute("""
        SELECT *
        FROM partners
        WHERE id = ?
        LIMIT 1
    """, (bill["vendor_id"],)).fetchone()

    vendor_account = safe(vendor["account_code"]) if vendor else ""
    if not vendor_account:
        vendor_account = safe(get_setting_value("vendor_control_account", "211100", conn=conn))
    input_vat_account = safe(get_setting_value("input_vat_account", "114100", conn=conn))
    wht_payable_account = safe(get_setting_value("wht_payable_account", "214200", conn=conn))

    bill_lines = []
    subtotal = Decimal("0.00")
    vat_amount = Decimal("0.00")
    wht_amount = Decimal("0.00")
    other_deductions = Decimal("0.00")
    net_amount = Decimal("0.00")
    credit_candidates = []

    for idx, line in enumerate(lines, start=1):
        account_code = safe(line["account_code"])
        debit = dec2(line["debit"])
        credit = dec2(line["credit"])
        line_desc = safe(line["line_description"]) or f"Bill line {idx}"

        if account_code == input_vat_account and debit > Decimal("0.00"):
            vat_amount += debit
            continue
        if account_code == wht_payable_account and credit > Decimal("0.00"):
            wht_amount += credit
            continue
        if account_code == vendor_account and credit > Decimal("0.00"):
            net_amount += credit
            continue

        if credit > Decimal("0.00"):
            credit_candidates.append(credit)
            other_deductions += credit
            continue

        if debit > Decimal("0.00"):
            subtotal += debit
            bill_lines.append({
                "line_no": len(bill_lines) + 1,
                "item_description": line_desc,
                "account_code": account_code,
                "qty": Decimal("1.0000000"),
                "unit_price": debit.quantize(Decimal("1.0000000"), rounding=ROUND_HALF_UP),
                "line_amount": debit.quantize(Decimal("1.0000000"), rounding=ROUND_HALF_UP),
            })

    if not bill_lines:
        raise Exception("Posted journal cannot sync to vendor bill because no bill lines were found")

    wht_amount += other_deductions
    total_amount = subtotal + vat_amount
    if net_amount <= Decimal("0.00"):
        fallback_payable = max(credit_candidates) if credit_candidates else Decimal("0.00")
        if fallback_payable > Decimal("0.00"):
            net_amount = fallback_payable
            wht_amount = (total_amount - net_amount).quantize(Decimal("1.00"), rounding=ROUND_HALF_UP)
            if wht_amount < Decimal("0.00"):
                wht_amount = Decimal("0.00")
        else:
            net_amount = total_amount - wht_amount

    vat_rate = Decimal("0.00") if subtotal <= Decimal("0.00") else ((vat_amount / subtotal) * Decimal("100")).quantize(Decimal("1.00"), rounding=ROUND_HALF_UP)
    wht_rate = Decimal("0.00") if subtotal <= Decimal("0.00") else ((wht_amount / subtotal) * Decimal("100")).quantize(Decimal("1.00"), rounding=ROUND_HALF_UP)

    ensure_column(conn, "vendor_bill_lines", "asset_category_id", "ALTER TABLE vendor_bill_lines ADD COLUMN asset_category_id INTEGER")
    ensure_column(conn, "vendor_bill_lines", "fixed_asset_id", "ALTER TABLE vendor_bill_lines ADD COLUMN fixed_asset_id INTEGER")

    old_line_assets = {}
    if table_exists(conn, "vendor_bill_lines"):
        for old in conn.execute("""
            SELECT line_no, asset_category_id, fixed_asset_id
            FROM vendor_bill_lines
            WHERE bill_id = ?
        """, (bill_id,)).fetchall():
            old_line_assets[int(old["line_no"] or 0)] = {
                "asset_category_id": old["asset_category_id"],
                "fixed_asset_id": old["fixed_asset_id"],
            }

    for line in bill_lines:
        meta = old_line_assets.get(int(line["line_no"] or 0), {})
        line["asset_category_id"] = meta.get("asset_category_id")
        line["fixed_asset_id"] = meta.get("fixed_asset_id")

    conn.execute("DELETE FROM vendor_bill_lines WHERE bill_id = ?", (bill_id,))

    for line in bill_lines:
        cur = conn.execute("""
            INSERT INTO vendor_bill_lines (
                bill_id, line_no, item_description, account_code,
                asset_category_id, fixed_asset_id, qty, unit_price, line_amount
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            bill_id,
            line["line_no"],
            line["item_description"],
            line["account_code"],
            line.get("asset_category_id"),
            line.get("fixed_asset_id"),
            float(line["qty"]),
            float(line["unit_price"]),
            float(line["line_amount"]),
        ))
        if safe(line.get("fixed_asset_id")) and table_exists(conn, "fixed_assets"):
            conn.execute("""
                UPDATE fixed_assets
                SET source_vendor_bill_line_id = ?,
                    acquisition_journal_id = ?
                WHERE id = ?
            """, (cur.lastrowid, journal_id, line.get("fixed_asset_id")))

    new_due_date = calc_due_date(entry["entry_date"], bill["payment_term_days"] or 0)
    conn.execute("""
        UPDATE vendor_bills
        SET bill_no = ?,
            bill_date = ?,
            due_date = ?,
            description = ?,
            subtotal = ?,
            vat_rate = ?,
            vat_amount = ?,
            wht_rate = ?,
            wht_amount = ?,
            total_amount = ?,
            net_amount = ?,
            status = 'posted'
        WHERE id = ?
    """, (
        safe(entry["reference"]) or safe(bill["bill_no"]),
        safe(entry["entry_date"]) or safe(bill["bill_date"]),
        new_due_date,
        safe(entry["description"]) or safe(bill["description"]),
        float(subtotal),
        float(vat_rate),
        float(vat_amount),
        float(wht_rate),
        float(wht_amount),
        float(total_amount),
        float(net_amount),
        bill_id,
    ))

    refresh_vendor_bill_payment_status(conn, bill_id)
    sync_vendor_bill_assets_after_final(conn, bill_id, journal_id)


def next_asset_code_for_conn(conn):
    row = conn.execute("""
        SELECT code
        FROM fixed_assets
        WHERE COALESCE(code, '') <> ''
        ORDER BY id DESC
        LIMIT 1
    """).fetchone()
    if not row or not safe(row["code"]):
        return "FA-00001"
    last = safe(row["code"])
    prefix = "FA"
    number = 0
    if "-" in last:
        prefix, raw = last.rsplit("-", 1)
        try:
            number = int(raw)
        except Exception:
            number = 0
    return f"{prefix}-{number + 1:05d}"


def sync_vendor_bill_assets_after_final(conn, bill_id: int, journal_id: int):
    if not table_exists(conn, "fixed_assets") or not table_exists(conn, "asset_categories"):
        return

    ensure_column(conn, "fixed_assets", "source_vendor_bill_id", "ALTER TABLE fixed_assets ADD COLUMN source_vendor_bill_id INTEGER")
    ensure_column(conn, "fixed_assets", "source_vendor_bill_line_id", "ALTER TABLE fixed_assets ADD COLUMN source_vendor_bill_line_id INTEGER")

    bill = conn.execute("SELECT * FROM vendor_bills WHERE id = ? LIMIT 1", (bill_id,)).fetchone()
    if not bill:
        return

    lines = conn.execute("""
        SELECT *
        FROM vendor_bill_lines
        WHERE bill_id = ?
          AND COALESCE(asset_category_id, 0) > 0
        ORDER BY line_no, id
    """, (bill_id,)).fetchall()

    for line in lines:
        amount = q2(line["line_amount"])
        if amount <= Decimal("0.00"):
            continue
        category = conn.execute("SELECT * FROM asset_categories WHERE id = ? LIMIT 1", (line["asset_category_id"],)).fetchone()
        if not category:
            continue
        asset_id = int(line["fixed_asset_id"] or 0) if safe(line["fixed_asset_id"]) else 0
        asset_name = safe(line["item_description"]) or f"{safe(bill['bill_no'])} line {safe(line['line_no'])}"
        asset_account = safe(line["account_code"]) or safe(category["asset_account_code"])
        offset_account = safe(get_setting_value("vendor_control_account", "211100", conn=conn))
        notes = f"Created from vendor bill {safe(bill['bill_no'])}"

        if asset_id > 0:
            conn.execute("""
                UPDATE fixed_assets
                SET name = ?,
                    category_id = ?,
                    purchase_date = ?,
                    in_service_date = COALESCE(NULLIF(in_service_date, ''), ?),
                    cost = ?,
                    status = 'running',
                    acquisition_account_code = ?,
                    offset_account_code = ?,
                    acquisition_journal_id = ?,
                    source_vendor_bill_id = ?,
                    source_vendor_bill_line_id = ?,
                    notes = ?
                WHERE id = ?
            """, (
                asset_name,
                line["asset_category_id"],
                safe(bill["bill_date"]),
                safe(bill["bill_date"]),
                float(amount),
                asset_account,
                offset_account,
                journal_id,
                bill_id,
                line["id"],
                notes,
                asset_id,
            ))
        else:
            cur = conn.execute("""
                INSERT INTO fixed_assets (
                    code, name, category_id, purchase_date, in_service_date,
                    cost, salvage_value, status, acquisition_account_code,
                    offset_account_code, acquisition_journal_id,
                    source_vendor_bill_id, source_vendor_bill_line_id, notes
                )
                VALUES (?, ?, ?, ?, ?, ?, 0, 'running', ?, ?, ?, ?, ?, ?)
            """, (
                next_asset_code_for_conn(conn),
                asset_name,
                line["asset_category_id"],
                safe(bill["bill_date"]),
                safe(bill["bill_date"]),
                float(amount),
                asset_account,
                offset_account,
                journal_id,
                bill_id,
                line["id"],
                notes,
            ))
            conn.execute("UPDATE vendor_bill_lines SET fixed_asset_id = ? WHERE id = ?", (cur.lastrowid, line["id"]))


def remove_vendor_bill_assets_for_reversal(conn, bill_id: int):
    if not bill_id or not table_exists(conn, "fixed_assets"):
        return

    ensure_column(conn, "fixed_assets", "source_vendor_bill_id", "ALTER TABLE fixed_assets ADD COLUMN source_vendor_bill_id INTEGER")
    ensure_column(conn, "vendor_bill_lines", "fixed_asset_id", "ALTER TABLE vendor_bill_lines ADD COLUMN fixed_asset_id INTEGER")

    conn.execute("""
        DELETE FROM fixed_assets
        WHERE source_vendor_bill_id = ?
          AND id NOT IN (
              SELECT COALESCE(asset_id, 0)
              FROM asset_depreciation_moves
          )
          AND id NOT IN (
              SELECT COALESCE(asset_id, 0)
              FROM asset_disposals
          )
    """, (bill_id,))

    conn.execute("""
        UPDATE fixed_assets
        SET status = 'reversed'
        WHERE source_vendor_bill_id = ?
    """, (bill_id,))

    conn.execute("""
        UPDATE vendor_bill_lines
        SET fixed_asset_id = NULL
        WHERE bill_id = ?
          AND fixed_asset_id NOT IN (
              SELECT id
              FROM fixed_assets
              WHERE source_vendor_bill_id = ?
          )
    """, (bill_id, bill_id))


def sync_source_document_from_journal(conn, journal_id: int):
    entry = get_entry(conn, journal_id)
    if not entry:
        return

    source_type = safe(entry["source_type"]).lower()
    source_id = int(entry["source_id"] or 0) if safe(entry["source_id"]) else 0
    if source_id <= 0:
        return

    if source_type in ("customer_invoice", "customer_bill"):
        sync_customer_invoice_from_journal(conn, journal_id, source_id)
    elif source_type == "vendor_bill":
        sync_vendor_bill_from_journal(conn, journal_id, source_id)


def post_journal_entry(conn, journal_id: int):
    entry = get_entry(conn, journal_id)
    if not entry:
        raise Exception("Journal entry not found")

    status = safe(entry["status"]).lower()
    if status not in ("draft", "pending_final_post"):
        raise Exception("Only draft or pending-final entries can be posted")

    lines = get_entry_lines(conn, journal_id)
    if not lines:
        raise Exception("Journal entry has no lines")

    for line in lines:
        if not safe(line["account_code"]):
            raise Exception("Account code is required on all lines")
        validate_account_for_posting(conn, safe(line["account_code"]))

    total_debit, total_credit, balanced = get_entry_totals(conn, journal_id)
    if total_debit <= Decimal("0.00") and total_credit <= Decimal("0.00"):
        raise Exception("Journal entry total cannot be zero")
    if not balanced:
        raise Exception(f"Journal not balanced: DR={total_debit}, CR={total_credit}")

    conn.execute("""
        UPDATE journal_entries
        SET status = 'posted'
        WHERE id = ?
    """, (journal_id,))
    sync_source_document_from_journal(conn, journal_id)


def reverse_journal_entry(conn, journal_id: int):
    entry = get_entry(conn, journal_id)
    if not entry:
        raise Exception("Journal entry not found")

    status = safe(entry["status"]).lower()
    if status != "posted":
        raise Exception("Only posted entries can be reversed")
    if entry["reversed_by_journal_id"]:
        raise Exception("Journal entry already reversed")

    original_lines = get_entry_lines(conn, journal_id)
    if not original_lines:
        raise Exception("Original journal has no lines")

    reverse_entry_no = next_entry_no(conn)

    cur = conn.execute("""
        INSERT INTO journal_entries (
            entry_no, entry_date, description, reference, status,
            source_type, source_id
        )
        VALUES (?, ?, ?, ?, 'posted', 'journal_reverse', ?)
    """, (
        reverse_entry_no,
        entry["entry_date"],
        f"Reversal of {entry['entry_no']} - {safe(entry['description'])}",
        f"REV-{safe(entry['entry_no'])}",
        journal_id,
    ))
    reverse_journal_id = cur.lastrowid

    for idx, line in enumerate(original_lines, start=1):
        conn.execute("""
            INSERT INTO journal_lines (
                journal_id, line_no, line_description, account_code,
                debit, credit, partner_type, partner_id, cost_center_id
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            reverse_journal_id,
            idx,
            f"Reverse - {safe(line['line_description'])}",
            safe(line["account_code"]),
            float(q2(line["credit"])),
            float(q2(line["debit"])),
            safe(line["partner_type"]),
            line["partner_id"],
            line["cost_center_id"],
        ))

    conn.execute("""
        UPDATE journal_entries
        SET reversed_by_journal_id = ?,
            reversed_by_id = ?
        WHERE id = ?
    """, (reverse_journal_id, reverse_journal_id, journal_id))

    conn.execute("""
        UPDATE journal_entries
        SET reversed_from_id = ?
        WHERE id = ?
    """, (journal_id, reverse_journal_id))

    if safe(entry["source_type"]).lower() == "vendor_bill" and safe(entry["source_id"]):
        bill_id = int(entry["source_id"])
        conn.execute("""
            UPDATE vendor_bills
            SET status = 'reversed',
                reversed_journal_id = ?,
                payment_status = 'cancelled'
            WHERE id = ?
        """, (reverse_journal_id, bill_id))
        remove_vendor_bill_assets_for_reversal(conn, bill_id)

    return reverse_journal_id


# =========================================================
# IMPORT / EXPORT HELPERS
# =========================================================
def import_journal_from_workbook(file_bytes: bytes):
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise Exception("Excel file is empty")

    header = [safe(x).lower() for x in rows[0]]
    expected = [
        "entry_date",
        "description",
        "reference",
        "line_no",
        "account_code",
        "line_description",
        "debit",
        "credit",
        "partner_type",
        "partner_code",
        "partner_name",
    ]
    if header != expected:
        raise Exception("Template columns are invalid")

    entry_date = ""
    description = ""
    reference = ""
    lines = []
    seen_dates = set()
    seen_descriptions = set()
    seen_references = set()

    for row in rows[1:]:
        if not row or not any(safe(cell) for cell in row):
            continue

        row_entry_date = safe(row[0])
        row_description = safe(row[1])
        row_reference = safe(row[2])
        line_no = safe(row[3])
        account_code = safe(row[4])
        line_description = safe(row[5])
        debit = q2(row[6] or 0)
        credit = q2(row[7] or 0)
        partner_type_raw = safe(row[8] if len(row) > 8 else "")
        partner_code = safe(row[9] if len(row) > 9 else "")
        partner_name = safe(row[10] if len(row) > 10 else "")

        if row_entry_date:
            seen_dates.add(row_entry_date)
            if not entry_date:
                entry_date = row_entry_date
        if row_description:
            seen_descriptions.add(row_description)
            if not description:
                description = row_description
        if row_reference:
            seen_references.add(row_reference)
            if not reference:
                reference = row_reference

        lines.append({
            "line_no": int(line_no) if safe(line_no) else len(lines) + 1,
            "account_code": account_code,
            "line_description": line_description,
            "debit": debit,
            "credit": credit,
            "partner_type": partner_type_raw,
            "partner_code": partner_code,
            "partner_name": partner_name,
            "partner_id": None,
        })

    if len(seen_dates) > 1:
        raise Exception("Import file must contain one entry date only.")
    if len(seen_descriptions) > 1:
        raise Exception("Import file must contain one description only.")
    if len(seen_references) > 1:
        raise Exception("Import file must contain one reference only.")

    created_ids = []
    conn = get_conn()
    try:
        if not description or not lines:
            return created_ids

        prepared_lines = []
        for line in lines:
            partner_type = safe(line.get("partner_type"))
            partner_code = safe(line.get("partner_code"))
            partner_name = safe(line.get("partner_name"))
            required_partner_type = infer_required_partner_type(conn, line["account_code"])
            effective_partner_type = partner_type or required_partner_type
            resolved_type = ""
            resolved_id = None
            if effective_partner_type or partner_code or partner_name:
                resolved_type, resolved_id = resolve_partner_reference(conn, effective_partner_type, partner_code, partner_name)
            if required_partner_type:
                if not resolved_id:
                    raise Exception(
                        f"Partner is required for account {safe(line['account_code'])}. "
                        f"Use partner_type={required_partner_type} with partner_code or partner_name."
                    )
                if resolved_type != required_partner_type:
                    raise Exception(
                        f"Account {safe(line['account_code'])} requires partner_type={required_partner_type}, "
                        f"but received partner_type={resolved_type or safe(partner_type) or '-'}"
                    )
            prepared_lines.append({
                "line_no": line["line_no"],
                "account_code": line["account_code"],
                "line_description": line["line_description"],
                "debit": line["debit"],
                "credit": line["credit"],
                "partner_type": resolved_type,
                "partner_id": resolved_id,
            })

        journal_id = create_journal_draft(conn, entry_date, description, reference, prepared_lines)
        if journal_id:
            created_ids.append(journal_id)
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    return created_ids


def build_journal_export_workbook(status: str = "all"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Journal Export"
    ws.append([
        "entry_no",
        "entry_date",
        "description",
        "reference",
        "status",
        "line_no",
        "account_code",
        "line_description",
        "debit",
        "credit",
        "partner_type",
        "partner_code",
        "partner_name",
    ])

    conn = get_conn()
    try:
        sql = """
            SELECT
                j.entry_no,
                j.entry_date,
                j.description,
                j.reference,
                j.status,
                l.line_no,
                l.account_code,
                l.line_description,
                l.debit,
                l.credit,
                l.partner_type,
                l.partner_id
            FROM journal_entries j
            LEFT JOIN journal_lines l ON l.journal_id = j.id
        """
        params = []
        if safe(status).lower() not in ["", "all"]:
            sql += " WHERE LOWER(COALESCE(j.status,'')) = ?"
            params.append(safe(status).lower())
        sql += " ORDER BY j.id DESC, COALESCE(l.line_no, 0), l.id"

        rows = conn.execute(sql, params).fetchall()
        for row in rows:
            partner_code, partner_name = get_partner_display(conn, row["partner_type"], row["partner_id"])
            ws.append([
                safe(row["entry_no"]),
                safe(row["entry_date"]),
                safe(row["description"]),
                safe(row["reference"]),
                safe(row["status"]),
                row["line_no"] or "",
                safe(row["account_code"]),
                safe(row["line_description"]),
                float(q2(row["debit"] or 0)),
                float(q2(row["credit"] or 0)),
                safe(row["partner_type"]),
                partner_code,
                partner_name,
            ])
    finally:
        conn.close()

    return wb


def journal_category_label(source_type: str, lang: str = "en"):
    key = safe(source_type).lower()
    labels = {
        "": tr(lang, "Manual", "يدوي"),
        "journal_reverse": tr(lang, "Reversal", "عكس قيد"),
        "customer_invoice": tr(lang, "Customer Invoice", "فاتورة عميل"),
        "vendor_bill": tr(lang, "Vendor Bill", "فاتورة مورد"),
        "customer_payment": tr(lang, "Customer Payment", "تحصيل عميل"),
        "vendor_payment": tr(lang, "Vendor Payment", "سداد مورد"),
        "fixed_asset": tr(lang, "Fixed Asset", "أصل ثابت"),
        "asset_depreciation": tr(lang, "Depreciation", "إهلاك"),
        "asset_disposal": tr(lang, "Asset Disposal", "استبعاد أصل"),
    }
    return labels.get(key, key.replace("_", " ").title() if key else tr(lang, "Manual", "يدوي"))


def journal_category_options(selected=None, lang: str = "en"):
    conn = get_conn()
    rows = conn.execute("""
        SELECT DISTINCT COALESCE(source_type, '') AS source_type
        FROM journal_entries
        ORDER BY COALESCE(source_type, '')
    """).fetchall()
    conn.close()

    html = f"<option value=''>{tr(lang, 'All Categories', 'كل الفئات')}</option>"
    found = set()
    for row in rows:
        key = safe(row["source_type"])
        if key in found:
            continue
        found.add(key)
        sel = "selected" if safe(selected).lower() == key.lower() else ""
        html += f"<option value='{key}' {sel}>{journal_category_label(key, lang)}</option>"
    return html


# =========================================================
# HTML HELPERS
# =========================================================
def render_line_rows(lines=None, readonly=False, lang: str = "en"):
    lines = lines or [
        {"line_description": "", "account_code": "", "debit": Decimal("0.00"), "credit": Decimal("0.00"), "partner_type": "", "partner_ref": ""}
    ]
    html = ""
    conn = get_conn()
    try:
        for idx, line in enumerate(lines, start=1):
            disabled = "disabled" if readonly else ""
            readonly_attr = "readonly" if readonly else ""
            account_code = safe(row_value(line, "account_code", ""))
            line_description = safe(row_value(line, "line_description", ""))
            debit = row_value(line, "debit", "0")
            credit = row_value(line, "credit", "0")
            partner_type = safe(row_value(line, "partner_type", ""))
            partner_ref = safe(row_value(line, "partner_ref", ""))
            if not partner_ref and safe(row_value(line, "partner_id", "")):
                partner_code, partner_name = get_partner_display(conn, partner_type, row_value(line, "partner_id", ""))
                partner_ref = partner_code or partner_name
            hidden = f"<input type='hidden' name='account_code' value='{account_code}'>" if readonly else ""
            hidden_partner = f"<input type='hidden' name='partner_type' value='{partner_type}'><input type='hidden' name='partner_ref' value='{partner_ref}'>" if readonly else ""
            remove_btn = "" if readonly else f"<button type='button' class='btn red' onclick='removeLine(this)'>{tr(lang, 'Remove', '???')}</button>"
            html += f"""
            <tr>
                <td>{idx}</td>
                <td><input type="text" name="line_description" value="{line_description}" {readonly_attr}></td>
                <td>
                    <select name="account_code" {disabled}>
                        {account_options(account_code)}
                    </select>
                    {hidden}
                </td>
                <td>
                    <select name="partner_type" {disabled}>
                        {partner_type_options(partner_type, lang)}
                    </select>
                    {hidden_partner if readonly else ''}
                </td>
            <td><input type="text" name="partner_ref" value="{partner_ref}" placeholder="{tr(lang, 'Code or name', 'كود أو اسم')}" {readonly_attr}></td>
                <td><input type="text" name="debit" value="{money(debit, 2)}" class="line-debit" {readonly_attr}></td>
                <td><input type="text" name="credit" value="{money(credit, 2)}" class="line-credit" {readonly_attr}></td>
                <td>{remove_btn}</td>
            </tr>
            """
    finally:
        conn.close()
    return html


def journal_form_html(action_url, values=None, lines=None, error_message="", lang: str = "en"):
    values = values or {}
    lines = lines or None
    entry_date = safe(values.get("entry_date")) or datetime.today().date().isoformat()
    description = safe(values.get("description"))
    reference = safe(values.get("reference"))

    msg = f'<div class="msg error">{error_message}</div>' if error_message else ""

    content = f"""
    {msg}
    <div class="card">
        <div class="toolbar">
            <h3 class="sub-title">{tr(lang, "Journal Entry", "قيد يومية")}</h3>
            <a href="{with_lang('/ui/accounting/journal', lang)}" class="btn gray">{tr(lang, "Back", "رجوع")}</a>
        </div>

        <form method="post" action="{action_url}">
            <div class="form-grid" style="margin-top:14px;">
                <div class="form-group">
                    <label>{tr(lang, "Entry Date", "تاريخ القيد")}</label>
                    <input type="date" name="entry_date" value="{entry_date}" required>
                </div>
                <div class="form-group" style="grid-column: span 2;">
                    <label>{tr(lang, "Description", "البيان")}</label>
                    <input name="description" value="{description}" required>
                </div>
                <div class="form-group">
                    <label>{tr(lang, "Reference", "المرجع")}</label>
                    <input name="reference" value="{reference}">
                </div>
            </div>

            <div class="card" style="margin-top:14px;">
                <div class="toolbar">
                    <h3 class="sub-title">{tr(lang, "Journal Lines", "سطور القيد")}</h3>
                    <button type="button" class="btn blue" onclick="addLine()">+ {tr(lang, "Add Line", "إضافة سطر")}</button>
                </div>
                <div class="table-wrap" style="margin-top:12px;">
                    <table>
                        <tr>
                            <th>#</th>
                            <th>{tr(lang, "Description", "البيان")}</th>
                            <th>{tr(lang, "Account", "الحساب")}</th>
                            <th>{tr(lang, "Partner Type", "نوع الطرف")}</th>
            <th>{tr(lang, "Partner / Code", "الطرف / الكود")}</th>
            <th>{tr(lang, "Debit", "مدين")}</th>
            <th>{tr(lang, "Credit", "دائن")}</th>
            <th>{tr(lang, "Action", "الإجراء")}</th>
                        </tr>
                        <tbody id="linesBody">
                            {render_line_rows(lines, lang=lang)}
                        </tbody>
                    </table>
                </div>
            </div>

            <div class="form-actions">
                <button class="btn green" type="submit">{tr(lang, "Save Draft", "حفظ كمسودة")}</button>
            </div>
        </form>
    </div>

    <script>
    function addLine() {{
        const body = document.getElementById("linesBody");
        const rowCount = body.querySelectorAll("tr").length + 1;
        const tr = document.createElement("tr");
        tr.innerHTML = `
            <td>${{rowCount}}</td>
            <td><input type="text" name="line_description"></td>
            <td>
                <select name="account_code">
                    {account_options("")}
                </select>
            </td>
            <td>
                <select name="partner_type">
                    {partner_type_options("", lang)}
                </select>
            </td>
            <td><input type="text" name="partner_ref" placeholder="{tr(lang, "Code or name", "كود أو اسم")}"></td>
            <td><input type="text" name="debit" value="0.00"></td>
            <td><input type="text" name="credit" value="0.00"></td>
            <td><button type="button" class="btn red" onclick="removeLine(this)">{tr(lang, "Remove", "حذف")}</button></td>
        `;
        body.appendChild(tr);
        if (window.makeSearchable) {{
            const select = tr.querySelector("select[name='account_code']");
            if (select) window.makeSearchable(select);
        }}
    }}

    function removeLine(btn) {{
        const tr = btn.closest("tr");
        tr.remove();
        const rows = document.querySelectorAll("#linesBody tr");
        rows.forEach((row, idx) => {{
            row.children[0].textContent = idx + 1;
        }});
    }}

    document.addEventListener("DOMContentLoaded", function() {{
        if (window.enhanceAllSearchableSelects) {{
            window.enhanceAllSearchableSelects(document);
        }}
    }});
    </script>
    """
    return content


# Override the old view renderer so unfinalized journal entries show a distinct final-post step.
def entry_view_html(entry, lines, total_debit, total_credit, balanced, lang: str = "en", allow_edit: bool = True, allow_post: bool = True):
    conn = get_conn()
    try:
        status = safe(entry["status"]).lower()
        status_badge = {
            "draft": f'<span class="badge orange">{tr(lang, "Draft", "مسودة")}</span>',
            "pending_final_post": f'<span class="badge blue">{tr(lang, "Pending Final Post", "في انتظار الترحيل النهائي")}</span>',
            "posted": f'<span class="badge green">{tr(lang, "Posted", "مرحل")}</span>',
            "reversed": f'<span class="badge red">{tr(lang, "Reversed", "معكوس")}</span>',
        }.get(status, f"<span>{safe(entry['status'])}</span>")

        actions = f'<a href="{with_lang("/ui/accounting/journal", lang)}" class="btn gray">{tr(lang, "Back", "رجوع")}</a>'
        if status == "draft":
            if allow_post:
                actions = f'<form method="post" action="{with_lang(f"/ui/accounting/journal/{entry["id"]}/post", lang)}" style="display:inline;"><button class="btn green" type="submit">{tr(lang, "Post", "ترحيل")}</button></form> ' + actions
            if allow_edit:
                actions = f'<form method="post" action="{with_lang(f"/ui/accounting/journal/{entry["id"]}/delete", lang)}" style="display:inline;" onsubmit="return confirm(\\\'{tr(lang, "Delete this draft journal?", "حذف هذا القيد غير المرحل؟")}\\\');"><button class="btn red" type="submit">{tr(lang, "Delete", "حذف")}</button></form> ' + actions
                actions = f'<a href="{with_lang(f"/ui/accounting/journal/{entry["id"]}/edit", lang)}" class="btn blue">{tr(lang, "Edit", "تعديل")}</a> ' + actions
        elif status == "pending_final_post":
            if allow_post:
                actions = f'<form method="post" action="{with_lang(f"/ui/accounting/journal/{entry["id"]}/post", lang)}" style="display:inline;"><button class="btn green" type="submit">{tr(lang, "Final Post", "ترحيل نهائي")}</button></form> ' + actions
            if allow_edit:
                actions = f'<form method="post" action="{with_lang(f"/ui/accounting/journal/{entry["id"]}/delete", lang)}" style="display:inline;" onsubmit="return confirm(\\\'{tr(lang, "Delete this draft journal?", "حذف هذا القيد غير المرحل؟")}\\\');"><button class="btn red" type="submit">{tr(lang, "Delete", "حذف")}</button></form> ' + actions
                actions = f'<a href="{with_lang(f"/ui/accounting/journal/{entry["id"]}/edit", lang)}" class="btn blue">{tr(lang, "Edit", "تعديل")}</a> ' + actions
        elif status == "posted":
            if allow_post:
                actions = f'<form method="post" action="{with_lang(f"/ui/accounting/journal/{entry["id"]}/reverse", lang)}" style="display:inline;"><button class="btn red" type="submit">{tr(lang, "Reverse", "عكس")}</button></form> ' + actions

        body = ""
        for line in lines:
            account_display = safe(line["account_code"])
            account_name = safe(line["account_name"]) if "account_name" in line.keys() else ""
            if account_name:
                account_display = f"{safe(line['account_code'])} - {safe(line['account_name'])}"
            partner_type = safe(line["partner_type"]) if "partner_type" in line.keys() else ""
            partner_type_display = partner_type_label(partner_type, lang)
            partner_display = get_partner_display_text(conn, partner_type, line["partner_id"] if "partner_id" in line.keys() else None)
            body += f"""
            <tr>
                <td>{line['line_no']}</td>
                <td>{safe(line['line_description'])}</td>
                <td>{account_display}</td>
                <td>{partner_type_display}</td>
                <td>{partner_display}</td>
                <td>{money(line['debit'])}</td>
                <td>{money(line['credit'])}</td>
            </tr>
            """
        if not body:
            body = f"<tr><td colspan='7'>{tr(lang, 'No lines found.', 'لا توجد سطور.')}</td></tr>"

        balance_badge = f'<span class="badge green">{tr(lang, "Balanced", "متوازن")}</span>' if balanced else f'<span class="badge red">{tr(lang, "Unbalanced", "غير متوازن")}</span>'

        return f"""
        <div class="card">
            <div class="toolbar">
                <h3 class="sub-title">{tr(lang, "Journal Entry", "قيد اليومية")} {safe(entry['entry_no'])}</h3>
                <div>{actions}</div>
            </div>

            <div class="form-grid" style="margin-top:14px;">
                <div class="form-group"><label>{tr(lang, "Entry Date", "تاريخ القيد")}</label><input value="{safe(entry['entry_date'])}" readonly></div>
                <div class="form-group"><label>{tr(lang, "Reference", "المرجع")}</label><input value="{safe(entry['reference'])}" readonly></div>
                <div class="form-group"><label>{tr(lang, "Status", "الحالة")}</label><div style="padding-top:10px;">{status_badge}</div></div>
                <div class="form-group"><label>{tr(lang, "Balance Check", "فحص التوازن")}</label><div style="padding-top:10px;">{balance_badge}</div></div>
                <div class="form-group" style="grid-column: span 4;"><label>{tr(lang, "Description", "البيان")}</label><input value="{safe(entry['description'])}" readonly></div>
            </div>
        </div>

        <div class="card">
            <h3 class="sub-title">{tr(lang, "Lines", "السطور")}</h3>
            <div class="table-wrap" style="margin-top:12px;">
                <table>
                    <tr>
                        <th>#</th>
                        <th>{tr(lang, "Description", "البيان")}</th>
                        <th>{tr(lang, "Account", "الحساب")}</th>
                        <th>{tr(lang, "Partner Type", "نوع الطرف")}</th>
                        <th>{tr(lang, "Partner / Code", "الطرف / الكود")}</th>
                        <th>{tr(lang, "Debit", "مدين")}</th>
                        <th>{tr(lang, "Credit", "دائن")}</th>
                    </tr>
                    {body}
                    <tr>
                        <th colspan="5" style="text-align:right;">{tr(lang, "Totals", "الإجماليات")}</th>
                        <th>{money(total_debit)}</th>
                        <th>{money(total_credit)}</th>
                    </tr>
                </table>
            </div>
        </div>
        """
    finally:
        conn.close()


# =========================================================
# ROUTES
# =========================================================
@router.get("/ui/accounting/journal/template")
def journal_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Journal Template"

    ws.append([
        "entry_date",
        "description",
        "reference",
        "line_no",
        "account_code",
        "line_description",
        "debit",
        "credit",
        "partner_type",
        "partner_code",
        "partner_name",
    ])
    ws.append(["2026-01-01", "Opening Entry", "OPEN-001", 1, "1020401", "Customer opening balance", 15000, 0, "customer", "CUST-0004", "شركة اورنج مصر للاتصالات"])
    ws.append(["2026-01-01", "Opening Entry", "OPEN-001", 2, "2010101", "Vendor opening balance", 0, 9000, "vendor", "", "اسم المورد"])
    ws.append(["2026-01-01", "Opening Entry", "OPEN-001", 3, "employee_advance_or_custody_account", "Employee advance opening", 5000, 0, "employee", "EMP-0001", "Ahmed Gamal"])
    ws.append(["2026-01-01", "Opening Entry", "OPEN-001", 4, "310000", "Opening balance offset", 0, 11000, "", "", ""])

    notes = wb.create_sheet("Notes")
    notes.append(["Field", "Guidance"])
    notes.append(["partner_type", "Optional. Use customer / vendor / employee for opening balances by party"])
    notes.append(["partner_code", "Recommended. The import matches code first"])
    notes.append(["partner_name", "Optional fallback if code is blank"])
    notes.append(["customer/vendor/employee", "Use these columns for customer statements, vendor statements, employee custody, and employee advances"])
    notes.append(["import behavior", "The whole Excel sheet is imported as one journal entry with many debit and credit lines."])

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=journal_template.xlsx"},
    )


@router.get("/ui/accounting/journal/export")
def journal_export(status: str = "all"):
    wb = build_journal_export_workbook(status=status)
    out = BytesIO()
    wb.save(out)
    out.seek(0)

    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=journal_export_{safe(status) or 'all'}.xlsx"},
    )


@router.get("/ui/accounting/journal/import", response_class=HTMLResponse)
def journal_import_page(request: Request):
    lang = get_lang(request)
    if not accounting_allowed(request, "create"):
        return permission_denied(lang, "You do not have permission to import journal entries.", "ليس لديك صلاحية استيراد قيود اليومية.")
    content = """
    <div class="card">
        <div class="toolbar">
            <h3 class="sub-title">Import Journal Entries</h3>
            <a class="btn gray" href="/ui/accounting/journal">Back</a>
        </div>

        <form method="post" action="/ui/accounting/journal/import" enctype="multipart/form-data" style="margin-top:14px;">
            <div class="msg info" style="margin-bottom:14px;">This import creates one journal entry only from the full file. Add all debit and credit lines in the same sheet under one entry header.</div>
            <div class="form-group">
                <label>Excel File</label>
                <input type="file" name="file" accept=".xlsx" required>
            </div>

            <div class="form-actions">
                <button class="btn green" type="submit">Import as Draft</button>
                <a class="btn blue" href="/ui/accounting/journal/template">Download Template</a>
            </div>
        </form>
    </div>
    """
    if lang == "ar":
        content = f"""
        <div class="card">
            <div class="toolbar">
                <h3 class="sub-title">استيراد قيود اليومية</h3>
                <a class="btn gray" href="{with_lang('/ui/accounting/journal', lang)}">رجوع</a>
            </div>

            <form method="post" action="{with_lang('/ui/accounting/journal/import', lang)}" enctype="multipart/form-data" style="margin-top:14px;">
                <div class="form-group">
                    <label>ملف إكسل</label>
                    <input type="file" name="file" accept=".xlsx" required>
                </div>

                <div class="form-actions">
                    <button class="btn green" type="submit">استيراد كمسودة</button>
                    <a class="btn blue" href="{with_lang('/ui/accounting/journal/template', lang)}">تحميل النموذج</a>
                </div>
            </form>
        </div>
        """
    return HTMLResponse(render_page(tr(lang, "Import Journal", "استيراد اليومية"), content, lang=lang, current_path=request.url.path))


@router.post("/ui/accounting/journal/import")
async def journal_import(request: Request, file: UploadFile = File(...)):
    lang = get_lang(request)
    if not accounting_allowed(request, "create"):
        return permission_denied(lang, "You do not have permission to import journal entries.", "ليس لديك صلاحية استيراد قيود اليومية.")
    if not safe(file.filename).lower().endswith(".xlsx"):
        return HTMLResponse("Only .xlsx files are allowed.", status_code=400)

    try:
        content = await file.read()
        created_ids = import_journal_from_workbook(content)
    except Exception as e:
        return HTMLResponse(f"Import error: {safe(e)}", status_code=400)

    if not created_ids:
        return HTMLResponse("No journal entries imported.", status_code=400)

    return RedirectResponse("/ui/accounting/journal", status_code=303)


@router.get("/ui/accounting/journal", response_class=HTMLResponse)
def journal_list(
    request: Request,
    status: str = "all",
    search: str = "",
    date_from: str = "",
    date_to: str = "",
    category: str = "",
    account_code: str = "",
    partner_type: str = "",
    partner_id: str = "",
):
    lang = get_lang(request)
    can_create = accounting_allowed(request, "create")
    conn = get_conn()

    sql = """
        SELECT DISTINCT je.*
        FROM journal_entries je
    """
    params = []
    where = []

    if safe(account_code) or safe(partner_type) or safe(partner_id):
        sql += " JOIN journal_lines account_filter_line ON account_filter_line.journal_id = je.id "

    if safe(account_code):
        where.append("COALESCE(account_filter_line.account_code, '') = ?")
        params.append(safe(account_code))

    if safe(partner_type):
        where.append("LOWER(COALESCE(account_filter_line.partner_type, '')) = ?")
        params.append(safe(partner_type).lower())

    try:
        partner_id_int = int(partner_id) if safe(partner_id) else 0
    except Exception:
        partner_id_int = 0

    if partner_id_int:
        where.append("COALESCE(account_filter_line.partner_id, 0) = ?")
        params.append(partner_id_int)

    if status and status.lower() != "all":
        where.append("LOWER(COALESCE(je.status,'')) = ?")
        params.append(status.lower())

    if safe(search):
        where.append("(LOWER(COALESCE(je.entry_no,'')) LIKE ? OR LOWER(COALESCE(je.description,'')) LIKE ? OR LOWER(COALESCE(je.reference,'')) LIKE ?)")
        like_value = f"%{safe(search).lower()}%"
        params.extend([like_value, like_value, like_value])

    if safe(date_from):
        where.append("COALESCE(je.entry_date,'') >= ?")
        params.append(safe(date_from))

    if safe(date_to):
        where.append("COALESCE(je.entry_date,'') <= ?")
        params.append(safe(date_to))

    category_values = [safe(x).lower() for x in request.query_params.getlist("category") if safe(x)]
    if not category_values and safe(category):
        category_values = [part.strip().lower() for part in safe(category).split(",") if part.strip()]

    if category_values:
        include_manual = "manual" in category_values
        named_categories = [x for x in category_values if x != "manual"]
        category_clauses = []
        if include_manual:
            category_clauses.append("COALESCE(je.source_type,'') = ''")
        if named_categories:
            placeholders = ", ".join(["?"] * len(named_categories))
            category_clauses.append(f"LOWER(COALESCE(je.source_type,'')) IN ({placeholders})")
            params.extend(named_categories)
        if category_clauses:
            where.append("(" + " OR ".join(category_clauses) + ")")

    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY je.id DESC"

    rows = conn.execute(sql, params).fetchall()

    body = ""
    total_debit_sum = Decimal("0.00")
    total_credit_sum = Decimal("0.00")
    for row in rows:
        total_debit, total_credit, balanced = get_entry_totals(conn, row["id"])
        total_debit_sum += total_debit
        total_credit_sum += total_credit
        badge = {
            "draft": f'<span class="badge orange">{tr(lang, "Draft", "مسودة")}</span>',
            "posted": f'<span class="badge green">{tr(lang, "Posted", "مرحل")}</span>',
            "reversed": f'<span class="badge red">{tr(lang, "Reversed", "معكوس")}</span>',
        }.get(safe(row["status"]).lower(), safe(row["status"]))
        category_badge = f'<span class="status-chip gray">{journal_category_label(row["source_type"], lang)}</span>'
        reverse_note = ""
        row_keys = row.keys()
        reversed_by_journal_id = safe_int(row["reversed_by_journal_id"]) if "reversed_by_journal_id" in row_keys else 0
        reversed_from_id = safe_int(row["reversed_from_id"]) if "reversed_from_id" in row_keys else 0
        if reversed_by_journal_id:
            reverse_row = conn.execute(
                "SELECT entry_no FROM journal_entries WHERE id = ?",
                (reversed_by_journal_id,),
            ).fetchone()
            reverse_no = safe(reverse_row["entry_no"]) if reverse_row else str(reversed_by_journal_id)
            reverse_note = (
                f'<div class="muted" style="font-size:12px;margin-top:4px;">'
                f'{tr(lang, "Reversed by", "تم عكسه بقيد")}: '
                f'<a href="{with_lang(f"/ui/accounting/journal/{reversed_by_journal_id}", lang)}">{reverse_no}</a>'
                f'</div>'
            )
        elif reversed_from_id:
            reverse_row = conn.execute(
                "SELECT entry_no FROM journal_entries WHERE id = ?",
                (reversed_from_id,),
            ).fetchone()
            reverse_no = safe(reverse_row["entry_no"]) if reverse_row else str(reversed_from_id)
            reverse_note = (
                f'<div class="muted" style="font-size:12px;margin-top:4px;">'
                f'{tr(lang, "Reversal of", "عكس للقيد")}: '
                f'<a href="{with_lang(f"/ui/accounting/journal/{reversed_from_id}", lang)}">{reverse_no}</a>'
                f'</div>'
            )
        action_html = f'<a class="action-btn blue" href="{with_lang(f"/ui/accounting/journal/{row["id"]}", lang)}">{tr(lang, "Open", "Open")}</a>'
        if can_edit_journal_entry(request, row):
            action_html += f"""
                <a class="action-btn gray" href="{with_lang(f'/ui/accounting/journal/{row["id"]}/edit', lang)}">{tr(lang, "Edit", "تعديل")}</a>
                <form method="post" action="{with_lang(f'/ui/accounting/journal/{row["id"]}/delete', lang)}" style="display:inline;"
                      onsubmit="return confirm('{tr(lang, "Delete this draft journal?", "حذف هذا القيد غير المرحل؟")}');">
                    <button class="action-btn red" type="submit">{tr(lang, "Delete", "حذف")}</button>
                </form>
            """

        body += f"""
        <tr>
            <td><span class="doc-no">{safe(row['entry_no'])}</span></td>
            <td>{safe(row['entry_date'])}</td>
            <td>{category_badge}</td>
            <td>{safe(row['description'])}</td>
            <td>{safe(row['reference'])}{reverse_note}</td>
            <td class="number-cell">{money(total_debit)}</td>
            <td class="number-cell">{money(total_credit)}</td>
            <td>{badge}</td>
            <td>{action_html}</td>
        </tr>
        """

    if not body:
        body = f"<tr><td colspan='9' class='empty-state'>{tr(lang, 'No journal entries found for the selected filters.', 'لا توجد قيود يومية مطابقة للفلاتر المحددة.')}</td></tr>"

    account_filter_note = ""
    if safe(account_code):
        account_filter_note = f'<span class="summary-pill">{tr(lang, "Account", "الحساب")}: {safe(account_code)}</span>'
    if safe(partner_type) or safe(partner_id):
        account_filter_note += f'<span class="summary-pill">{tr(lang, "Partner", "الشريك")}: {safe(partner_type)} {safe(partner_id)}</span>'

    def tab(x, title):
        href = f"/ui/accounting/journal?status={x}&search={safe(search)}&date_from={safe(date_from)}&date_to={safe(date_to)}&category={safe(category)}&account_code={safe(account_code)}&partner_type={safe(partner_type)}&partner_id={safe(partner_id)}"
        return f'<a class="page-tab {"active" if status == x else ""}" href="{with_lang(href, lang)}">{title}</a>'

    content = f"""
    <div class="list-shell">
        <div class="card">
            <div class="list-header">
                <div class="list-title">
                    <h2>{tr(lang, "Journal", "دفتر اليومية")}</h2>
                </div>
                <div style="display:flex;gap:10px;flex-wrap:wrap;">
                    {"<a href='" + with_lang('/ui/accounting/journal/template', lang) + "' class='btn gray'>" + tr(lang, "Template", "النموذج") + "</a>" if can_create else ""}
                    {"<a href='" + with_lang('/ui/accounting/journal/import', lang) + "' class='btn blue'>" + tr(lang, "Import Excel", "استيراد إكسل") + "</a>" if can_create else ""}
                    <a href="{with_lang(f'/ui/accounting/journal/export?status={safe(status) or "all"}', lang)}" class="btn gray">{tr(lang, "Export Excel", "تصدير إكسل")}</a>
                    {"<a href='" + with_lang('/ui/accounting/journal/new', lang) + "' class='btn green'>+ " + tr(lang, "New Entry", "قيد جديد") + "</a>" if can_create else ""}
                </div>
            </div>

            <div class="page-tabs" style="margin-top:14px;">
                {tab("all", tr(lang, "All", "الكل"))}
                {tab("draft", tr(lang, "Draft", "مسودة"))}
                {tab("posted", tr(lang, "Posted", "مرحل"))}
                {tab("reversed", tr(lang, "Reversed", "معكوس"))}
            </div>
        </div>

        <div class="card">
            <div class="toolbar" style="margin-bottom:14px;">
                <div>
                    <h3 class="sub-title">{tr(lang, "Filters", "الفلاتر")}</h3>
                </div>
                <div class="table-summary">
                    <span class="summary-pill">{tr(lang, "Entries", "القيود")}: {len(rows)}</span>
                    {account_filter_note}
                    <span class="summary-pill">{tr(lang, "Debit", "مدين")}: {money(total_debit_sum)}</span>
                    <span class="summary-pill">{tr(lang, "Credit", "دائن")}: {money(total_credit_sum)}</span>
                </div>
            </div>

            <form method="get">
                <input type="hidden" name="lang" value="{lang}">
                <input type="hidden" name="status" value="{safe(status)}">
                <input type="hidden" name="account_code" value="{safe(account_code)}">
                <input type="hidden" name="partner_type" value="{safe(partner_type)}">
                <input type="hidden" name="partner_id" value="{safe(partner_id)}">
                <div class="filter-grid" style="grid-template-columns: 1.4fr 1fr 1fr 1fr;">
                    <div class="form-group">
                        <label>{tr(lang, "Search", "بحث")}</label>
                        <input type="text" name="search" value="{safe(search)}" placeholder="{tr(lang, 'Entry no, description, reference...', 'رقم القيد أو البيان أو المرجع...')}">
                    </div>
                    <div class="form-group">
                        <label>{tr(lang, "From Date", "من تاريخ")}</label>
                        <input type="date" name="date_from" value="{safe(date_from)}">
                    </div>
                    <div class="form-group">
                        <label>{tr(lang, "To Date", "إلى تاريخ")}</label>
                        <input type="date" name="date_to" value="{safe(date_to)}">
                    </div>
                    <div class="form-group">
                        <label>{tr(lang, "Category", "الفئة")}</label>
                        <select name="category" multiple size="4" data-searchable="1">
                            {journal_category_options(category_values, lang)}
                        </select>
            <div class="muted" style="margin-top:6px;font-size:12px;">{tr(lang, "You can choose more than one category.", "يمكن اختيار أكثر من تصنيف.")}</div>
                    </div>
                </div>

                <div class="filter-actions">
                    <button class="btn blue" type="submit">{tr(lang, "Filter", "تصفية")}</button>
                    <a class="btn gray" href="{with_lang('/ui/accounting/journal', lang)}">{tr(lang, "Clear", "مسح")}</a>
                </div>
            </form>
        </div>

        <div class="card">
            <div class="toolbar" style="margin-bottom:16px;">
                <div>
                    <h3 class="sub-title">{tr(lang, "Journal Entries", "قيود اليومية")}</h3>
                </div>
            </div>

            <div class="table-wrap">
                <table>
                    <tr>
                        <th>{tr(lang, "Entry No", "رقم القيد")}</th>
                        <th>{tr(lang, "Date", "التاريخ")}</th>
                        <th>{tr(lang, "Category", "الفئة")}</th>
                        <th>{tr(lang, "Description", "البيان")}</th>
                        <th>{tr(lang, "Reference", "المرجع")}</th>
                        <th class="text-right">{tr(lang, "Total Debit", "إجمالي المدين")}</th>
                        <th class="text-right">{tr(lang, "Total Credit", "إجمالي الدائن")}</th>
                        <th>{tr(lang, "Status", "الحالة")}</th>
                        <th>{tr(lang, "Actions", "الإجراءات")}</th>
                    </tr>
                    {body}
                </table>
            </div>
        </div>
    </div>
    """
    conn.close()
    return HTMLResponse(render_page(tr(lang, "Journal", "دفتر اليومية"), content, lang=lang, current_path=request.url.path))


@router.get("/ui/accounting/journal/new", response_class=HTMLResponse)
def journal_new(request: Request):
    lang = get_lang(request)
    if not accounting_allowed(request, "create"):
        return permission_denied(lang, "You do not have permission to create journal entries.", "ليس لديك صلاحية إنشاء قيود يومية.")
    return HTMLResponse(render_page(tr(lang, "New Journal Entry", "قيد يومية جديد"), journal_form_html(with_lang("/ui/accounting/journal/new", lang), lang=lang), lang=lang, current_path=request.url.path))


@router.post("/ui/accounting/journal/new")
async def journal_create(request: Request):
    lang = get_lang(request)
    if not accounting_allowed(request, "create"):
        return permission_denied(lang, "You do not have permission to create journal entries.", "ليس لديك صلاحية إنشاء قيود يومية.")
    form = await request.form()
    entry_date = safe(form.get("entry_date"))
    description = safe(form.get("description"))
    reference = safe(form.get("reference"))
    lines = parse_lines_from_form(form)

    if not description:
        return HTMLResponse(render_page(tr(lang, "New Journal Entry", "قيد يومية جديد"), journal_form_html(with_lang("/ui/accounting/journal/new", lang), values=dict(form), lines=lines, error_message=tr(lang, "Description is required.", "البيان مطلوب."), lang=lang), lang=lang, current_path=request.url.path), status_code=400)
    if not lines:
        return HTMLResponse(render_page(tr(lang, "New Journal Entry", "قيد يومية جديد"), journal_form_html(with_lang("/ui/accounting/journal/new", lang), values=dict(form), lines=lines, error_message=tr(lang, "At least one line is required.", "مطلوب سطر واحد على الأقل."), lang=lang), lang=lang, current_path=request.url.path), status_code=400)

    conn = get_conn()
    try:
        journal_id = create_journal_draft(conn, entry_date, description, reference, lines)
        safe_log_action(
            "journal_entry",
            journal_id,
            "Created",
            done_by=actor_name_from_request(request),
            notes=f"Draft journal created. Description: {description or '-'} | Lines: {len(lines)}",
            conn=conn,
        )
        conn.commit()
        return RedirectResponse(with_lang(f"/ui/accounting/journal/{journal_id}", lang), status_code=303)
    except Exception as e:
        conn.rollback()
        return HTMLResponse(render_page(tr(lang, "New Journal Entry", "قيد يومية جديد"), journal_form_html(with_lang("/ui/accounting/journal/new", lang), values=dict(form), lines=lines, error_message=str(e), lang=lang), lang=lang, current_path=request.url.path), status_code=400)
    finally:
        conn.close()


@router.get("/ui/accounting/journal/{journal_id}", response_class=HTMLResponse)
def journal_view(request: Request, journal_id: int):
    lang = get_lang(request)
    conn = get_conn()
    entry = get_entry(conn, journal_id)
    if not entry:
        conn.close()
        return HTMLResponse(tr(lang, "Journal entry not found.", "قيد اليومية غير موجود."), status_code=404)

    lines = get_entry_lines(conn, journal_id)
    total_debit, total_credit, balanced = get_entry_totals(conn, journal_id)
    content = entry_view_html(
        entry,
        lines,
        total_debit,
        total_credit,
        balanced,
        lang=lang,
        allow_edit=can_edit_journal_entry(request, entry),
        allow_post=accounting_allowed(request, "post"),
    )
    content += render_audit_log_card("journal_entry", journal_id)
    conn.close()
    return HTMLResponse(render_page(f'{tr(lang, "Journal", "دفتر اليومية")} {safe(entry["entry_no"])}', content, lang=lang, current_path=request.url.path))


@router.get("/ui/accounting/journal/{journal_id}/edit", response_class=HTMLResponse)
def journal_edit(request: Request, journal_id: int):
    lang = get_lang(request)
    conn = get_conn()
    entry = get_entry(conn, journal_id)
    if not entry:
        conn.close()
        return HTMLResponse(tr(lang, "Journal entry not found.", "قيد اليومية غير موجود."), status_code=404)
    if not can_edit_journal_entry(request, entry):
        conn.close()
        return permission_denied(
            lang,
            "You do not have permission to edit this journal entry.",
            "ليس لديك صلاحية تعديل هذا القيد.",
        )
    if safe(entry["status"]).lower() not in ["draft", "pending_final_post"]:
        conn.close()
        return RedirectResponse(with_lang(f"/ui/accounting/journal/{journal_id}", lang), status_code=303)

    lines = get_entry_lines(conn, journal_id)
    content = journal_form_html(
        with_lang(f"/ui/accounting/journal/{journal_id}/edit", lang),
        values=dict(entry),
        lines=lines,
        lang=lang,
    )
    conn.close()
    return HTMLResponse(render_page(tr(lang, "Edit Journal Entry", "تعديل قيد اليومية"), content, lang=lang, current_path=request.url.path))


@router.post("/ui/accounting/journal/{journal_id}/edit")
async def journal_update(request: Request, journal_id: int):
    lang = get_lang(request)
    form = await request.form()
    entry_date = safe(form.get("entry_date"))
    description = safe(form.get("description"))
    reference = safe(form.get("reference"))
    lines = []

    conn = get_conn()
    try:
        lines = parse_lines_from_form(form)
        entry = get_entry(conn, journal_id)
        if not entry:
            conn.close()
            return HTMLResponse(tr(lang, "Journal entry not found.", "قيد اليومية غير موجود."), status_code=404)
        if not can_edit_journal_entry(request, entry):
            conn.close()
            return permission_denied(
                lang,
                "You do not have permission to edit this journal entry.",
                "ليس لديك صلاحية تعديل هذا القيد.",
            )

        status = safe(entry["status"]).lower()
        if status in ("draft", "pending_final_post"):
            update_journal_draft(conn, journal_id, entry_date, description, reference, lines)
            audit_action = "Updated"
            audit_notes = f"Unfinalized journal updated. Description: {description or '-'} | Lines: {len(lines)}"
        else:
            raise Exception("Only draft or pending-final journal entries can be edited")

        safe_log_action(
            "journal_entry",
            journal_id,
            audit_action,
            done_by=actor_name_from_request(request),
            notes=audit_notes,
            conn=conn,
        )
        conn.commit()
        return RedirectResponse(with_lang(f"/ui/accounting/journal/{journal_id}", lang), status_code=303)
    except Exception as e:
        conn.rollback()
        return HTMLResponse(render_page(tr(lang, "Edit Journal Entry", "تعديل قيد اليومية"), journal_form_html(with_lang(f"/ui/accounting/journal/{journal_id}/edit", lang), values=dict(form), lines=lines, error_message=str(e), lang=lang), lang=lang, current_path=request.url.path), status_code=400)
    finally:
        conn.close()


@router.post("/ui/accounting/journal/{journal_id}/delete")
def journal_delete(request: Request, journal_id: int):
    lang = get_lang(request)
    conn = get_conn()
    try:
        entry = get_entry(conn, journal_id)
        if not entry:
            raise Exception("Journal entry not found")
        if not (accounting_allowed(request, "delete") or can_edit_journal_entry(request, entry)):
            return permission_denied(lang, "You do not have permission to delete journal entries.", "ليس لديك صلاحية حذف قيود اليومية.")
        if safe(entry["status"]).lower() not in ("draft", "pending_final_post"):
            raise Exception("Only draft or pending-final journal entries can be deleted")
        entry_no = safe(entry["entry_no"])
        delete_unfinalized_journal(conn, journal_id)
        safe_log_action(
            "journal_entry",
            journal_id,
            "Deleted",
            done_by=actor_name_from_request(request),
            notes=f"Unfinalized journal deleted: {entry_no or journal_id}",
            conn=conn,
        )
        conn.commit()
        return RedirectResponse(with_lang("/ui/accounting/journal?status=draft", lang), status_code=303)
    except Exception as e:
        conn.rollback()
        return HTMLResponse(f"{tr(lang, 'Delete error', 'خطأ في الحذف')}: {safe(e)}", status_code=400)
    finally:
        conn.close()


@router.post("/ui/accounting/journal/{journal_id}/post")
def journal_post(request: Request, journal_id: int):
    lang = get_lang(request)
    if not accounting_allowed(request, "post"):
        return permission_denied(lang, "You do not have permission to post journal entries.", "ليس لديك صلاحية ترحيل قيود اليومية.")
    conn = get_conn()
    try:
        entry = get_entry(conn, journal_id)
        if not entry:
            raise Exception("Journal entry not found")
        old_status = safe(entry["status"]).lower()
        post_journal_entry(conn, journal_id)
        safe_log_action(
            "journal_entry",
            journal_id,
            "Posted",
            done_by=actor_name_from_request(request),
            notes=f"Status changed from {old_status or 'unknown'} to posted.",
            conn=conn,
        )
        conn.commit()
    except Exception as e:
        conn.rollback()
        return HTMLResponse(f"{tr(lang, 'Post error', 'خطأ في الترحيل')}: {safe(e)}", status_code=400)
    finally:
        conn.close()
    return RedirectResponse(with_lang(f"/ui/accounting/journal/{journal_id}", lang), status_code=303)


@router.post("/ui/accounting/journal/{journal_id}/reverse")
def journal_reverse(request: Request, journal_id: int):
    lang = get_lang(request)
    if not accounting_allowed(request, "post"):
        return permission_denied(lang, "You do not have permission to reverse posted journal entries.", "ليس لديك صلاحية عكس القيود المرحلة.")
    conn = get_conn()
    try:
        reverse_journal_entry(conn, journal_id)
        safe_log_action(
            "journal_entry",
            journal_id,
            "Reversed",
            done_by=actor_name_from_request(request),
            notes="Status changed from Posted to Reversed.",
            conn=conn,
        )
        conn.commit()
    except Exception as e:
        conn.rollback()
        return HTMLResponse(f"{tr(lang, 'Reverse error', 'خطأ في العكس')}: {safe(e)}", status_code=400)
    finally:
        conn.close()
    return RedirectResponse(with_lang(f"/ui/accounting/journal/{journal_id}", lang), status_code=303)
