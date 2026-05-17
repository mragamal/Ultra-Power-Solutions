import io
from datetime import datetime
from html import escape
from urllib.parse import quote

from fastapi import APIRouter, Form, Request
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse

try:
    from openpyxl import Workbook
except Exception:
    Workbook = None

from db import get_conn
from i18n import get_lang
from layout import render_page
from modules.accounting.employee_advances_statement import advance_statement_ui
from modules.hr.employees import ensure_employees_table, safe, to_float

router = APIRouter()
BASE_ROUTE = "/ui/accounting/employee-advances"
LEGACY_ROUTE = "/ui/hr/advances"


def money(value):
    try:
        return f"{float(value or 0):,.2f}"
    except Exception:
        return "0.00"


def tr(request: Request, en: str, ar: str) -> str:
    return ar if get_lang(request) == "ar" else en


def with_lang(request: Request, path: str) -> str:
    lang = get_lang(request)
    separator = "&" if "?" in path else "?"
    return f"{path}{separator}lang={lang}" if lang == "ar" else path


def with_msg(request: Request, path: str, msg: str) -> str:
    url = with_lang(request, path)
    separator = "&" if "?" in url else "?"
    return f"{url}{separator}msg={quote(msg)}"


def status_label(request: Request, status: str) -> str:
    key = safe(status).lower() or "active"
    labels = {
        "active": tr(request, "Active", "ظ†ط´ط·ط©"),
        "open": tr(request, "Open", "ظ…ظپطھظˆط­ط©"),
        "closed": tr(request, "Closed", "ظ…ظ‚ظپظ„ط©"),
        "cancelled": tr(request, "Cancelled", "ظ…ظ„ط؛ظٹط©"),
    }
    return labels.get(key, safe(status) or tr(request, "Active", "ظ†ط´ط·ط©"))


def normalize_deduction_type(value: str) -> str:
    return "direct" if safe(value).strip().lower() == "direct" else "installment"


def deduction_type_label(request: Request, value: str) -> str:
    if normalize_deduction_type(value) == "direct":
        return tr(request, "Direct Salary Deduction", "خصم مباشر من المرتب")
    return tr(request, "Installments", "أقساط")


def to_int(value, default=0) -> int:
    try:
        text = safe(value).strip()
        if not text:
            return default
        return int(float(text))
    except Exception:
        return default


def deduction_period_from_inputs(advance_date: str, start_month, start_year) -> tuple[int, int]:
    month = to_int(start_month)
    year = to_int(start_year)
    if 1 <= month <= 12 and year >= 1900:
        return month, year

    try:
        parsed = datetime.strptime(safe(advance_date), "%Y-%m-%d")
        return parsed.month, parsed.year
    except Exception:
        today = datetime.now()
        return today.month, today.year


def can_edit_or_delete_advance(row) -> bool:
    if not row:
        return False
    if safe(row.get("journal_line_id")).strip():
        return False
    return safe(row.get("status")).lower() in ("active", "open")


def advance_row_get(row, key, default=None):
    if not row:
        return default
    if isinstance(row, dict):
        return row.get(key, default)
    try:
        return row[key]
    except Exception:
        return default


def advance_has_posted_journal(conn, row) -> bool:
    journal_line_id = safe(advance_row_get(row, "journal_line_id")).strip()
    if not journal_line_id:
        return False

    linked = conn.execute(
        """
        SELECT j.status
        FROM journal_lines jl
        JOIN journal_entries j ON j.id = jl.journal_id
        WHERE jl.id = ?
        LIMIT 1
        """,
        (journal_line_id,),
    ).fetchone()
    if linked:
        return safe(linked["status"]).lower() == "posted"

    return True


def can_delete_advance(conn, row) -> bool:
    if not row:
        return False
    status = safe(advance_row_get(row, "status")).lower() or "active"
    if status in ("closed", "cancelled", "paid", "posted", "reversed"):
        return False
    if advance_has_posted_journal(conn, row):
        return False

    advance_id = advance_row_get(row, "id")
    if not advance_id:
        return False
    has_payroll_deduction = conn.execute(
        "SELECT 1 FROM employee_advance_deductions WHERE advance_id = ? LIMIT 1",
        (advance_id,),
    ).fetchone()
    return has_payroll_deduction is None


def can_edit_advance_schedule(row) -> bool:
    if not row:
        return False
    return safe(row.get("status")).lower() in ("active", "open")


def current_period_key() -> int:
    today = datetime.now()
    return period_key(today.month, today.year)


def ensure_column(conn, table_name, column_name, alter_sql):
    cols = conn.execute(f"PRAGMA table_info({table_name})").fetchall()
    names = [c["name"] for c in cols]
    if column_name not in names:
        conn.execute(alter_sql)


def normalize_employee_name(value):
    return "".join(safe(value).lower().split())


def resolve_employee_id_from_journal_line(conn, line):
    partner_id = int(line["partner_id"] or 0)
    if partner_id > 0:
        emp = conn.execute("SELECT id FROM employees WHERE id = ? LIMIT 1", (partner_id,)).fetchone()
        if emp:
            return partner_id

    raw_name = safe(line["line_description"] or "")
    if not raw_name:
        return partner_id

    target = normalize_employee_name(raw_name)
    employees = conn.execute("SELECT id, name FROM employees").fetchall()
    for emp in employees:
        if normalize_employee_name(emp["name"]) == target:
            return int(emp["id"])

    return partner_id


def ensure_advances_tables():
    ensure_employees_table()
    conn = get_conn()
    try:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS employee_advances (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                advance_no TEXT UNIQUE,
                advance_date TEXT,
                employee_id INTEGER NOT NULL,
                amount REAL DEFAULT 0,
                deduction_type TEXT DEFAULT 'installment',
                installment_amount REAL DEFAULT 0,
                start_month INTEGER,
                start_year INTEGER,
                notes TEXT,
                status TEXT DEFAULT 'active',
                journal_line_id INTEGER,
                is_opening_balance INTEGER DEFAULT 0,
                paid_before_start REAL DEFAULT 0,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
            """
        )

        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS employee_advance_deductions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                advance_id INTEGER NOT NULL,
                payroll_run_id INTEGER NOT NULL,
                payroll_line_id INTEGER,
                deduction_month INTEGER,
                deduction_year INTEGER,
                amount REAL DEFAULT 0,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
            """
        )

        conn.execute(
            """
            CREATE UNIQUE INDEX IF NOT EXISTS idx_advance_deduction_unique
            ON employee_advance_deductions (advance_id, payroll_run_id, payroll_line_id)
            """
        )

        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS employee_advance_installments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                advance_id INTEGER NOT NULL,
                installment_month INTEGER,
                installment_year INTEGER,
                planned_amount REAL DEFAULT 0,
                paid_amount REAL DEFAULT 0,
                status TEXT DEFAULT 'pending',
                is_deferred INTEGER DEFAULT 0,
                deferred_reason TEXT,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
            """
        )

        # Table to track deferral requests (exceptions)
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS employee_advance_deferrals (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                advance_id INTEGER NOT NULL,
                installment_id INTEGER,
                original_month INTEGER,
                original_year INTEGER,
                deferred_to_month INTEGER,
                deferred_to_year INTEGER,
                amount REAL DEFAULT 0,
                reason TEXT,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
            """
        )

        ensure_column(conn, "employee_advances", "advance_no", "ALTER TABLE employee_advances ADD COLUMN advance_no TEXT")
        ensure_column(conn, "employee_advances", "advance_date", "ALTER TABLE employee_advances ADD COLUMN advance_date TEXT")
        ensure_column(conn, "employee_advances", "employee_id", "ALTER TABLE employee_advances ADD COLUMN employee_id INTEGER")
        ensure_column(conn, "employee_advances", "amount", "ALTER TABLE employee_advances ADD COLUMN amount REAL DEFAULT 0")
        ensure_column(conn, "employee_advances", "deduction_type", "ALTER TABLE employee_advances ADD COLUMN deduction_type TEXT DEFAULT 'installment'")
        ensure_column(conn, "employee_advances", "installment_amount", "ALTER TABLE employee_advances ADD COLUMN installment_amount REAL DEFAULT 0")
        ensure_column(conn, "employee_advances", "start_month", "ALTER TABLE employee_advances ADD COLUMN start_month INTEGER")
        ensure_column(conn, "employee_advances", "start_year", "ALTER TABLE employee_advances ADD COLUMN start_year INTEGER")
        ensure_column(conn, "employee_advances", "notes", "ALTER TABLE employee_advances ADD COLUMN notes TEXT")
        ensure_column(conn, "employee_advances", "status", "ALTER TABLE employee_advances ADD COLUMN status TEXT DEFAULT 'active'")
        ensure_column(conn, "employee_advances", "journal_line_id", "ALTER TABLE employee_advances ADD COLUMN journal_line_id INTEGER")
        ensure_column(conn, "employee_advances", "created_at", "ALTER TABLE employee_advances ADD COLUMN created_at TEXT DEFAULT CURRENT_TIMESTAMP")
        ensure_column(conn, "employee_advances", "is_opening_balance", "ALTER TABLE employee_advances ADD COLUMN is_opening_balance INTEGER DEFAULT 0")
        ensure_column(conn, "employee_advances", "paid_before_start", "ALTER TABLE employee_advances ADD COLUMN paid_before_start REAL DEFAULT 0")

        ensure_column(conn, "employee_advance_deductions", "advance_id", "ALTER TABLE employee_advance_deductions ADD COLUMN advance_id INTEGER")
        ensure_column(conn, "employee_advance_deductions", "payroll_run_id", "ALTER TABLE employee_advance_deductions ADD COLUMN payroll_run_id INTEGER")
        ensure_column(conn, "employee_advance_deductions", "payroll_line_id", "ALTER TABLE employee_advance_deductions ADD COLUMN payroll_line_id INTEGER")
        ensure_column(conn, "employee_advance_deductions", "deduction_month", "ALTER TABLE employee_advance_deductions ADD COLUMN deduction_month INTEGER")
        ensure_column(conn, "employee_advance_deductions", "deduction_year", "ALTER TABLE employee_advance_deductions ADD COLUMN deduction_year INTEGER")
        ensure_column(conn, "employee_advance_deductions", "amount", "ALTER TABLE employee_advance_deductions ADD COLUMN amount REAL DEFAULT 0")
        ensure_column(conn, "employee_advance_deductions", "created_at", "ALTER TABLE employee_advance_deductions ADD COLUMN created_at TEXT DEFAULT CURRENT_TIMESTAMP")

        conn.commit()
    finally:
        conn.close()


def regenerate_installment_schedule(conn, advance_id):
    advance = conn.execute("SELECT * FROM employee_advances WHERE id = ?", (advance_id,)).fetchone()
    if not advance:
        return

    paid_total = advance_paid_amount(conn, advance_id)
    # For opening balance advances, subtract the amount already paid before system start
    paid_before_start = float(advance["paid_before_start"] or 0)
    remaining = float(advance["amount"] or 0) - paid_total - paid_before_start
    remaining = max(remaining, 0)

    # Delete pending installments (keep paid/deferred ones intact)
    conn.execute("DELETE FROM employee_advance_installments WHERE advance_id = ? AND status = 'pending'", (advance_id,))

    if remaining <= 0:
        return

    deduction_type = normalize_deduction_type(advance["deduction_type"] if "deduction_type" in advance.keys() else "")
    installment_amount = remaining if deduction_type == "direct" else float(advance["installment_amount"] or 0)
    if installment_amount <= 0:
        installment_amount = remaining  # One shot

    m = int(advance["start_month"] or 1)
    y = int(advance["start_year"] or 2024)

    current_key = current_period_key()

    while remaining > 0.001:
        # Skip months that already have a paid deduction
        paid_in_period = conn.execute("""
            SELECT SUM(amount) as total FROM employee_advance_deductions
            WHERE advance_id = ? AND deduction_month = ? AND deduction_year = ?
        """, (advance_id, m, y)).fetchone()

        if paid_in_period and paid_in_period["total"] and paid_in_period["total"] > 0:
            m += 1
            if m > 12:
                m = 1
                y += 1
            continue

        # Skip months that already have an installment (paid/deferred/manual).
        existing = conn.execute("""
            SELECT id FROM employee_advance_installments
            WHERE advance_id = ? AND installment_month = ? AND installment_year = ?
        """, (advance_id, m, y)).fetchone()

        if existing:
            m += 1
            if m > 12:
                m = 1
                y += 1
            continue

        amt = min(remaining, installment_amount)
        status = "pending"
        paid_amount = 0
        conn.execute("""
            INSERT INTO employee_advance_installments (
                advance_id, installment_month, installment_year, planned_amount, paid_amount, status
            )
            VALUES (?, ?, ?, ?, ?, ?)
        """, (advance_id, m, y, amt, paid_amount, status))

        remaining -= amt
        m += 1
        if m > 12:
            m = 1
            y += 1


def defer_installment(conn, advance_id, installment_id, reason=""):
    """
    Defer a pending installment to the month after the current final schedule month.
    This extends the advance term instead of moving the skipped installment into
    the middle of the existing schedule.
    """
    inst = conn.execute(
        "SELECT * FROM employee_advance_installments WHERE id = ? AND advance_id = ? AND status = 'pending'",
        (installment_id, advance_id)
    ).fetchone()
    if not inst:
        return False, "Installment not found or already paid."

    orig_month = int(inst["installment_month"] or 0)
    orig_year = int(inst["installment_year"] or 0)
    amount = float(inst["planned_amount"] or 0)

    last_period = conn.execute(
        """
        SELECT installment_month, installment_year
        FROM employee_advance_installments
        WHERE advance_id = ?
          AND id <> ?
        ORDER BY installment_year DESC, installment_month DESC, id DESC
        LIMIT 1
        """,
        (advance_id, installment_id),
    ).fetchone()

    if last_period:
        nm = int(last_period["installment_month"] or orig_month)
        ny = int(last_period["installment_year"] or orig_year)
        nm += 1
    else:
        nm = orig_month + 1
        ny = orig_year
    if nm > 12:
        nm = 1
        ny += 1

    # Keep advancing until we find a free month after the current final schedule.
    for _ in range(120):  # max 10 years safety
        occupied = conn.execute("""
            SELECT id FROM employee_advance_installments
            WHERE advance_id = ? AND installment_month = ? AND installment_year = ?
        """, (advance_id, nm, ny)).fetchone()
        if not occupied:
            break
        nm += 1
        if nm > 12:
            nm = 1
            ny += 1

    # Mark original as deferred
    conn.execute(
        "UPDATE employee_advance_installments SET status = 'deferred', is_deferred = 1, deferred_reason = ? WHERE id = ?",
        (reason, installment_id)
    )

    # Insert deferred installment in new month
    conn.execute("""
        INSERT INTO employee_advance_installments
            (advance_id, installment_month, installment_year, planned_amount, status, is_deferred, deferred_reason)
        VALUES (?, ?, ?, ?, 'pending', 1, ?)
    """, (advance_id, nm, ny, amount, reason))

    # Record deferral history
    conn.execute("""
        INSERT INTO employee_advance_deferrals
            (advance_id, installment_id, original_month, original_year, deferred_to_month, deferred_to_year, amount, reason)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    """, (advance_id, installment_id, orig_month, orig_year, nm, ny, amount, reason))

    return True, f"Installment deferred from {orig_month:02d}/{orig_year} to {nm:02d}/{ny}. Advance term extended."

def next_advance_no():
    ensure_advances_tables()
    conn = get_conn()
    try:
        row = conn.execute(
            """
            SELECT advance_no
            FROM employee_advances
            WHERE COALESCE(advance_no, '') <> ''
            ORDER BY id DESC
            LIMIT 1
            """
        ).fetchone()
        last = safe(row["advance_no"]) if row else ""
        if not last:
            return "ADV-0001"
        try:
            num = int(last.split("-")[-1])
        except Exception:
            num = 0
        return f"ADV-{num + 1:04d}"
    finally:
        conn.close()


def employee_options_html(selected_id=""):
    ensure_employees_table()
    conn = get_conn()
    try:
        rows = conn.execute(
            """
            SELECT id, code, name
            FROM employees
            WHERE COALESCE(is_active, 1) = 1
            ORDER BY code, name
            """
        ).fetchall()

        html = '<option value="">-- Select Employee --</option>'
        selected_text = safe(selected_id)
        for row in rows:
            rid = str(row["id"])
            sel = "selected" if rid == selected_text else ""
            label = f"{safe(row['code'])} - {safe(row['name'])}" if safe(row["code"]) else safe(row["name"])
            html += f'<option value="{rid}" {sel}>{escape(label)}</option>'
        return html
    finally:
        conn.close()


def advance_paid_amount(conn, advance_id):
    deduction_row = conn.execute(
        "SELECT COALESCE(SUM(amount), 0) AS total FROM employee_advance_deductions WHERE advance_id = ?",
        (advance_id,),
    ).fetchone()
    installment_row = conn.execute(
        """
        SELECT COALESCE(SUM(COALESCE(ei.paid_amount, 0)), 0) AS total
        FROM employee_advance_installments ei
        WHERE ei.advance_id = ?
          AND LOWER(COALESCE(ei.status, '')) = 'paid'
          AND NOT EXISTS (
              SELECT 1
              FROM employee_advance_deductions d
              WHERE d.advance_id = ei.advance_id
                AND d.deduction_month = ei.installment_month
                AND d.deduction_year = ei.installment_year
          )
        """,
        (advance_id,),
    ).fetchone()
    return float(deduction_row["total"] or 0) + float(installment_row["total"] or 0)


def advance_remaining_amount(conn, advance_row):
    paid_before_start = 0.0
    try:
        paid_before_start = float(advance_row["paid_before_start"] or 0)
    except Exception:
        paid_before_start = 0.0
    return max(float(advance_row["amount"] or 0) - advance_paid_amount(conn, advance_row["id"]) - paid_before_start, 0)


def period_key(month, year):
    return (int(year or 0) * 100) + int(month or 0)


def is_due_in_period(start_month, start_year, payroll_month, payroll_year):
    if int(start_month or 0) <= 0 or int(start_year or 0) <= 0:
        return True
    return period_key(start_month, start_year) <= period_key(payroll_month, payroll_year)


def get_employee_due_advances(conn, employee_id, payroll_month, payroll_year):
    # Primary source: scheduled installments for this payroll period.
    rows = conn.execute(
        """
        SELECT ei.*, ea.advance_no, ea.employee_id
        FROM employee_advance_installments ei
        JOIN employee_advances ea ON ea.id = ei.advance_id
        WHERE ea.employee_id = ?
          AND ei.installment_month = ?
          AND ei.installment_year = ?
          AND ei.status = 'pending'
          AND LOWER(COALESCE(ea.status, 'active')) IN ('active', 'open')
        """,
        (employee_id, payroll_month, payroll_year),
    ).fetchall()

    due_rows = []
    for row in rows:
        due_rows.append(
            {
                "id": int(row["advance_id"]),
                "installment_id": int(row["id"]),
                "advance_no": safe(row["advance_no"]),
                "due_amount": float(row["planned_amount"] or 0),
            }
        )

    # Backward-compatibility fallback:
    # Some historical advances may not have generated installment rows.
    # In that case, derive due amount from advance setup so payroll still deducts.
    if due_rows:
        return due_rows

    legacy_advances = conn.execute(
        """
        SELECT *
        FROM employee_advances
        WHERE employee_id = ?
          AND LOWER(COALESCE(status, 'active')) IN ('active', 'open')
        ORDER BY id
        """,
        (employee_id,),
    ).fetchall()

    for advance in legacy_advances:
        if not is_due_in_period(
            advance["start_month"],
            advance["start_year"],
            payroll_month,
            payroll_year,
        ):
            continue

        remaining = advance_remaining_amount(conn, advance)
        if remaining <= 0:
            continue

        deduction_type = normalize_deduction_type(advance["deduction_type"] if "deduction_type" in advance.keys() else "")
        installment = remaining if deduction_type == "direct" else float(advance["installment_amount"] or 0)
        due_amount = installment if installment > 0 else remaining
        due_amount = min(due_amount, remaining)
        if due_amount <= 0:
            continue

        due_rows.append(
            {
                "id": int(advance["id"]),
                "installment_id": None,
                "advance_no": safe(advance["advance_no"]),
                "due_amount": due_amount,
            }
        )

    return due_rows


def get_employee_due_advance_total(conn, employee_id, payroll_month, payroll_year):
    return sum(row["due_amount"] for row in get_employee_due_advances(conn, employee_id, payroll_month, payroll_year))


def sync_advance_status(conn, advance_id):
    row = conn.execute("SELECT * FROM employee_advances WHERE id = ? LIMIT 1", (advance_id,)).fetchone()
    if not row:
        return
    if safe(row["status"]).lower() == "cancelled":
        return
    remaining = advance_remaining_amount(conn, row)
    status = "closed" if remaining <= 0.0001 else "active"
    conn.execute("UPDATE employee_advances SET status = ? WHERE id = ?", (status, advance_id))


def sync_all_advance_statuses(conn):
    rows = conn.execute("SELECT id FROM employee_advances").fetchall()
    for row in rows:
        sync_advance_status(conn, row["id"])


def allocate_payroll_advance_deductions(conn, run_id):
    run = conn.execute("SELECT * FROM payroll_runs WHERE id = ? LIMIT 1", (run_id,)).fetchone()
    if not run:
        return

    payroll_month = int(run["payroll_month"] or 0)
    payroll_year = int(run["payroll_year"] or 0)
    lines = conn.execute(
        """
        SELECT *
        FROM payroll_lines
        WHERE payroll_run_id = ?
        ORDER BY id
        """,
        (run_id,),
    ).fetchall()

    conn.execute("DELETE FROM employee_advance_deductions WHERE payroll_run_id = ?", (run_id,))

    for line in lines:
        remaining_to_allocate = float(line["advance_deduction"] or 0)
        if remaining_to_allocate <= 0:
            continue

        due_advances = get_employee_due_advances(conn, line["employee_id"], payroll_month, payroll_year)
        for advance in due_advances:
            if remaining_to_allocate <= 0:
                break
            allocation = min(remaining_to_allocate, float(advance["due_amount"] or 0))
            if allocation <= 0:
                continue
            conn.execute(
                """
                INSERT INTO employee_advance_deductions (
                    advance_id, payroll_run_id, payroll_line_id, deduction_month, deduction_year, amount
                )
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (
                    advance["id"],
                    run_id,
                    line["id"],
                    payroll_month,
                    payroll_year,
                    allocation,
                ),
            )
            remaining_to_allocate -= allocation
            if advance.get("installment_id"):
                conn.execute(
                    """
                    UPDATE employee_advance_installments
                    SET paid_amount = MIN(COALESCE(planned_amount, 0), COALESCE(paid_amount, 0) + ?),
                        status = CASE
                            WHEN COALESCE(paid_amount, 0) + ? >= COALESCE(planned_amount, 0) - 0.001
                            THEN 'paid'
                            ELSE status
                        END
                    WHERE id = ?
                    """,
                    (allocation, allocation, advance["installment_id"]),
                )
            sync_advance_status(conn, advance["id"])

        if remaining_to_allocate > 0:
            for advance in due_advances:
                if remaining_to_allocate <= 0:
                    break
                row = conn.execute("SELECT * FROM employee_advances WHERE id = ? LIMIT 1", (advance["id"],)).fetchone()
                if not row:
                    continue
                remaining_balance = advance_remaining_amount(conn, row)
                extra_allocation = min(remaining_to_allocate, remaining_balance)
                if extra_allocation <= 0:
                    continue
                exists = conn.execute(
                    """
                    SELECT id, amount
                    FROM employee_advance_deductions
                    WHERE advance_id = ? AND payroll_run_id = ? AND payroll_line_id = ?
                    LIMIT 1
                    """,
                    (advance["id"], run_id, line["id"]),
                ).fetchone()
                if exists:
                    conn.execute(
                        "UPDATE employee_advance_deductions SET amount = ? WHERE id = ?",
                        (float(exists["amount"] or 0) + extra_allocation, exists["id"]),
                    )
                else:
                    conn.execute(
                        """
                        INSERT INTO employee_advance_deductions (
                            advance_id, payroll_run_id, payroll_line_id, deduction_month, deduction_year, amount
                        )
                        VALUES (?, ?, ?, ?, ?, ?)
                        """,
                        (
                            advance["id"],
                            run_id,
                            line["id"],
                            payroll_month,
                            payroll_year,
                            extra_allocation,
                        ),
                    )
                remaining_to_allocate -= extra_allocation
                sync_advance_status(conn, advance["id"])

    advance_ids = conn.execute("SELECT id FROM employee_advances").fetchall()
    for row in advance_ids:
        sync_advance_status(conn, row["id"])


ensure_advances_tables()


@router.get(LEGACY_ROUTE)
def legacy_advances_redirect(request: Request):
    return RedirectResponse(with_lang(request, BASE_ROUTE), status_code=302)


@router.get(f"{BASE_ROUTE}/sync-from-journal")
def advances_sync_from_journal(request: Request):
    conn = get_conn()
    try:
        # Find journal lines for employees in accounts that look like advances
        # which are NOT already in employee_advances
        # We look for debit lines (new advances or opening balances)
        lines = conn.execute("""
            SELECT l.id, l.partner_id, l.debit, l.credit, j.entry_date, j.entry_no, j.description, l.line_description, a.name as account_name
            FROM journal_lines l
            JOIN journal_entries j ON j.id = l.journal_id
            JOIN accounts a ON a.code = l.account_code
            WHERE l.partner_type = 'employee'
              AND l.debit > 0
              AND (
                a.code IN (SELECT value FROM accounting_settings WHERE key = 'employee_advance_account')
                OR a.name LIKE '%ط³ظ„ظپ%'
                OR a.name LIKE '%ظ‚ط±ط¶%'
                OR a.name LIKE '%advance%'
                OR a.name LIKE '%loan%'
              )
              AND l.id NOT IN (SELECT journal_line_id FROM employee_advances WHERE journal_line_id IS NOT NULL)
              AND j.status = 'posted'
        """).fetchall()

        count = 0
        for l in lines:
            # Create advance
            # We need a unique advance_no for each
            row = conn.execute("SELECT advance_no FROM employee_advances ORDER BY id DESC LIMIT 1").fetchone()
            last = safe(row["advance_no"]) if row else ""
            if not last:
                adv_no = f"ADV-S-{l['id']:04d}"
            else:
                try:
                    num = int(last.split("-")[-1])
                    adv_no = f"ADV-{num + 1 + count:04d}"
                except:
                    adv_no = f"ADV-S-{l['id']:04d}"

            desc = l["line_description"] or l["description"] or f"Synced from {l['entry_no']}"

            employee_id = resolve_employee_id_from_journal_line(conn, l)

            conn.execute("""
                INSERT INTO employee_advances (
                    advance_no, advance_date, employee_id, amount, installment_amount, notes, status, journal_line_id
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                adv_no, l["entry_date"], employee_id, l["debit"], 0, desc, 'open', l["id"]
            ))
            count += 1
        
        conn.commit()
        msg = tr(request, f"Synced {count} advances from journal.", f"طھظ… ظ…ط²ط§ظ…ظ†ط© {count} ط³ظ„ظپط© ظ…ظ† ظ‚ظٹظˆط¯ ط§ظ„ظٹظˆظ…ظٹط©.")
        return RedirectResponse(with_msg(request, BASE_ROUTE, msg), status_code=302)
    except Exception as e:
        return HTMLResponse(f"Error: {str(e)}", status_code=500)
    finally:
        conn.close()


@router.post(f"{BASE_ROUTE}/{{advance_id}}/mark-disbursed")
def advance_mark_disbursed(request: Request, advance_id: int):
    conn = get_conn()
    try:
        conn.execute("UPDATE employee_advances SET status = 'open' WHERE id = ?", (advance_id,))
        conn.commit()
    finally:
        conn.close()
    msg = tr(request, "Advance marked as disbursed.", "طھظ… طھط­ط¯ظٹط¯ ط§ظ„ط³ظ„ظپط© ظƒظ…طµط±ظˆظپط©.")
    return RedirectResponse(with_msg(request, BASE_ROUTE, msg), status_code=302)


@router.get(f"{BASE_ROUTE}/{{advance_id}}/edit", response_class=HTMLResponse)
def advances_edit_ui(request: Request, advance_id: int):
    lang = get_lang(request)
    conn = get_conn()
    try:
        advance = conn.execute("SELECT * FROM employee_advances WHERE id = ?", (advance_id,)).fetchone()
        if not advance:
            return HTMLResponse("Not found", status_code=404)
        
        if not can_edit_advance_schedule(dict(advance)):
            return HTMLResponse(
                tr(
                    request,
                    "Cannot edit this advance after it is linked to journal or closed/cancelled.",
                    "ظ„ط§ ظٹظ…ظƒظ† طھط¹ط¯ظٹظ„ ظ‡ط°ظ‡ ط§ظ„ط³ظ„ظپط© ط¨ط¹ط¯ ط±ط¨ط·ظ‡ط§ ط¨ط§ظ„ظ‚ظٹظˆط¯ ط£ظˆ ط¨ط¹ط¯ ط¥ظ‚ظپط§ظ„ظ‡ط§/ط¥ظ„ط؛ط§ط¦ظ‡ط§.",
                ),
                status_code=400,
            )
        
        is_ob = int(advance["is_opening_balance"] or 0)
        paid_ob_val = float(advance["paid_before_start"] or 0)
        employee_label = f"{safe(advance['employee_id'])}"
        # Get employee name for display
        emp_row = conn.execute("SELECT code, name FROM employees WHERE id = ?", (advance["employee_id"],)).fetchone()
        emp_display = f"{safe(emp_row['code'])} - {safe(emp_row['name'])}" if emp_row else str(advance["employee_id"])

        ob_badge = f' <span class="status-chip blue" style="font-size:11px;">{tr(request, "Opening Balance", "ط±طµظٹط¯ ط§ظپطھطھط§ط­ظٹ")}</span>' if is_ob else ""
        start_month_value = int(advance["start_month"] or datetime.now().month)
        start_year_value = int(advance["start_year"] or datetime.now().year)
        installment_value = max(float(advance["installment_amount"] or 0), 0)
        deduction_type_value = normalize_deduction_type(advance["deduction_type"] if "deduction_type" in advance.keys() else "")
        direct_selected = "selected" if deduction_type_value == "direct" else ""
        installment_selected = "selected" if deduction_type_value == "installment" else ""
        installment_display = "block" if deduction_type_value == "installment" else "none"
        installment_required = "required" if deduction_type_value == "installment" else ""

        html = f"""
        <div class="card">
            <h2>{tr(request, "Edit Installment Schedule", "طھط¹ط¯ظٹظ„ ط·ط±ظٹظ‚ط© ط§ظ„ط³ط¯ط§ط¯")} â€” {escape(safe(advance['advance_no']))}{ob_badge}</h2>

            <div style="background:#f9fafb; border-radius:8px; padding:14px; margin-bottom:20px; border:1px solid #e5e7eb;">
                <div style="font-weight:600; color:#6b7280; margin-bottom:10px; font-size:13px;">
                    {tr(request, "Advance Info (Read Only)", "ط¨ظٹط§ظ†ط§طھ ط§ظ„ط³ظ„ظپط© â€” ظ„ظ„ط§ط·ظ„ط§ط¹ ظپظ‚ط·")}
                </div>
                <div style="display:grid; grid-template-columns: repeat(auto-fit, minmax(200px,1fr)); gap:12px;">
                    <div>
                        <div style="font-size:12px;color:#9ca3af;">{tr(request, "Advance No", "ط±ظ‚ظ… ط§ظ„ط³ظ„ظپط©")}</div>
                        <div style="font-weight:600;">{escape(safe(advance['advance_no']))}</div>
                    </div>
                    <div>
                        <div style="font-size:12px;color:#9ca3af;">{tr(request, "Date", "ط§ظ„طھط§ط±ظٹط®")}</div>
                        <div style="font-weight:600;">{escape(safe(advance['advance_date']))}</div>
                    </div>
                    <div>
                        <div style="font-size:12px;color:#9ca3af;">{tr(request, "Employee", "ط§ظ„ظ…ظˆط¸ظپ")}</div>
                        <div style="font-weight:600;">{escape(emp_display)}</div>
                    </div>
                    <div>
                        <div style="font-size:12px;color:#9ca3af;">{tr(request, "Advance Amount", "ظ‚ظٹظ…ط© ط§ظ„ط³ظ„ظپط©")}</div>
                        <div style="font-weight:600; color:#1d4ed8;">{money(advance['amount'])}</div>
                    </div>
                    {f'''<div>
                        <div style="font-size:12px;color:#9ca3af;">{tr(request, "Paid Before System", "ظ…ط³ط¯ط¯ ظ‚ط¨ظ„ ط§ظ„ظ†ط¸ط§ظ…")}</div>
                        <div style="font-weight:600; color:#7c3aed;">{money(paid_ob_val)}</div>
                    </div>''' if is_ob and paid_ob_val > 0 else ""}
                </div>
            </div>

            <form method="post" action="{with_lang(request, f'{BASE_ROUTE}/{advance_id}/edit')}">
                <input type="hidden" name="advance_date" value="{safe(advance['advance_date'])}">
                <input type="hidden" name="employee_id" value="{advance['employee_id']}">
                <input type="hidden" name="amount" value="{advance['amount']}">
                <input type="hidden" name="is_opening_balance" value="{'1' if is_ob else '0'}">
                <input type="hidden" name="paid_before_start" value="{paid_ob_val}">

                <div style="font-weight:600; color:#374151; margin-bottom:14px;">
                    {tr(request, "Payment Schedule Settings", "ط¥ط¹ط¯ط§ط¯ط§طھ ط¬ط¯ظˆظ„ ط§ظ„ط³ط¯ط§ط¯")}
                </div>

                <div class="row">
                    <div class="col">
                        <label>{tr(request, "Deduction Type", "نوع الخصم")}</label>
                        <select name="deduction_type" id="deductionType" onchange="toggleInstallmentAmount()">
                            <option value="direct" {direct_selected}>{tr(request, "Direct Salary Deduction", "خصم مباشر من المرتب")}</option>
                            <option value="installment" {installment_selected}>{tr(request, "Installments", "أقساط")}</option>
                        </select>
                    </div>
                    <div class="col" id="installmentAmountBox" style="display:{installment_display};">
                        <label>{tr(request, "Monthly Installment", "ط§ظ„ظ‚ط³ط· ط§ظ„ط´ظ‡ط±ظٹ")}</label>
                        <input type="number" step="0.01" min="0" name="installment_amount" id="installmentAmount" value="{installment_value}" {installment_required}>
                    </div>
                </div>

                <div class="row" style="margin-top:14px;">
                    <div class="col">
                        <label>{tr(request, "Deduction Start Month", "ط´ظ‡ط± ط¨ط¯ط§ظٹط© ط§ظ„ط®طµظ…")}</label>
                        <input type="number" min="1" max="12" name="start_month" value="{start_month_value}" required>
                    </div>
                </div>

                <div class="row" style="margin-top:14px;">
                    <div class="col">
                        <label>{tr(request, "Deduction Start Year", "ط³ظ†ط© ط¨ط¯ط§ظٹط© ط§ظ„ط®طµظ…")}</label>
                        <input type="number" min="2020" max="2100" name="start_year" value="{start_year_value}" required>
                    </div>
                    <div class="col">
                        <label>{tr(request, "Notes", "ظ…ظ„ط§ط­ط¸ط§طھ")}</label>
                        <input name="notes" value="{escape(safe(advance['notes']))}">
                    </div>
                </div>

                <div style="margin-top:18px;">
                    <button class="btn green" type="submit">{tr(request, "Update Schedule", "طھط­ط¯ظٹط« ط¬ط¯ظˆظ„ ط§ظ„ط³ط¯ط§ط¯")}</button>
                    <a class="btn gray" href="{with_lang(request, f'{BASE_ROUTE}/{advance_id}')}">{tr(request, "Back", "ط±ط¬ظˆط¹")}</a>
                </div>
            </form>
            <script>
                function toggleInstallmentAmount() {{
                    const type = document.getElementById('deductionType').value;
                    const box = document.getElementById('installmentAmountBox');
                    const input = document.getElementById('installmentAmount');
                    const isInstallment = type === 'installment';
                    box.style.display = isInstallment ? 'block' : 'none';
                    input.required = isInstallment;
                    if (!isInstallment) input.value = '';
                }}
                toggleInstallmentAmount();
            </script>
        </div>
        """
        return HTMLResponse(render_page(tr(request, "Edit Installment Schedule", "طھط¹ط¯ظٹظ„ ط·ط±ظٹظ‚ط© ط§ظ„ط³ط¯ط§ط¯"), html, lang, current_path=request.url.path))
    finally:
        conn.close()


@router.post(f"{BASE_ROUTE}/{{advance_id}}/edit")
def advances_update(
    request: Request,
    advance_id: int,
    advance_date: str = Form(""),
    employee_id: int = Form(...),
    amount: str = Form("0"),
    deduction_type: str = Form("installment"),
    installment_amount: str = Form("0"),
    start_month: str = Form(""),
    start_year: str = Form(""),
    notes: str = Form(""),
    is_opening_balance: str = Form(""),
    paid_before_start: str = Form("0"),
):
    conn = get_conn()
    advance = conn.execute("SELECT * FROM employee_advances WHERE id = ?", (advance_id,)).fetchone()
    if not advance or not can_edit_advance_schedule(dict(advance)):
        conn.close()
        return HTMLResponse(
            tr(
                request,
                "Cannot update this advance after it is linked to journal or closed/cancelled.",
                "ظ„ط§ ظٹظ…ظƒظ† طھط­ط¯ظٹط« ظ‡ط°ظ‡ ط§ظ„ط³ظ„ظپط© ط¨ط¹ط¯ ط±ط¨ط·ظ‡ط§ ط¨ط§ظ„ظ‚ظٹظˆط¯ ط£ظˆ ط¨ط¹ط¯ ط¥ظ‚ظپط§ظ„ظ‡ط§/ط¥ظ„ط؛ط§ط¦ظ‡ط§.",
            ),
            status_code=400,
        )

    is_ob = 1 if is_opening_balance == "1" else 0
    paid_ob = max(to_float(paid_before_start), 0) if is_ob else 0.0
    deduction_type = normalize_deduction_type(deduction_type)
    amount_value = to_float(advance["amount"] if advance["journal_line_id"] else amount)
    monthly_installment = amount_value if deduction_type == "direct" else max(to_float(installment_amount), 0)
    if amount_value <= 0:
        conn.close()
        return RedirectResponse(with_msg(request, f"{BASE_ROUTE}/{advance_id}/edit", tr(request, "Advance amount must be greater than zero.", "قيمة السلفة يجب أن تكون أكبر من صفر.")), status_code=302)
    if deduction_type == "installment" and monthly_installment <= 0:
        conn.close()
        return RedirectResponse(with_msg(request, f"{BASE_ROUTE}/{advance_id}/edit", tr(request, "Monthly installment is required for installment advances.", "قيمة القسط مطلوبة عند اختيار الأقساط.")), status_code=302)
    deduction_month, deduction_year = deduction_period_from_inputs(advance_date, start_month, start_year)

    conn.execute(
        """
        UPDATE employee_advances
        SET advance_date = ?, employee_id = ?, amount = ?, deduction_type = ?, installment_amount = ?,
            start_month = ?, start_year = ?, notes = ?, is_opening_balance = ?, paid_before_start = ?
        WHERE id = ?
        """,
        (
            safe(advance["advance_date"] if advance["journal_line_id"] else advance_date),
            int(advance["employee_id"] if advance["journal_line_id"] else employee_id),
            amount_value,
            deduction_type,
            monthly_installment,
            deduction_month,
            deduction_year,
            safe(notes),
            is_ob,
            paid_ob,
            advance_id
        ),
    )
    regenerate_installment_schedule(conn, advance_id)
    sync_advance_status(conn, advance_id)
    conn.commit()
    conn.close()
    msg = tr(request, "Advance updated successfully.", "طھظ… طھط­ط¯ظٹط« ط§ظ„ط³ظ„ظپط© ط¨ظ†ط¬ط§ط­.")
    return RedirectResponse(with_msg(request, BASE_ROUTE, msg), status_code=302)


@router.post(f"{BASE_ROUTE}/{{advance_id}}/delete")
def advances_delete(request: Request, advance_id: int):
    conn = get_conn()
    advance = conn.execute("SELECT id, status, journal_line_id FROM employee_advances WHERE id = ?", (advance_id,)).fetchone()
    if not advance or not can_delete_advance(conn, advance):
        conn.close()
        return HTMLResponse(
            tr(
                request,
                "Cannot delete this advance after it has posted journal or payroll deductions.",
                "لا يمكن حذف هذه السلفة بعد وجود قيد مرحل أو خصومات مرتبات عليها.",
            ),
            status_code=400,
        )

    conn.execute("DELETE FROM employee_advance_deferrals WHERE advance_id = ?", (advance_id,))
    conn.execute("DELETE FROM employee_advance_installments WHERE advance_id = ?", (advance_id,))
    conn.execute("DELETE FROM employee_advance_deductions WHERE advance_id = ?", (advance_id,))
    conn.execute("DELETE FROM employee_advances WHERE id = ?", (advance_id,))
    conn.commit()
    conn.close()
    msg = tr(request, "Advance deleted successfully.", "تم حذف السلفة بنجاح.")
    return RedirectResponse(with_msg(request, BASE_ROUTE, msg), status_code=302)


@router.get(f"{BASE_ROUTE}/statement", response_class=HTMLResponse)
def advance_statement_router(request: Request):
    from modules.accounting.employee_advances_statement import advance_statement_ui
    return advance_statement_ui(request)


@router.get(f"{BASE_ROUTE}/statement", response_class=HTMLResponse)
def advances_statement_route(request: Request):
    return advance_statement_ui(request)


@router.get(BASE_ROUTE, response_class=HTMLResponse)
@router.get(LEGACY_ROUTE, response_class=HTMLResponse)
def advances_list(request: Request):
    ensure_advances_tables()
    conn = get_conn()
    sync_all_advance_statuses(conn)
    conn.commit()
    rows = conn.execute(
        """
        SELECT ea.*, e.code AS employee_code, e.name AS employee_name,
               jl.line_description AS journal_employee_name
        FROM employee_advances ea
        LEFT JOIN employees e ON e.id = ea.employee_id
        LEFT JOIN journal_lines jl ON jl.id = ea.journal_line_id
        ORDER BY ea.id DESC
        """
    ).fetchall()
    conn.close()

    msg = safe(request.query_params.get("msg"))
    msg_html = f'<div class="msg success">{escape(msg)}</div>' if msg else ""

    body = ""
    for row in rows:
        local_conn = get_conn()
        paid = advance_paid_amount(local_conn, row["id"])
        paid_before_start = float(row["paid_before_start"] or 0)
        balance = max(float(row["amount"] or 0) - paid - paid_before_start, 0)
        displayed_paid = paid + paid_before_start
        can_delete = can_delete_advance(local_conn, row)
        local_conn.close()
        status_cls = "green" if safe(row["status"]).lower() == "closed" else "orange"
        employee_label = f"{safe(row['employee_code'])} - {safe(row['employee_name'])}" if safe(row["employee_code"]) else (safe(row["employee_name"]) or safe(row["journal_employee_name"]) or f"Employee #{safe(row['employee_id'])}")
        deduction_type_value = normalize_deduction_type(row["deduction_type"] if "deduction_type" in row.keys() else "")
        installment_cell = f"{money(row['installment_amount'])}<br><small>{escape(deduction_type_label(request, deduction_type_value))}</small>"
        body += f"""
        <tr>
            <td>{escape(safe(row['advance_no']))}</td>
            <td>{escape(safe(row['advance_date']))}</td>
            <td>
                <div style="font-weight:800;color:#0b2d5c;white-space:nowrap;">{escape(employee_label)}</div>
                <small>
                    <a href="{with_lang(request, f'{BASE_ROUTE}/statement?employee_id={row["employee_id"]}')}" style="color:blue;text-decoration:none;">
                        {tr(request, "View Statement", "ظƒط´ظپ ط­ط³ط§ط¨")}
                    </a>
                </small>
            </td>
            <td class="number-cell">{money(row['amount'])}</td>
            <td class="number-cell">{installment_cell}</td>
            <td class="number-cell">{money(displayed_paid)}</td>
            <td class="number-cell">{money(balance)}</td>
            <td><span class="status-chip {status_cls}">{escape(status_label(request, safe(row['status'])))}</span></td>
            <td style="white-space:nowrap;">
                <a class="btn orange" href="{with_lang(request, f'{BASE_ROUTE}/{row["id"]}/edit')}">{tr(request, "Distribute", "طھظˆط²ظٹط¹")}</a>
                <a class="btn blue" href="{with_lang(request, f'{BASE_ROUTE}/{row["id"]}')}">{tr(request, "Open", "ظپطھط­")}</a>
                <a class="btn green" href="{with_lang(request, f'{BASE_ROUTE}/{row["id"]}/installments')}">{tr(request, "Edit Installments", "طھط¹ط¯ظٹظ„ ط§ظ„ط£ظ‚ط³ط§ط·")}</a>
                {f'<form method="post" action="{with_lang(request, f"{BASE_ROUTE}/{row["id"]}/delete")}" style="display:inline;" onsubmit="return confirm(\'{tr(request, "Are you sure you want to delete this advance?", "هل أنت متأكد من حذف هذه السلفة؟")}\')"><button class="btn red" type="submit">{tr(request, "Delete", "حذف")}</button></form>' if can_delete else ""}
            </td>
        </tr>
        """

    if not body:
        body = f"<tr><td colspan='9' style='text-align:center;'>{tr(request, 'No employee advances found.', 'ظ„ط§ طھظˆط¬ط¯ ط³ظ„ظپ ظ…ظˆط¸ظپظٹظ† ظ…ط³ط¬ظ„ط©.')}</td></tr>"

    lang = get_lang(request)
    html = f"""
    <div class="card">
        {msg_html}
        <div style="display:flex;justify-content:space-between;align-items:center;gap:10px;flex-wrap:wrap;">
            <h2>{tr(request, "Employee Advances", "ط³ظ„ظپ ط§ظ„ظ…ظˆط¸ظپظٹظ†")}</h2>
            <div style="display:flex;gap:10px;">
                <a class="btn gray" href="{with_lang(request, f'{BASE_ROUTE}/statement')}">
                    {tr(request, "Advances Statement", "ظƒط´ظپ ط­ط³ط§ط¨ ط§ظ„ط³ظ„ظپ")}
                </a>
                <a class="btn blue" href="{with_lang(request, f'{BASE_ROUTE}/export/excel')}">
                    {tr(request, "Export Excel", "طھطµط¯ظٹط± ط¥ظƒط³ظ„")}
                </a>
                <a class="btn blue" href="{with_lang(request, f'{BASE_ROUTE}/sync-from-journal')}">
                    {tr(request, "Sync from Journal", "ظ…ط²ط§ظ…ظ†ط© ظ…ظ† ط§ظ„ظ‚ظٹظˆط¯")}
                </a>
                <a class="btn green" href="{with_lang(request, f'{BASE_ROUTE}/new')}">+ {tr(request, "New Advance", "ط³ظ„ظپط© ط¬ط¯ظٹط¯ط©")}</a>
            </div>
        </div>
        <p class="section-note">{tr(request, "Employee advances are tracked separately from custody and can be deducted automatically in payroll.", "ط³ظ„ظپ ط§ظ„ظ…ظˆط¸ظپظٹظ† طھظڈط³ط¬ظ„ ط¨ط´ظƒظ„ ظ…ط³طھظ‚ظ„ ط¹ظ† ط§ظ„ط¹ظ‡ط¯ط©طŒ ظˆظٹطھظ… ط®طµظ…ظ‡ط§ طھظ„ظ‚ط§ط¦ظٹظ‹ط§ ظ…ظ† ط§ظ„ظ…ط±طھط¨ط§طھ ط­ط³ط¨ ط§ظ„ط¬ط¯ظˆظ„ ط§ظ„ظ…ط³طھط­ظ‚.")}</p>
    </div>

    <div class="card">
        <table>
            <tr>
                <th>{tr(request, "Advance No", "ط±ظ‚ظ… ط§ظ„ط³ظ„ظپط©")}</th>
                <th>{tr(request, "Date", "ط§ظ„طھط§ط±ظٹط®")}</th>
                <th>{tr(request, "Employee", "ط§ظ„ظ…ظˆط¸ظپ")}</th>
                <th>{tr(request, "Total", "ط§ظ„ط¥ط¬ظ…ط§ظ„ظٹ")}</th>
                <th>{tr(request, "Deduction", "الخصم")}</th>
                <th>{tr(request, "Deducted", "ط§ظ„ظ…ط®طµظˆظ…")}</th>
                <th>{tr(request, "Balance", "ط§ظ„ط±طµظٹط¯")}</th>
                <th>{tr(request, "Status", "ط§ظ„ط­ط§ظ„ط©")}</th>
                <th>{tr(request, "Action", "ط§ظ„ط¥ط¬ط±ط§ط،")}</th>
            </tr>
            {body}
        </table>
    </div>
    """
    return HTMLResponse(render_page(tr(request, "Employee Advances", "ط³ظ„ظپ ط§ظ„ظ…ظˆط¸ظپظٹظ†"), html, lang, current_path=request.url.path))


@router.get(f"{LEGACY_ROUTE}/new")
def legacy_advances_new_redirect(request: Request):
    return RedirectResponse(with_lang(request, f"{BASE_ROUTE}/new"), status_code=302)


@router.get(f"{BASE_ROUTE}/export/excel")
def export_advances_excel(request: Request):
    if Workbook is None:
        return RedirectResponse(with_lang(request, BASE_ROUTE) + f"?msg={quote(tr(request, 'Excel export is not available. Install openpyxl.', 'طھطµط¯ظٹط± ط¥ظƒط³ظ„ ط؛ظٹط± ظ…طھط§ط­. ظ‚ظ… ط¨طھط«ط¨ظٹطھ openpyxl.'))}")
    
    ensure_advances_tables()
    conn = get_conn()
    
    try:
        rows = conn.execute("""
            SELECT 
                ea.id,
                ea.employee_id,
                e.code as employee_code,
                e.name as employee_name,
                ea.advance_no,
                ea.amount,
                ea.advance_date,
                ea.start_month,
                ea.start_year,
                ea.installment_amount,
                ea.total_installments,
                ea.paid_installments,
                ea.status,
                ea.notes
            FROM employee_advances ea
            LEFT JOIN employees e ON e.id = ea.employee_id
            ORDER BY ea.id DESC
        """).fetchall()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Employee Advances Data"
        
        headers = [
            'ID', 'Employee ID', 'Employee Code', 'Employee Name', 'Advance No',
            'Amount', 'Advance Date', 'Start Month', 'Start Year', 'Installment Amount',
            'Total Installments', 'Paid Installments', 'Status', 'Notes'
        ]
        
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)
        
        for row_num, adv in enumerate(rows, 2):
            ws.cell(row=row_num, column=1, value=adv[0])
            ws.cell(row=row_num, column=2, value=adv[1])
            ws.cell(row=row_num, column=3, value=adv[2])
            ws.cell(row=row_num, column=4, value=adv[3])
            ws.cell(row=row_num, column=5, value=adv[4])
            ws.cell(row=row_num, column=6, value=adv[5])
            ws.cell(row=row_num, column=7, value=adv[6])
            ws.cell(row=row_num, column=8, value=adv[7])
            ws.cell(row=row_num, column=9, value=adv[8])
            ws.cell(row=row_num, column=10, value=adv[9])
            ws.cell(row=row_num, column=11, value=adv[10])
            ws.cell(row=row_num, column=12, value=adv[11])
            ws.cell(row=row_num, column=13, value=adv[12])
            ws.cell(row=row_num, column=14, value=adv[13])
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"employee_advances_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return StreamingResponse(
            io.BytesIO(output.getvalue()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        return RedirectResponse(with_lang(request, BASE_ROUTE) + f"?msg={quote(tr(request, f'Export error: {str(e)}', f'ط®ط·ط£ ظپظٹ ط§ظ„طھطµط¯ظٹط±: {str(e)}'))}")
    finally:
        conn.close()


@router.get(f"{BASE_ROUTE}/new", response_class=HTMLResponse)
def advances_new(request: Request):
    lang = get_lang(request)
    today = datetime.now()
    html = f"""
    <div class="card">
        <h2>{tr(request, "New Employee Advance", "ط³ظ„ظپط© ظ…ظˆط¸ظپ ط¬ط¯ظٹط¯ط©")}</h2>
        <form method="post" action="{with_lang(request, f'{BASE_ROUTE}/new')}">
            <div class="row">
                <div class="col">
                    <label>{tr(request, "Advance No", "ط±ظ‚ظ… ط§ظ„ط³ظ„ظپط©")}</label>
                    <input name="advance_no" value="{next_advance_no()}" readonly>
                </div>
                <div class="col">
                    <label>{tr(request, "Advance Date", "طھط§ط±ظٹط® ط§ظ„ط³ظ„ظپط©")}</label>
                    <input type="date" name="advance_date" required>
                </div>
            </div>

            <div class="row" style="margin-top:14px;">
                <div class="col">
                    <label>{tr(request, "Employee", "ط§ظ„ظ…ظˆط¸ظپ")}</label>
                    <select name="employee_id" required>
                        {employee_options_html()}
                    </select>
                </div>
                <div class="col">
                    <label>{tr(request, "Advance Amount", "ظ‚ظٹظ…ط© ط§ظ„ط³ظ„ظپط©")}</label>
                    <input type="number" step="0.01" min="0" name="amount" required>
                </div>
            </div>

            <div class="row" style="margin-top:14px;">
                <div class="col">
                    <label>{tr(request, "Deduction Type", "نوع الخصم")}</label>
                    <select name="deduction_type" id="deductionType" onchange="toggleInstallmentAmount()">
                        <option value="direct" selected>{tr(request, "Direct Salary Deduction", "خصم مباشر من المرتب")}</option>
                        <option value="installment">{tr(request, "Installments", "أقساط")}</option>
                    </select>
                </div>
                <div class="col" id="installmentAmountBox" style="display:none;">
                    <label>{tr(request, "Monthly Installment", "ط§ظ„ظ‚ط³ط· ط§ظ„ط´ظ‡ط±ظٹ")}</label>
                    <input type="number" step="0.01" min="0" name="installment_amount" id="installmentAmount">
                </div>
            </div>

            <div class="row" style="margin-top:14px;">
                <div class="col">
                    <label>{tr(request, "Deduction Start Month", "ط´ظ‡ط± ط¨ط¯ط§ظٹط© ط§ظ„ط®طµظ…")}</label>
                    <input type="number" min="1" max="12" name="start_month" value="{today.month}">
                </div>
                <div class="col">
                    <label>{tr(request, "Deduction Start Year", "ط³ظ†ط© ط¨ط¯ط§ظٹط© ط§ظ„ط®طµظ…")}</label>
                    <input type="number" min="2020" max="2100" name="start_year" value="{today.year}">
                </div>
            </div>

            <div class="row" style="margin-top:14px;">
                <div class="col">
                    <label>{tr(request, "Notes", "ظ…ظ„ط§ط­ط¸ط§طھ")}</label>
                    <input name="notes">
                </div>
            </div>

            <div style="margin-top:18px; padding:14px; background:#f0f9ff; border-radius:8px; border:1px solid #bae6fd;">
                <div style="font-weight:600; margin-bottom:10px; color:#0369a1;">
                    {tr(request, "Opening Balance Options", "ط®ظٹط§ط±ط§طھ ط§ظ„ط±طµظٹط¯ ط§ظ„ط§ظپطھطھط§ط­ظٹ")}
                </div>
                <div class="row">
                    <div class="col">
                        <label style="display:flex;align-items:center;gap:8px;cursor:pointer;">
                            <input type="checkbox" name="is_opening_balance" value="1" id="isOpeningBalance"
                                onchange="document.getElementById('openingBalanceFields').style.display=this.checked?'block':'none'">
                            {tr(request, "This is an Opening Balance Advance", "ظ‡ط°ظ‡ ط§ظ„ط³ظ„ظپط© ظ…ظ† ط§ظ„ط±طµظٹط¯ ط§ظ„ط§ظپطھطھط§ط­ظٹ")}
                        </label>
                        <small style="color:#6b7280;">
                            {tr(request, "Use this for advances that existed before the system start date.", "ط§ط³طھط®ط¯ظ… ظ‡ط°ط§ ط§ظ„ط®ظٹط§ط± ظ„ظ„ط³ظ„ظپ ط§ظ„طھظٹ ظƒط§ظ†طھ ظ…ظˆط¬ظˆط¯ط© ظ‚ط¨ظ„ ط¨ط¯ط، ط§ط³طھط®ط¯ط§ظ… ط§ظ„ظ†ط¸ط§ظ….")}
                        </small>
                    </div>
                </div>
                <div id="openingBalanceFields" style="display:none; margin-top:12px;">
                    <div class="row">
                        <div class="col">
                            <label>{tr(request, "Amount Already Paid (Before System Start)", "ط§ظ„ظ…ط¨ظ„ط؛ ط§ظ„ظ…ط³ط¯ط¯ ظ‚ط¨ظ„ ط¨ط¯ط، ط§ظ„ظ†ط¸ط§ظ…")}</label>
                            <input type="number" step="0.01" min="0" name="paid_before_start" value="0">
                            <small style="color:#6b7280;">
                                {tr(request, "This amount will be deducted from the total to calculate remaining installments.", "ظ‡ط°ط§ ط§ظ„ظ…ط¨ظ„ط؛ ط³ظٹظڈط­ط³ظ… ظ…ظ† ط§ظ„ط¥ط¬ظ…ط§ظ„ظٹ ظ„ط­ط³ط§ط¨ ط§ظ„ط£ظ‚ط³ط§ط· ط§ظ„ظ…طھط¨ظ‚ظٹط©.")}
                            </small>
                        </div>
                    </div>
                </div>
            </div>

            <div style="margin-top:18px;">
                <button class="btn green" type="submit">{tr(request, "Save Advance", "ط­ظپط¸ ط§ظ„ط³ظ„ظپط©")}</button>
                <a class="btn gray" href="{with_lang(request, BASE_ROUTE)}">{tr(request, "Back", "ط±ط¬ظˆط¹")}</a>
            </div>
        </form>
        <script>
            function toggleInstallmentAmount() {{
                const type = document.getElementById('deductionType').value;
                const box = document.getElementById('installmentAmountBox');
                const input = document.getElementById('installmentAmount');
                const isInstallment = type === 'installment';
                box.style.display = isInstallment ? 'block' : 'none';
                input.required = isInstallment;
                if (!isInstallment) input.value = '';
            }}
            toggleInstallmentAmount();
        </script>
    </div>
    """
    return HTMLResponse(render_page(tr(request, "New Employee Advance", "ط³ظ„ظپط© ظ…ظˆط¸ظپ ط¬ط¯ظٹط¯ط©"), html, lang, current_path=request.url.path))


@router.post(f"{LEGACY_ROUTE}/new")
@router.post(f"{BASE_ROUTE}/new")
def advances_create(
    request: Request,
    advance_no: str = Form(""),
    advance_date: str = Form(""),
    employee_id: int = Form(...),
    amount: str = Form("0"),
    deduction_type: str = Form("installment"),
    installment_amount: str = Form("0"),
    start_month: str = Form(""),
    start_year: str = Form(""),
    notes: str = Form(""),
    is_opening_balance: str = Form(""),
    paid_before_start: str = Form("0"),
):
    ensure_advances_tables()
    advance_total = to_float(amount)
    deduction_type = normalize_deduction_type(deduction_type)
    monthly_installment = advance_total if deduction_type == "direct" else max(to_float(installment_amount), 0)
    if advance_total <= 0:
        return RedirectResponse(with_msg(request, BASE_ROUTE, tr(request, "Advance amount must be greater than zero.", "ظ‚ظٹظ…ط© ط§ظ„ط³ظ„ظپط© ظٹط¬ط¨ ط£ظ† طھظƒظˆظ† ط£ظƒط¨ط± ظ…ظ† طµظپط±.")), status_code=302)
    if deduction_type == "installment" and monthly_installment <= 0:
        return RedirectResponse(with_msg(request, BASE_ROUTE, tr(request, "Monthly installment is required for installment advances.", "قيمة القسط مطلوبة عند اختيار الأقساط.")), status_code=302)
    deduction_month, deduction_year = deduction_period_from_inputs(advance_date, start_month, start_year)

    is_ob = 1 if is_opening_balance == "1" else 0
    paid_ob = max(to_float(paid_before_start), 0) if is_ob else 0.0

    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        """
        INSERT INTO employee_advances (
            advance_no, advance_date, employee_id, amount, deduction_type, installment_amount,
            start_month, start_year, notes, status, is_opening_balance, paid_before_start
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'active', ?, ?)
        """,
        (
            safe(advance_no) or next_advance_no(),
            safe(advance_date),
            int(employee_id),
            advance_total,
            deduction_type,
            monthly_installment,
            deduction_month,
            deduction_year,
            safe(notes),
            is_ob,
            paid_ob,
        ),
    )
    new_id = cur.lastrowid
    regenerate_installment_schedule(conn, new_id)
    conn.commit()
    conn.close()
    return RedirectResponse(with_msg(request, BASE_ROUTE, tr(request, "Employee advance created successfully.", "طھظ… ط¥ظ†ط´ط§ط، ط³ظ„ظپط© ط§ظ„ظ…ظˆط¸ظپ ط¨ظ†ط¬ط§ط­.")), status_code=302)


@router.get(f"{LEGACY_ROUTE}" + "/{advance_id}")
def legacy_advance_open_redirect(request: Request, advance_id: int):
    return RedirectResponse(with_lang(request, f"{BASE_ROUTE}/{advance_id}"), status_code=302)


@router.get(f"{BASE_ROUTE}" + "/{advance_id}/installments", response_class=HTMLResponse)
def advance_installments_edit(request: Request, advance_id: int):
    ensure_advances_tables()
    conn = get_conn()
    
    # Get advance details
    advance = conn.execute("""
        SELECT ea.*, e.code as employee_code, e.name as employee_name,
               jl.line_description AS journal_employee_name
        FROM employee_advances ea
        LEFT JOIN employees e ON e.id = ea.employee_id
        LEFT JOIN journal_lines jl ON jl.id = ea.journal_line_id
        WHERE ea.id = ? LIMIT 1
    """, (advance_id,)).fetchone()
    
    if not advance:
        return HTMLResponse("Advance not found", status_code=404)
    
    # Get installments
    installments = conn.execute("""
        SELECT ei.*,
               CASE
                   WHEN LOWER(COALESCE(ei.status, '')) = 'paid'
                        OR COALESCE(ei.paid_amount, 0) > 0 THEN 'paid'
                   WHEN LOWER(COALESCE(ei.status, '')) = 'deferred' THEN 'deferred'
                   ELSE 'pending'
               END as display_status
        FROM employee_advance_installments ei
        WHERE ei.advance_id = ?
        ORDER BY ei.installment_year, ei.installment_month
    """, (advance_id,)).fetchall()
    
    conn.close()
    
    # Build installments table
    body = ""
    for inst in installments:
        status_cls = {
            'pending': 'orange',
            'paid': 'green',
            'deferred': 'red'
        }.get(inst['display_status'], 'gray')
        
        status_label_map = {
            'pending': tr(request, "Pending", "ظ…ط¹ظ„ظ‚"),
            'paid': tr(request, "Paid", "ظ…ط¯ظپظˆط¹"),
            'deferred': tr(request, "Deferred", "ظ…ط¤ط¬ظ„")
        }
        defer_action = ""
        is_due_installment = (
            period_key(inst["installment_month"], inst["installment_year"]) <= current_period_key()
        )
        if inst["display_status"] == "pending" and is_due_installment:
            defer_action = f"""
                <div style="display:flex;gap:6px;align-items:center;flex-wrap:wrap;">
                    <input name="defer_reason_{inst['id']}" placeholder="{tr(request, 'Reason', 'السبب')}"
                           style="padding:8px;border:1px solid #d6e0ef;border-radius:8px;width:140px;">
                    <button class="btn orange" type="submit"
                            name="defer_installment_id" value="{inst['id']}">
                        {tr(request, "Defer to End", "تأجيل لنهاية المدة")}
                    </button>
                </div>
            """
        
        body += f"""
        <tr>
            <td>
                <div style="display:flex; gap:8px; align-items:center; flex-wrap:wrap;">
                    <input type="number" min="1" max="12" name="installment_month_{inst['id']}"
                           value="{safe(inst['installment_month'] or '')}" class="form-input"
                           style="width:86px;" title="{tr(request, 'Month', 'ط§ظ„ط´ظ‡ط±')}">
                    <input type="number" min="2020" max="2100" name="installment_year_{inst['id']}"
                           value="{safe(inst['installment_year'] or '')}" class="form-input"
                           style="width:110px;" title="{tr(request, 'Year', 'ط§ظ„ط³ظ†ط©')}">
                </div>
            </td>
            <td class="number-cell">
                <input type="number" step="0.01" name="planned_amount_{inst['id']}" 
                       value="{safe(inst['planned_amount'] or '0')}" class="form-input">
            </td>
            <td class="number-cell">{money(inst['paid_amount'])}</td>
            <td>
                <select name="status_{inst['id']}" class="form-select">
                    <option value="pending" {'selected' if inst['display_status'] == 'pending' else ''}>
                        {tr(request, "Pending", "ظ…ط¹ظ„ظ‚")}
                    </option>
                    {'<option value="deferred" selected disabled>' + tr(request, "Deferred", "ظ…ط¤ط¬ظ„") + '</option>' if inst['display_status'] == 'deferred' else ''}
                    <option value="paid" {'selected' if inst['display_status'] == 'paid' else ''}>
                        {tr(request, "Paid", "ظ…ط¯ظپظˆط¹")}
                    </option>
                </select>
            </td>
            <td>{escape(safe(inst['is_deferred'] or '0'))}</td>
            <td>{escape(safe(inst['deferred_reason'] or ''))}</td>
            <td>{defer_action}</td>
        </tr>
        """
    
    if not body:
        body = f"<tr><td colspan='7' style='text-align:center;'>{tr(request, 'No installments found.', 'ظ„ط§ طھظˆط¬ط¯ ط£ظ‚ط³ط§ط· ظ…ط³ط¬ظ„ط©.')}</td></tr>"
    
    setup_html = ""
    if not installments:
        default_month = int(advance["start_month"] or datetime.now().month)
        default_year = int(advance["start_year"] or datetime.now().year)
        default_installment = max(float(advance["installment_amount"] or 0), 0)
        setup_html = f"""
        <div class="card">
            <h3>{tr(request, "Distribute Advance", "طھظˆط²ظٹط¹ ط§ظ„ط³ظ„ظپط©")}</h3>
            <p style="color:#6b7280;margin-bottom:16px;">
                {tr(request, "Set the monthly installment and deduction start period to generate the schedule.", "ط­ط¯ط¯ ظ‚ظٹظ…ط© ط§ظ„ظ‚ط³ط· ط§ظ„ط´ظ‡ط±ظٹ ظˆط´ظ‡ط± ط¨ط¯ط§ظٹط© ط§ظ„ط®طµظ… ظ„ط¥ظ†ط´ط§ط، ط¬ط¯ظˆظ„ ط§ظ„ط£ظ‚ط³ط§ط·.")}
            </p>
            <form method="post" action="{with_lang(request, f'{BASE_ROUTE}/{advance_id}/edit')}">
                <input type="hidden" name="advance_date" value="{safe(advance['advance_date'])}">
                <input type="hidden" name="employee_id" value="{safe(advance['employee_id'])}">
                <input type="hidden" name="amount" value="{safe(advance['amount'])}">
                <input type="hidden" name="is_opening_balance" value="{safe(advance['is_opening_balance'] if 'is_opening_balance' in advance.keys() else '0')}">
                <input type="hidden" name="paid_before_start" value="{safe(advance['paid_before_start'] if 'paid_before_start' in advance.keys() else '0')}">
                <div class="row">
                    <div class="col">
                        <label>{tr(request, "Monthly Installment", "ط§ظ„ظ‚ط³ط· ط§ظ„ط´ظ‡ط±ظٹ")}</label>
                        <input type="number" step="0.01" min="0.01" name="installment_amount" value="{default_installment}" required>
                    </div>
                    <div class="col">
                        <label>{tr(request, "Start Month", "ط´ظ‡ط± ط§ظ„ط¨ط¯ط§ظٹط©")}</label>
                        <input type="number" min="1" max="12" name="start_month" value="{default_month}" required>
                    </div>
                    <div class="col">
                        <label>{tr(request, "Start Year", "ط³ظ†ط© ط§ظ„ط¨ط¯ط§ظٹط©")}</label>
                        <input type="number" min="2020" max="2100" name="start_year" value="{default_year}" required>
                    </div>
                    <div class="col">
                        <label>{tr(request, "Notes", "ظ…ظ„ط§ط­ط¸ط§طھ")}</label>
                        <input name="notes" value="{escape(safe(advance['notes']))}">
                    </div>
                </div>
                <div style="margin-top:18px;">
                    <button class="btn green" type="submit">{tr(request, "Generate Installments", "ط¥ظ†ط´ط§ط، ط§ظ„ط£ظ‚ط³ط§ط·")}</button>
                </div>
            </form>
        </div>
        """

    lang = get_lang(request)
    html = f"""
    <div class="card">
        <h3>{tr(request, "Edit Installments", "طھط¹ط¯ظٹظ„ ط§ظ„ط£ظ‚ط³ط§ط·")} - {escape(safe(advance['advance_no']))}</h3>
        <p><strong>{tr(request, "Employee", "ط§ظ„ظ…ظˆط¸ظپ")}:</strong> {escape((f"{safe(advance['employee_code'])} - {safe(advance['employee_name'])}" if safe(advance['employee_code']) else (safe(advance['employee_name']) or safe(advance['journal_employee_name']) or f"Employee #{safe(advance['employee_id'])}")))}</p>
        <p><strong>{tr(request, "Total Amount", "ط§ظ„ط¥ط¬ظ…ط§ظ„ظٹ")}:</strong> {money(advance['amount'])} | <strong>{tr(request, "Installment Amount", "ظ‚ط³ط·")}:</strong> {money(advance['installment_amount'])}</p>
    </div>
    
    {setup_html}

    <div class="card">
        <form method="post" action="{with_lang(request, f'{BASE_ROUTE}/{advance_id}/installments/update')}">
            <table>
                <tr>
                    <th>{tr(request, "Period", "ط§ظ„ظپطھط±ط©")}</th>
                    <th>{tr(request, "Planned Amount", "ط§ظ„ظ…ط¨ظ„ط؛ ط§ظ„ظ…ط®ط·ط·")}</th>
                    <th>{tr(request, "Paid Amount", "ط§ظ„ظ…ط¨ظ„ط؛ ط§ظ„ظ…ط¯ظپظˆط¹")}</th>
                    <th>{tr(request, "Status", "ط§ظ„ط­ط§ظ„ط©")}</th>
                    <th>{tr(request, "Deferred", "ظ…ط¤ط¬ظ„")}</th>
                    <th>{tr(request, "Reason", "ط§ظ„ط³ط¨ط¨")}</th>
                    <th>{tr(request, "Action", "الإجراء")}</th>
                </tr>
                {body}
            </table>
            <div style="margin-top:20px;">
                <button type="submit" class="btn green">{tr(request, "Update Installments", "طھط­ط¯ظٹط« ط§ظ„ط£ظ‚ط³ط§ط·")}</button>
                <a href="{with_lang(request, f'{BASE_ROUTE}/{advance_id}')}" class="btn gray">{tr(request, "Back", "ط±ط¬ظˆط¹")}</a>
            </div>
        </form>
    </div>
    """
    
    return HTMLResponse(render_page(tr(request, "Edit Installments", "طھط¹ط¯ظٹظ„ ط§ظ„ط£ظ‚ط³ط§ط·"), html, lang, current_path=request.url.path))


@router.post(f"{BASE_ROUTE}" + "/{advance_id}/installments/update")
async def advance_installments_update(request: Request, advance_id: int):
    ensure_advances_tables()
    conn = get_conn()
    
    try:
        # Get form data
        form_data = await request.form()
        defer_installment_id = int(to_float(form_data.get("defer_installment_id")) or 0)
        if defer_installment_id > 0:
            reason = safe(form_data.get(f"defer_reason_{defer_installment_id}"))
            ok, msg = defer_installment(conn, advance_id, defer_installment_id, reason)
            if not ok:
                raise ValueError(msg)
            conn.commit()
            return RedirectResponse(with_msg(request, f"{BASE_ROUTE}/{advance_id}/installments", msg), status_code=302)
        
        # Get all installments for this advance
        installments = conn.execute("""
            SELECT id FROM employee_advance_installments 
            WHERE advance_id = ?
        """, (advance_id,)).fetchall()
        
        updated_count = 0
        for inst in installments:
            inst_id = inst['id']
            
            # Get form values
            planned_amount = to_float(form_data.get(f"planned_amount_{inst_id}"))
            installment_month = int(to_float(form_data.get(f"installment_month_{inst_id}")) or 0)
            installment_year = int(to_float(form_data.get(f"installment_year_{inst_id}")) or 0)
            current = conn.execute(
                """
                SELECT status, deferred_reason
                FROM employee_advance_installments
                WHERE id = ? AND advance_id = ?
                """,
                (inst_id, advance_id),
            ).fetchone()
            current_status = safe(current["status"] if current else "pending").lower()
            current_reason = safe(current["deferred_reason"] if current else "")
            new_status = safe(form_data.get(f"status_{inst_id}")).lower()
            if new_status not in ("pending", "paid"):
                new_status = current_status if current_status in ("pending", "paid", "deferred") else "pending"

            if installment_month < 1 or installment_month > 12:
                raise ValueError("Installment month must be between 1 and 12.")
            if installment_year < 2020 or installment_year > 2100:
                raise ValueError("Installment year must be between 2020 and 2100.")
            
            paid_amount = 0.0
            is_deferred = 0
            deferred_reason = None
            if new_status == "paid":
                paid_amount = planned_amount
            elif new_status == "deferred":
                is_deferred = 1
                deferred_reason = current_reason or None

            # Update installment
            conn.execute("""
                UPDATE employee_advance_installments 
                SET installment_month = ?,
                    installment_year = ?,
                    planned_amount = ?,
                    paid_amount = ?,
                    status = ?,
                    is_deferred = ?,
                    deferred_reason = ?
                WHERE id = ?
            """, (
                installment_month,
                installment_year,
                planned_amount,
                paid_amount,
                new_status,
                is_deferred,
                deferred_reason,
                inst_id,
            ))
            
            updated_count += 1
        
        sync_advance_status(conn, advance_id)
        conn.commit()
        msg = f"Updated {updated_count} installments successfully"
        
    except Exception as e:
        conn.rollback()
        msg = f"Error updating installments: {str(e)}"
    finally:
        conn.close()
    
    return RedirectResponse(with_msg(request, f"{BASE_ROUTE}/{advance_id}/installments", msg), status_code=302)


@router.get(f"{BASE_ROUTE}" + "/{advance_id}", response_class=HTMLResponse)
def advance_open(request: Request, advance_id: int):
    ensure_advances_tables()
    conn = get_conn()
    advance = conn.execute(
        """
        SELECT ea.*, e.code AS employee_code, e.name AS employee_name,
               jl.line_description AS journal_employee_name
        FROM employee_advances ea
        LEFT JOIN employees e ON e.id = ea.employee_id
        LEFT JOIN journal_lines jl ON jl.id = ea.journal_line_id
        WHERE ea.id = ?
        LIMIT 1
        """,
        (advance_id,),
    ).fetchone()
    if not advance:
        conn.close()
        return HTMLResponse(tr(request, "Employee advance not found", "ط³ظ„ظپط© ط§ظ„ظ…ظˆط¸ظپ ط؛ظٹط± ظ…ظˆط¬ظˆط¯ط©"), status_code=404)

    deductions = conn.execute(
        """
        SELECT ead.*, pr.payroll_no, pr.payroll_month, pr.payroll_year
        FROM employee_advance_deductions ead
        LEFT JOIN payroll_runs pr ON pr.id = ead.payroll_run_id
        WHERE ead.advance_id = ?
        ORDER BY ead.id DESC
        """,
        (advance_id,),
    ).fetchall()

    # Installment schedule
    installments = conn.execute(
        """
        SELECT * FROM employee_advance_installments
        WHERE advance_id = ?
        ORDER BY installment_year, installment_month
        """,
        (advance_id,),
    ).fetchall()

    # Deferral history
    deferrals = conn.execute(
        """
        SELECT * FROM employee_advance_deferrals WHERE advance_id = ?
        ORDER BY id DESC
        """,
        (advance_id,),
    ).fetchall()

    paid = advance_paid_amount(conn, advance_id)
    paid_before_start = float(advance["paid_before_start"] or 0)
    balance = max(float(advance["amount"] or 0) - paid - paid_before_start, 0)
    is_ob = int(advance["is_opening_balance"] or 0)
    can_modify = can_delete_advance(conn, advance)
    conn.close()

    deduction_rows = ""
    for row in deductions:
        deduction_rows += f"""
        <tr>
            <td>{escape(safe(row['payroll_no']))}</td>
            <td>{int(row['deduction_month'] or 0):02d}/{row['deduction_year'] or ''}</td>
            <td class="number-cell">{money(row['amount'])}</td>
        </tr>
        """

    if not deduction_rows:
        deduction_rows = f"<tr><td colspan='3' style='text-align:center;'>{tr(request, 'No payroll deductions recorded yet.', 'ظ„ط§ طھظˆط¬ط¯ ط®طµظˆظ…ط§طھ ظ…ط±طھط¨ط§طھ ظ…ط³ط¬ظ„ط© ط¹ظ„ظ‰ ظ‡ط°ظ‡ ط§ظ„ط³ظ„ظپط© ط­طھظ‰ ط§ظ„ط¢ظ†.')}</td></tr>"

    # Build installment schedule rows
    inst_rows = ""
    is_active = safe(advance["status"]).lower() in ("active", "open")
    for inst in installments:
        st = safe(inst["status"]).lower()
        if st == "paid":
            st_badge = f'<span class="status-chip green">{tr(request, "Paid", "ظ…ط¯ظپظˆط¹")}</span>'
            action_cell = ""
        elif st == "deferred":
            reason_txt = escape(safe(inst["deferred_reason"]))
            st_badge = f'<span class="status-chip orange">{tr(request, "Deferred", "ظ…ط¤ط¬ظ„")}</span>'
            action_cell = f'<small style="color:#9ca3af;">{reason_txt}</small>'
        else:
            st_badge = f'<span class="status-chip blue">{tr(request, "Pending", "ظ‚ظٹط¯ ط§ظ„ط§ظ†طھط¸ط§ط±")}</span>'
            deferred_label = tr(request, "Defer to End", "تأجيل لنهاية المدة")
            is_due_installment = (
                period_key(inst["installment_month"], inst["installment_year"]) <= current_period_key()
            )
            if is_active and is_due_installment:
                action_cell = f"""
                <form method="post" action="{with_lang(request, f'{BASE_ROUTE}/{advance_id}/defer-installment')}" style="display:inline-flex;gap:6px;align-items:center;">
                    <input type="hidden" name="installment_id" value="{inst['id']}">
                    <input name="reason" placeholder="{tr(request, 'Reason', 'ط§ظ„ط³ط¨ط¨')}" style="padding:3px 6px;border:1px solid #ddd;border-radius:4px;font-size:12px;width:120px;">
                    <button class="btn orange" type="submit" style="padding:2px 8px;font-size:12px;">{deferred_label}</button>
                </form>
                """
            else:
                action_cell = ""

        deferred_icon = " ًں”پ" if int(inst["is_deferred"] or 0) and st != "deferred" else ""
        inst_rows += f"""
        <tr>
            <td>{int(inst['installment_month'] or 0):02d}/{inst['installment_year'] or ''}</td>
            <td class="number-cell">{money(inst['planned_amount'])}{deferred_icon}</td>
            <td>{st_badge}</td>
            <td>{action_cell}</td>
        </tr>
        """

    if not inst_rows:
        inst_rows = f"<tr><td colspan='4' style='text-align:center;'>{tr(request, 'No installment schedule generated.', 'ظ„ط§ ظٹظˆط¬ط¯ ط¬ط¯ظˆظ„ ط£ظ‚ط³ط§ط·.')}</td></tr>"

    # Deferral history rows
    deferral_rows = ""
    for d in deferrals:
        deferral_rows += f"""
        <tr>
            <td>{int(d['original_month'] or 0):02d}/{d['original_year'] or ''}</td>
            <td>{int(d['deferred_to_month'] or 0):02d}/{d['deferred_to_year'] or ''}</td>
            <td class="number-cell">{money(d['amount'])}</td>
            <td>{escape(safe(d['reason']))}</td>
        </tr>
        """

    opening_balance_badge = ""
    if is_ob:
        opening_balance_badge = f' <span class="status-chip blue" style="font-size:11px;">{tr(request, "Opening Balance", "ط±طµظٹط¯ ط§ظپطھطھط§ط­ظٹ")}</span>'

    opening_balance_kpi = ""
    if is_ob and paid_before_start > 0:
        opening_balance_kpi = f"""
        <div class="kpi-card">
            <div class="kpi-label">{tr(request, "Paid Before System", "ظ…ط¯ظپظˆط¹ ظ‚ط¨ظ„ ط§ظ„ظ†ط¸ط§ظ…")}</div>
            <div class="kpi-value" style="color:#7c3aed;">{money(paid_before_start)}</div>
        </div>
        """

    employee_label = f"{safe(advance['employee_code'])} - {safe(advance['employee_name'])}" if safe(advance["employee_code"]) else (safe(advance["employee_name"]) or safe(advance["journal_employee_name"]) or f"Employee #{safe(advance['employee_id'])}")
    status_cls = "green" if safe(advance["status"]).lower() == "closed" else "orange"
    can_edit_schedule = can_edit_advance_schedule(dict(advance))
    lang = get_lang(request)
    html = f"""
    <div class="card">
        <div style="display:flex;justify-content:space-between;align-items:flex-start;gap:16px;flex-wrap:wrap;">
            <div>
                <h2>{tr(request, "Employee Advance", "ط³ظ„ظپط© ط§ظ„ظ…ظˆط¸ظپ")} {escape(safe(advance['advance_no']))}{opening_balance_badge}</h2>
                <p><b>{tr(request, "Employee:", "ط§ظ„ظ…ظˆط¸ظپ:")}</b> {escape(employee_label)}</p>
                <p><b>{tr(request, "Date:", "ط§ظ„طھط§ط±ظٹط®:")}</b> {escape(safe(advance['advance_date']))}</p>
                <p><b>{tr(request, "Deduction Start:", "ط¨ط¯ط§ظٹط© ط§ظ„ط®طµظ…:")}</b> {int(advance['start_month'] or 0):02d}/{advance['start_year'] or ''}</p>
                <p><b>{tr(request, "Status:", "ط§ظ„ط­ط§ظ„ط©:")}</b> <span class="status-chip {status_cls}">{escape(status_label(request, safe(advance['status'])))}</span></p>
                <p><b>{tr(request, "Notes:", "ظ…ظ„ط§ط­ط¸ط§طھ:")}</b> {escape(safe(advance['notes']))}</p>
            </div>
            <div class="kpi-grid" style="min-width:280px;">
                <div class="kpi-card">
                    <div class="kpi-label">{tr(request, "Advance Total", "ط¥ط¬ظ…ط§ظ„ظٹ ط§ظ„ط³ظ„ظپط©")}</div>
                    <div class="kpi-value">{money(advance['amount'])}</div>
                </div>
                <div class="kpi-card">
                    <div class="kpi-label">{tr(request, "Installment", "ط§ظ„ظ‚ط³ط·")}</div>
                    <div class="kpi-value">{money(advance['installment_amount'])}</div>
                </div>
                {opening_balance_kpi}
                <div class="kpi-card">
                    <div class="kpi-label">{tr(request, "Deducted", "ط§ظ„ظ…ط®طµظˆظ…")}</div>
                    <div class="kpi-value">{money(paid)}</div>
                </div>
                <div class="kpi-card">
                    <div class="kpi-label">{tr(request, "Balance", "ط§ظ„ط±طµظٹط¯")}</div>
                    <div class="kpi-value">{money(balance)}</div>
                </div>
            </div>
        </div>
        <div style="margin-top:16px; display:flex; gap:10px;">
            <a class="btn gray" href="{with_lang(request, BASE_ROUTE)}">{tr(request, "Back", "ط±ط¬ظˆط¹")}</a>
            {f'<a class="btn green" href="/ui/accounting/cash-payments/new?party_type=employee&employee_id={advance["employee_id"]}&employee_trans_type=advance&advance_id={advance["id"]}&amount={advance["amount"]}">{tr(request, "Disburse Advance", "طµط±ظپ ط§ظ„ط³ظ„ظپط©")}</a>' if safe(advance["status"]).lower() == "active" else ""}
            {f'<a class="btn blue" href="/ui/accounting/cash-payments/new?party_type=employee&employee_id={advance["employee_id"]}&employee_trans_type=advance&advance_id={advance["id"]}&amount={advance["amount"]}&expense_payment_source=custody">{tr(request, "Disburse From Custody", "صرف من عهدة")}</a>' if safe(advance["status"]).lower() == "active" else ""}
            {f'<a class="btn orange" href="{with_lang(request, f"{BASE_ROUTE}/{advance["id"]}/edit")}">{tr(request, "Edit Schedule", "طھظˆط²ظٹط¹ ط§ظ„ط³ظ„ظپط©")}</a>' if can_edit_schedule else ""}
            {f'<form method="post" action="{with_lang(request, f"{BASE_ROUTE}/{advance["id"]}/delete")}" style="display:inline;" onsubmit="return confirm(\'{tr(request, "Are you sure you want to delete this advance?", "ظ‡ظ„ ط£ظ†طھ ظ…طھط£ظƒط¯ ظ…ظ† ط­ط°ظپ ظ‡ط°ظ‡ ط§ظ„ط³ظ„ظپط©طں")}\')"><button class="btn red" type="submit">{tr(request, "Delete", "ط­ط°ظپ")}</button></form>' if can_modify else ""}
            {f'<form method="post" action="{with_lang(request, f"{BASE_ROUTE}/{advance["id"]}/mark-disbursed")}" style="display:inline;"><button class="btn blue" type="submit">{tr(request, "Mark as Disbursed", "طھط­ط¯ظٹط¯ ظƒظ…طµط±ظˆظپط©")}</button></form>' if safe(advance["status"]).lower() == "active" else ""}
        </div>
    </div>

    <div class="card">
        <h3>{tr(request, "Installment Schedule", "ط¬ط¯ظˆظ„ ط§ظ„ط£ظ‚ط³ط§ط·")}</h3>
        <p style="color:#6b7280;font-size:13px;">{tr(request, "You can defer a pending installment to the end of the schedule, extending the advance term by one month.", "ظٹظ…ظƒظ†ظƒ طھط£ط¬ظٹظ„ ظ‚ط³ط· ظ…ط¹ظ„ظ‚ ط¥ظ„ظ‰ ط£ظˆظ„ ط´ظ‡ط± ظ…طھط§ط­.")}</p>
        <table>
            <thead><tr>
                <th>{tr(request, "Month", "ط§ظ„ط´ظ‡ط±")}</th>
                <th>{tr(request, "Amount", "ط§ظ„ظ…ط¨ظ„ط؛")}</th>
                <th>{tr(request, "Status", "ط§ظ„ط­ط§ظ„ط©")}</th>
                <th>{tr(request, "Action", "ط§ظ„ط¥ط¬ط±ط§ط،")}</th>
            </tr></thead>
            <tbody>{inst_rows}</tbody>
        </table>
    </div>

    <div class="card">
        <h3>{tr(request, "Deduction Ledger", "ظ„ظٹط¯ط¬ط± ط§ظ„ط®طµظˆظ…ط§طھ")}</h3>
        <table>
            <tr>
                <th>{tr(request, "Payroll Run", "ظ…ط³ظٹط± ط§ظ„ظ…ط±طھط¨")}</th>
                <th>{tr(request, "Period", "ط§ظ„ظپطھط±ط©")}</th>
                <th>{tr(request, "Amount", "ط§ظ„ظ…ط¨ظ„ط؛")}</th>
            </tr>
            {deduction_rows}
        </table>
    </div>

    {f'''
    <div class="card">
        <h3>{tr(request, "Deferral History", "ط³ط¬ظ„ ط§ظ„طھط£ط¬ظٹظ„ط§طھ")}</h3>
        <table>
            <thead><tr>
                <th>{tr(request, "Original Month", "ط§ظ„ط´ظ‡ط± ط§ظ„ط£طµظ„ظٹ")}</th>
                <th>{tr(request, "Deferred To", "ظ…ط¤ط¬ظ„ ط¥ظ„ظ‰")}</th>
                <th>{tr(request, "Amount", "ط§ظ„ظ…ط¨ظ„ط؛")}</th>
                <th>{tr(request, "Reason", "ط§ظ„ط³ط¨ط¨")}</th>
            </tr></thead>
            <tbody>{deferral_rows}</tbody>
        </table>
    </div>
    ''' if deferrals else ""}
    """
    return HTMLResponse(render_page(tr(request, "Employee Advance", "ط³ظ„ظپط© ط§ظ„ظ…ظˆط¸ظپ"), html, lang, current_path=request.url.path))


@router.post(f"{BASE_ROUTE}/{{advance_id}}/defer-installment")
def advance_defer_installment(
    request: Request,
    advance_id: int,
    installment_id: int = Form(...),
    reason: str = Form(""),
):
    conn = get_conn()
    ok, msg_text = defer_installment(conn, advance_id, installment_id, safe(reason))
    if ok:
        conn.commit()
    conn.close()
    msg = tr(request, msg_text, msg_text)
    return RedirectResponse(with_lang(request, f"{BASE_ROUTE}/{advance_id}") + f"&msg={quote(msg)}", status_code=302)

