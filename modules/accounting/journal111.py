from io import BytesIO
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from datetime import datetime

import openpyxl
from fastapi import APIRouter, Request, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse

from db import get_conn
from layout import render_page

router = APIRouter()


# =========================================================
# HELPERS
# =========================================================
def safe(x):
    return "" if x is None else str(x).strip()


def to_decimal(value, default="0"):
    try:
        text = safe(value).replace(",", "")
        if text in ["", ".", "-", "-."]:
            text = default
        return Decimal(text)
    except (InvalidOperation, ValueError, TypeError):
        return Decimal(default)


def money(value, places=2):
    try:
        d = Decimal(str(value or 0))
    except Exception:
        d = Decimal("0")
    q = Decimal("1." + ("0" * places))
    d = d.quantize(q, rounding=ROUND_HALF_UP)
    return f"{d:,.{places}f}"


def ensure_column(conn, table_name, column_name, alter_sql):
    cols = conn.execute(f"PRAGMA table_info({table_name})").fetchall()
    names = [c["name"] for c in cols]
    if column_name not in names:
        conn.execute(alter_sql)


def get_columns(conn, table_name):
    rows = conn.execute(f"PRAGMA table_info({table_name})").fetchall()
    return [r["name"] for r in rows]


def get_setting_value(key: str, default=None):
    conn = get_conn()
    try:
        row = conn.execute("""
            SELECT value
            FROM accounting_settings
            WHERE key = ?
            LIMIT 1
        """, (key,)).fetchone()
        if row and row["value"] not in [None, ""]:
            return row["value"]
    except Exception:
        pass
    finally:
        conn.close()

    fallback = {
        "journal_prefix": "JV",
    }
    return fallback.get(key, default)


# =========================================================
# DB SCHEMA + MIGRATION
# =========================================================
def ensure_tables():
    conn = get_conn()

    conn.execute("""
        CREATE TABLE IF NOT EXISTS journal_entries (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            entry_no TEXT,
            entry_date TEXT,
            description TEXT,
            reference TEXT,
            status TEXT DEFAULT 'draft',
            source_type TEXT,
            source_id INTEGER,
            reversed_by_journal_id INTEGER,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS journal_lines (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            journal_id INTEGER,
            line_no INTEGER DEFAULT 1,
            line_description TEXT,
            account_code TEXT,
            debit REAL DEFAULT 0,
            credit REAL DEFAULT 0,
            partner_type TEXT,
            partner_id INTEGER,
            cost_center_id INTEGER
        )
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS accounts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT UNIQUE,
            name TEXT NOT NULL,
            type TEXT,
            parent_id INTEGER,
            level1 TEXT,
            level2 TEXT,
            statement_type TEXT,
            is_group INTEGER DEFAULT 0,
            allow_posting INTEGER DEFAULT 1,
            is_active INTEGER DEFAULT 1
        )
    """)

    # journal_entries
    ensure_column(conn, "journal_entries", "entry_no", "ALTER TABLE journal_entries ADD COLUMN entry_no TEXT")
    ensure_column(conn, "journal_entries", "entry_date", "ALTER TABLE journal_entries ADD COLUMN entry_date TEXT")
    ensure_column(conn, "journal_entries", "description", "ALTER TABLE journal_entries ADD COLUMN description TEXT")
    ensure_column(conn, "journal_entries", "reference", "ALTER TABLE journal_entries ADD COLUMN reference TEXT")
    ensure_column(conn, "journal_entries", "status", "ALTER TABLE journal_entries ADD COLUMN status TEXT DEFAULT 'draft'")
    ensure_column(conn, "journal_entries", "source_type", "ALTER TABLE journal_entries ADD COLUMN source_type TEXT")
    ensure_column(conn, "journal_entries", "source_id", "ALTER TABLE journal_entries ADD COLUMN source_id INTEGER")
    ensure_column(conn, "journal_entries", "reversed_by_journal_id", "ALTER TABLE journal_entries ADD COLUMN reversed_by_journal_id INTEGER")
    ensure_column(conn, "journal_entries", "created_at", "ALTER TABLE journal_entries ADD COLUMN created_at TEXT DEFAULT CURRENT_TIMESTAMP")

    # journal_lines
    ensure_column(conn, "journal_lines", "journal_id", "ALTER TABLE journal_lines ADD COLUMN journal_id INTEGER")
    ensure_column(conn, "journal_lines", "line_no", "ALTER TABLE journal_lines ADD COLUMN line_no INTEGER DEFAULT 1")
    ensure_column(conn, "journal_lines", "line_description", "ALTER TABLE journal_lines ADD COLUMN line_description TEXT")
    ensure_column(conn, "journal_lines", "account_code", "ALTER TABLE journal_lines ADD COLUMN account_code TEXT")
    ensure_column(conn, "journal_lines", "debit", "ALTER TABLE journal_lines ADD COLUMN debit REAL DEFAULT 0")
    ensure_column(conn, "journal_lines", "credit", "ALTER TABLE journal_lines ADD COLUMN credit REAL DEFAULT 0")
    ensure_column(conn, "journal_lines", "partner_type", "ALTER TABLE journal_lines ADD COLUMN partner_type TEXT")
    ensure_column(conn, "journal_lines", "partner_id", "ALTER TABLE journal_lines ADD COLUMN partner_id INTEGER")
    ensure_column(conn, "journal_lines", "cost_center_id", "ALTER TABLE journal_lines ADD COLUMN cost_center_id INTEGER")

    # accounts
    ensure_column(conn, "accounts", "code", "ALTER TABLE accounts ADD COLUMN code TEXT")
    ensure_column(conn, "accounts", "name", "ALTER TABLE accounts ADD COLUMN name TEXT")
    ensure_column(conn, "accounts", "type", "ALTER TABLE accounts ADD COLUMN type TEXT")
    ensure_column(conn, "accounts", "parent_id", "ALTER TABLE accounts ADD COLUMN parent_id INTEGER")
    ensure_column(conn, "accounts", "level1", "ALTER TABLE accounts ADD COLUMN level1 TEXT")
    ensure_column(conn, "accounts", "level2", "ALTER TABLE accounts ADD COLUMN level2 TEXT")
    ensure_column(conn, "accounts", "statement_type", "ALTER TABLE accounts ADD COLUMN statement_type TEXT")
    ensure_column(conn, "accounts", "is_group", "ALTER TABLE accounts ADD COLUMN is_group INTEGER DEFAULT 0")
    ensure_column(conn, "accounts", "allow_posting", "ALTER TABLE accounts ADD COLUMN allow_posting INTEGER DEFAULT 1")
    ensure_column(conn, "accounts", "is_active", "ALTER TABLE accounts ADD COLUMN is_active INTEGER DEFAULT 1")

    # -----------------------------------------------------
    # MIGRATION: old journal_lines.entry_id -> journal_id
    # -----------------------------------------------------
    cols = get_columns(conn, "journal_lines")
    if "entry_id" in cols and "journal_id" in cols:
        conn.execute("""
            UPDATE journal_lines
            SET journal_id = entry_id
            WHERE journal_id IS NULL
              AND entry_id IS NOT NULL
        """)

    conn.commit()
    conn.close()


# =========================================================
# MASTER FUNCTIONS
# =========================================================
def next_entry_no(conn):
    prefix = get_setting_value("journal_prefix", "JV")
    row = conn.execute("""
        SELECT entry_no
        FROM journal_entries
        WHERE COALESCE(entry_no, '') <> ''
        ORDER BY id DESC
        LIMIT 1
    """).fetchone()

    if not row or not row["entry_no"]:
        return f"{prefix}-0000001"

    try:
        last_num = int(str(row["entry_no"]).split("-")[-1])
    except Exception:
        last_num = 0

    return f"{prefix}-{last_num + 1:07d}"


def account_options(selected_code=""):
    conn = get_conn()
    rows = conn.execute("""
        SELECT code, name
        FROM accounts
        WHERE COALESCE(is_active, 1) = 1
          AND COALESCE(is_group, 0) = 0
          AND COALESCE(allow_posting, 1) = 1
        ORDER BY code, name
    """).fetchall()
    conn.close()

    html = "<option value=''>-- Select Account --</option>"
    for row in rows:
        sel = "selected" if safe(selected_code) == safe(row["code"]) else ""
        html += f"<option value='{safe(row['code'])}' {sel}>{safe(row['code'])} - {safe(row['name'])}</option>"
    return html


def validate_account_for_posting(conn, account_code: str):
    row = conn.execute("""
        SELECT code, name,
               COALESCE(is_group, 0) AS is_group,
               COALESCE(allow_posting, 1) AS allow_posting,
               COALESCE(is_active, 1) AS is_active
        FROM accounts
        WHERE code = ?
        LIMIT 1
    """, (account_code,)).fetchone()

    if not row:
        raise Exception(f"Account {account_code} not found")

    if int(row["is_group"] or 0) == 1:
        raise Exception(f"Account {row['code']} - {row['name']} is a group account")

    if int(row["allow_posting"] or 0) == 0:
        raise Exception(f"Account {row['code']} - {row['name']} is not allowed for posting")

    if int(row["is_active"] or 0) == 0:
        raise Exception(f"Account {row['code']} - {row['name']} is inactive")


def get_entry_totals(conn, journal_id: int):
    row = conn.execute("""
        SELECT
            COALESCE(SUM(debit), 0) AS total_debit,
            COALESCE(SUM(credit), 0) AS total_credit
        FROM journal_lines
        WHERE journal_id = ?
    """, (journal_id,)).fetchone()

    total_debit = Decimal(str(row["total_debit"] or 0)).quantize(Decimal("1.00"), rounding=ROUND_HALF_UP)
    total_credit = Decimal(str(row["total_credit"] or 0)).quantize(Decimal("1.00"), rounding=ROUND_HALF_UP)
    balanced = total_debit == total_credit
    return total_debit, total_credit, balanced


def get_entry(conn, journal_id: int):
    return conn.execute("""
        SELECT *
        FROM journal_entries
        WHERE id = ?
        LIMIT 1
    """, (journal_id,)).fetchone()


def get_entry_lines(conn, journal_id: int):
    return conn.execute("""
        SELECT *
        FROM journal_lines
        WHERE journal_id = ?
        ORDER BY line_no, id
    """, (journal_id,)).fetchall()


def parse_lines_from_form(form):
    account_codes = form.getlist("account_code")
    line_descriptions = form.getlist("line_description")
    debits = form.getlist("debit")
    credits = form.getlist("credit")

    lines = []
    max_len = max(len(account_codes), len(line_descriptions), len(debits), len(credits), 0)

    for i in range(max_len):
        account_code = safe(account_codes[i]) if i < len(account_codes) else ""
        line_description = safe(line_descriptions[i]) if i < len(line_descriptions) else ""
        debit = to_decimal(debits[i] if i < len(debits) else "0")
        credit = to_decimal(credits[i] if i < len(credits) else "0")

        if account_code == "" and debit == Decimal("0") and credit == Decimal("0") and line_description == "":
            continue

        lines.append({
            "line_no": i + 1,
            "account_code": account_code,
            "line_description": line_description,
            "debit": debit.quantize(Decimal("1.00"), rounding=ROUND_HALF_UP),
            "credit": credit.quantize(Decimal("1.00"), rounding=ROUND_HALF_UP),
        })

    return lines


def create_journal_draft(conn, entry_date: str, description: str, reference: str, lines: list):
    entry_no = next_entry_no(conn)

    cur = conn.execute("""
        INSERT INTO journal_entries (
            entry_no, entry_date, description, reference, status
        )
        VALUES (?, ?, ?, ?, 'draft')
    """, (
        entry_no,
        safe(entry_date),
        safe(description),
        safe(reference),
    ))
    journal_id = cur.lastrowid

    for line in lines:
        conn.execute("""
            INSERT INTO journal_lines (
                journal_id, line_no, line_description, account_code, debit, credit
            )
            VALUES (?, ?, ?, ?, ?, ?)
        """, (
            journal_id,
            line["line_no"],
            line["line_description"],
            line["account_code"],
            float(line["debit"]),
            float(line["credit"]),
        ))

    return journal_id


def update_journal_draft(conn, journal_id: int, entry_date: str, description: str, reference: str, lines: list):
    entry = get_entry(conn, journal_id)
    if not entry:
        raise Exception("Journal entry not found")

    if safe(entry["status"]).lower() != "draft":
        raise Exception("Only draft entries can be edited")

    conn.execute("""
        UPDATE journal_entries
        SET entry_date = ?, description = ?, reference = ?
        WHERE id = ?
    """, (
        safe(entry_date),
        safe(description),
        safe(reference),
        journal_id,
    ))

    conn.execute("DELETE FROM journal_lines WHERE journal_id = ?", (journal_id,))

    for line in lines:
        conn.execute("""
            INSERT INTO journal_lines (
                journal_id, line_no, line_description, account_code, debit, credit
            )
            VALUES (?, ?, ?, ?, ?, ?)
        """, (
            journal_id,
            line["line_no"],
            line["line_description"],
            line["account_code"],
            float(line["debit"]),
            float(line["credit"]),
        ))


def post_journal_entry(conn, journal_id: int):
    entry = get_entry(conn, journal_id)
    if not entry:
        raise Exception("Journal entry not found")

    status = safe(entry["status"]).lower()
    if status != "draft":
        raise Exception("Only draft entries can be posted")

    lines = get_entry_lines(conn, journal_id)
    if not lines:
        raise Exception("Journal entry has no lines")

    for line in lines:
        if not safe(line["account_code"]):
            raise Exception("Account code is required on all lines")
        validate_account_for_posting(conn, safe(line["account_code"]))

    total_debit, total_credit, balanced = get_entry_totals(conn, journal_id)

    if total_debit <= Decimal("0") and total_credit <= Decimal("0"):
        raise Exception("Journal entry total cannot be zero")

    if not balanced:
        raise Exception(f"Journal not balanced: DR={total_debit}, CR={total_credit}")

    conn.execute("""
        UPDATE journal_entries
        SET status = 'posted'
        WHERE id = ?
    """, (journal_id,))


def reverse_journal_entry(conn, journal_id: int):
    entry = get_entry(conn, journal_id)
    if not entry:
        raise Exception("Journal entry not found")

    status = safe(entry["status"]).lower()
    if status != "posted":
        raise Exception("Only posted entries can be reversed")

    if entry["reversed_by_journal_id"]:
        raise Exception("Journal entry already reversed")

    original_lines = get_entry_lines(conn, journal_id)
    if not original_lines:
        raise Exception("Original journal has no lines")

    reverse_entry_no = next_entry_no(conn)

    cur = conn.execute("""
        INSERT INTO journal_entries (
            entry_no, entry_date, description, reference, status,
            source_type, source_id
        )
        VALUES (?, ?, ?, ?, 'posted', 'journal_reverse', ?)
    """, (
        reverse_entry_no,
        entry["entry_date"],
        f"Reversal of {entry['entry_no']} - {safe(entry['description'])}",
        f"REV-{safe(entry['entry_no'])}",
        journal_id,
    ))
    reverse_journal_id = cur.lastrowid

    for idx, line in enumerate(original_lines, start=1):
        conn.execute("""
            INSERT INTO journal_lines (
                journal_id, line_no, line_description, account_code,
                debit, credit, partner_type, partner_id, cost_center_id
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            reverse_journal_id,
            idx,
            f"Reverse - {safe(line['line_description'])}",
            safe(line["account_code"]),
            float(Decimal(str(line["credit"] or 0)).quantize(Decimal("1.00"), rounding=ROUND_HALF_UP)),
            float(Decimal(str(line["debit"] or 0)).quantize(Decimal("1.00"), rounding=ROUND_HALF_UP)),
            safe(line["partner_type"]),
            line["partner_id"],
            line["cost_center_id"],
        ))

    conn.execute("""
        UPDATE journal_entries
        SET status = 'reversed',
            reversed_by_journal_id = ?
        WHERE id = ?
    """, (reverse_journal_id, journal_id))

    return reverse_journal_id


# =========================================================
# IMPORT HELPERS
# =========================================================
def import_journal_from_workbook(file_bytes: bytes):
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise Exception("Excel file is empty")

    header = [safe(x).lower() for x in rows[0]]
    expected = [
        "entry_date",
        "description",
        "reference",
        "line_no",
        "account_code",
        "line_description",
        "debit",
        "credit",
    ]

    if header != expected:
        raise Exception("Template columns are invalid")

    data_rows = rows[1:]
    grouped = {}

    for r in data_rows:
        entry_date = safe(r[0])
        description = safe(r[1])
        reference = safe(r[2])
        line_no = safe(r[3])
        account_code = safe(r[4])
        line_description = safe(r[5])
        debit = to_decimal(r[6], "0")
        credit = to_decimal(r[7], "0")

        group_key = (entry_date, description, reference)

        if group_key not in grouped:
            grouped[group_key] = []

        grouped[group_key].append({
            "line_no": int(line_no) if safe(line_no) else len(grouped[group_key]) + 1,
            "account_code": account_code,
            "line_description": line_description,
            "debit": debit.quantize(Decimal("1.00"), rounding=ROUND_HALF_UP),
            "credit": credit.quantize(Decimal("1.00"), rounding=ROUND_HALF_UP),
        })

    conn = get_conn()
    created_ids = []

    try:
        for (entry_date, description, reference), lines in grouped.items():
            if not lines:
                continue
            journal_id = create_journal_draft(conn, entry_date, description, reference, lines)
            created_ids.append(journal_id)

        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    return created_ids


# =========================================================
# UI HELPERS
# =========================================================
def render_line_rows(lines=None):
    lines = lines or [{
        "line_no": 1,
        "account_code": "",
        "line_description": "",
        "debit": Decimal("0.00"),
        "credit": Decimal("0.00"),
    }]

    html = ""
    for idx, line in enumerate(lines, start=1):
        account_code = safe(line.get("account_code", ""))
        line_description = safe(line.get("line_description", ""))
        debit = Decimal(str(line.get("debit", 0))).quantize(Decimal("1.00"), rounding=ROUND_HALF_UP)
        credit = Decimal(str(line.get("credit", 0))).quantize(Decimal("1.00"), rounding=ROUND_HALF_UP)

        html += f"""
        <tr>
            <td>{idx}</td>
            <td>
                <select name="account_code" class="line-account">
                    {account_options(account_code)}
                </select>
            </td>
            <td>
                <input type="text" name="line_description" value="{line_description}">
            </td>
            <td>
                <input type="text" inputmode="decimal" name="debit" value="{debit}" class="line-debit">
            </td>
            <td>
                <input type="text" inputmode="decimal" name="credit" value="{credit}" class="line-credit">
            </td>
            <td>
                <button type="button" class="btn red" onclick="removeLine(this)">Remove</button>
            </td>
        </tr>
        """
    return html


def journal_form(values=None, lines=None, row_id=None, readonly=False):
    values = values or {}
    lines = lines or []

    add_line_button = "" if readonly else '<button type="button" class="btn green" onclick="addLine()">+ Add Line</button>'
    save_button = "" if readonly else '<button class="btn green" type="submit">Save Draft</button>'
    entry_date_rule = "readonly" if readonly else "required"
    text_readonly = "readonly" if readonly else ""
    action = f"/ui/accounting/journal/{row_id}/edit" if row_id else "/ui/accounting/journal/new"
    title = "View Journal" if readonly else ("Edit Journal" if row_id else "New Journal")

    account_options_html = account_options().replace("\\", "\\\\").replace("`", "\\`")

    html = f"""
    <div class="card">
        <h2>{title}</h2>

        <form method="post" action="{action}">
            <div class="form-grid">
                <div class="form-group">
                    <label>Entry No</label>
                    <input type="text" value="{safe(values.get('entry_no', 'Auto'))}" readonly>
                </div>

                <div class="form-group">
                    <label>Entry Date</label>
                    <input type="date" name="entry_date" value="{safe(values.get('entry_date', ''))}" {entry_date_rule}>
                </div>

                <div class="form-group">
                    <label>Description</label>
                    <input type="text" name="description" value="{safe(values.get('description', ''))}" {text_readonly}>
                </div>

                <div class="form-group">
                    <label>Reference</label>
                    <input type="text" name="reference" value="{safe(values.get('reference', ''))}" {text_readonly}>
                </div>

                <div class="form-group">
                    <label>Status</label>
                    <input type="text" value="{safe(values.get('status', 'draft'))}" readonly>
                </div>

                <div class="form-group">
                    <label>Reversed By</label>
                    <input type="text" value="{safe(values.get('reversed_by_journal_id', ''))}" readonly>
                </div>
            </div>

            <div style="margin-top:20px;">
                <div class="table-header">
                    <h3>Journal Lines</h3>
                    {add_line_button}
                </div>

                <table id="lines-table">
                    <thead>
                        <tr>
                            <th style="width:60px;">#</th>
                            <th>Account</th>
                            <th>Description</th>
                            <th style="width:160px;">Debit</th>
                            <th style="width:160px;">Credit</th>
                            <th style="width:120px;">Action</th>
                        </tr>
                    </thead>
                    <tbody>
                        {render_line_rows(lines)}
                    </tbody>
                </table>
            </div>

            <div style="margin-top:20px; max-width:420px; margin-left:auto;">
                <table>
                    <tr>
                        <th>Total Debit</th>
                        <td><input type="text" id="total_debit_view" readonly value="0.00"></td>
                    </tr>
                    <tr>
                        <th>Total Credit</th>
                        <td><input type="text" id="total_credit_view" readonly value="0.00"></td>
                    </tr>
                    <tr>
                        <th>Balanced</th>
                        <td><input type="text" id="balanced_view" readonly value="No"></td>
                    </tr>
                </table>
            </div>

            <div class="form-actions">
                {save_button}
                <a class="btn gray" href="/ui/accounting/journal">Back</a>
            </div>
        </form>
    </div>

    <script>
    (function() {{
        const defaultAccountOptions = `{account_options_html}`;
        const isReadonly = {"true" if readonly else "false"};

        function sanitizeDecimalInput(v) {{
            if (v === null || v === undefined) return "";
            v = String(v).replace(/,/g, "");
            v = v.replace(/[^0-9.\\-]/g, "");

            const minusCount = (v.match(/-/g) || []).length;
            if (minusCount > 1) v = v.replace(/-/g, "");
            if (v.indexOf("-") > 0) v = v.replace(/-/g, "");
            if (minusCount >= 1 && !v.startsWith("-")) v = "-" + v.replace(/-/g, "");

            const parts = v.split(".");
            if (parts.length > 2) {{
                v = parts[0] + "." + parts.slice(1).join("");
            }}

            return v;
        }}

        function parseNum(v) {{
            v = sanitizeDecimalInput(v);
            if (v === "" || v === "." || v === "-" || v === "-.") return 0;
            const n = parseFloat(v);
            return isNaN(n) ? 0 : n;
        }}

        function fmt(v) {{
            return parseNum(v).toFixed(2);
        }}

        function renumberLines() {{
            const rows = document.querySelectorAll("#lines-table tbody tr");
            rows.forEach((row, idx) => {{
                const firstCell = row.querySelector("td");
                if (firstCell) firstCell.textContent = idx + 1;
            }});
        }}

        function recalcTotals() {{
            let totalDebit = 0;
            let totalCredit = 0;

            document.querySelectorAll("#lines-table tbody tr").forEach((row) => {{
                const debitInput = row.querySelector(".line-debit");
                const creditInput = row.querySelector(".line-credit");
                totalDebit += parseNum(debitInput ? debitInput.value : 0);
                totalCredit += parseNum(creditInput ? creditInput.value : 0);
            }});

            const d = document.getElementById("total_debit_view");
            const c = document.getElementById("total_credit_view");
            const b = document.getElementById("balanced_view");

            if (d) d.value = fmt(totalDebit);
            if (c) c.value = fmt(totalCredit);
            if (b) b.value = Math.abs(totalDebit - totalCredit) < 0.0001 ? "Yes" : "No";
        }}

        function bindLineEvents(scope) {{
            scope.querySelectorAll(".line-debit, .line-credit").forEach((el) => {{
                el.addEventListener("input", function() {{
                    const oldStart = el.selectionStart;
                    const oldLen = el.value.length;
                    el.value = sanitizeDecimalInput(el.value);
                    const newLen = el.value.length;
                    if (oldStart !== null) {{
                        const nextPos = oldStart + (newLen - oldLen);
                        try {{
                            el.setSelectionRange(nextPos, nextPos);
                        }} catch (e) {{}}
                    }}
                    recalcTotals();
                }});
            }});
        }}

        window.addLine = function() {{
            const tbody = document.querySelector("#lines-table tbody");
            if (!tbody) return;

            const tr = document.createElement("tr");
            tr.innerHTML =
                "<td></td>" +
                "<td><select name='account_code' class='line-account'>" + defaultAccountOptions + "</select></td>" +
                "<td><input type='text' name='line_description' value=''></td>" +
                "<td><input type='text' inputmode='decimal' name='debit' value='0.00' class='line-debit'></td>" +
                "<td><input type='text' inputmode='decimal' name='credit' value='0.00' class='line-credit'></td>" +
                "<td><button type='button' class='btn red' onclick='removeLine(this)'>Remove</button></td>";

            tbody.appendChild(tr);
            bindLineEvents(tr);
            renumberLines();
            recalcTotals();
        }}

        window.removeLine = function(btn) {{
            const tbody = document.querySelector("#lines-table tbody");
            if (!tbody) return;

            if (tbody.querySelectorAll("tr").length <= 1) {{
                alert("Journal must contain at least one line.");
                return;
            }}

            btn.closest("tr").remove();
            renumberLines();
            recalcTotals();
        }}

        document.addEventListener("DOMContentLoaded", function() {{
            document.querySelectorAll("#lines-table tbody tr").forEach(bindLineEvents);
            renumberLines();
            recalcTotals();

            if (isReadonly) {{
                document.querySelectorAll("#lines-table tbody select").forEach(el => el.setAttribute("disabled", "disabled"));
                document.querySelectorAll("#lines-table tbody input").forEach(el => el.setAttribute("readonly", "readonly"));
            }}
        }});
    }})();
    </script>
    """

    return html


# =========================================================
# TEMPLATE / IMPORT PAGE
# =========================================================
@router.get("/ui/accounting/journal/template")
def journal_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Journal Template"

    ws.append([
        "entry_date",
        "description",
        "reference",
        "line_no",
        "account_code",
        "line_description",
        "debit",
        "credit",
    ])

    ws.append(["2026-01-01", "Opening Entry", "OPEN-001", 1, "111100", "Opening Cash", 10000, 0])
    ws.append(["2026-01-01", "Opening Entry", "OPEN-001", 2, "310000", "Opening Capital", 0, 10000])

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=journal_template.xlsx"},
    )


@router.get("/ui/accounting/journal/import", response_class=HTMLResponse)
def journal_import_page(request: Request):
    content = """
    <div class="card">
        <h2>Import Journal Entries</h2>
        <form method="post" action="/ui/accounting/journal/import" enctype="multipart/form-data">
            <div class="form-group">
                <label>Excel File</label>
                <input type="file" name="file" accept=".xlsx" required>
            </div>

            <div class="form-actions">
                <button class="btn green" type="submit">Import as Draft</button>
                <a class="btn gray" href="/ui/accounting/journal/template">Download Template</a>
                <a class="btn gray" href="/ui/accounting/journal">Back</a>
            </div>
        </form>
    </div>
    """
    return HTMLResponse(render_page("Import Journal", content, current_path=str(request.url.path)))


@router.post("/ui/accounting/journal/import")
async def journal_import(file: UploadFile = File(...)):
    if not file.filename.lower().endswith(".xlsx"):
        return HTMLResponse("Only .xlsx files are allowed", status_code=400)

    try:
        content = await file.read()
        created_ids = import_journal_from_workbook(content)
    except Exception as e:
        return HTMLResponse(f"Import error: {safe(e)}", status_code=400)

    if not created_ids:
        return HTMLResponse("No journal entries imported", status_code=400)

    return RedirectResponse("/ui/accounting/journal", status_code=303)


# =========================================================
# ROUTES
# =========================================================
ensure_tables()


@router.get("/ui/accounting/journal", response_class=HTMLResponse)
def journal_list(request: Request):
    conn = get_conn()
    rows = conn.execute("""
        SELECT
            j.*,
            COALESCE(SUM(l.debit), 0) AS total_debit,
            COALESCE(SUM(l.credit), 0) AS total_credit
        FROM journal_entries j
        LEFT JOIN journal_lines l ON l.journal_id = j.id
        GROUP BY j.id
        ORDER BY j.id DESC
    """).fetchall()
    conn.close()

    rows_html = ""
    for row in rows:
        edit_btn = ""
        post_btn = ""
        reverse_btn = ""

        if safe(row["status"]).lower() == "draft":
            edit_btn = f"<a class='btn green' href='/ui/accounting/journal/{row['id']}/edit'>Edit</a>"
            post_btn = (
                f"<form method='post' action='/ui/accounting/journal/{row['id']}/post' style='display:inline;'>"
                f"<button class='btn green' type='submit'>Post</button></form>"
            )

        if safe(row["status"]).lower() == "posted" and not row["reversed_by_journal_id"]:
            reverse_btn = (
                f"<form method='post' action='/ui/accounting/journal/{row['id']}/reverse' style='display:inline;'>"
                f"<button class='btn red' type='submit'>Reverse</button></form>"
            )

        rows_html += f"""
        <tr>
            <td>{safe(row['entry_no'])}</td>
            <td>{safe(row['entry_date'])}</td>
            <td>{safe(row['description'])}</td>
            <td>{safe(row['reference'])}</td>
            <td>{money(row['total_debit'])}</td>
            <td>{money(row['total_credit'])}</td>
            <td>{safe(row['status'])}</td>
            <td>{safe(row['reversed_by_journal_id'])}</td>
            <td>
                <a class="btn blue" href="/ui/accounting/journal/{row['id']}/view">View</a>
                {edit_btn}
                {post_btn}
                {reverse_btn}
            </td>
        </tr>
        """

    if not rows_html:
        rows_html = "<tr><td colspan='9'>No journal entries found.</td></tr>"

    content = f"""
    <div class="table-header">
        <h3>Journal Entries</h3>
        <div style="display:flex;gap:8px;flex-wrap:wrap;">
            <a class="btn green" href="/ui/accounting/journal/new">+ New Journal</a>
            <a class="btn gray" href="/ui/accounting/journal/import">Import</a>
            <a class="btn gray" href="/ui/accounting/journal/template">Download Template</a>
        </div>
    </div>

    <table>
        <tr>
            <th>No</th>
            <th>Date</th>
            <th>Description</th>
            <th>Reference</th>
            <th>Total Debit</th>
            <th>Total Credit</th>
            <th>Status</th>
            <th>Reversed By</th>
            <th>Action</th>
        </tr>
        {rows_html}
    </table>
    """
    return HTMLResponse(render_page("Journal", content, current_path=str(request.url.path)))


@router.get("/ui/accounting/journal/new", response_class=HTMLResponse)
def journal_new(request: Request):
    today = datetime.today().date().isoformat()
    values = {
        "entry_no": "Auto",
        "entry_date": today,
        "description": "",
        "reference": "",
        "status": "draft",
    }
    lines = [{
        "line_no": 1,
        "account_code": "",
        "line_description": "",
        "debit": Decimal("0.00"),
        "credit": Decimal("0.00"),
    }]
    return HTMLResponse(render_page("New Journal", journal_form(values, lines=lines), current_path=str(request.url.path)))


@router.post("/ui/accounting/journal/new")
async def journal_create(request: Request):
    form = await request.form()

    entry_date = safe(form.get("entry_date"))
    description = safe(form.get("description"))
    reference = safe(form.get("reference"))
    lines = parse_lines_from_form(form)

    if not lines:
        return HTMLResponse("Journal must contain at least one line", status_code=400)

    conn = get_conn()
    try:
        journal_id = create_journal_draft(conn, entry_date, description, reference, lines)
        conn.commit()
    except Exception as e:
        conn.rollback()
        conn.close()
        return HTMLResponse(f"Save error: {safe(e)}", status_code=400)

    conn.close()
    return RedirectResponse(f"/ui/accounting/journal/{journal_id}/view", status_code=303)


@router.get("/ui/accounting/journal/{row_id}/edit", response_class=HTMLResponse)
def journal_edit(request: Request, row_id: int):
    conn = get_conn()
    entry = get_entry(conn, row_id)
    if not entry:
        conn.close()
        return HTMLResponse("Journal entry not found", status_code=404)

    if safe(entry["status"]).lower() != "draft":
        conn.close()
        return HTMLResponse("Only draft entries can be edited", status_code=400)

    lines = get_entry_lines(conn, row_id)
    conn.close()

    return HTMLResponse(
        render_page(
            "Edit Journal",
            journal_form(dict(entry), lines=[dict(x) for x in lines], row_id=row_id),
            current_path=str(request.url.path),
        )
    )


@router.post("/ui/accounting/journal/{row_id}/edit")
async def journal_update(request: Request, row_id: int):
    form = await request.form()

    entry_date = safe(form.get("entry_date"))
    description = safe(form.get("description"))
    reference = safe(form.get("reference"))
    lines = parse_lines_from_form(form)

    if not lines:
        return HTMLResponse("Journal must contain at least one line", status_code=400)

    conn = get_conn()
    try:
        update_journal_draft(conn, row_id, entry_date, description, reference, lines)
        conn.commit()
    except Exception as e:
        conn.rollback()
        conn.close()
        return HTMLResponse(f"Update error: {safe(e)}", status_code=400)

    conn.close()
    return RedirectResponse(f"/ui/accounting/journal/{row_id}/view", status_code=303)


@router.post("/ui/accounting/journal/{row_id}/post")
def journal_post(row_id: int):
    conn = get_conn()
    try:
        post_journal_entry(conn, row_id)
        conn.commit()
    except Exception as e:
        conn.rollback()
        conn.close()
        return HTMLResponse(f"Post error: {safe(e)}", status_code=400)

    conn.close()
    return RedirectResponse(f"/ui/accounting/journal/{row_id}/view", status_code=303)


@router.post("/ui/accounting/journal/{row_id}/reverse")
def journal_reverse(row_id: int):
    conn = get_conn()
    try:
        reverse_journal_entry(conn, row_id)
        conn.commit()
    except Exception as e:
        conn.rollback()
        conn.close()
        return HTMLResponse(f"Reverse error: {safe(e)}", status_code=400)

    conn.close()
    return RedirectResponse("/ui/accounting/journal", status_code=303)


@router.get("/ui/accounting/journal/{row_id}/view", response_class=HTMLResponse)
def journal_view(request: Request, row_id: int):
    conn = get_conn()
    entry = get_entry(conn, row_id)
    if not entry:
        conn.close()
        return HTMLResponse("Journal entry not found", status_code=404)

    lines = get_entry_lines(conn, row_id)
    total_debit, total_credit, balanced = get_entry_totals(conn, row_id)
    conn.close()

    content = journal_form(
        values=dict(entry),
        lines=[dict(x) for x in lines],
        row_id=row_id,
        readonly=True,
    )

    extra_buttons = ""
    if safe(entry["status"]).lower() == "draft":
        extra_buttons += f"<a class='btn green' href='/ui/accounting/journal/{row_id}/edit'>Edit Draft</a>"
        extra_buttons += (
            f"<form method='post' action='/ui/accounting/journal/{row_id}/post' style='display:inline;'>"
            f"<button class='btn green' type='submit'>Post</button></form>"
        )

    if safe(entry["status"]).lower() == "posted" and not entry["reversed_by_journal_id"]:
        extra_buttons += (
            f"<form method='post' action='/ui/accounting/journal/{row_id}/reverse' style='display:inline;'>"
            f"<button class='btn red' type='submit'>Reverse</button></form>"
        )

    extra = f"""
    <div class="card" style="margin-top:20px;">
        <h3>Entry Summary</h3>
        <p><b>Total Debit:</b> {money(total_debit)}</p>
        <p><b>Total Credit:</b> {money(total_credit)}</p>
        <p><b>Balanced:</b> {"Yes" if balanced else "No"}</p>
    </div>

    <div class="form-actions" style="margin-top:16px;">
        {extra_buttons}
        <a class="btn gray" href="/ui/accounting/journal">Back to Journal</a>
    </div>
    """

    return HTMLResponse(render_page("View Journal", content + extra, current_path=str(request.url.path)))