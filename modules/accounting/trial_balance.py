from fastapi import APIRouter, Request
from fastapi.responses import HTMLResponse, StreamingResponse
from io import BytesIO

from db import get_conn
from layout import render_page

router = APIRouter()


def money(x):
    try:
        return f"{float(x or 0):,.2f}"
    except Exception:
        return "0.00"


def ensure_column(conn, table_name, column_name, alter_sql):
    cols = conn.execute(f"PRAGMA table_info({table_name})").fetchall()
    names = [c["name"] for c in cols]
    if column_name not in names:
        conn.execute(alter_sql)


def ensure_tables():
    conn = get_conn()

    conn.execute("""
        CREATE TABLE IF NOT EXISTS accounts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT UNIQUE,
            name TEXT NOT NULL,
            type TEXT,
            parent_id INTEGER,
            is_group INTEGER DEFAULT 0,
            is_active INTEGER DEFAULT 1,
            allow_posting INTEGER DEFAULT 1
        )
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS journal_entries (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            entry_no TEXT,
            entry_date TEXT,
            description TEXT,
            reference TEXT,
            status TEXT DEFAULT 'draft',
            source_type TEXT,
            source_id INTEGER,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS journal_lines (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            journal_id INTEGER NOT NULL,
            line_no INTEGER DEFAULT 1,
            line_description TEXT,
            account_code TEXT,
            debit REAL DEFAULT 0,
            credit REAL DEFAULT 0,
            partner_type TEXT,
            partner_id INTEGER
        )
    """)

    ensure_column(conn, "accounts", "code", "ALTER TABLE accounts ADD COLUMN code TEXT")
    ensure_column(conn, "accounts", "name", "ALTER TABLE accounts ADD COLUMN name TEXT")
    ensure_column(conn, "accounts", "type", "ALTER TABLE accounts ADD COLUMN type TEXT")
    ensure_column(conn, "accounts", "parent_id", "ALTER TABLE accounts ADD COLUMN parent_id INTEGER")
    ensure_column(conn, "accounts", "is_group", "ALTER TABLE accounts ADD COLUMN is_group INTEGER DEFAULT 0")
    ensure_column(conn, "accounts", "is_active", "ALTER TABLE accounts ADD COLUMN is_active INTEGER DEFAULT 1")
    ensure_column(conn, "accounts", "allow_posting", "ALTER TABLE accounts ADD COLUMN allow_posting INTEGER DEFAULT 1")

    ensure_column(conn, "journal_entries", "entry_no", "ALTER TABLE journal_entries ADD COLUMN entry_no TEXT")
    ensure_column(conn, "journal_entries", "entry_date", "ALTER TABLE journal_entries ADD COLUMN entry_date TEXT")
    ensure_column(conn, "journal_entries", "description", "ALTER TABLE journal_entries ADD COLUMN description TEXT")
    ensure_column(conn, "journal_entries", "reference", "ALTER TABLE journal_entries ADD COLUMN reference TEXT")
    ensure_column(conn, "journal_entries", "status", "ALTER TABLE journal_entries ADD COLUMN status TEXT DEFAULT 'draft'")
    ensure_column(conn, "journal_entries", "source_type", "ALTER TABLE journal_entries ADD COLUMN source_type TEXT")
    ensure_column(conn, "journal_entries", "source_id", "ALTER TABLE journal_entries ADD COLUMN source_id INTEGER")
    ensure_column(conn, "journal_entries", "created_at", "ALTER TABLE journal_entries ADD COLUMN created_at TEXT DEFAULT CURRENT_TIMESTAMP")

    ensure_column(conn, "journal_lines", "journal_id", "ALTER TABLE journal_lines ADD COLUMN journal_id INTEGER")
    ensure_column(conn, "journal_lines", "line_no", "ALTER TABLE journal_lines ADD COLUMN line_no INTEGER DEFAULT 1")
    ensure_column(conn, "journal_lines", "line_description", "ALTER TABLE journal_lines ADD COLUMN line_description TEXT")
    ensure_column(conn, "journal_lines", "account_code", "ALTER TABLE journal_lines ADD COLUMN account_code TEXT")
    ensure_column(conn, "journal_lines", "debit", "ALTER TABLE journal_lines ADD COLUMN debit REAL DEFAULT 0")
    ensure_column(conn, "journal_lines", "credit", "ALTER TABLE journal_lines ADD COLUMN credit REAL DEFAULT 0")
    ensure_column(conn, "journal_lines", "partner_type", "ALTER TABLE journal_lines ADD COLUMN partner_type TEXT")
    ensure_column(conn, "journal_lines", "partner_id", "ALTER TABLE journal_lines ADD COLUMN partner_id INTEGER")

    conn.commit()
    conn.close()


ensure_tables()


def get_opening_balance(conn, account_code, date_from):
    if not date_from:
        return 0.0

    row = conn.execute("""
        SELECT
            COALESCE(SUM(jl.debit), 0) AS total_debit,
            COALESCE(SUM(jl.credit), 0) AS total_credit
        FROM journal_lines jl
        JOIN journal_entries je ON je.id = jl.journal_id
        WHERE LOWER(COALESCE(je.status, '')) = 'posted'
          AND COALESCE(jl.account_code, '') = ?
          AND COALESCE(je.entry_date, '') < ?
    """, (account_code, date_from)).fetchone()

    return float(row["total_debit"] or 0) - float(row["total_credit"] or 0)


def get_period_totals(conn, account_code, date_from="", date_to=""):
    sql = """
        SELECT
            COALESCE(SUM(jl.debit), 0) AS total_debit,
            COALESCE(SUM(jl.credit), 0) AS total_credit
        FROM journal_lines jl
        JOIN journal_entries je ON je.id = jl.journal_id
        WHERE LOWER(COALESCE(je.status, '')) = 'posted'
          AND COALESCE(jl.account_code, '') = ?
    """
    params = [account_code]

    if date_from:
        sql += " AND COALESCE(je.entry_date, '') >= ?"
        params.append(date_from)

    if date_to:
        sql += " AND COALESCE(je.entry_date, '') <= ?"
        params.append(date_to)

    row = conn.execute(sql, params).fetchone()
    return float(row["total_debit"] or 0), float(row["total_credit"] or 0)


def get_accounts(conn):
    return conn.execute("""
        SELECT DISTINCT a.*
        FROM accounts a
        JOIN journal_lines jl ON jl.account_code = a.code
        JOIN journal_entries je ON je.id = jl.journal_id
        WHERE COALESCE(a.is_active, 1) = 1
          AND COALESCE(a.is_group, 0) = 0
          AND LOWER(COALESCE(je.status, '')) = 'posted'
        ORDER BY a.code
    """).fetchall()


def build_trial_balance_data(conn, date_from="", date_to=""):
    accounts = get_accounts(conn)
    rows = []
    totals = {
        "opening_debit": 0.0,
        "opening_credit": 0.0,
        "period_debit": 0.0,
        "period_credit": 0.0,
        "closing_debit": 0.0,
        "closing_credit": 0.0,
    }

    for acc in accounts:
        opening = get_opening_balance(conn, acc["code"], date_from)
        period_debit, period_credit = get_period_totals(conn, acc["code"], date_from, date_to)
        closing = opening + period_debit - period_credit

        opening_debit = opening if opening > 0 else 0.0
        opening_credit = abs(opening) if opening < 0 else 0.0
        closing_debit = closing if closing > 0 else 0.0
        closing_credit = abs(closing) if closing < 0 else 0.0

        if (
            abs(opening_debit) < 0.0001 and
            abs(opening_credit) < 0.0001 and
            abs(period_debit) < 0.0001 and
            abs(period_credit) < 0.0001 and
            abs(closing_debit) < 0.0001 and
            abs(closing_credit) < 0.0001
        ):
            continue

        row = {
            "code": acc["code"] or "",
            "name": acc["name"] or "",
            "type": acc["type"] or "",
            "opening_debit": opening_debit,
            "opening_credit": opening_credit,
            "period_debit": period_debit,
            "period_credit": period_credit,
            "closing_debit": closing_debit,
            "closing_credit": closing_credit,
        }
        rows.append(row)
        for key in totals:
            totals[key] += row[key]

    return rows, totals


@router.get("/ui/accounting/trial-balance", response_class=HTMLResponse)
@router.get("/ui/accounting/trial_balance", response_class=HTMLResponse)
def trial_balance(request: Request, date_from: str = "", date_to: str = "", embed: int = 0):
    conn = get_conn()
    rows, totals = build_trial_balance_data(conn, date_from, date_to)

    rows_html = ""
    visible_count = len(rows)
    total_opening_debit = totals["opening_debit"]
    total_opening_credit = totals["opening_credit"]
    total_period_debit = totals["period_debit"]
    total_period_credit = totals["period_credit"]
    total_closing_debit = totals["closing_debit"]
    total_closing_credit = totals["closing_credit"]

    for row in rows:
        account_code = row["code"]
        ledger_href = f"/ui/accounting/general-ledger?account_code={account_code}&date_from={date_from}&date_to={date_to}"
        journals_href = f"/ui/accounting/journal?status=posted&account_code={account_code}&date_from={date_from}&date_to={date_to}"
        rows_html += f"""
        <tr>
            <td>
                <details class="tb-account-menu">
                    <summary>{row['code']}</summary>
                    <div class="tb-account-actions">
                        <a href="{ledger_href}">Ledger</a>
                        <a href="{journals_href}">Journals</a>
                    </div>
                </details>
            </td>
            <td>
                <details class="tb-account-menu">
                    <summary>{row['name']}</summary>
                    <div class="tb-account-actions">
                        <a href="{ledger_href}">Ledger</a>
                        <a href="{journals_href}">Journals</a>
                    </div>
                </details>
            </td>
            <td>{row['type']}</td>
            <td class="text-right">{money(row['opening_debit'])}</td>
            <td class="text-right">{money(row['opening_credit'])}</td>
            <td class="text-right">{money(row['period_debit'])}</td>
            <td class="text-right">{money(row['period_credit'])}</td>
            <td class="text-right">{money(row['closing_debit'])}</td>
            <td class="text-right">{money(row['closing_credit'])}</td>
        </tr>
        """

    conn.close()

    if not rows_html:
        rows_html = """
        <tr>
            <td colspan="9" class="empty-state">No posted balances found for the selected period.</td>
        </tr>
        """

    report_tabs_html = ""
    if int(embed or 0) != 1:
        try:
            from modules.accounting.reports import report_tabs
            report_tabs_html = f'<div style="margin-top:18px;">{report_tabs("tb")}</div>'
        except Exception:
            report_tabs_html = ""

    balanced_check = "OK" if abs(total_closing_debit - total_closing_credit) < 0.01 else "Review"

    filter_html = f"""
    <div class="card">
        <div class="toolbar" style="margin-bottom:16px;">
            <div>
                <div style="display:inline-flex;align-items:center;gap:8px;padding:6px 10px;border-radius:999px;background:#eef4ff;color:#295db8;font-size:12px;font-weight:800;margin-bottom:10px;">Financial Position</div>
                <h2 style="font-size:28px;line-height:1.1;color:#13315c;margin-bottom:8px;">Trial Balance</h2>
                <div class="section-note">Opening, period, and closing balances in a cleaner list-style report layout.</div>
            </div>
            <div class="table-summary">
                <span class="summary-pill">Accounts: {visible_count}</span>
                <span class="summary-pill">Period: {date_from or 'Beginning'} to {date_to or 'Today'}</span>
            </div>
        </div>
        {report_tabs_html}
    </div>

    <div class="card">
        <div class="toolbar" style="margin-bottom:16px;">
            <div>
                <h3 class="sub-title">Filters</h3>
                <div class="section-note">Select a date range, refresh the balances, then review the summary and detailed table below.</div>
            </div>
            <div class="table-summary">
                <span class="summary-pill">Balanced Check: {balanced_check}</span>
            </div>
        </div>
        <form method="get">
            <div class="filter-grid" style="grid-template-columns: 1fr 1fr;">
                <div class="form-group">
                    <label>From Date</label>
                    <input type="date" name="date_from" value="{date_from}">
                </div>
                <div class="form-group">
                    <label>To Date</label>
                    <input type="date" name="date_to" value="{date_to}">
                </div>
            </div>

            <div class="filter-actions">
                <button class="btn blue" type="submit">Show Trial Balance</button>
                <a class="btn gray" href="/ui/accounting/trial-balance">Clear</a>
                <a class="btn blue" href="/ui/accounting/trial-balance/export.xlsx?date_from={date_from}&date_to={date_to}">Export Excel</a>
            </div>
        </form>
    </div>
    """

    summary_html = f"""
    <div class="card">
        <div class="toolbar" style="margin-bottom:16px;">
            <div>
                <h3 class="sub-title">Summary</h3>
                <div class="section-note">Quick totals to review the overall movement before going into the detailed rows.</div>
            </div>
        </div>
        <div class="kpi-grid" style="grid-template-columns:repeat(3,minmax(0,1fr));">
            <div class="kpi-card">
                <div class="kpi-label">Opening Debit</div>
                <div class="kpi-value">{money(total_opening_debit)}</div>
                <div class="kpi-foot">Opening Credit: {money(total_opening_credit)}</div>
            </div>
            <div class="kpi-card">
                <div class="kpi-label">Period Debit</div>
                <div class="kpi-value">{money(total_period_debit)}</div>
                <div class="kpi-foot">Period Credit: {money(total_period_credit)}</div>
            </div>
            <div class="kpi-card">
                <div class="kpi-label">Closing Debit</div>
                <div class="kpi-value">{money(total_closing_debit)}</div>
                <div class="kpi-foot">Closing Credit: {money(total_closing_credit)}</div>
            </div>
        </div>
    </div>
    """

    table_html = f"""
    <div class="card">
        <div class="toolbar" style="margin-bottom:16px;">
            <div>
                <h3 class="sub-title">Balance Details</h3>
                <div class="section-note">Each account shows opening balance, current-period movement, and final closing position.</div>
            </div>
            <div class="table-summary">
                <span class="summary-pill">Opening Net: {money(total_opening_debit - total_opening_credit)}</span>
                <span class="summary-pill">Closing Net: {money(total_closing_debit - total_closing_credit)}</span>
            </div>
        </div>
        <div class="table-wrap">
            <table>
                <thead>
                    <tr>
                        <th rowspan="2">Code</th>
                        <th rowspan="2">Account Name</th>
                        <th rowspan="2">Type</th>
                        <th colspan="2" class="text-right">Opening</th>
                        <th colspan="2" class="text-right">Period</th>
                        <th colspan="2" class="text-right">Closing</th>
                    </tr>
                    <tr>
                        <th class="text-right">Debit</th>
                        <th class="text-right">Credit</th>
                        <th class="text-right">Debit</th>
                        <th class="text-right">Credit</th>
                        <th class="text-right">Debit</th>
                        <th class="text-right">Credit</th>
                    </tr>
                </thead>
                <tbody>
                    {rows_html}
                    <tr class="summary-row">
                        <td colspan="3">TOTAL</td>
                        <td class="text-right">{money(total_opening_debit)}</td>
                        <td class="text-right">{money(total_opening_credit)}</td>
                        <td class="text-right">{money(total_period_debit)}</td>
                        <td class="text-right">{money(total_period_credit)}</td>
                        <td class="text-right">{money(total_closing_debit)}</td>
                        <td class="text-right">{money(total_closing_credit)}</td>
                    </tr>
                </tbody>
            </table>
        </div>
    </div>
    """

    menu_style = """
    <style>
        .tb-account-menu { position: relative; display: inline-block; }
        .tb-account-menu summary { cursor: pointer; color: #2563eb; font-weight: 800; list-style: none; }
        .tb-account-menu summary::-webkit-details-marker { display: none; }
        .tb-account-actions {
            position: absolute;
            z-index: 20;
            min-width: 132px;
            margin-top: 6px;
            padding: 6px;
            border: 1px solid #dbe4f0;
            border-radius: 8px;
            background: #fff;
            box-shadow: 0 12px 30px rgba(15, 23, 42, .14);
        }
        .tb-account-actions a {
            display: block;
            padding: 8px 10px;
            border-radius: 6px;
            color: #0f2e5f;
            text-decoration: none;
            font-weight: 800;
            white-space: nowrap;
        }
        .tb-account-actions a:hover { background: #eef4ff; }
    </style>
    <script>
        document.addEventListener('toggle', function (event) {
            const opened = event.target;
            if (!opened.matches('.tb-account-menu') || !opened.open) return;
            document.querySelectorAll('.tb-account-menu[open]').forEach(function (menu) {
                if (menu !== opened) menu.open = false;
            });
        }, true);
        document.addEventListener('click', function (event) {
            if (event.target.closest('.tb-account-menu')) return;
            document.querySelectorAll('.tb-account-menu[open]').forEach(function (menu) {
                menu.open = false;
            });
        });
    </script>
    """

    content = f'<div class="list-shell">{menu_style + filter_html + summary_html + table_html}</div>'

    if int(embed or 0) == 1:
        return HTMLResponse(content)

    return HTMLResponse(
        render_page("Trial Balance", content, "en", request.url.path)
    )


@router.get("/ui/accounting/trial-balance/export.xlsx")
def trial_balance_export_xlsx(date_from: str = "", date_to: str = ""):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except Exception:
        return HTMLResponse("Excel export is not available. Install openpyxl.", status_code=500)

    conn = get_conn()
    rows, totals = build_trial_balance_data(conn, date_from, date_to)
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Trial Balance"

    title = "Trial Balance"
    period = f"Period: {date_from or 'Beginning'} to {date_to or 'Today'}"
    ws.append([title])
    ws.append([period])
    ws.append([])
    ws.append([
        "Code", "Account Name", "Type",
        "Opening Debit", "Opening Credit",
        "Period Debit", "Period Credit",
        "Closing Debit", "Closing Credit",
    ])

    for row in rows:
        ws.append([
            row["code"],
            row["name"],
            row["type"],
            row["opening_debit"],
            row["opening_credit"],
            row["period_debit"],
            row["period_credit"],
            row["closing_debit"],
            row["closing_credit"],
        ])

    ws.append([
        "TOTAL", "", "",
        totals["opening_debit"],
        totals["opening_credit"],
        totals["period_debit"],
        totals["period_credit"],
        totals["closing_debit"],
        totals["closing_credit"],
    ])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    total_fill = PatternFill("solid", fgColor="EAF7EA")
    thin = Side(style="thin", color="D9E2EF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"].font = Font(bold=True, size=16)
    ws["A2"].font = Font(italic=True, color="52657A")

    for cell in ws[4]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    total_row = ws.max_row
    for cell in ws[total_row]:
        cell.font = Font(bold=True)
        cell.fill = total_fill
        cell.border = border

    for row in ws.iter_rows(min_row=5, max_row=total_row, min_col=1, max_col=9):
        for cell in row:
            cell.border = border
        for cell in row[3:]:
            cell.number_format = '#,##0.00'

    widths = [14, 34, 16, 16, 16, 16, 16, 16, 16]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"trial_balance_{date_from or 'beginning'}_{date_to or 'today'}.xlsx".replace("/", "-")
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
