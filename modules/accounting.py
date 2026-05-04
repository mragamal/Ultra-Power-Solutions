from fastapi import APIRouter, Request, Form, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
import openpyxl
from io import BytesIO

from auth import current_user
from modules.layout import render_layout
from db import get_conn

router = APIRouter()


def require_login(request: Request):
    user = current_user(request)
    if not user:
        return None
    return user


def get_account_columns():
    conn = get_conn()
    try:
        rows = conn.execute("PRAGMA table_info(accounts)").fetchall()
        return [row["name"] for row in rows]
    finally:
        conn.close()


def account_type_options(selected: str = ""):
    types = [
        ("asset", "Asset"),
        ("liability", "Liability"),
        ("equity", "Equity"),
        ("revenue", "Revenue"),
        ("expense", "Expense"),
    ]

    html = ""
    for value, label in types:
        is_selected = "selected" if selected == value else ""
        html += f'<option value="{value}" {is_selected}>{label}</option>'
    return html


def parent_account_options(selected_id=None, current_id=None):
    conn = get_conn()
    try:
        rows = conn.execute("""
            SELECT id, code, name
            FROM accounts
            ORDER BY code, name
        """).fetchall()
    finally:
        conn.close()

    html = '<option value="">No Parent</option>'
    for row in rows:
        if current_id is not None and row["id"] == current_id:
            continue

        selected = "selected" if selected_id and str(selected_id) == str(row["id"]) else ""
        code = row["code"] if row["code"] else "-"
        html += f'<option value="{row["id"]}" {selected}>{code} - {row["name"]}</option>'
    return html


@router.get("/ui/accounting")
def accounting_home(request: Request):
    user = require_login(request)
    if not user:
        return RedirectResponse(url="/login", status_code=302)

    conn = get_conn()
    columns = get_account_columns()

    try:
        accounts_count = conn.execute("SELECT COUNT(*) FROM accounts").fetchone()[0]

        if "is_active" in columns:
            active_accounts = conn.execute(
                "SELECT COUNT(*) FROM accounts WHERE is_active = 1"
            ).fetchone()[0]
        else:
            active_accounts = accounts_count

        if "is_group" in columns:
            group_accounts = conn.execute(
                "SELECT COUNT(*) FROM accounts WHERE is_group = 1"
            ).fetchone()[0]
        else:
            group_accounts = 0
    finally:
        conn.close()

    content = f"""
    <div style="
        display:grid;
        grid-template-columns:repeat(3,minmax(220px,1fr));
        gap:20px;
        margin-bottom:24px;
    ">
        <div style="background:white;padding:22px;border-radius:16px;box-shadow:0 8px 24px rgba(0,0,0,0.06);">
            <div style="font-size:14px;color:#6b7280;margin-bottom:8px;">Total Accounts</div>
            <div style="font-size:34px;font-weight:bold;color:#1d4ed8;">{accounts_count}</div>
        </div>

        <div style="background:white;padding:22px;border-radius:16px;box-shadow:0 8px 24px rgba(0,0,0,0.06);">
            <div style="font-size:14px;color:#6b7280;margin-bottom:8px;">Active Accounts</div>
            <div style="font-size:34px;font-weight:bold;color:#1d4ed8;">{active_accounts}</div>
        </div>

        <div style="background:white;padding:22px;border-radius:16px;box-shadow:0 8px 24px rgba(0,0,0,0.06);">
            <div style="font-size:14px;color:#6b7280;margin-bottom:8px;">Group Accounts</div>
            <div style="font-size:34px;font-weight:bold;color:#1d4ed8;">{group_accounts}</div>
        </div>
    </div>

    <div style="background:white;padding:24px;border-radius:16px;box-shadow:0 8px 24px rgba(0,0,0,0.06);">
        <h2 style="margin-top:0;">Accounting Module</h2>
        <p style="color:#6b7280;">Start by managing your chart of accounts.</p>

        <div style="margin-top:18px;">
            <a href="/ui/accounting/accounts" style="
                display:inline-block;
                padding:12px 18px;
                background:#2563eb;
                color:white;
                text-decoration:none;
                border-radius:10px;
                margin-right:10px;
            ">View Accounts</a>

            <a href="/ui/accounting/accounts/new" style="
                display:inline-block;
                padding:12px 18px;
                background:#0f766e;
                color:white;
                text-decoration:none;
                border-radius:10px;
                margin-right:10px;
            ">Add New Account</a>

            <a href="/ui/accounting/accounts/import" style="
                display:inline-block;
                padding:12px 18px;
                background:#059669;
                color:white;
                text-decoration:none;
                border-radius:10px;
            ">Import Excel</a>
        </div>
    </div>
    """

    return HTMLResponse(render_page(content, active="accounting", user=user))


@router.get("/ui/accounting/customers")
def accounting_customers_redirect():
    return RedirectResponse(url="/ui/accounting", status_code=302)


@router.get("/ui/accounting/accounts", response_class=HTMLResponse)
def accounts_list(request: Request):
    user = require_login(request)
    if not user:
        return RedirectResponse(url="/login", status_code=302)

    columns = get_account_columns()
    has_is_group = "is_group" in columns
    has_is_active = "is_active" in columns

    select_parts = [
        "a.id",
        "a.code",
        "a.name",
        "a.type",
        "a.parent_id",
        "a.opening_balance",
        "p.name as parent_name",
    ]

    if has_is_group:
        select_parts.append("a.is_group")
    else:
        select_parts.append("0 as is_group")

    if has_is_active:
        select_parts.append("a.is_active")
    else:
        select_parts.append("1 as is_active")

    sql = f"""
        SELECT
            {", ".join(select_parts)}
        FROM accounts a
        LEFT JOIN accounts p ON a.parent_id = p.id
        ORDER BY a.code, a.id
    """

    conn = get_conn()
    try:
        rows = conn.execute(sql).fetchall()
    finally:
        conn.close()

    table_rows = ""

    for row in rows:
        code = row["code"] if row["code"] else "-"
        parent_name = row["parent_name"] if row["parent_name"] else "-"
        is_group = "Yes" if row["is_group"] == 1 else "No"
        is_active = "Yes" if row["is_active"] == 1 else "No"

        table_rows += f"""
        <tr>
            <td style="padding:14px;border-bottom:1px solid #e5e7eb;">{row["id"]}</td>
            <td style="padding:14px;border-bottom:1px solid #e5e7eb;">{code}</td>
            <td style="padding:14px;border-bottom:1px solid #e5e7eb;">{row["name"]}</td>
            <td style="padding:14px;border-bottom:1px solid #e5e7eb;">{row["type"]}</td>
            <td style="padding:14px;border-bottom:1px solid #e5e7eb;">{parent_name}</td>
            <td style="padding:14px;border-bottom:1px solid #e5e7eb;">{row["opening_balance"]}</td>
            <td style="padding:14px;border-bottom:1px solid #e5e7eb;">{is_group}</td>
            <td style="padding:14px;border-bottom:1px solid #e5e7eb;">{is_active}</td>
            <td style="padding:14px;border-bottom:1px solid #e5e7eb;">
                <a href="/ui/accounting/accounts/{row["id"]}/edit" style="
                    display:inline-block;
                    padding:8px 12px;
                    background:#f59e0b;
                    color:white;
                    text-decoration:none;
                    border-radius:8px;
                    margin-right:6px;
                ">Edit</a>

                <form method="post" action="/ui/accounting/accounts/{row["id"]}/delete" style="display:inline;">
                    <button type="submit" style="
                        padding:8px 12px;
                        background:#dc2626;
                        color:white;
                        border:none;
                        border-radius:8px;
                        cursor:pointer;
                    ">Delete</button>
                </form>
            </td>
        </tr>
        """

    if not rows:
        table_rows = """
        <tr>
            <td colspan="9" style="text-align:center;padding:20px;color:#6b7280;">
                No accounts found
            </td>
        </tr>
        """

    content = f"""
    <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:18px;">
        <h2 style="margin:0;">Chart of Accounts</h2>
        <div>
            <a href="/ui/accounting/accounts/new" style="
                display:inline-block;
                padding:12px 18px;
                background:#2563eb;
                color:white;
                text-decoration:none;
                border-radius:10px;
                margin-right:10px;
            ">Add New Account</a>

            <a href="/ui/accounting/accounts/import" style="
                display:inline-block;
                padding:12px 18px;
                background:#059669;
                color:white;
                text-decoration:none;
                border-radius:10px;
                margin-right:10px;
            ">Import Excel</a>

            <a href="/ui/accounting/accounts/template" style="
                display:inline-block;
                padding:12px 18px;
                background:#7c3aed;
                color:white;
                text-decoration:none;
                border-radius:10px;
            ">Download Template</a>
        </div>
    </div>

    <div style="background:white;border-radius:16px;box-shadow:0 8px 24px rgba(0,0,0,0.06);overflow:auto;">
        <table style="width:100%;border-collapse:collapse;">
            <thead>
                <tr style="background:#eff6ff;">
                    <th style="padding:14px;text-align:left;">ID</th>
                    <th style="padding:14px;text-align:left;">Code</th>
                    <th style="padding:14px;text-align:left;">Name</th>
                    <th style="padding:14px;text-align:left;">Type</th>
                    <th style="padding:14px;text-align:left;">Parent</th>
                    <th style="padding:14px;text-align:left;">Opening Balance</th>
                    <th style="padding:14px;text-align:left;">Group</th>
                    <th style="padding:14px;text-align:left;">Active</th>
                    <th style="padding:14px;text-align:left;">Actions</th>
                </tr>
            </thead>
            <tbody>
                {table_rows}
            </tbody>
        </table>
    </div>
    """

    return HTMLResponse(render_page(content, active="accounting", user=user))


@router.get("/ui/accounting/accounts/new", response_class=HTMLResponse)
def new_account_form(request: Request):
    user = require_login(request)
    if not user:
        return RedirectResponse(url="/login", status_code=302)

    columns = get_account_columns()
    has_is_group = "is_group" in columns
    has_is_active = "is_active" in columns

    group_html = ""
    active_html = ""

    if has_is_group:
        group_html = """
        <label style="display:block;margin-bottom:8px;font-weight:bold;">
            <input type="checkbox" name="is_group" value="1"> Group Account
        </label>
        """

    if has_is_active:
        active_html = """
        <label style="display:block;margin-bottom:8px;font-weight:bold;">
            <input type="checkbox" name="is_active" value="1" checked> Active
        </label>
        """

    content = f"""
    <div style="
        background:white;
        padding:24px;
        border-radius:16px;
        box-shadow:0 8px 24px rgba(0,0,0,0.06);
        max-width:760px;
    ">
        <h2 style="margin-top:0;">Add New Account</h2>

        <form method="post" action="/ui/accounting/accounts/new">
            <div style="margin-bottom:16px;">
                <label style="display:block;margin-bottom:6px;font-weight:bold;">Code</label>
                <input type="text" name="code" style="width:100%;padding:12px;border:1px solid #d1d5db;border-radius:10px;">
            </div>

            <div style="margin-bottom:16px;">
                <label style="display:block;margin-bottom:6px;font-weight:bold;">Name</label>
                <input type="text" name="name" required style="width:100%;padding:12px;border:1px solid #d1d5db;border-radius:10px;">
            </div>

            <div style="margin-bottom:16px;">
                <label style="display:block;margin-bottom:6px;font-weight:bold;">Type</label>
                <select name="type" required style="width:100%;padding:12px;border:1px solid #d1d5db;border-radius:10px;">
                    {account_type_options()}
                </select>
            </div>

            <div style="margin-bottom:16px;">
                <label style="display:block;margin-bottom:6px;font-weight:bold;">Parent Account</label>
                <select name="parent_id" style="width:100%;padding:12px;border:1px solid #d1d5db;border-radius:10px;">
                    {parent_account_options()}
                </select>
            </div>

            <div style="margin-bottom:16px;">
                <label style="display:block;margin-bottom:6px;font-weight:bold;">Opening Balance</label>
                <input type="number" step="0.01" name="opening_balance" value="0" style="width:100%;padding:12px;border:1px solid #d1d5db;border-radius:10px;">
            </div>

            <div style="margin-bottom:16px;">
                {group_html}
                {active_html}
            </div>

            <div style="margin-top:20px;">
                <button type="submit" style="
                    padding:12px 18px;
                    background:#2563eb;
                    color:white;
                    border:none;
                    border-radius:10px;
                    cursor:pointer;
                    margin-right:8px;
                ">Save</button>

                <a href="/ui/accounting/accounts" style="
                    display:inline-block;
                    padding:12px 18px;
                    background:#6b7280;
                    color:white;
                    text-decoration:none;
                    border-radius:10px;
                ">Back</a>
            </div>
        </form>
    </div>
    """

    return HTMLResponse(render_page(content, active="accounting", user=user))


@router.post("/ui/accounting/accounts/new")
def create_account(
    request: Request,
    code: str = Form(""),
    name: str = Form(...),
    type: str = Form(...),
    parent_id: str = Form(""),
    opening_balance: float = Form(0),
    is_group: str = Form("0"),
    is_active: str = Form("0"),
):
    user = require_login(request)
    if not user:
        return RedirectResponse(url="/login", status_code=302)

    columns = get_account_columns()

    insert_columns = ["code", "name", "type", "parent_id", "opening_balance"]
    values = [
        code.strip() if code.strip() else None,
        name.strip(),
        type.strip(),
        int(parent_id) if parent_id.strip() else None,
        float(opening_balance),
    ]

    if "is_group" in columns:
        insert_columns.append("is_group")
        values.append(1 if is_group == "1" else 0)

    if "is_active" in columns:
        insert_columns.append("is_active")
        values.append(1 if is_active == "1" else 0)

    placeholders = ", ".join(["?"] * len(insert_columns))
    cols_sql = ", ".join(insert_columns)

    conn = get_conn()
    try:
        conn.execute(
            f"INSERT INTO accounts ({cols_sql}) VALUES ({placeholders})",
            values
        )
        conn.commit()
    finally:
        conn.close()

    return RedirectResponse(url="/ui/accounting/accounts", status_code=302)


@router.get("/ui/accounting/accounts/import", response_class=HTMLResponse)
def import_accounts_form(request: Request):
    user = require_login(request)
    if not user:
        return RedirectResponse(url="/login", status_code=302)

    content = """
    <div style="
        background:white;
        padding:24px;
        border-radius:16px;
        box-shadow:0 8px 24px rgba(0,0,0,0.06);
        max-width:600px;
    ">
        <h2 style="margin-top:0;">Import Accounts from Excel</h2>

        <p style="color:#6b7280;margin-bottom:15px;">
            Required columns in Excel:
        </p>

        <div style="
            background:#f8fafc;
            border:1px solid #e5e7eb;
            padding:14px;
            border-radius:10px;
            margin-bottom:18px;
            line-height:1.8;
        ">
            code | name | type | parent_code | opening_balance
        </div>

        <p style="color:#6b7280;margin-bottom:18px;">
            Example:
            <br>1000 | Assets | asset | | 0
            <br>1100 | Cash | asset | 1000 | 0
        </p>

        <div style="margin-bottom:18px;">
            <a href="/ui/accounting/accounts/template" style="
                display:inline-block;
                padding:12px 18px;
                background:#7c3aed;
                color:white;
                text-decoration:none;
                border-radius:10px;
            ">Download Template</a>
        </div>

        <form method="post" action="/ui/accounting/accounts/import" enctype="multipart/form-data">
            <input type="file" name="file" required style="margin-bottom:15px;">
            <br>

            <button type="submit" style="
                padding:12px 18px;
                background:#059669;
                color:white;
                border:none;
                border-radius:10px;
                cursor:pointer;
                margin-right:8px;
            ">
                Upload Excel
            </button>

            <a href="/ui/accounting/accounts" style="
                display:inline-block;
                padding:12px 18px;
                background:#6b7280;
                color:white;
                text-decoration:none;
                border-radius:10px;
            ">Back</a>
        </form>
    </div>
    """

    return HTMLResponse(render_page(content, active="accounting", user=user))


@router.get("/ui/accounting/accounts/template")
def download_accounts_template(request: Request):
    user = require_login(request)
    if not user:
        return RedirectResponse(url="/login", status_code=302)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Accounts Template"

    ws.append(["code", "name", "type", "parent_code", "opening_balance"])
    ws.append(["1000", "Assets", "asset", "", 0])
    ws.append(["1100", "Cash", "asset", "1000", 0])
    ws.append(["1200", "Bank", "asset", "1000", 0])
    ws.append(["1300", "Customers", "asset", "1000", 0])
    ws.append(["1400", "Inventory", "asset", "1000", 0])
    ws.append(["2000", "Liabilities", "liability", "", 0])
    ws.append(["2100", "Vendors", "liability", "2000", 0])
    ws.append(["3000", "Equity", "equity", "", 0])
    ws.append(["3100", "Capital", "equity", "3000", 0])
    ws.append(["4000", "Revenue", "revenue", "", 0])
    ws.append(["4100", "Sales Revenue", "revenue", "4000", 0])
    ws.append(["5000", "Expenses", "expense", "", 0])
    ws.append(["5100", "Operating Expenses", "expense", "5000", 0])

    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = str(cell.value) if cell.value is not None else ""
            if len(value) > max_length:
                max_length = len(value)
        ws.column_dimensions[column_letter].width = max_length + 4

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    headers = {
        "Content-Disposition": 'attachment; filename="accounts_template.xlsx"'
    }

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )


@router.post("/ui/accounting/accounts/import")
def import_accounts(request: Request, file: UploadFile = File(...)):
    user = require_login(request)
    if not user:
        return RedirectResponse(url="/login", status_code=302)

    wb = openpyxl.load_workbook(file.file)
    sheet = wb.active
    columns = get_account_columns()

    conn = get_conn()
    cur = conn.cursor()

    try:
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row:
                continue

            code = row[0] if len(row) > 0 else None
            name = row[1] if len(row) > 1 else None
            acc_type = row[2] if len(row) > 2 else None
            opening = row[4] if len(row) > 4 else 0

            if not code or not name or not acc_type:
                continue

            existing = cur.execute(
                "SELECT id FROM accounts WHERE code = ?",
                (str(code),)
            ).fetchone()

            if existing:
                continue

            insert_columns = ["code", "name", "type", "parent_id", "opening_balance"]
            values = [str(code), str(name), str(acc_type), None, float(opening or 0)]

            if "is_group" in columns:
                insert_columns.append("is_group")
                values.append(0)

            if "is_active" in columns:
                insert_columns.append("is_active")
                values.append(1)

            placeholders = ", ".join(["?"] * len(insert_columns))
            cols_sql = ", ".join(insert_columns)

            cur.execute(
                f"INSERT INTO accounts ({cols_sql}) VALUES ({placeholders})",
                values
            )

        conn.commit()

        code_to_id = {}
        rows = conn.execute("SELECT id, code FROM accounts").fetchall()
        for r in rows:
            code_to_id[str(r["code"])] = r["id"]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row:
                continue

            code = row[0] if len(row) > 0 else None
            parent_code = row[3] if len(row) > 3 else None

            if not code:
                continue

            if parent_code and str(parent_code) in code_to_id:
                cur.execute("""
                    UPDATE accounts
                    SET parent_id = ?
                    WHERE code = ?
                """, (
                    code_to_id[str(parent_code)],
                    str(code)
                ))

        conn.commit()
    finally:
        conn.close()

    return RedirectResponse(url="/ui/accounting/accounts", status_code=302)


@router.get("/ui/accounting/accounts/{account_id}/edit", response_class=HTMLResponse)
def edit_account_form(request: Request, account_id: int):
    user = require_login(request)
    if not user:
        return RedirectResponse(url="/login", status_code=302)

    columns = get_account_columns()
    has_is_group = "is_group" in columns
    has_is_active = "is_active" in columns

    conn = get_conn()
    try:
        row = conn.execute(
            "SELECT * FROM accounts WHERE id = ?",
            (account_id,)
        ).fetchone()
    finally:
        conn.close()

    if not row:
        return HTMLResponse("Account not found", status_code=404)

    checked_group = "checked" if has_is_group and row["is_group"] == 1 else ""
    checked_active = "checked" if has_is_active and row["is_active"] == 1 else ""

    group_html = ""
    active_html = ""

    if has_is_group:
        group_html = f"""
        <label style="display:block;margin-bottom:8px;font-weight:bold;">
            <input type="checkbox" name="is_group" value="1" {checked_group}> Group Account
        </label>
        """

    if has_is_active:
        active_html = f"""
        <label style="display:block;margin-bottom:8px;font-weight:bold;">
            <input type="checkbox" name="is_active" value="1" {checked_active}> Active
        </label>
        """

    content = f"""
    <div style="
        background:white;
        padding:24px;
        border-radius:16px;
        box-shadow:0 8px 24px rgba(0,0,0,0.06);
        max-width:760px;
    ">
        <h2 style="margin-top:0;">Edit Account</h2>

        <form method="post" action="/ui/accounting/accounts/{account_id}/edit">
            <div style="margin-bottom:16px;">
                <label style="display:block;margin-bottom:6px;font-weight:bold;">Code</label>
                <input type="text" name="code" value="{row["code"] or ''}" style="width:100%;padding:12px;border:1px solid #d1d5db;border-radius:10px;">
            </div>

            <div style="margin-bottom:16px;">
                <label style="display:block;margin-bottom:6px;font-weight:bold;">Name</label>
                <input type="text" name="name" value="{row["name"]}" required style="width:100%;padding:12px;border:1px solid #d1d5db;border-radius:10px;">
            </div>

            <div style="margin-bottom:16px;">
                <label style="display:block;margin-bottom:6px;font-weight:bold;">Type</label>
                <select name="type" required style="width:100%;padding:12px;border:1px solid #d1d5db;border-radius:10px;">
                    {account_type_options(row["type"])}
                </select>
            </div>

            <div style="margin-bottom:16px;">
                <label style="display:block;margin-bottom:6px;font-weight:bold;">Parent Account</label>
                <select name="parent_id" style="width:100%;padding:12px;border:1px solid #d1d5db;border-radius:10px;">
                    {parent_account_options(row["parent_id"], account_id)}
                </select>
            </div>

            <div style="margin-bottom:16px;">
                <label style="display:block;margin-bottom:6px;font-weight:bold;">Opening Balance</label>
                <input type="number" step="0.01" name="opening_balance" value="{row["opening_balance"]}" style="width:100%;padding:12px;border:1px solid #d1d5db;border-radius:10px;">
            </div>

            <div style="margin-bottom:16px;">
                {group_html}
                {active_html}
            </div>

            <div style="margin-top:20px;">
                <button type="submit" style="
                    padding:12px 18px;
                    background:#2563eb;
                    color:white;
                    border:none;
                    border-radius:10px;
                    cursor:pointer;
                    margin-right:8px;
                ">Update</button>

                <a href="/ui/accounting/accounts" style="
                    display:inline-block;
                    padding:12px 18px;
                    background:#6b7280;
                    color:white;
                    text-decoration:none;
                    border-radius:10px;
                ">Back</a>
            </div>
        </form>
    </div>
    """

    return HTMLResponse(render_page(content, active="accounting", user=user))


@router.post("/ui/accounting/accounts/{account_id}/edit")
def update_account(
    request: Request,
    account_id: int,
    code: str = Form(""),
    name: str = Form(...),
    type: str = Form(...),
    parent_id: str = Form(""),
    opening_balance: float = Form(0),
    is_group: str = Form("0"),
    is_active: str = Form("0"),
):
    user = require_login(request)
    if not user:
        return RedirectResponse(url="/login", status_code=302)

    columns = get_account_columns()

    set_parts = [
        "code = ?",
        "name = ?",
        "type = ?",
        "parent_id = ?",
        "opening_balance = ?",
    ]

    values = [
        code.strip() if code.strip() else None,
        name.strip(),
        type.strip(),
        int(parent_id) if parent_id.strip() else None,
        float(opening_balance),
    ]

    if "is_group" in columns:
        set_parts.append("is_group = ?")
        values.append(1 if is_group == "1" else 0)

    if "is_active" in columns:
        set_parts.append("is_active = ?")
        values.append(1 if is_active == "1" else 0)

    values.append(account_id)

    conn = get_conn()
    try:
        conn.execute(
            f"UPDATE accounts SET {', '.join(set_parts)} WHERE id = ?",
            values
        )
        conn.commit()
    finally:
        conn.close()

    return RedirectResponse(url="/ui/accounting/accounts", status_code=302)


@router.post("/ui/accounting/accounts/{account_id}/delete")
def delete_account(request: Request, account_id: int):
    user = require_login(request)
    if not user:
        return RedirectResponse(url="/login", status_code=302)

    conn = get_conn()
    try:
        conn.execute("DELETE FROM accounts WHERE id = ?", (account_id,))
        conn.commit()
    finally:
        conn.close()

    return RedirectResponse(url="/ui/accounting/accounts", status_code=302)