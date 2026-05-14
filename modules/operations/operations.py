import csv
import io
import os
import uuid
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from urllib.parse import quote

from fastapi import APIRouter, File, Form, Request, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse

from audit import actor_name_from_request, render_audit_log_card, safe_log_action
from db import get_conn
from i18n import get_lang
from layout import render_page

try:
    from openpyxl import Workbook, load_workbook
except Exception:
    Workbook = None
    load_workbook = None

router = APIRouter()


SERVICE_CATEGORIES = [
    ("field_maintenance", "Field Maintenance"),
    ("workshop_repair", "Workshop Repair"),
    ("cabinet_assembly", "Cabinet Assembly"),
]

ZONE_LEVELS = [
    ("zone_1", "Zone 1"),
    ("zone_2", "Zone 2"),
    ("zone_3", "Zone 3"),
    ("zone_4", "Zone 4"),
]

WORK_ORDER_STATUSES = [
    ("new", "New"),
    ("assigned", "Assigned"),
    ("in_progress", "In Progress"),
    ("submitted", "Submitted"),
    ("approved", "Approved"),
    ("rejected", "Rejected"),
    ("closed", "Closed"),
]

WORKFLOW_TYPES = [
    ("workshop_repair", "Workshop Repair"),
    ("field_service", "Field Service"),
    ("planning", "Orange Planning"),
]

FIELD_ACTION_TYPES = [
    ("repair", "Repair"),
    ("technical_visit", "Technical Visit"),
    ("swap", "Swap"),
    ("install", "Install"),
]

ROLLOUT_STATUSES = [
    ("not_required", "Not Required"),
    ("pending", "Pending"),
    ("approved", "Approved"),
    ("rejected", "Rejected"),
]

TRIP_STATUSES = [
    ("draft", "Draft"),
    ("dispatched", "Dispatched"),
    ("completed", "Completed"),
    ("approved", "Approved"),
]

DRIVER_SOURCE_OPTIONS = [
    ("company_driver", "Company Driver"),
    ("supplier_driver", "Supplier Driver"),
]

OPS_DEPARTMENTS = [
    ("operation", "Operation"),
    ("planning", "Planning"),
]

CUSTODY_WAREHOUSE_TYPES = [
    ("operation", "Operation Warehouse"),
    ("planning", "Planning Warehouse"),
    ("repair", "Repair Warehouse"),
]

CUSTOMER_MODULE_STATUSES = [
    ("faulty", "Faulty"),
    ("under_repair", "Under Repair"),
    ("working", "Working"),
    ("installed", "Installed"),
    ("returned", "Returned"),
    ("scrap", "Scrap"),
]

CUSTODY_MOVEMENT_TYPES = [
    ("receipt", "Receive From Customer"),
    ("issue_to_repair", "Issue To Repair"),
    ("return_from_repair", "Return From Repair"),
    ("site_issue", "Issue To Site"),
    ("site_return", "Return From Site"),
    ("swap_removed", "Swap Removed Faulty Unit"),
    ("adjustment", "Adjustment"),
]


def safe(value):
    return "" if value is None else str(value).strip()


def safe_int(value, default=0):
    try:
        if value is None or str(value).strip() == "":
            return default
        return int(float(value))
    except Exception:
        return default


def to_decimal(value, default="0"):
    try:
        text = safe(value).replace(",", "")
        if text in ["", ".", "-", "-."]:
            text = default
        return Decimal(text)
    except (InvalidOperation, ValueError, TypeError):
        return Decimal(default)


def q2(value):
    return to_decimal(value).quantize(Decimal("1.00"), rounding=ROUND_HALF_UP)


def money(value):
    return f"{q2(value):,.2f}"


def ensure_column(conn, table_name, column_name, alter_sql):
    cols = conn.execute(f"PRAGMA table_info({table_name})").fetchall()
    names = [c["name"] for c in cols]
    if column_name not in names:
        conn.execute(alter_sql)


def ensure_tables():
    conn = get_conn()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS ops_contract_companies (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT,
            name TEXT NOT NULL,
            contact_person TEXT,
            phone TEXT,
            email TEXT,
            notes TEXT,
            is_active INTEGER DEFAULT 1,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS ops_contracts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            contract_no TEXT,
            company_id INTEGER NOT NULL,
            contract_name TEXT NOT NULL,
            pricing_method TEXT DEFAULT 'standard',
            start_date TEXT,
            end_date TEXT,
            status TEXT DEFAULT 'active',
            notes TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS ops_rental_suppliers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT,
            name TEXT NOT NULL,
            contact_person TEXT,
            phone TEXT,
            notes TEXT,
            is_active INTEGER DEFAULT 1,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS ops_fault_types (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT,
            name TEXT NOT NULL,
            service_category TEXT,
            default_service_price REAL DEFAULT 0,
            default_incentive REAL DEFAULT 0,
            notes TEXT,
            is_active INTEGER DEFAULT 1,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS ops_regions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT,
            name TEXT NOT NULL,
            zone_level TEXT,
            allowance_amount REAL DEFAULT 0,
            notes TEXT,
            is_active INTEGER DEFAULT 1,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS ops_service_catalog (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT,
            name TEXT NOT NULL,
            service_category TEXT,
            unit_price REAL DEFAULT 0,
            technician_incentive REAL DEFAULT 0,
            default_region_level TEXT,
            default_duration_hours REAL DEFAULT 0,
            notes TEXT,
            is_active INTEGER DEFAULT 1,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS ops_vehicle_rates (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT,
            rental_supplier_id INTEGER,
            slab_name TEXT,
            vehicle_type TEXT NOT NULL,
            ticket_open_price REAL DEFAULT 0,
            second_slab_to_km REAL DEFAULT 300,
            km_rate_101_300 REAL DEFAULT 0,
            km_rate_over_300 REAL DEFAULT 0,
            km_rate REAL DEFAULT 0,
            waiting_hour_rate REAL DEFAULT 0,
            notes TEXT,
            is_active INTEGER DEFAULT 1,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS ops_vehicles (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT,
            vehicle_name TEXT,
            vehicle_type TEXT NOT NULL,
            rental_supplier_id INTEGER,
            vehicle_rate_id INTEGER,
            pricing_slab_name TEXT,
            plate_no TEXT,
            driver_source TEXT DEFAULT 'company_driver',
            supplier_driver_name TEXT,
            notes TEXT,
            is_active INTEGER DEFAULT 1,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS ops_action_price_versions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            contract_id INTEGER NOT NULL,
            action_id INTEGER NOT NULL,
            version_name TEXT,
            effective_from TEXT NOT NULL,
            effective_to TEXT,
            fuel_reference TEXT,
            action_price REAL DEFAULT 0,
            technician_incentive REAL DEFAULT 0,
            region_allowance REAL DEFAULT 0,
            is_active INTEGER DEFAULT 1,
            notes TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS ops_vehicle_price_versions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            contract_id INTEGER NOT NULL,
            vehicle_rate_id INTEGER NOT NULL,
            version_name TEXT,
            effective_from TEXT NOT NULL,
            effective_to TEXT,
            fuel_reference TEXT,
            ticket_open_price REAL DEFAULT 0,
            second_slab_to_km REAL DEFAULT 300,
            km_rate_101_300 REAL DEFAULT 0,
            km_rate_over_300 REAL DEFAULT 0,
            waiting_hour_rate REAL DEFAULT 0,
            is_active INTEGER DEFAULT 1,
            notes TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS ops_action_materials (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            action_id INTEGER NOT NULL,
            line_no INTEGER DEFAULT 1,
            item_id INTEGER,
            item_code TEXT,
            item_name TEXT,
            uom TEXT,
            qty REAL DEFAULT 0,
            unit_cost REAL DEFAULT 0,
            notes TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS ops_work_orders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            work_order_no TEXT,
            company_id INTEGER,
            fault_type_id INTEGER,
            service_id INTEGER,
            region_id INTEGER,
            request_date TEXT,
            site_code TEXT,
            site_name TEXT,
            complaint_details TEXT,
            required_materials TEXT,
            technician_id INTEGER,
            manager_id INTEGER,
            priority TEXT DEFAULT 'normal',
            trip_required INTEGER DEFAULT 1,
            service_price REAL DEFAULT 0,
            technician_incentive REAL DEFAULT 0,
            region_allowance REAL DEFAULT 0,
            department TEXT DEFAULT 'operation',
            workflow_type TEXT DEFAULT 'field_service',
            action_type TEXT,
            customer_warehouse_id INTEGER,
            ticket_id INTEGER,
            requested_qty REAL DEFAULT 0,
            issued_qty REAL DEFAULT 0,
            completed_qty REAL DEFAULT 0,
            returned_qty REAL DEFAULT 0,
            rollout_status TEXT DEFAULT 'not_required',
            rollout_notes TEXT,
            rollout_by TEXT,
            rollout_at TEXT,
            actual_actions TEXT,
            status TEXT DEFAULT 'new',
            created_by TEXT,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
            completed_at TEXT,
            closure_notes TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS ops_tickets (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ticket_no TEXT,
            company_id INTEGER NOT NULL,
            contract_id INTEGER,
            ticket_date TEXT,
            fault_type_id INTEGER,
            site_code TEXT,
            site_name TEXT,
            priority TEXT DEFAULT 'normal',
            request_channel TEXT,
            complaint_details TEXT,
            status TEXT DEFAULT 'open',
            created_by TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS ops_trip_tickets (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            trip_no TEXT,
            trip_date TEXT,
            vehicle_id INTEGER NOT NULL,
            rental_supplier_id INTEGER,
            vehicle_type TEXT,
            vehicle_rate_id INTEGER,
            driver_source TEXT DEFAULT 'company_driver',
            driver_employee_id INTEGER,
            supplier_driver_name TEXT,
            status TEXT DEFAULT 'draft',
            start_odometer REAL DEFAULT 0,
            start_photo_path TEXT,
            end_odometer REAL DEFAULT 0,
            end_photo_path TEXT,
            waiting_hours REAL DEFAULT 0,
            total_km REAL DEFAULT 0,
            ticket_open_price REAL DEFAULT 0,
            second_slab_to_km REAL DEFAULT 300,
            km_rate_101_300 REAL DEFAULT 0,
            km_rate_over_300 REAL DEFAULT 0,
            waiting_hour_rate REAL DEFAULT 0,
            total_cost REAL DEFAULT 0,
            allocated_cost_per_work_order REAL DEFAULT 0,
            driver_commission_pct REAL DEFAULT 0,
            driver_commission_amount REAL DEFAULT 0,
            notes TEXT,
            movement_notes TEXT,
            accounting_notes TEXT,
            created_by TEXT,
            movement_closed_by TEXT,
            approved_by TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            completed_at TEXT,
            approved_at TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS ops_trip_work_orders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            trip_id INTEGER NOT NULL,
            work_order_id INTEGER NOT NULL,
            line_no INTEGER DEFAULT 1,
            allocated_cost REAL DEFAULT 0,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS ops_technician_reports (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            work_order_id INTEGER NOT NULL,
            report_date TEXT,
            arrival_time TEXT,
            completion_time TEXT,
            issue_found TEXT,
            action_taken TEXT,
            materials_used TEXT,
            technician_notes TEXT,
            customer_notes TEXT,
            report_status TEXT DEFAULT 'draft',
            service_price REAL DEFAULT 0,
            technician_incentive REAL DEFAULT 0,
            region_allowance REAL DEFAULT 0,
            review_notes TEXT,
            submitted_by TEXT,
            submitted_at TEXT,
            reviewed_by TEXT,
            reviewed_at TEXT,
            status TEXT DEFAULT 'new',
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS ops_customer_custody_warehouses (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT,
            company_id INTEGER NOT NULL,
            department TEXT DEFAULT 'operation',
            warehouse_type TEXT DEFAULT 'operation',
            name TEXT NOT NULL,
            location TEXT,
            notes TEXT,
            is_active INTEGER DEFAULT 1,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS ops_customer_custody_stock (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            company_id INTEGER NOT NULL,
            department TEXT DEFAULT 'operation',
            warehouse_id INTEGER NOT NULL,
            item_id INTEGER DEFAULT 0,
            module_code TEXT,
            module_name TEXT,
            serial_no TEXT,
            status TEXT DEFAULT 'faulty',
            qty REAL DEFAULT 0,
            uom TEXT,
            notes TEXT,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS ops_customer_custody_transactions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            transaction_no TEXT,
            transaction_date TEXT,
            company_id INTEGER NOT NULL,
            department TEXT DEFAULT 'operation',
            warehouse_id INTEGER NOT NULL,
            ticket_id INTEGER,
            work_order_id INTEGER,
            item_id INTEGER DEFAULT 0,
            module_code TEXT,
            module_name TEXT,
            serial_no TEXT,
            movement_type TEXT,
            from_status TEXT,
            to_status TEXT,
            qty REAL DEFAULT 0,
            uom TEXT,
            notes TEXT,
            created_by TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """)
    ensure_column(conn, "ops_customer_custody_warehouses", "code", "ALTER TABLE ops_customer_custody_warehouses ADD COLUMN code TEXT")
    ensure_column(conn, "ops_customer_custody_warehouses", "company_id", "ALTER TABLE ops_customer_custody_warehouses ADD COLUMN company_id INTEGER")
    ensure_column(conn, "ops_customer_custody_warehouses", "department", "ALTER TABLE ops_customer_custody_warehouses ADD COLUMN department TEXT DEFAULT 'operation'")
    ensure_column(conn, "ops_customer_custody_warehouses", "warehouse_type", "ALTER TABLE ops_customer_custody_warehouses ADD COLUMN warehouse_type TEXT DEFAULT 'operation'")
    ensure_column(conn, "ops_customer_custody_warehouses", "name", "ALTER TABLE ops_customer_custody_warehouses ADD COLUMN name TEXT")
    ensure_column(conn, "ops_customer_custody_warehouses", "location", "ALTER TABLE ops_customer_custody_warehouses ADD COLUMN location TEXT")
    ensure_column(conn, "ops_customer_custody_warehouses", "notes", "ALTER TABLE ops_customer_custody_warehouses ADD COLUMN notes TEXT")
    ensure_column(conn, "ops_customer_custody_warehouses", "is_active", "ALTER TABLE ops_customer_custody_warehouses ADD COLUMN is_active INTEGER DEFAULT 1")
    ensure_column(conn, "ops_customer_custody_stock", "company_id", "ALTER TABLE ops_customer_custody_stock ADD COLUMN company_id INTEGER")
    ensure_column(conn, "ops_customer_custody_stock", "department", "ALTER TABLE ops_customer_custody_stock ADD COLUMN department TEXT DEFAULT 'operation'")
    ensure_column(conn, "ops_customer_custody_stock", "warehouse_id", "ALTER TABLE ops_customer_custody_stock ADD COLUMN warehouse_id INTEGER")
    ensure_column(conn, "ops_customer_custody_stock", "item_id", "ALTER TABLE ops_customer_custody_stock ADD COLUMN item_id INTEGER DEFAULT 0")
    ensure_column(conn, "ops_customer_custody_stock", "module_code", "ALTER TABLE ops_customer_custody_stock ADD COLUMN module_code TEXT")
    ensure_column(conn, "ops_customer_custody_stock", "module_name", "ALTER TABLE ops_customer_custody_stock ADD COLUMN module_name TEXT")
    ensure_column(conn, "ops_customer_custody_stock", "serial_no", "ALTER TABLE ops_customer_custody_stock ADD COLUMN serial_no TEXT")
    ensure_column(conn, "ops_customer_custody_stock", "status", "ALTER TABLE ops_customer_custody_stock ADD COLUMN status TEXT DEFAULT 'faulty'")
    ensure_column(conn, "ops_customer_custody_stock", "qty", "ALTER TABLE ops_customer_custody_stock ADD COLUMN qty REAL DEFAULT 0")
    ensure_column(conn, "ops_customer_custody_stock", "uom", "ALTER TABLE ops_customer_custody_stock ADD COLUMN uom TEXT")
    ensure_column(conn, "ops_customer_custody_stock", "notes", "ALTER TABLE ops_customer_custody_stock ADD COLUMN notes TEXT")
    ensure_column(conn, "ops_customer_custody_stock", "updated_at", "ALTER TABLE ops_customer_custody_stock ADD COLUMN updated_at TEXT DEFAULT CURRENT_TIMESTAMP")
    ensure_column(conn, "ops_customer_custody_transactions", "transaction_no", "ALTER TABLE ops_customer_custody_transactions ADD COLUMN transaction_no TEXT")
    ensure_column(conn, "ops_customer_custody_transactions", "transaction_date", "ALTER TABLE ops_customer_custody_transactions ADD COLUMN transaction_date TEXT")
    ensure_column(conn, "ops_customer_custody_transactions", "company_id", "ALTER TABLE ops_customer_custody_transactions ADD COLUMN company_id INTEGER")
    ensure_column(conn, "ops_customer_custody_transactions", "department", "ALTER TABLE ops_customer_custody_transactions ADD COLUMN department TEXT DEFAULT 'operation'")
    ensure_column(conn, "ops_customer_custody_transactions", "warehouse_id", "ALTER TABLE ops_customer_custody_transactions ADD COLUMN warehouse_id INTEGER")
    ensure_column(conn, "ops_customer_custody_transactions", "ticket_id", "ALTER TABLE ops_customer_custody_transactions ADD COLUMN ticket_id INTEGER")
    ensure_column(conn, "ops_customer_custody_transactions", "work_order_id", "ALTER TABLE ops_customer_custody_transactions ADD COLUMN work_order_id INTEGER")
    ensure_column(conn, "ops_customer_custody_transactions", "item_id", "ALTER TABLE ops_customer_custody_transactions ADD COLUMN item_id INTEGER DEFAULT 0")
    ensure_column(conn, "ops_customer_custody_transactions", "module_code", "ALTER TABLE ops_customer_custody_transactions ADD COLUMN module_code TEXT")
    ensure_column(conn, "ops_customer_custody_transactions", "module_name", "ALTER TABLE ops_customer_custody_transactions ADD COLUMN module_name TEXT")
    ensure_column(conn, "ops_customer_custody_transactions", "serial_no", "ALTER TABLE ops_customer_custody_transactions ADD COLUMN serial_no TEXT")
    ensure_column(conn, "ops_customer_custody_transactions", "movement_type", "ALTER TABLE ops_customer_custody_transactions ADD COLUMN movement_type TEXT")
    ensure_column(conn, "ops_customer_custody_transactions", "from_status", "ALTER TABLE ops_customer_custody_transactions ADD COLUMN from_status TEXT")
    ensure_column(conn, "ops_customer_custody_transactions", "to_status", "ALTER TABLE ops_customer_custody_transactions ADD COLUMN to_status TEXT")
    ensure_column(conn, "ops_customer_custody_transactions", "qty", "ALTER TABLE ops_customer_custody_transactions ADD COLUMN qty REAL DEFAULT 0")
    ensure_column(conn, "ops_customer_custody_transactions", "uom", "ALTER TABLE ops_customer_custody_transactions ADD COLUMN uom TEXT")
    ensure_column(conn, "ops_customer_custody_transactions", "notes", "ALTER TABLE ops_customer_custody_transactions ADD COLUMN notes TEXT")
    ensure_column(conn, "ops_customer_custody_transactions", "created_by", "ALTER TABLE ops_customer_custody_transactions ADD COLUMN created_by TEXT")
    ensure_column(conn, "ops_work_orders", "request_date", "ALTER TABLE ops_work_orders ADD COLUMN request_date TEXT")
    ensure_column(conn, "ops_work_orders", "site_code", "ALTER TABLE ops_work_orders ADD COLUMN site_code TEXT")
    ensure_column(conn, "ops_work_orders", "site_name", "ALTER TABLE ops_work_orders ADD COLUMN site_name TEXT")
    ensure_column(conn, "ops_work_orders", "complaint_details", "ALTER TABLE ops_work_orders ADD COLUMN complaint_details TEXT")
    ensure_column(conn, "ops_work_orders", "required_materials", "ALTER TABLE ops_work_orders ADD COLUMN required_materials TEXT")
    ensure_column(conn, "ops_work_orders", "technician_id", "ALTER TABLE ops_work_orders ADD COLUMN technician_id INTEGER")
    ensure_column(conn, "ops_work_orders", "manager_id", "ALTER TABLE ops_work_orders ADD COLUMN manager_id INTEGER")
    ensure_column(conn, "ops_work_orders", "priority", "ALTER TABLE ops_work_orders ADD COLUMN priority TEXT DEFAULT 'normal'")
    ensure_column(conn, "ops_work_orders", "trip_required", "ALTER TABLE ops_work_orders ADD COLUMN trip_required INTEGER DEFAULT 1")
    ensure_column(conn, "ops_work_orders", "service_price", "ALTER TABLE ops_work_orders ADD COLUMN service_price REAL DEFAULT 0")
    ensure_column(conn, "ops_work_orders", "technician_incentive", "ALTER TABLE ops_work_orders ADD COLUMN technician_incentive REAL DEFAULT 0")
    ensure_column(conn, "ops_work_orders", "region_allowance", "ALTER TABLE ops_work_orders ADD COLUMN region_allowance REAL DEFAULT 0")
    ensure_column(conn, "ops_work_orders", "department", "ALTER TABLE ops_work_orders ADD COLUMN department TEXT DEFAULT 'operation'")
    ensure_column(conn, "ops_work_orders", "workflow_type", "ALTER TABLE ops_work_orders ADD COLUMN workflow_type TEXT DEFAULT 'field_service'")
    ensure_column(conn, "ops_work_orders", "action_type", "ALTER TABLE ops_work_orders ADD COLUMN action_type TEXT")
    ensure_column(conn, "ops_work_orders", "customer_warehouse_id", "ALTER TABLE ops_work_orders ADD COLUMN customer_warehouse_id INTEGER")
    ensure_column(conn, "ops_work_orders", "ticket_id", "ALTER TABLE ops_work_orders ADD COLUMN ticket_id INTEGER")
    ensure_column(conn, "ops_work_orders", "requested_qty", "ALTER TABLE ops_work_orders ADD COLUMN requested_qty REAL DEFAULT 0")
    ensure_column(conn, "ops_work_orders", "issued_qty", "ALTER TABLE ops_work_orders ADD COLUMN issued_qty REAL DEFAULT 0")
    ensure_column(conn, "ops_work_orders", "completed_qty", "ALTER TABLE ops_work_orders ADD COLUMN completed_qty REAL DEFAULT 0")
    ensure_column(conn, "ops_work_orders", "returned_qty", "ALTER TABLE ops_work_orders ADD COLUMN returned_qty REAL DEFAULT 0")
    ensure_column(conn, "ops_work_orders", "rollout_status", "ALTER TABLE ops_work_orders ADD COLUMN rollout_status TEXT DEFAULT 'not_required'")
    ensure_column(conn, "ops_work_orders", "rollout_notes", "ALTER TABLE ops_work_orders ADD COLUMN rollout_notes TEXT")
    ensure_column(conn, "ops_work_orders", "rollout_by", "ALTER TABLE ops_work_orders ADD COLUMN rollout_by TEXT")
    ensure_column(conn, "ops_work_orders", "rollout_at", "ALTER TABLE ops_work_orders ADD COLUMN rollout_at TEXT")
    ensure_column(conn, "ops_work_orders", "actual_actions", "ALTER TABLE ops_work_orders ADD COLUMN actual_actions TEXT")
    ensure_column(conn, "ops_work_orders", "created_by", "ALTER TABLE ops_work_orders ADD COLUMN created_by TEXT")
    ensure_column(conn, "ops_work_orders", "updated_at", "ALTER TABLE ops_work_orders ADD COLUMN updated_at TEXT DEFAULT CURRENT_TIMESTAMP")
    ensure_column(conn, "ops_work_orders", "completed_at", "ALTER TABLE ops_work_orders ADD COLUMN completed_at TEXT")
    ensure_column(conn, "ops_work_orders", "closure_notes", "ALTER TABLE ops_work_orders ADD COLUMN closure_notes TEXT")
    ensure_column(conn, "ops_technician_reports", "report_date", "ALTER TABLE ops_technician_reports ADD COLUMN report_date TEXT")
    ensure_column(conn, "ops_technician_reports", "arrival_time", "ALTER TABLE ops_technician_reports ADD COLUMN arrival_time TEXT")
    ensure_column(conn, "ops_technician_reports", "completion_time", "ALTER TABLE ops_technician_reports ADD COLUMN completion_time TEXT")
    ensure_column(conn, "ops_technician_reports", "issue_found", "ALTER TABLE ops_technician_reports ADD COLUMN issue_found TEXT")
    ensure_column(conn, "ops_technician_reports", "action_taken", "ALTER TABLE ops_technician_reports ADD COLUMN action_taken TEXT")
    ensure_column(conn, "ops_technician_reports", "materials_used", "ALTER TABLE ops_technician_reports ADD COLUMN materials_used TEXT")
    ensure_column(conn, "ops_technician_reports", "technician_notes", "ALTER TABLE ops_technician_reports ADD COLUMN technician_notes TEXT")
    ensure_column(conn, "ops_technician_reports", "customer_notes", "ALTER TABLE ops_technician_reports ADD COLUMN customer_notes TEXT")
    ensure_column(conn, "ops_technician_reports", "report_status", "ALTER TABLE ops_technician_reports ADD COLUMN report_status TEXT DEFAULT 'draft'")
    ensure_column(conn, "ops_technician_reports", "service_price", "ALTER TABLE ops_technician_reports ADD COLUMN service_price REAL DEFAULT 0")
    ensure_column(conn, "ops_technician_reports", "technician_incentive", "ALTER TABLE ops_technician_reports ADD COLUMN technician_incentive REAL DEFAULT 0")
    ensure_column(conn, "ops_technician_reports", "region_allowance", "ALTER TABLE ops_technician_reports ADD COLUMN region_allowance REAL DEFAULT 0")
    ensure_column(conn, "ops_technician_reports", "review_notes", "ALTER TABLE ops_technician_reports ADD COLUMN review_notes TEXT")
    ensure_column(conn, "ops_technician_reports", "submitted_by", "ALTER TABLE ops_technician_reports ADD COLUMN submitted_by TEXT")
    ensure_column(conn, "ops_technician_reports", "submitted_at", "ALTER TABLE ops_technician_reports ADD COLUMN submitted_at TEXT")
    ensure_column(conn, "ops_technician_reports", "reviewed_by", "ALTER TABLE ops_technician_reports ADD COLUMN reviewed_by TEXT")
    ensure_column(conn, "ops_technician_reports", "reviewed_at", "ALTER TABLE ops_technician_reports ADD COLUMN reviewed_at TEXT")
    ensure_column(conn, "ops_action_materials", "action_id", "ALTER TABLE ops_action_materials ADD COLUMN action_id INTEGER")
    ensure_column(conn, "ops_action_materials", "line_no", "ALTER TABLE ops_action_materials ADD COLUMN line_no INTEGER DEFAULT 1")
    ensure_column(conn, "ops_action_materials", "item_id", "ALTER TABLE ops_action_materials ADD COLUMN item_id INTEGER")
    ensure_column(conn, "ops_action_materials", "item_code", "ALTER TABLE ops_action_materials ADD COLUMN item_code TEXT")
    ensure_column(conn, "ops_action_materials", "item_name", "ALTER TABLE ops_action_materials ADD COLUMN item_name TEXT")
    ensure_column(conn, "ops_action_materials", "uom", "ALTER TABLE ops_action_materials ADD COLUMN uom TEXT")
    ensure_column(conn, "ops_action_materials", "qty", "ALTER TABLE ops_action_materials ADD COLUMN qty REAL DEFAULT 0")
    ensure_column(conn, "ops_action_materials", "unit_cost", "ALTER TABLE ops_action_materials ADD COLUMN unit_cost REAL DEFAULT 0")
    ensure_column(conn, "ops_action_materials", "notes", "ALTER TABLE ops_action_materials ADD COLUMN notes TEXT")
    ensure_column(conn, "ops_vehicle_rates", "ticket_open_price", "ALTER TABLE ops_vehicle_rates ADD COLUMN ticket_open_price REAL DEFAULT 0")
    ensure_column(conn, "ops_vehicle_rates", "rental_supplier_id", "ALTER TABLE ops_vehicle_rates ADD COLUMN rental_supplier_id INTEGER")
    ensure_column(conn, "ops_vehicle_rates", "slab_name", "ALTER TABLE ops_vehicle_rates ADD COLUMN slab_name TEXT")
    ensure_column(conn, "ops_vehicle_rates", "second_slab_to_km", "ALTER TABLE ops_vehicle_rates ADD COLUMN second_slab_to_km REAL DEFAULT 300")
    ensure_column(conn, "ops_vehicle_rates", "km_rate_101_300", "ALTER TABLE ops_vehicle_rates ADD COLUMN km_rate_101_300 REAL DEFAULT 0")
    ensure_column(conn, "ops_vehicle_rates", "km_rate_over_300", "ALTER TABLE ops_vehicle_rates ADD COLUMN km_rate_over_300 REAL DEFAULT 0")
    ensure_column(conn, "ops_rental_suppliers", "code", "ALTER TABLE ops_rental_suppliers ADD COLUMN code TEXT")
    ensure_column(conn, "ops_rental_suppliers", "contact_person", "ALTER TABLE ops_rental_suppliers ADD COLUMN contact_person TEXT")
    ensure_column(conn, "ops_rental_suppliers", "phone", "ALTER TABLE ops_rental_suppliers ADD COLUMN phone TEXT")
    ensure_column(conn, "ops_rental_suppliers", "notes", "ALTER TABLE ops_rental_suppliers ADD COLUMN notes TEXT")
    ensure_column(conn, "ops_rental_suppliers", "is_active", "ALTER TABLE ops_rental_suppliers ADD COLUMN is_active INTEGER DEFAULT 1")
    ensure_column(conn, "ops_vehicles", "code", "ALTER TABLE ops_vehicles ADD COLUMN code TEXT")
    ensure_column(conn, "ops_vehicles", "vehicle_name", "ALTER TABLE ops_vehicles ADD COLUMN vehicle_name TEXT")
    ensure_column(conn, "ops_vehicles", "vehicle_type", "ALTER TABLE ops_vehicles ADD COLUMN vehicle_type TEXT")
    ensure_column(conn, "ops_vehicles", "rental_supplier_id", "ALTER TABLE ops_vehicles ADD COLUMN rental_supplier_id INTEGER")
    ensure_column(conn, "ops_vehicles", "vehicle_rate_id", "ALTER TABLE ops_vehicles ADD COLUMN vehicle_rate_id INTEGER")
    ensure_column(conn, "ops_vehicles", "pricing_slab_name", "ALTER TABLE ops_vehicles ADD COLUMN pricing_slab_name TEXT")
    ensure_column(conn, "ops_vehicles", "plate_no", "ALTER TABLE ops_vehicles ADD COLUMN plate_no TEXT")
    ensure_column(conn, "ops_vehicles", "driver_source", "ALTER TABLE ops_vehicles ADD COLUMN driver_source TEXT DEFAULT 'company_driver'")
    ensure_column(conn, "ops_vehicles", "supplier_driver_name", "ALTER TABLE ops_vehicles ADD COLUMN supplier_driver_name TEXT")
    ensure_column(conn, "ops_vehicles", "notes", "ALTER TABLE ops_vehicles ADD COLUMN notes TEXT")
    ensure_column(conn, "ops_vehicles", "is_active", "ALTER TABLE ops_vehicles ADD COLUMN is_active INTEGER DEFAULT 1")
    ensure_column(conn, "ops_contracts", "contract_no", "ALTER TABLE ops_contracts ADD COLUMN contract_no TEXT")
    ensure_column(conn, "ops_contracts", "company_id", "ALTER TABLE ops_contracts ADD COLUMN company_id INTEGER")
    ensure_column(conn, "ops_contracts", "contract_name", "ALTER TABLE ops_contracts ADD COLUMN contract_name TEXT")
    ensure_column(conn, "ops_contracts", "pricing_method", "ALTER TABLE ops_contracts ADD COLUMN pricing_method TEXT DEFAULT 'standard'")
    ensure_column(conn, "ops_contracts", "start_date", "ALTER TABLE ops_contracts ADD COLUMN start_date TEXT")
    ensure_column(conn, "ops_contracts", "end_date", "ALTER TABLE ops_contracts ADD COLUMN end_date TEXT")
    ensure_column(conn, "ops_contracts", "status", "ALTER TABLE ops_contracts ADD COLUMN status TEXT DEFAULT 'active'")
    ensure_column(conn, "ops_contracts", "notes", "ALTER TABLE ops_contracts ADD COLUMN notes TEXT")
    ensure_column(conn, "ops_action_price_versions", "contract_id", "ALTER TABLE ops_action_price_versions ADD COLUMN contract_id INTEGER")
    ensure_column(conn, "ops_action_price_versions", "action_id", "ALTER TABLE ops_action_price_versions ADD COLUMN action_id INTEGER")
    ensure_column(conn, "ops_action_price_versions", "version_name", "ALTER TABLE ops_action_price_versions ADD COLUMN version_name TEXT")
    ensure_column(conn, "ops_action_price_versions", "effective_from", "ALTER TABLE ops_action_price_versions ADD COLUMN effective_from TEXT")
    ensure_column(conn, "ops_action_price_versions", "effective_to", "ALTER TABLE ops_action_price_versions ADD COLUMN effective_to TEXT")
    ensure_column(conn, "ops_action_price_versions", "fuel_reference", "ALTER TABLE ops_action_price_versions ADD COLUMN fuel_reference TEXT")
    ensure_column(conn, "ops_action_price_versions", "action_price", "ALTER TABLE ops_action_price_versions ADD COLUMN action_price REAL DEFAULT 0")
    ensure_column(conn, "ops_action_price_versions", "technician_incentive", "ALTER TABLE ops_action_price_versions ADD COLUMN technician_incentive REAL DEFAULT 0")
    ensure_column(conn, "ops_action_price_versions", "region_allowance", "ALTER TABLE ops_action_price_versions ADD COLUMN region_allowance REAL DEFAULT 0")
    ensure_column(conn, "ops_action_price_versions", "is_active", "ALTER TABLE ops_action_price_versions ADD COLUMN is_active INTEGER DEFAULT 1")
    ensure_column(conn, "ops_action_price_versions", "notes", "ALTER TABLE ops_action_price_versions ADD COLUMN notes TEXT")
    ensure_column(conn, "ops_vehicle_price_versions", "contract_id", "ALTER TABLE ops_vehicle_price_versions ADD COLUMN contract_id INTEGER")
    ensure_column(conn, "ops_vehicle_price_versions", "vehicle_rate_id", "ALTER TABLE ops_vehicle_price_versions ADD COLUMN vehicle_rate_id INTEGER")
    ensure_column(conn, "ops_vehicle_price_versions", "version_name", "ALTER TABLE ops_vehicle_price_versions ADD COLUMN version_name TEXT")
    ensure_column(conn, "ops_vehicle_price_versions", "effective_from", "ALTER TABLE ops_vehicle_price_versions ADD COLUMN effective_from TEXT")
    ensure_column(conn, "ops_vehicle_price_versions", "effective_to", "ALTER TABLE ops_vehicle_price_versions ADD COLUMN effective_to TEXT")
    ensure_column(conn, "ops_vehicle_price_versions", "fuel_reference", "ALTER TABLE ops_vehicle_price_versions ADD COLUMN fuel_reference TEXT")
    ensure_column(conn, "ops_vehicle_price_versions", "ticket_open_price", "ALTER TABLE ops_vehicle_price_versions ADD COLUMN ticket_open_price REAL DEFAULT 0")
    ensure_column(conn, "ops_vehicle_price_versions", "second_slab_to_km", "ALTER TABLE ops_vehicle_price_versions ADD COLUMN second_slab_to_km REAL DEFAULT 300")
    ensure_column(conn, "ops_vehicle_price_versions", "km_rate_101_300", "ALTER TABLE ops_vehicle_price_versions ADD COLUMN km_rate_101_300 REAL DEFAULT 0")
    ensure_column(conn, "ops_vehicle_price_versions", "km_rate_over_300", "ALTER TABLE ops_vehicle_price_versions ADD COLUMN km_rate_over_300 REAL DEFAULT 0")
    ensure_column(conn, "ops_vehicle_price_versions", "waiting_hour_rate", "ALTER TABLE ops_vehicle_price_versions ADD COLUMN waiting_hour_rate REAL DEFAULT 0")
    ensure_column(conn, "ops_vehicle_price_versions", "is_active", "ALTER TABLE ops_vehicle_price_versions ADD COLUMN is_active INTEGER DEFAULT 1")
    ensure_column(conn, "ops_vehicle_price_versions", "notes", "ALTER TABLE ops_vehicle_price_versions ADD COLUMN notes TEXT")
    ensure_column(conn, "ops_tickets", "ticket_no", "ALTER TABLE ops_tickets ADD COLUMN ticket_no TEXT")
    ensure_column(conn, "ops_tickets", "company_id", "ALTER TABLE ops_tickets ADD COLUMN company_id INTEGER")
    ensure_column(conn, "ops_tickets", "contract_id", "ALTER TABLE ops_tickets ADD COLUMN contract_id INTEGER")
    ensure_column(conn, "ops_tickets", "ticket_date", "ALTER TABLE ops_tickets ADD COLUMN ticket_date TEXT")
    ensure_column(conn, "ops_tickets", "fault_type_id", "ALTER TABLE ops_tickets ADD COLUMN fault_type_id INTEGER")
    ensure_column(conn, "ops_tickets", "site_code", "ALTER TABLE ops_tickets ADD COLUMN site_code TEXT")
    ensure_column(conn, "ops_tickets", "site_name", "ALTER TABLE ops_tickets ADD COLUMN site_name TEXT")
    ensure_column(conn, "ops_tickets", "priority", "ALTER TABLE ops_tickets ADD COLUMN priority TEXT DEFAULT 'normal'")
    ensure_column(conn, "ops_tickets", "request_channel", "ALTER TABLE ops_tickets ADD COLUMN request_channel TEXT")
    ensure_column(conn, "ops_tickets", "complaint_details", "ALTER TABLE ops_tickets ADD COLUMN complaint_details TEXT")
    ensure_column(conn, "ops_tickets", "status", "ALTER TABLE ops_tickets ADD COLUMN status TEXT DEFAULT 'open'")
    ensure_column(conn, "ops_tickets", "created_by", "ALTER TABLE ops_tickets ADD COLUMN created_by TEXT")
    ensure_column(conn, "ops_trip_tickets", "trip_no", "ALTER TABLE ops_trip_tickets ADD COLUMN trip_no TEXT")
    ensure_column(conn, "ops_trip_tickets", "trip_date", "ALTER TABLE ops_trip_tickets ADD COLUMN trip_date TEXT")
    ensure_column(conn, "ops_trip_tickets", "vehicle_id", "ALTER TABLE ops_trip_tickets ADD COLUMN vehicle_id INTEGER")
    ensure_column(conn, "ops_trip_tickets", "rental_supplier_id", "ALTER TABLE ops_trip_tickets ADD COLUMN rental_supplier_id INTEGER")
    ensure_column(conn, "ops_trip_tickets", "vehicle_type", "ALTER TABLE ops_trip_tickets ADD COLUMN vehicle_type TEXT")
    ensure_column(conn, "ops_trip_tickets", "vehicle_rate_id", "ALTER TABLE ops_trip_tickets ADD COLUMN vehicle_rate_id INTEGER")
    ensure_column(conn, "ops_trip_tickets", "driver_source", "ALTER TABLE ops_trip_tickets ADD COLUMN driver_source TEXT DEFAULT 'company_driver'")
    ensure_column(conn, "ops_trip_tickets", "driver_employee_id", "ALTER TABLE ops_trip_tickets ADD COLUMN driver_employee_id INTEGER")
    ensure_column(conn, "ops_trip_tickets", "supplier_driver_name", "ALTER TABLE ops_trip_tickets ADD COLUMN supplier_driver_name TEXT")
    ensure_column(conn, "ops_trip_tickets", "status", "ALTER TABLE ops_trip_tickets ADD COLUMN status TEXT DEFAULT 'draft'")
    ensure_column(conn, "ops_trip_tickets", "start_odometer", "ALTER TABLE ops_trip_tickets ADD COLUMN start_odometer REAL DEFAULT 0")
    ensure_column(conn, "ops_trip_tickets", "start_photo_path", "ALTER TABLE ops_trip_tickets ADD COLUMN start_photo_path TEXT")
    ensure_column(conn, "ops_trip_tickets", "end_odometer", "ALTER TABLE ops_trip_tickets ADD COLUMN end_odometer REAL DEFAULT 0")
    ensure_column(conn, "ops_trip_tickets", "end_photo_path", "ALTER TABLE ops_trip_tickets ADD COLUMN end_photo_path TEXT")
    ensure_column(conn, "ops_trip_tickets", "waiting_hours", "ALTER TABLE ops_trip_tickets ADD COLUMN waiting_hours REAL DEFAULT 0")
    ensure_column(conn, "ops_trip_tickets", "total_km", "ALTER TABLE ops_trip_tickets ADD COLUMN total_km REAL DEFAULT 0")
    ensure_column(conn, "ops_trip_tickets", "ticket_open_price", "ALTER TABLE ops_trip_tickets ADD COLUMN ticket_open_price REAL DEFAULT 0")
    ensure_column(conn, "ops_trip_tickets", "second_slab_to_km", "ALTER TABLE ops_trip_tickets ADD COLUMN second_slab_to_km REAL DEFAULT 300")
    ensure_column(conn, "ops_trip_tickets", "km_rate_101_300", "ALTER TABLE ops_trip_tickets ADD COLUMN km_rate_101_300 REAL DEFAULT 0")
    ensure_column(conn, "ops_trip_tickets", "km_rate_over_300", "ALTER TABLE ops_trip_tickets ADD COLUMN km_rate_over_300 REAL DEFAULT 0")
    ensure_column(conn, "ops_trip_tickets", "waiting_hour_rate", "ALTER TABLE ops_trip_tickets ADD COLUMN waiting_hour_rate REAL DEFAULT 0")
    ensure_column(conn, "ops_trip_tickets", "total_cost", "ALTER TABLE ops_trip_tickets ADD COLUMN total_cost REAL DEFAULT 0")
    ensure_column(conn, "ops_trip_tickets", "allocated_cost_per_work_order", "ALTER TABLE ops_trip_tickets ADD COLUMN allocated_cost_per_work_order REAL DEFAULT 0")
    ensure_column(conn, "ops_trip_tickets", "driver_commission_pct", "ALTER TABLE ops_trip_tickets ADD COLUMN driver_commission_pct REAL DEFAULT 0")
    ensure_column(conn, "ops_trip_tickets", "driver_commission_amount", "ALTER TABLE ops_trip_tickets ADD COLUMN driver_commission_amount REAL DEFAULT 0")
    ensure_column(conn, "ops_trip_tickets", "notes", "ALTER TABLE ops_trip_tickets ADD COLUMN notes TEXT")
    ensure_column(conn, "ops_trip_tickets", "movement_notes", "ALTER TABLE ops_trip_tickets ADD COLUMN movement_notes TEXT")
    ensure_column(conn, "ops_trip_tickets", "accounting_notes", "ALTER TABLE ops_trip_tickets ADD COLUMN accounting_notes TEXT")
    ensure_column(conn, "ops_trip_tickets", "created_by", "ALTER TABLE ops_trip_tickets ADD COLUMN created_by TEXT")
    ensure_column(conn, "ops_trip_tickets", "movement_closed_by", "ALTER TABLE ops_trip_tickets ADD COLUMN movement_closed_by TEXT")
    ensure_column(conn, "ops_trip_tickets", "approved_by", "ALTER TABLE ops_trip_tickets ADD COLUMN approved_by TEXT")
    ensure_column(conn, "ops_trip_tickets", "completed_at", "ALTER TABLE ops_trip_tickets ADD COLUMN completed_at TEXT")
    ensure_column(conn, "ops_trip_tickets", "approved_at", "ALTER TABLE ops_trip_tickets ADD COLUMN approved_at TEXT")
    ensure_column(conn, "ops_trip_work_orders", "trip_id", "ALTER TABLE ops_trip_work_orders ADD COLUMN trip_id INTEGER")
    ensure_column(conn, "ops_trip_work_orders", "work_order_id", "ALTER TABLE ops_trip_work_orders ADD COLUMN work_order_id INTEGER")
    ensure_column(conn, "ops_trip_work_orders", "line_no", "ALTER TABLE ops_trip_work_orders ADD COLUMN line_no INTEGER DEFAULT 1")
    ensure_column(conn, "ops_trip_work_orders", "allocated_cost", "ALTER TABLE ops_trip_work_orders ADD COLUMN allocated_cost REAL DEFAULT 0")
    ensure_column(conn, "employees", "is_driver", "ALTER TABLE employees ADD COLUMN is_driver INTEGER DEFAULT 0")
    ensure_column(conn, "employees", "trip_commission_pct", "ALTER TABLE employees ADD COLUMN trip_commission_pct REAL DEFAULT 0")
    conn.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_ops_report_work_order ON ops_technician_reports(work_order_id)")
    conn.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_ops_trip_work_order_unique ON ops_trip_work_orders(trip_id, work_order_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_ops_custody_wh_company_department ON ops_customer_custody_warehouses(company_id, department)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_ops_custody_stock_lookup ON ops_customer_custody_stock(company_id, department, warehouse_id, status)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_ops_custody_tx_lookup ON ops_customer_custody_transactions(company_id, department, warehouse_id, transaction_date)")
    conn.commit()
    conn.close()


ensure_tables()


def next_code(table_name: str, prefix: str):
    conn = get_conn()
    row = conn.execute(
        f"SELECT code FROM {table_name} WHERE COALESCE(code, '') <> '' ORDER BY id DESC LIMIT 1"
    ).fetchone()
    conn.close()
    if not row or not row["code"]:
        return f"{prefix}-0001"
    last = safe(row["code"])
    try:
        num = int(last.split("-")[-1])
    except Exception:
        num = 0
    return f"{prefix}-{num + 1:04d}"


def next_number(table_name: str, field_name: str, prefix: str):
    conn = get_conn()
    row = conn.execute(
        f"SELECT {field_name} AS serial_value FROM {table_name} WHERE COALESCE({field_name}, '') <> '' ORDER BY id DESC LIMIT 1"
    ).fetchone()
    conn.close()
    if not row or not row["serial_value"]:
        return f"{prefix}-0001"
    last = safe(row["serial_value"])
    try:
        num = int(last.split("-")[-1])
    except Exception:
        num = 0
    return f"{prefix}-{num + 1:04d}"


def count_rows(table_name: str):
    conn = get_conn()
    row = conn.execute(f"SELECT COUNT(*) AS cnt FROM {table_name}").fetchone()
    conn.close()
    return int(row["cnt"] or 0)


def current_form_notice(notice: str):
    if safe(notice).lower() == "saved":
        return '<div class="msg success">Saved successfully.</div>'
    if safe(notice).lower() == "updated":
        return '<div class="msg success">Updated successfully.</div>'
    if safe(notice):
        return f'<div class="msg success">{safe(notice)}</div>'
    return ""


def status_options(selected=""):
    html = ""
    for value, label in WORK_ORDER_STATUSES:
        sel = "selected" if safe(selected) == value else ""
        html += f"<option value='{value}' {sel}>{label}</option>"
    return html


def status_chip(status: str):
    state = safe(status).lower()
    color = "blue"
    if state in ["approved", "closed"]:
        color = "green"
    elif state in ["rejected"]:
        color = "red"
    elif state in ["submitted", "in_progress"]:
        color = "orange"
    elif state in ["assigned"]:
        color = "blue"
    return f"<span class='status-chip {color}'>{safe(status).replace('_', ' ').title()}</span>"


def rollout_status_chip(status: str):
    state = safe(status).lower() or "not_required"
    color = "blue"
    if state == "approved":
        color = "green"
    elif state == "rejected":
        color = "red"
    elif state == "pending":
        color = "orange"
    return f"<span class='status-chip {color}'>{state.replace('_', ' ').title()}</span>"


def simple_options(options, selected="", empty_label=""):
    html = f"<option value=''>{safe(empty_label)}</option>" if empty_label else ""
    for value, label in options:
        sel = "selected" if safe(selected) == value else ""
        html += f"<option value='{safe(value)}' {sel}>{safe(label)}</option>"
    return html


def log_ops_event(request: Request, entity_type: str, entity_id: int, action: str, notes: str = "", conn=None):
    safe_log_action(
        entity_type,
        entity_id,
        action,
        done_by=actor_name_from_request(request),
        notes=notes,
        conn=conn,
        module="operations",
        path=str(request.url.path),
        method=request.method,
    )


def parse_csv_rows(file_bytes):
    text = file_bytes.decode("utf-8-sig", errors="ignore")
    reader = csv.DictReader(io.StringIO(text))
    return [dict(r) for r in reader]


def parse_xlsx_rows(file_bytes):
    if load_workbook is None:
        raise Exception("Excel import is not available right now.")
    workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [safe(h) for h in rows[0]]
    result = []
    for data_row in rows[1:]:
        item = {}
        for i, header in enumerate(headers):
            if not header:
                continue
            item[header] = "" if i >= len(data_row) or data_row[i] is None else str(data_row[i])
        result.append(item)
    return result


def next_action_code():
    return next_code("ops_service_catalog", "ACT")


def next_ticket_no():
    return next_number("ops_tickets", "ticket_no", "TKT")


def next_contract_no():
    return next_number("ops_contracts", "contract_no", "CTR")


def next_trip_no():
    return next_number("ops_trip_tickets", "trip_no", "TRP")


def next_rental_supplier_code():
    return next_code("ops_rental_suppliers", "SUP")


def next_vehicle_code():
    return next_code("ops_vehicles", "CAR")


def next_custody_warehouse_code():
    return next_code("ops_customer_custody_warehouses", "CCW")


def next_custody_transaction_no():
    return next_number("ops_customer_custody_transactions", "transaction_no", "CCT")


def normalize_action_import_row(row):
    return {
        "code": safe(row.get("code") or row.get("action_code") or row.get("item_code")),
        "name": safe(row.get("name") or row.get("action_name") or row.get("action")),
        "service_category": safe(row.get("service_category") or row.get("action_category") or "field_maintenance"),
        "unit_price": safe(row.get("unit_price") or row.get("price") or "0"),
        "technician_incentive": safe(row.get("technician_incentive") or row.get("incentive") or "0"),
        "default_region_level": safe(row.get("default_region_level") or row.get("region_level") or "zone_1"),
        "default_duration_hours": safe(row.get("default_duration_hours") or row.get("duration_hours") or "0"),
        "notes": safe(row.get("notes")),
        "is_active": safe(row.get("is_active") or "1"),
    }


def get_item_rows():
    conn = get_conn()
    rows = conn.execute("""
        SELECT id, code, name, uom
        FROM items
        WHERE COALESCE(is_active, 1) = 1
          AND LOWER(COALESCE(item_type, 'stock_item')) = 'stock_item'
        ORDER BY code, name
    """).fetchall()
    conn.close()
    return rows


def company_options(selected_id=0):
    conn = get_conn()
    rows = conn.execute("""
        SELECT id, code, name
        FROM ops_contract_companies
        WHERE COALESCE(is_active, 1) = 1
        ORDER BY name
    """).fetchall()
    conn.close()
    html = "<option value=''>-- Select Company --</option>"
    for row in rows:
        sel = "selected" if str(selected_id or "") == str(row["id"]) else ""
        label = f"{safe(row['code'])} - {safe(row['name'])}" if safe(row["code"]) else safe(row["name"])
        html += f"<option value='{row['id']}' {sel}>{label}</option>"
    return html


def contract_options(selected_id=0, company_id=0):
    conn = get_conn()
    sql = """
        SELECT c.id, c.contract_no, c.contract_name, co.name AS company_name
        FROM ops_contracts c
        LEFT JOIN ops_contract_companies co ON co.id = c.company_id
        WHERE LOWER(COALESCE(c.status, 'active')) = 'active'
    """
    params = []
    if safe_int(company_id) > 0:
        sql += " AND c.company_id = ?"
        params.append(safe_int(company_id))
    sql += " ORDER BY c.id DESC"
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    html = "<option value=''>-- Select Contract --</option>"
    for row in rows:
        sel = "selected" if str(selected_id or "") == str(row["id"]) else ""
        label = f"{safe(row['contract_no'])} - {safe(row['contract_name'])}"
        html += f"<option value='{row['id']}' {sel}>{label}</option>"
    return html


def action_options(selected_id=0):
    conn = get_conn()
    rows = conn.execute("""
        SELECT id, code, name
        FROM ops_service_catalog
        WHERE COALESCE(is_active, 1) = 1
        ORDER BY name
    """).fetchall()
    conn.close()
    html = "<option value=''>-- Select Action --</option>"
    for row in rows:
        sel = "selected" if str(selected_id or "") == str(row["id"]) else ""
        label = f"{safe(row['code'])} - {safe(row['name'])}"
        html += f"<option value='{row['id']}' {sel}>{label}</option>"
    return html


def fault_options(selected_id=0):
    conn = get_conn()
    rows = conn.execute("""
        SELECT id, code, name
        FROM ops_fault_types
        WHERE COALESCE(is_active, 1) = 1
        ORDER BY name
    """).fetchall()
    conn.close()
    html = "<option value=''>-- Select Fault --</option>"
    for row in rows:
        sel = "selected" if str(selected_id or "") == str(row["id"]) else ""
        label = f"{safe(row['code'])} - {safe(row['name'])}"
        html += f"<option value='{row['id']}' {sel}>{label}</option>"
    return html


def item_options(selected_id=0):
    html = "<option value=''>-- Select Item --</option>"
    for row in get_item_rows():
        sel = "selected" if str(selected_id or "") == str(row["id"]) else ""
        label = f"{safe(row['code'])} - {safe(row['name'])} ({safe(row['uom'])})"
        html += f"<option value='{row['id']}' {sel}>{label}</option>"
    return html


def get_item_snapshot(item_id=0):
    if safe_int(item_id) <= 0:
        return {"id": 0, "code": "", "name": "", "uom": ""}
    conn = get_conn()
    row = conn.execute("SELECT id, code, name, uom FROM items WHERE id = ? LIMIT 1", (safe_int(item_id),)).fetchone()
    conn.close()
    if not row:
        return {"id": 0, "code": "", "name": "", "uom": ""}
    return {"id": row["id"], "code": safe(row["code"]), "name": safe(row["name"]), "uom": safe(row["uom"])}


def custody_warehouse_options(selected_id=0, company_id=0, department=""):
    conn = get_conn()
    sql = """
        SELECT w.id, w.code, w.name, w.department, w.warehouse_type, c.name AS company_name
        FROM ops_customer_custody_warehouses w
        LEFT JOIN ops_contract_companies c ON c.id = w.company_id
        WHERE COALESCE(w.is_active, 1) = 1
    """
    params = []
    if safe_int(company_id) > 0:
        sql += " AND w.company_id = ?"
        params.append(safe_int(company_id))
    if safe(department):
        sql += " AND LOWER(COALESCE(w.department, '')) = ?"
        params.append(safe(department).lower())
    sql += " ORDER BY c.name, w.department, w.name"
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    html = "<option value=''>-- Select Customer Warehouse --</option>"
    for row in rows:
        sel = "selected" if str(selected_id or "") == str(row["id"]) else ""
        label = f"{safe(row['code'])} - {safe(row['company_name'])} / {safe(row['department']).title()} / {safe(row['name'])}"
        html += f"<option value='{row['id']}' {sel}>{label}</option>"
    return html


def work_order_options(selected_id=0):
    conn = get_conn()
    rows = conn.execute("""
        SELECT id, work_order_no, site_code, site_name, status
        FROM ops_work_orders
        ORDER BY id DESC
        LIMIT 250
    """).fetchall()
    conn.close()
    html = "<option value=''>-- Select Work Order --</option>"
    for row in rows:
        sel = "selected" if str(selected_id or "") == str(row["id"]) else ""
        label = safe(row["work_order_no"]) or f"WO-{row['id']}"
        extras = [x for x in [safe(row["site_code"]), safe(row["site_name"]), safe(row["status"])] if x]
        if extras:
            label = f"{label} - {' / '.join(extras)}"
        html += f"<option value='{row['id']}' {sel}>{label}</option>"
    return html


def ticket_options(selected_id=0):
    conn = get_conn()
    rows = conn.execute("""
        SELECT id, ticket_no, site_code, site_name, status
        FROM ops_tickets
        ORDER BY id DESC
        LIMIT 250
    """).fetchall()
    conn.close()
    html = "<option value=''>-- Select Ticket --</option>"
    for row in rows:
        sel = "selected" if str(selected_id or "") == str(row["id"]) else ""
        label = safe(row["ticket_no"]) or f"TKT-{row['id']}"
        extras = [x for x in [safe(row["site_code"]), safe(row["site_name"]), safe(row["status"])] if x]
        if extras:
            label = f"{label} - {' / '.join(extras)}"
        html += f"<option value='{row['id']}' {sel}>{label}</option>"
    return html


def sync_work_order_qty_from_custody(conn, work_order_id, request: Request = None):
    work_order_id = safe_int(work_order_id)
    if work_order_id <= 0:
        return
    row = conn.execute("SELECT * FROM ops_work_orders WHERE id = ? LIMIT 1", (work_order_id,)).fetchone()
    if not row:
        return
    sums = conn.execute("""
        SELECT
            SUM(CASE WHEN movement_type IN ('issue_to_repair', 'site_issue') THEN COALESCE(qty, 0) ELSE 0 END) AS issued_qty,
            SUM(CASE WHEN movement_type = 'return_from_repair' THEN COALESCE(qty, 0) ELSE 0 END) AS repair_completed_qty,
            SUM(CASE WHEN movement_type IN ('return_from_repair', 'site_return', 'swap_removed') THEN COALESCE(qty, 0) ELSE 0 END) AS returned_qty
        FROM ops_customer_custody_transactions
        WHERE work_order_id = ?
    """, (work_order_id,)).fetchone()
    issued_qty = float(sums["issued_qty"] or 0)
    repair_completed_qty = float(sums["repair_completed_qty"] or 0)
    returned_qty = float(sums["returned_qty"] or 0)
    requested_qty = float(row["requested_qty"] or 0)
    completed_qty = float(row["completed_qty"] or 0)
    if safe(row["workflow_type"]) == "workshop_repair":
        completed_qty = repair_completed_qty
    next_status = safe(row["status"]) or "new"
    if requested_qty > 0 and completed_qty >= requested_qty:
        next_status = "closed"
    elif issued_qty > 0 or completed_qty > 0 or returned_qty > 0:
        next_status = "in_progress"
    conn.execute("""
        UPDATE ops_work_orders
        SET issued_qty = ?, returned_qty = ?, completed_qty = ?, status = ?, updated_at = CURRENT_TIMESTAMP
        WHERE id = ?
    """, (issued_qty, returned_qty, completed_qty, next_status, work_order_id))
    if request:
        log_ops_event(
            request,
            "ops_work_order",
            work_order_id,
            "Progress Sync",
            f"Issued {money(issued_qty)}, completed {money(completed_qty)}, returned {money(returned_qty)}",
            conn=conn,
        )


def custody_stock_status_cell(status):
    color = {
        "faulty": "#fee2e2",
        "under_repair": "#fef3c7",
        "working": "#dcfce7",
        "installed": "#dbeafe",
        "returned": "#ede9fe",
        "scrap": "#f3f4f6",
    }.get(safe(status), "#eef2ff")
    return f"<span class='summary-pill' style='background:{color};'>{safe(status).replace('_', ' ').title()}</span>"


def adjust_customer_custody_stock(
    conn,
    company_id,
    department,
    warehouse_id,
    item_id,
    module_code,
    module_name,
    serial_no,
    status,
    qty_delta,
    uom,
    notes="",
):
    status = safe(status) or "faulty"
    delta = float(qty_delta or 0)
    item_id = safe_int(item_id)
    row = conn.execute("""
        SELECT *
        FROM ops_customer_custody_stock
        WHERE company_id = ?
          AND COALESCE(department, '') = ?
          AND warehouse_id = ?
          AND COALESCE(item_id, 0) = ?
          AND COALESCE(module_code, '') = ?
          AND COALESCE(module_name, '') = ?
          AND COALESCE(serial_no, '') = ?
          AND COALESCE(status, '') = ?
        ORDER BY id DESC
        LIMIT 1
    """, (
        safe_int(company_id),
        safe(department) or "operation",
        safe_int(warehouse_id),
        item_id,
        safe(module_code),
        safe(module_name),
        safe(serial_no),
        status,
    )).fetchone()
    current_qty = float(row["qty"] or 0) if row else 0.0
    new_qty = current_qty + delta
    if new_qty < -0.00001:
        raise Exception(f"Not enough customer custody stock for {safe(module_name) or safe(module_code)} in status {status}.")
    if row:
        conn.execute("""
            UPDATE ops_customer_custody_stock
            SET qty = ?, module_name = ?, uom = ?, notes = ?, updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
        """, (new_qty, safe(module_name), safe(uom), safe(notes), row["id"]))
    else:
        conn.execute("""
            INSERT INTO ops_customer_custody_stock (
                company_id, department, warehouse_id, item_id, module_code, module_name,
                serial_no, status, qty, uom, notes
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            safe_int(company_id),
            safe(department) or "operation",
            safe_int(warehouse_id),
            item_id,
            safe(module_code),
            safe(module_name),
            safe(serial_no),
            status,
            new_qty,
            safe(uom),
            safe(notes),
        ))


def apply_customer_custody_movement(conn, movement_type, company_id, department, warehouse_id, item_id, module_code, module_name, serial_no, from_status, to_status, qty, uom, notes=""):
    qty_value = float(q2(qty))
    if qty_value <= 0:
        raise Exception("Quantity must be greater than zero.")
    movement = safe(movement_type) or "receipt"
    if movement == "receipt":
        adjust_customer_custody_stock(conn, company_id, department, warehouse_id, item_id, module_code, module_name, serial_no, to_status or "faulty", qty_value, uom, notes)
    elif movement == "issue_to_repair":
        adjust_customer_custody_stock(conn, company_id, department, warehouse_id, item_id, module_code, module_name, serial_no, from_status or "faulty", -qty_value, uom, notes)
        adjust_customer_custody_stock(conn, company_id, department, warehouse_id, item_id, module_code, module_name, serial_no, to_status or "under_repair", qty_value, uom, notes)
    elif movement == "return_from_repair":
        adjust_customer_custody_stock(conn, company_id, department, warehouse_id, item_id, module_code, module_name, serial_no, from_status or "under_repair", -qty_value, uom, notes)
        adjust_customer_custody_stock(conn, company_id, department, warehouse_id, item_id, module_code, module_name, serial_no, to_status or "working", qty_value, uom, notes)
    elif movement == "site_issue":
        adjust_customer_custody_stock(conn, company_id, department, warehouse_id, item_id, module_code, module_name, serial_no, from_status or "working", -qty_value, uom, notes)
        adjust_customer_custody_stock(conn, company_id, department, warehouse_id, item_id, module_code, module_name, serial_no, to_status or "installed", qty_value, uom, notes)
    elif movement == "site_return":
        adjust_customer_custody_stock(conn, company_id, department, warehouse_id, item_id, module_code, module_name, serial_no, from_status or "installed", -qty_value, uom, notes)
        adjust_customer_custody_stock(conn, company_id, department, warehouse_id, item_id, module_code, module_name, serial_no, to_status or "working", qty_value, uom, notes)
    elif movement == "swap_removed":
        adjust_customer_custody_stock(conn, company_id, department, warehouse_id, item_id, module_code, module_name, serial_no, to_status or "faulty", qty_value, uom, notes)
    elif movement == "adjustment":
        adjust_customer_custody_stock(conn, company_id, department, warehouse_id, item_id, module_code, module_name, serial_no, to_status or from_status or "faulty", qty_value, uom, notes)
    else:
        raise Exception("Unsupported customer custody movement type.")


def employee_display_name(row):
    return safe(
        row["employee_name"]
        if "employee_name" in row.keys() and row["employee_name"]
        else row["name"]
        if "name" in row.keys() and row["name"]
        else row["full_name"]
        if "full_name" in row.keys() and row["full_name"]
        else row["code"]
        if "code" in row.keys() and row["code"]
        else ""
    )


def employee_options(selected_id=0, driver_only=False):
    conn = get_conn()
    rows = conn.execute("""
        SELECT
            id,
            code,
            name,
            employee_name,
            full_name,
            department,
            job_title,
            COALESCE(is_driver, 0) AS is_driver
        FROM employees
        WHERE COALESCE(is_active, 1) = 1
        ORDER BY COALESCE(employee_name, name, full_name, code)
    """).fetchall()
    conn.close()
    html = "<option value=''>-- Select Employee --</option>"
    filtered_rows = []
    for row in rows:
        if not driver_only:
            filtered_rows.append(row)
            continue
        job_title = safe(row["job_title"]).lower()
        department = safe(row["department"]).lower()
        if int(row["is_driver"] or 0) == 1 or "driver" in job_title or "transport" in department or "حركة" in department:
            filtered_rows.append(row)
    if driver_only and not filtered_rows:
        filtered_rows = rows
    for row in filtered_rows:
        sel = "selected" if str(selected_id or "") == str(row["id"]) else ""
        label = employee_display_name(row)
        if safe(row["code"]):
            label = f"{safe(row['code'])} - {label}"
        if safe(row["job_title"]):
            label = f"{label} ({safe(row['job_title'])})"
        html += f"<option value='{row['id']}' {sel}>{label}</option>"
    return html


def rental_supplier_options(selected_id=0):
    conn = get_conn()
    rows = conn.execute("""
        SELECT id, code, name
        FROM ops_rental_suppliers
        WHERE COALESCE(is_active, 1) = 1
        ORDER BY name
    """).fetchall()
    conn.close()
    html = "<option value=''>-- Select Rental Office --</option>"
    for row in rows:
        sel = "selected" if str(selected_id or "") == str(row["id"]) else ""
        label = f"{safe(row['code'])} - {safe(row['name'])}" if safe(row["code"]) else safe(row["name"])
        html += f"<option value='{row['id']}' {sel}>{label}</option>"
    return html


def vehicle_rate_options(selected_id=0):
    conn = get_conn()
    rows = conn.execute("""
        SELECT
            vr.id,
            vr.code,
            vr.vehicle_type,
            vr.slab_name,
            rs.name AS supplier_name
        FROM ops_vehicle_rates vr
        LEFT JOIN ops_rental_suppliers rs ON rs.id = vr.rental_supplier_id
        WHERE COALESCE(vr.is_active, 1) = 1
        ORDER BY COALESCE(rs.name, ''), vr.vehicle_type, vr.slab_name, vr.code
    """).fetchall()
    conn.close()
    html = "<option value=''>-- Select Vehicle Rate --</option>"
    for row in rows:
        sel = "selected" if str(selected_id or "") == str(row["id"]) else ""
        label = safe(row["code"])
        extras = [x for x in [safe(row["supplier_name"]), safe(row["vehicle_type"]), safe(row["slab_name"])] if x]
        if extras:
            label = f"{label} - {' / '.join(extras)}"
        html += f"<option value='{row['id']}' {sel}>{label}</option>"
    return html


def vehicle_options(selected_id=0):
    conn = get_conn()
    rows = conn.execute("""
        SELECT
            v.id,
            v.code,
            v.vehicle_name,
            v.vehicle_type,
            v.plate_no,
            rs.name AS supplier_name
        FROM ops_vehicles v
        LEFT JOIN ops_rental_suppliers rs ON rs.id = v.rental_supplier_id
        WHERE COALESCE(v.is_active, 1) = 1
        ORDER BY v.code, v.vehicle_name, v.vehicle_type
    """).fetchall()
    conn.close()
    html = "<option value=''>-- Select Vehicle --</option>"
    for row in rows:
        sel = "selected" if str(selected_id or "") == str(row["id"]) else ""
        label = safe(row["code"]) or safe(row["vehicle_name"]) or safe(row["vehicle_type"])
        extras = []
        if safe(row["vehicle_type"]):
            extras.append(safe(row["vehicle_type"]))
        if safe(row["supplier_name"]):
            extras.append(safe(row["supplier_name"]))
        if safe(row["plate_no"]):
            extras.append(safe(row["plate_no"]))
        if extras:
            label = f"{label} ({' / '.join(extras)})"
        html += f"<option value='{row['id']}' {sel}>{label}</option>"
    return html


def driver_source_options(selected="company_driver"):
    html = ""
    for value, label in DRIVER_SOURCE_OPTIONS:
        sel = "selected" if safe(selected) == value else ""
        html += f"<option value='{value}' {sel}>{label}</option>"
    return html


def trip_status_chip(status: str):
    state = safe(status).lower()
    color = "gray"
    if state == "approved":
        color = "green"
    elif state == "completed":
        color = "orange"
    elif state == "dispatched":
        color = "blue"
    return f"<span class='status-chip {color}'>{safe(status).replace('_', ' ').title()}</span>"


def trip_upload_dir():
    return os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "static", "uploads", "trip-meters"))


async def save_trip_photo(file: UploadFile, prefix: str):
    filename = safe(file.filename)
    if not filename:
        return ""
    ext = os.path.splitext(filename)[1] or ".jpg"
    folder = trip_upload_dir()
    os.makedirs(folder, exist_ok=True)
    saved_name = f"{prefix}_{uuid.uuid4().hex[:12]}{ext}"
    target = os.path.join(folder, saved_name)
    with open(target, "wb") as output:
        output.write(await file.read())
    return f"/static/uploads/trip-meters/{saved_name}"


def get_vehicle_row(vehicle_id: int):
    conn = get_conn()
    row = conn.execute("""
        SELECT
            v.*,
            rs.name AS supplier_name,
            vr.id AS vehicle_rate_id_resolved,
            vr.code AS vehicle_rate_code,
            vr.slab_name AS rate_slab_name,
            vr.ticket_open_price,
            vr.second_slab_to_km,
            vr.km_rate_101_300,
            vr.km_rate_over_300,
            vr.waiting_hour_rate
        FROM ops_vehicles v
        LEFT JOIN ops_rental_suppliers rs ON rs.id = v.rental_supplier_id
        LEFT JOIN ops_vehicle_rates vr ON vr.id = (
            SELECT vr2.id
            FROM ops_vehicle_rates vr2
            WHERE COALESCE(vr2.is_active, 1) = 1
              AND COALESCE(vr2.rental_supplier_id, 0) = COALESCE(v.rental_supplier_id, 0)
              AND LOWER(COALESCE(vr2.vehicle_type, '')) = LOWER(COALESCE(v.vehicle_type, ''))
              AND LOWER(COALESCE(vr2.slab_name, 'standard')) = LOWER(COALESCE(CASE WHEN COALESCE(v.pricing_slab_name, '') = '' THEN 'standard' ELSE v.pricing_slab_name END, 'standard'))
            ORDER BY vr2.id DESC
            LIMIT 1
        )
        WHERE v.id = ?
        LIMIT 1
    """, (vehicle_id,)).fetchone()
    if row and not row["vehicle_rate_id_resolved"] and safe_int(row["vehicle_rate_id"]) > 0:
        row = conn.execute("""
            SELECT
                v.*,
                rs.name AS supplier_name,
                vr.id AS vehicle_rate_id_resolved,
                vr.code AS vehicle_rate_code,
                vr.slab_name AS rate_slab_name,
                vr.ticket_open_price,
                vr.second_slab_to_km,
                vr.km_rate_101_300,
                vr.km_rate_over_300,
                vr.waiting_hour_rate
            FROM ops_vehicles v
            LEFT JOIN ops_rental_suppliers rs ON rs.id = v.rental_supplier_id
            LEFT JOIN ops_vehicle_rates vr ON vr.id = v.vehicle_rate_id
            WHERE v.id = ?
            LIMIT 1
        """, (vehicle_id,)).fetchone()
    conn.close()
    return row


def get_driver_row(employee_id: int):
    conn = get_conn()
    row = conn.execute("""
        SELECT
            id,
            code,
            name,
            employee_name,
            full_name,
            job_title,
            COALESCE(trip_commission_pct, 0) AS trip_commission_pct
        FROM employees
        WHERE id = ?
        LIMIT 1
    """, (employee_id,)).fetchone()
    conn.close()
    return row


def work_order_choices(selected_ids=None):
    selected_ids = selected_ids or []
    conn = get_conn()
    rows = conn.execute("""
        SELECT
            wo.id,
            wo.work_order_no,
            wo.site_name,
            wo.status,
            wo.request_date,
            c.name AS company_name
        FROM ops_work_orders wo
        LEFT JOIN ops_contract_companies c ON c.id = wo.company_id
        WHERE COALESCE(wo.trip_required, 1) = 1
        ORDER BY wo.id DESC
    """).fetchall()
    conn.close()
    html = ""
    for row in rows:
        checked = "checked" if str(row["id"]) in [str(x) for x in selected_ids] else ""
        label = safe(row["work_order_no"]) or f"WO-{row['id']}"
        meta = " / ".join([x for x in [safe(row["company_name"]), safe(row["site_name"]), safe(row["status"]).title()] if x])
        html += f"""
        <label class="checkbox-card">
            <input type="checkbox" name="work_order_ids" value="{row['id']}" {checked}>
            <span><b>{label}</b><br><small>{meta or 'No extra details'}</small></span>
        </label>
        """
    if not html:
        html = "<div class='empty-note'>No work orders available yet. Create work orders first, then attach them to the trip.</div>"
    return html


def calculate_trip_cost(total_km, waiting_hours, ticket_open_price, second_slab_to_km, km_rate_101_300, km_rate_over_300, waiting_hour_rate):
    km_value = q2(total_km)
    waiting_value = q2(waiting_hours)
    waiting_cost = waiting_value * q2(waiting_hour_rate)
    second_slab_limit = q2(second_slab_to_km)
    if second_slab_limit < Decimal("100.00"):
        second_slab_limit = Decimal("300.00")
    if km_value <= Decimal("100.00"):
        total = q2(ticket_open_price) + waiting_cost
    elif km_value <= second_slab_limit:
        total = q2(ticket_open_price) + ((km_value - Decimal("100.00")) * q2(km_rate_101_300)) + waiting_cost
    else:
        total = (km_value * q2(km_rate_over_300)) + waiting_cost
    return q2(total)


def get_item_by_id(item_id: int):
    conn = get_conn()
    row = conn.execute("""
        SELECT id, code, name, uom
        FROM items
        WHERE id = ?
        LIMIT 1
    """, (item_id,)).fetchone()
    conn.close()
    return row


def get_item_last_cost(item_id: int):
    conn = get_conn()
    row = conn.execute("""
        SELECT unit_cost
        FROM stock_ledger
        WHERE item_id = ?
          AND COALESCE(unit_cost, 0) > 0
        ORDER BY id DESC
        LIMIT 1
    """, (item_id,)).fetchone()
    conn.close()
    return q2(row["unit_cost"] if row else 0)


def get_action_materials(conn, action_id: int):
    return conn.execute("""
        SELECT *
        FROM ops_action_materials
        WHERE action_id = ?
        ORDER BY line_no, id
    """, (action_id,)).fetchall()


def service_category_options(selected=""):
    html = ""
    for value, label in SERVICE_CATEGORIES:
        sel = "selected" if safe(selected) == value else ""
        html += f"<option value='{value}' {sel}>{label}</option>"
    return html


def zone_level_options(selected=""):
    html = ""
    for value, label in ZONE_LEVELS:
        sel = "selected" if safe(selected) == value else ""
        html += f"<option value='{value}' {sel}>{label}</option>"
    return html


def active_checkbox(checked=True):
    return "checked" if checked else ""


def module_cards(cards):
    html = '<div class="card-grid">'
    for title, href, icon, desc in cards:
        icon_html = f'<img src="{icon}" alt="{title} icon">' if icon.startswith("/static/") else icon
        html += f"""
        <a href="{href}" class="module-card">
            <div class="module-card-icon">{icon_html}</div>
            <div class="module-card-title">{title}</div>
            <div class="module-card-sub">{desc}</div>
        </a>
        """
    html += "</div>"
    return html


def render_ops_page(request: Request, title: str, content: str):
    return HTMLResponse(render_page(title, content, current_path=str(request.url.path)))


@router.get("/ui/projects")
def projects_legacy_redirect():
    return RedirectResponse("/ui/operations", status_code=302)


@router.get("/ui/operations", response_class=HTMLResponse)
def operations_root(request: Request):
    lang = get_lang(request)

    def label(en, ar):
        return ar if lang == "ar" else en

    setup_cards = [
        (label("Tickets", "التذاكر"), "/ui/operations/tickets", "/static/icons/journal.svg", label("Open customer operational tickets before work order assignment.", "فتح تذاكر تشغيل العميل قبل إنشاء أمر الشغل.")),
        (label("Contracts", "العقود"), "/ui/operations/contracts", "/static/icons/customers.svg", label("Different commercial agreements and pricing methods by company.", "اتفاقيات وأساليب تسعير مختلفة حسب الشركة.")),
        (label("Pricing Versions", "إصدارات التسعير"), "/ui/operations/pricing-versions", "/static/icons/reports.svg", label("Versioned action and vehicle prices by contract and effective date.", "أسعار الأعمال والسيارات حسب العقد وتاريخ السريان.")),
        (label("Trip Tickets", "أوامر الرحلات"), "/ui/operations/trips", "/static/icons/goods-receipts.svg", label("Draft, movement, and accounting approval for vehicle trips linked to work orders.", "مسودات وحركة واعتماد محاسبي لرحلات السيارات المرتبطة بأوامر الشغل.")),
        (label("Vehicles", "السيارات"), "/ui/operations/vehicles", "/static/icons/goods-receipts.svg", label("Vehicle master with code, rental office, driver source, and linked rate.", "بيانات السيارات مع الكود ومكتب التأجير والسائق والتسعير المرتبط.")),
        (label("Rental Offices", "مكاتب التأجير"), "/ui/operations/rental-suppliers", "/static/icons/vendors.svg", label("Rental suppliers or car offices that provide vehicles by location.", "موردي أو مكاتب السيارات حسب الموقع.")),
        (label("Work Orders", "أوامر الشغل"), "/ui/operations/work-orders", "/static/icons/journal.svg", label("Create, assign, track, and review field maintenance and workshop jobs.", "إنشاء وتوزيع ومتابعة ومراجعة أعمال الصيانة والمواقع.")),
        (label("Workflow Guide", "دليل دورة التشغيل"), "/ui/operations/workflow-guide", "/static/icons/reports.svg", label("Complete operating cycles for workshop repair, field service, and Orange planning.", "دورات التشغيل الكاملة للتصليح والمواقع وتخطيط أورانج.")),
        (label("Customer Warehouses", "مخازن العميل"), "/ui/operations/customer-custody-warehouses", "/static/icons/warehouses.svg", label("Separate customer custody warehouses by customer department: Operation, Planning, or Repair.", "مخازن عهدة العميل حسب القسم: التشغيل أو التخطيط أو الصيانة.")),
        (label("Customer Stock", "أرصدة مخازن العميل"), "/ui/operations/customer-custody-stock", "/static/icons/stock-balance.svg", label("Operational module balances by customer warehouse and lifecycle status without accounting valuation.", "أرصدة الموديولات حسب مخزن العميل والحالة بدون تقييم محاسبي.")),
        (label("Custody Movements", "حركات عهدة العميل"), "/ui/operations/customer-custody-transactions", "/static/icons/stock-ledger.svg", label("Receive, issue, repair-return, swap, and site movement history for customer-owned modules.", "استلام وصرف وإرجاع صيانة وسواب وحركة مواقع للموديولات المملوكة للعميل.")),
        (label("Contract Companies", "شركات التعاقد"), "/ui/operations/companies", "/static/icons/customers.svg", label("Companies that send faults, modules, assemblies, and customer custody stock.", "الشركات التي ترسل أعطال وموديولات وتجميعات ومخزون عهدة.")),
        (label("Fault Types", "أنواع الأعطال"), "/ui/operations/fault-types", "/static/icons/reports.svg", label("Known fault codes used when opening tickets and classifying field issues.", "أكواد الأعطال المستخدمة في فتح التذاكر وتصنيف مشاكل المواقع.")),
        (label("Regions", "المناطق"), "/ui/operations/regions", "/static/icons/reports.svg", label("Zone allowances for field visits and technician bonus by area.", "بدلات المناطق للزيارات وحوافز الفني حسب المنطقة.")),
        (label("Action Catalog", "كتالوج الأعمال"), "/ui/operations/action-catalog", "/static/icons/customer-invoices.svg", label("Priced actions with import template and fixed raw materials for work order costing.", "الأعمال المسعرة مع قالب الاستيراد والخامات الثابتة لتكلفة أمر الشغل.")),
        (label("Vehicle Rates", "تسعير السيارات"), "/ui/operations/vehicle-rates", "/static/icons/goods-receipts.svg", label("Kilometer and waiting-hour rates for trip tickets.", "أسعار الكيلومتر وساعات الانتظار لأوامر الرحلات.")),
    ]
    summary_labels = [
        (label("Companies", "الشركات"), count_rows('ops_contract_companies')),
        (label("Contracts", "العقود"), count_rows('ops_contracts')),
        (label("Tickets", "التذاكر"), count_rows('ops_tickets')),
        (label("Fault Types", "أنواع الأعطال"), count_rows('ops_fault_types')),
        (label("Regions", "المناطق"), count_rows('ops_regions')),
        (label("Actions", "الأعمال"), count_rows('ops_service_catalog')),
        (label("Rental Offices", "مكاتب التأجير"), count_rows('ops_rental_suppliers')),
        (label("Vehicles", "السيارات"), count_rows('ops_vehicles')),
        (label("Vehicle Rates", "تسعير السيارات"), count_rows('ops_vehicle_rates')),
        (label("Trips", "الرحلات"), count_rows('ops_trip_tickets')),
        (label("Work Orders", "أوامر الشغل"), count_rows('ops_work_orders')),
        (label("Customer Warehouses", "مخازن العميل"), count_rows('ops_customer_custody_warehouses')),
        (label("Custody Movements", "حركات العهدة"), count_rows('ops_customer_custody_transactions')),
    ]
    summary_pills = "".join(
        f'<span class="summary-pill">{title}: {value}</span>'
        for title, value in summary_labels
    )
    summary = f"""
    <div class="card">
        <div class="toolbar">
            <div>
                <h2 style="margin:0;">{label("Operations", "التشغيل")}</h2>
                <div class="section-note">{label("Field maintenance, workshop repairs, cabinet assembly, and customer custody stock in one workflow.", "صيانة المواقع وتصليح الورشة وتجميع الكابينت ومخزون عهدة العميل في دورة تشغيل واحدة.")}</div>
            </div>
            <div class="table-summary">
                {summary_pills}
            </div>
        </div>
    </div>
    """
    workflow = f"""
    <div class="card">
        <h3 class="sub-title">{label("Operational Flow", "دورة التشغيل")}</h3>
        <div class="section-note">{label("Trip workflow is now live first: draft by technical manager, completion by movement manager, and final approval by accounting before transport cost is distributed.", "دورة الرحلات تعمل حاليا: مسودة من المدير الفني، إتمام من مسؤول الحركة، واعتماد محاسبي قبل توزيع تكلفة النقل.")}</div>
        <div class="form-grid">
            <div class="form-group"><label>{label("Field Maintenance", "صيانة المواقع")}</label><input value="{label("Ticket -> Work Order -> Trip Ticket -> Technician Report -> Manager Review -> Invoice", "تذكرة -> أمر شغل -> أمر رحلة -> تقرير فني -> مراجعة مدير -> فاتورة")}" readonly></div>
            <div class="form-group"><label>{label("Workshop Repairs", "تصليح الورشة")}</label><input value="{label("Customer Intake -> Repair Order -> Parts Usage -> Test -> Return to Customer", "استلام من العميل -> أمر صيانة -> صرف قطع -> اختبار -> رجوع للعميل")}" readonly></div>
            <div class="form-group"><label>{label("Cabinet Assembly", "تجميع الكابينت")}</label><input value="{label("Assembly Request -> Components -> Build -> Completion -> Delivery", "طلب تجميع -> مكونات -> تصنيع -> إنهاء -> تسليم")}" readonly></div>
            <div class="form-group"><label>{label("Customer Custody", "عهدة العميل")}</label><input value="{label("Custody Receipt -> Issue to Work Order -> Balance Statement by Customer", "استلام عهدة -> صرف لأمر شغل -> كشف رصيد حسب العميل")}" readonly></div>
        </div>
    </div>
    """
    content = summary + f'<div class="card"><h3 class="sub-title">{label("Master Data", "البيانات الأساسية")}</h3>' + module_cards(setup_cards) + "</div>" + workflow
    return render_ops_page(request, label("Operations", "التشغيل"), content)


@router.get("/ui/operations/workflow-guide", response_class=HTMLResponse)
def workflow_guide_page(request: Request):
    html = """
    <div class="card">
        <div class="toolbar">
            <div>
                <h2 style="margin:0;">Operations Workflow Guide</h2>
                <div class="section-note">Use the current Operations screens in the same order below. Customer-owned stock is tracking only; company spare parts remain company inventory.</div>
            </div>
            <a class="btn gray" href="/ui/operations">Operations</a>
        </div>
    </div>
    <div class="card">
        <h3 class="sub-title">1. Workshop Repair - Orange Operation</h3>
        <div class="form-grid">
            <div class="form-group"><label>Receive Faulty Modules</label><input value="Customer Warehouses -> Custody Movements -> Receive From Customer as Faulty" readonly></div>
            <div class="form-group"><label>Create Ticket</label><input value="Tickets -> select Orange + fault/site/request details" readonly></div>
            <div class="form-group"><label>Create Work Order</label><input value="Work Orders -> Workflow Workshop Repair + Orange Operation warehouse + requested qty" readonly></div>
            <div class="form-group"><label>Issue To Maintenance</label><input value="Open WO -> Issue To Repair; modules become Under Repair" readonly></div>
            <div class="form-group"><label>Repair And Partial Return</label><input value="Open WO -> Return From Repair; repaired qty becomes Working" readonly></div>
            <div class="form-group"><label>Progress</label><input value="WO tracks requested, issued, completed, remaining, and activity log" readonly></div>
        </div>
    </div>
    <div class="card">
        <h3 class="sub-title">2. Field Service - Technical Visit / Swap / Install</h3>
        <div class="form-grid">
            <div class="form-group"><label>Open Work Order</label><input value="Ticket -> Work Order -> Workflow Field Service + action repair/swap/install/technical visit" readonly></div>
            <div class="form-group"><label>Assign Resources</label><input value="Use existing technician, region, trip ticket, vehicle, vehicle rates, technician reports" readonly></div>
            <div class="form-group"><label>Issue Materials</label><input value="Open WO -> Issue To Site from customer warehouse; shortage later from company warehouse workflow" readonly></div>
            <div class="form-group"><label>Swap</label><input value="Installed unit becomes Installed; removed faulty unit goes back as Faulty" readonly></div>
            <div class="form-group"><label>Closure</label><input value="Engineer closes WO with completed qty, actual actions, and notes" readonly></div>
            <div class="form-group"><label>Incentive</label><input value="Action incentive + region allowance are tracked on the WO and reports" readonly></div>
        </div>
    </div>
    <div class="card">
        <h3 class="sub-title">3. Orange Planning</h3>
        <div class="form-grid">
            <div class="form-group"><label>Project Request</label><input value="Ticket includes site code and job type from the official request" readonly></div>
            <div class="form-group"><label>Standard Kit</label><input value="Action Catalog materials represent the PM-defined standard kit" readonly></div>
            <div class="form-group"><label>Planning Warehouse</label><input value="Customer warehouse department Planning is separate from Operation" readonly></div>
            <div class="form-group"><label>Rollout Approval</label><input value="Open WO -> Rollout Confirm; notes are mandatory before execution" readonly></div>
            <div class="form-group"><label>Technician Execution</label><input value="Materials are issued to technician, returned to same source, then WO is closed" readonly></div>
            <div class="form-group"><label>Audit</label><input value="Every movement, approval, and closure is written to Activity Log" readonly></div>
        </div>
    </div>
    """
    return render_ops_page(request, "Operations Workflow Guide", html)


@router.get("/ui/operations/customer-custody-warehouses", response_class=HTMLResponse)
def customer_custody_warehouses_page(request: Request, edit_id: int = 0, notice: str = ""):
    conn = get_conn()
    rows = conn.execute("""
        SELECT w.*, c.code AS company_code, c.name AS company_name
        FROM ops_customer_custody_warehouses w
        LEFT JOIN ops_contract_companies c ON c.id = w.company_id
        ORDER BY c.name, w.department, w.name
    """).fetchall()
    edit_row = conn.execute("SELECT * FROM ops_customer_custody_warehouses WHERE id = ? LIMIT 1", (edit_id,)).fetchone() if edit_id else None
    conn.close()
    form_values = dict(edit_row) if edit_row else {
        "id": 0,
        "code": next_custody_warehouse_code(),
        "company_id": "",
        "department": "operation",
        "warehouse_type": "operation",
        "name": "",
        "location": "",
        "notes": "",
        "is_active": 1,
    }
    body = ""
    for row in rows:
        body += f"""
        <tr>
            <td>{safe(row['code'])}</td>
            <td>{safe(row['company_code'])} - {safe(row['company_name'])}</td>
            <td>{safe(row['department']).title()}</td>
            <td>{safe(row['warehouse_type']).replace('_', ' ').title()}</td>
            <td>{safe(row['name'])}</td>
            <td>{safe(row['location'])}</td>
            <td>{'Active' if int(row['is_active'] or 0) == 1 else 'Inactive'}</td>
            <td><a class="btn blue" href="/ui/operations/customer-custody-warehouses?edit_id={row['id']}">Edit</a></td>
        </tr>
        """
    if not body:
        body = "<tr><td colspan='8' style='text-align:center;'>No customer custody warehouses added yet.</td></tr>"
    html = f"""
    {current_form_notice(notice)}
    <div class="card">
        <div class="toolbar">
            <div>
                <h2 style="margin:0;">Customer Custody Warehouses</h2>
                <div class="section-note">One customer can have separate Operation, Planning, and Repair warehouses.</div>
            </div>
            <a class="btn gray" href="/ui/operations">Back to Operations</a>
        </div>
        <form method="post" action="/ui/operations/customer-custody-warehouses/save">
            <input type="hidden" name="row_id" value="{safe(form_values.get('id', 0))}">
            <div class="form-grid">
                <div class="form-group"><label>Code</label><input name="code" value="{safe(form_values.get('code'))}" required></div>
                <div class="form-group"><label>Customer</label><select name="company_id" required>{company_options(form_values.get('company_id'))}</select></div>
                <div class="form-group"><label>Department</label><select name="department" required>{simple_options(OPS_DEPARTMENTS, form_values.get('department'))}</select></div>
                <div class="form-group"><label>Warehouse Type</label><select name="warehouse_type" required>{simple_options(CUSTODY_WAREHOUSE_TYPES, form_values.get('warehouse_type'))}</select></div>
                <div class="form-group"><label>Warehouse Name</label><input name="name" value="{safe(form_values.get('name'))}" required></div>
                <div class="form-group"><label>Location</label><input name="location" value="{safe(form_values.get('location'))}"></div>
                <div class="form-group" style="grid-column: span 2;"><label>Notes</label><input name="notes" value="{safe(form_values.get('notes'))}"></div>
                <div class="form-group"><label>Active</label><div style="padding-top:10px;"><label><input type="checkbox" name="is_active" value="1" {active_checkbox(int(form_values.get('is_active', 1) or 0) == 1)}> Active</label></div></div>
            </div>
            <div class="form-actions">
                <button class="btn green" type="submit">{"Update Warehouse" if safe_int(form_values.get('id')) > 0 else "Save Warehouse"}</button>
                <a class="btn gray" href="/ui/operations/customer-custody-warehouses">Clear</a>
            </div>
        </form>
    </div>
    <div class="card">
        <h3 class="sub-title">Warehouses</h3>
        <table>
            <tr><th>Code</th><th>Customer</th><th>Department</th><th>Type</th><th>Name</th><th>Location</th><th>Status</th><th>Action</th></tr>
            {body}
        </table>
    </div>
    """
    return render_ops_page(request, "Customer Custody Warehouses", html)


@router.post("/ui/operations/customer-custody-warehouses/save")
def save_customer_custody_warehouse(
    request: Request,
    row_id: int = Form(0),
    code: str = Form(""),
    company_id: int = Form(0),
    department: str = Form("operation"),
    warehouse_type: str = Form("operation"),
    name: str = Form(""),
    location: str = Form(""),
    notes: str = Form(""),
    is_active: str = Form(""),
):
    conn = get_conn()
    active_flag = 1 if safe(is_active) == "1" else 0
    values = (
        safe(code) or next_custody_warehouse_code(),
        safe_int(company_id),
        safe(department) or "operation",
        safe(warehouse_type) or "operation",
        safe(name),
        safe(location),
        safe(notes),
        active_flag,
    )
    if row_id > 0:
        conn.execute("""
            UPDATE ops_customer_custody_warehouses
            SET code = ?, company_id = ?, department = ?, warehouse_type = ?, name = ?, location = ?, notes = ?, is_active = ?
            WHERE id = ?
        """, values + (row_id,))
        notice = "updated"
        entity_id = row_id
    else:
        cur = conn.execute("""
            INSERT INTO ops_customer_custody_warehouses (
                code, company_id, department, warehouse_type, name, location, notes, is_active
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, values)
        notice = "saved"
        entity_id = cur.lastrowid
    log_ops_event(request, "ops_customer_custody_warehouse", safe_int(entity_id), "saved", f"Warehouse {safe(code) or safe(name)}", conn=conn)
    conn.commit()
    conn.close()
    return RedirectResponse(f"/ui/operations/customer-custody-warehouses?notice={notice}", status_code=303)


@router.get("/ui/operations/customer-custody-stock", response_class=HTMLResponse)
def customer_custody_stock_page(request: Request, company_id: int = 0, department: str = "", warehouse_id: int = 0, status: str = ""):
    conn = get_conn()
    sql = """
        SELECT s.*, c.code AS company_code, c.name AS company_name, w.code AS warehouse_code, w.name AS warehouse_name
        FROM ops_customer_custody_stock s
        LEFT JOIN ops_contract_companies c ON c.id = s.company_id
        LEFT JOIN ops_customer_custody_warehouses w ON w.id = s.warehouse_id
        WHERE ABS(COALESCE(s.qty, 0)) > 0.00001
    """
    params = []
    if safe_int(company_id) > 0:
        sql += " AND s.company_id = ?"
        params.append(safe_int(company_id))
    if safe(department):
        sql += " AND COALESCE(s.department, '') = ?"
        params.append(safe(department))
    if safe_int(warehouse_id) > 0:
        sql += " AND s.warehouse_id = ?"
        params.append(safe_int(warehouse_id))
    if safe(status):
        sql += " AND COALESCE(s.status, '') = ?"
        params.append(safe(status))
    sql += " ORDER BY c.name, s.department, w.name, s.module_name, s.status"
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    body = ""
    total_qty = Decimal("0")
    for row in rows:
        total_qty += to_decimal(row["qty"])
        module_label = safe(row["module_name"]) or safe(row["module_code"])
        if safe(row["module_code"]):
            module_label = f"{safe(row['module_code'])} - {module_label}" if safe(row["module_name"]) else safe(row["module_code"])
        body += f"""
        <tr>
            <td>{safe(row['company_code'])} - {safe(row['company_name'])}</td>
            <td>{safe(row['department']).title()}</td>
            <td>{safe(row['warehouse_code'])} - {safe(row['warehouse_name'])}</td>
            <td>{module_label}</td>
            <td>{safe(row['serial_no'])}</td>
            <td>{custody_stock_status_cell(row['status'])}</td>
            <td>{money(row['qty'])}</td>
            <td>{safe(row['uom'])}</td>
        </tr>
        """
    if not body:
        body = "<tr><td colspan='8' style='text-align:center;'>No customer custody stock found.</td></tr>"
    html = f"""
    <div class="card">
        <div class="toolbar">
            <div>
                <h2 style="margin:0;">Customer Custody Stock</h2>
                <div class="section-note">Operational balances only. Customer-owned modules do not affect accounting inventory valuation.</div>
            </div>
            <a class="btn gray" href="/ui/operations">Back to Operations</a>
        </div>
        <form method="get" action="/ui/operations/customer-custody-stock">
            <div class="form-grid">
                <div class="form-group"><label>Customer</label><select name="company_id">{company_options(company_id)}</select></div>
                <div class="form-group"><label>Department</label><select name="department">{simple_options(OPS_DEPARTMENTS, department, "-- All Departments --")}</select></div>
                <div class="form-group"><label>Warehouse</label><select name="warehouse_id">{custody_warehouse_options(warehouse_id, company_id, department)}</select></div>
                <div class="form-group"><label>Status</label><select name="status">{simple_options(CUSTOMER_MODULE_STATUSES, status, "-- All Statuses --")}</select></div>
            </div>
            <div class="form-actions">
                <button class="btn green" type="submit">Filter</button>
                <a class="btn gray" href="/ui/operations/customer-custody-stock">Clear</a>
                <a class="btn blue" href="/ui/operations/customer-custody-transactions">New Movement</a>
            </div>
        </form>
    </div>
    <div class="card">
        <div class="toolbar">
            <h3 class="sub-title">Stock Balances</h3>
            <span class="summary-pill">Total Qty: {money(total_qty)}</span>
        </div>
        <table>
            <tr><th>Customer</th><th>Department</th><th>Warehouse</th><th>Module</th><th>Serial</th><th>Status</th><th>Qty</th><th>UOM</th></tr>
            {body}
        </table>
    </div>
    """
    return render_ops_page(request, "Customer Custody Stock", html)


@router.get("/ui/operations/customer-custody-transactions", response_class=HTMLResponse)
def customer_custody_transactions_page(
    request: Request,
    notice: str = "",
    company_id: int = 0,
    department: str = "",
    warehouse_id: int = 0,
    ticket_id: int = 0,
    work_order_id: int = 0,
    movement_type: str = "receipt",
    module_code: str = "",
    module_name: str = "",
    serial_no: str = "",
    from_status: str = "",
    to_status: str = "faulty",
    qty: str = "",
    uom: str = "PCS",
    notes: str = "",
):
    conn = get_conn()
    sql = """
        SELECT t.*, c.code AS company_code, c.name AS company_name, w.code AS warehouse_code, w.name AS warehouse_name,
               wo.work_order_no, tk.ticket_no
        FROM ops_customer_custody_transactions t
        LEFT JOIN ops_contract_companies c ON c.id = t.company_id
        LEFT JOIN ops_customer_custody_warehouses w ON w.id = t.warehouse_id
        LEFT JOIN ops_work_orders wo ON wo.id = t.work_order_id
        LEFT JOIN ops_tickets tk ON tk.id = t.ticket_id
        WHERE 1 = 1
    """
    params = []
    if safe_int(company_id) > 0:
        sql += " AND t.company_id = ?"
        params.append(safe_int(company_id))
    if safe(department):
        sql += " AND t.department = ?"
        params.append(safe(department))
    if safe_int(warehouse_id) > 0:
        sql += " AND t.warehouse_id = ?"
        params.append(safe_int(warehouse_id))
    sql += " ORDER BY t.id DESC LIMIT 250"
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    body = ""
    for row in rows:
        module_label = safe(row["module_name"]) or safe(row["module_code"])
        if safe(row["module_code"]):
            module_label = f"{safe(row['module_code'])} - {module_label}" if safe(row["module_name"]) else safe(row["module_code"])
        body += f"""
        <tr>
            <td>{safe(row['transaction_no'])}</td>
            <td>{safe(row['transaction_date'])}</td>
            <td>{safe(row['company_code'])} - {safe(row['company_name'])}</td>
            <td>{safe(row['department']).title()}</td>
            <td>{safe(row['warehouse_code'])} - {safe(row['warehouse_name'])}</td>
            <td>{safe(row['movement_type']).replace('_', ' ').title()}</td>
            <td>{module_label}</td>
            <td>{safe(row['from_status']).replace('_', ' ').title()} -> {safe(row['to_status']).replace('_', ' ').title()}</td>
            <td>{money(row['qty'])}</td>
            <td>{safe(row['work_order_no']) or safe(row['ticket_no'])}</td>
        </tr>
        """
    if not body:
        body = "<tr><td colspan='10' style='text-align:center;'>No custody movements found.</td></tr>"
    html = f"""
    {current_form_notice(notice)}
    <div class="card">
        <div class="toolbar">
            <div>
                <h2 style="margin:0;">Customer Custody Movement</h2>
                <div class="section-note">Use this for Orange Operation and Planning warehouse movements before linking deeper work-order flows.</div>
            </div>
            <a class="btn gray" href="/ui/operations">Back to Operations</a>
        </div>
        <form method="post" action="/ui/operations/customer-custody-transactions/save">
            <div class="form-grid">
                <div class="form-group"><label>No</label><input name="transaction_no" value="{next_custody_transaction_no()}" required></div>
                <div class="form-group"><label>Date</label><input type="date" name="transaction_date" required></div>
                <div class="form-group"><label>Customer</label><select name="company_id" required>{company_options(company_id)}</select></div>
                <div class="form-group"><label>Department</label><select name="department" required>{simple_options(OPS_DEPARTMENTS, department or 'operation')}</select></div>
                <div class="form-group"><label>Customer Warehouse</label><select name="warehouse_id" required>{custody_warehouse_options(warehouse_id, company_id, department)}</select></div>
                <div class="form-group"><label>Movement Type</label><select name="movement_type" required>{simple_options(CUSTODY_MOVEMENT_TYPES, movement_type or 'receipt')}</select></div>
                <div class="form-group"><label>Item Master (Optional)</label><select name="item_id">{item_options()}</select></div>
                <div class="form-group"><label>Module Code</label><input name="module_code" value="{safe(module_code)}"></div>
                <div class="form-group"><label>Module Name</label><input name="module_name" value="{safe(module_name)}" required></div>
                <div class="form-group"><label>Serial No</label><input name="serial_no" value="{safe(serial_no)}"></div>
                <div class="form-group"><label>From Status</label><select name="from_status">{simple_options(CUSTOMER_MODULE_STATUSES, from_status or 'faulty', "-- No From Status --")}</select></div>
                <div class="form-group"><label>To Status</label><select name="to_status">{simple_options(CUSTOMER_MODULE_STATUSES, to_status or 'faulty')}</select></div>
                <div class="form-group"><label>Qty</label><input type="number" step="0.01" name="qty" value="{safe(qty)}" required></div>
                <div class="form-group"><label>UOM</label><input name="uom" value="{safe(uom) or 'PCS'}"></div>
                <div class="form-group"><label>Ticket</label><select name="ticket_id">{ticket_options(ticket_id)}</select></div>
                <div class="form-group"><label>Work Order</label><select name="work_order_id">{work_order_options(work_order_id)}</select></div>
                <div class="form-group" style="grid-column: span 2;"><label>Notes</label><input name="notes" value="{safe(notes)}"></div>
            </div>
            <div class="form-actions">
                <button class="btn green" type="submit">Save Movement</button>
                <a class="btn gray" href="/ui/operations/customer-custody-transactions">Clear</a>
            </div>
        </form>
    </div>
    <div class="card">
        <h3 class="sub-title">Latest Movements</h3>
        <table>
            <tr><th>No</th><th>Date</th><th>Customer</th><th>Department</th><th>Warehouse</th><th>Movement</th><th>Module</th><th>Status Flow</th><th>Qty</th><th>Linked Doc</th></tr>
            {body}
        </table>
    </div>
    """
    return render_ops_page(request, "Customer Custody Movement", html)


@router.post("/ui/operations/customer-custody-transactions/save")
def save_customer_custody_transaction(
    request: Request,
    transaction_no: str = Form(""),
    transaction_date: str = Form(""),
    company_id: int = Form(0),
    department: str = Form("operation"),
    warehouse_id: int = Form(0),
    ticket_id: int = Form(0),
    work_order_id: int = Form(0),
    item_id: int = Form(0),
    module_code: str = Form(""),
    module_name: str = Form(""),
    serial_no: str = Form(""),
    movement_type: str = Form("receipt"),
    from_status: str = Form(""),
    to_status: str = Form("faulty"),
    qty: str = Form("0"),
    uom: str = Form("PCS"),
    notes: str = Form(""),
):
    conn = get_conn()
    item = get_item_snapshot(item_id)
    final_module_code = safe(module_code) or item["code"]
    final_module_name = safe(module_name) or item["name"] or final_module_code
    final_uom = safe(uom) or item["uom"] or "PCS"
    try:
        apply_customer_custody_movement(
            conn,
            movement_type,
            company_id,
            safe(department) or "operation",
            warehouse_id,
            safe_int(item_id),
            final_module_code,
            final_module_name,
            serial_no,
            from_status,
            to_status,
            qty,
            final_uom,
            notes,
        )
        cur = conn.execute("""
            INSERT INTO ops_customer_custody_transactions (
                transaction_no, transaction_date, company_id, department, warehouse_id,
                ticket_id, work_order_id, item_id, module_code, module_name, serial_no,
                movement_type, from_status, to_status, qty, uom, notes, created_by
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            safe(transaction_no) or next_custody_transaction_no(),
            safe(transaction_date),
            safe_int(company_id),
            safe(department) or "operation",
            safe_int(warehouse_id),
            safe_int(ticket_id),
            safe_int(work_order_id),
            safe_int(item_id),
            final_module_code,
            final_module_name,
            safe(serial_no),
            safe(movement_type) or "receipt",
            safe(from_status),
            safe(to_status),
            float(q2(qty)),
            final_uom,
            safe(notes),
            actor_name_from_request(request),
        ))
        entity_id = cur.lastrowid
        log_ops_event(request, "ops_customer_custody_transaction", safe_int(entity_id), "created", f"{safe(movement_type)} {final_module_name} qty {money(qty)}", conn=conn)
        if safe_int(work_order_id) > 0:
            sync_work_order_qty_from_custody(conn, work_order_id, request)
        conn.commit()
        notice = "saved"
    except Exception as exc:
        conn.rollback()
        notice = f"Error: {safe(exc)}"
    finally:
        conn.close()
    return RedirectResponse(f"/ui/operations/customer-custody-transactions?notice={quote(notice)}", status_code=303)


@router.get("/ui/operations/companies", response_class=HTMLResponse)
def companies_page(request: Request, edit_id: int = 0, notice: str = ""):
    conn = get_conn()
    rows = conn.execute("SELECT * FROM ops_contract_companies ORDER BY id DESC").fetchall()
    edit_row = conn.execute("SELECT * FROM ops_contract_companies WHERE id = ? LIMIT 1", (edit_id,)).fetchone() if edit_id else None
    conn.close()
    form_values = dict(edit_row) if edit_row else {
        "id": 0,
        "code": next_code("ops_contract_companies", "COMP"),
        "name": "",
        "contact_person": "",
        "phone": "",
        "email": "",
        "notes": "",
        "is_active": 1,
    }
    body = ""
    for row in rows:
        status = "Active" if int(row["is_active"] or 0) == 1 else "Inactive"
        body += f"""
        <tr>
            <td>{safe(row['code'])}</td>
            <td>{safe(row['name'])}</td>
            <td>{safe(row['contact_person'])}</td>
            <td>{safe(row['phone'])}</td>
            <td>{safe(row['email'])}</td>
            <td>{status}</td>
            <td><a class="btn blue" href="/ui/operations/companies?edit_id={row['id']}">Edit</a></td>
        </tr>
        """
    if not body:
        body = "<tr><td colspan='7' style='text-align:center;'>No contract companies added yet.</td></tr>"
    html = f"""
    {current_form_notice(notice)}
    <div class="card">
        <div class="toolbar">
            <div>
                <h2 style="margin:0;">Contract Companies</h2>
                <div class="section-note">Companies that open faults, send modules for repair, request cabinet assembly, or hand over custody stock.</div>
            </div>
            <a class="btn gray" href="/ui/operations">Back to Operations</a>
        </div>
        <form method="post" action="/ui/operations/companies/save">
            <input type="hidden" name="row_id" value="{safe(form_values.get('id', 0))}">
            <div class="form-grid">
                <div class="form-group"><label>Code</label><input name="code" value="{safe(form_values.get('code'))}" required></div>
                <div class="form-group"><label>Company Name</label><input name="name" value="{safe(form_values.get('name'))}" required></div>
                <div class="form-group"><label>Contact Person</label><input name="contact_person" value="{safe(form_values.get('contact_person'))}"></div>
                <div class="form-group"><label>Phone</label><input name="phone" value="{safe(form_values.get('phone'))}"></div>
                <div class="form-group"><label>Email</label><input name="email" value="{safe(form_values.get('email'))}"></div>
                <div class="form-group"><label>Active</label><div style="padding-top:10px;"><label><input type="checkbox" name="is_active" value="1" {active_checkbox(int(form_values.get('is_active', 1) or 0) == 1)}> Active Company</label></div></div>
                <div class="form-group" style="grid-column: span 2;"><label>Notes</label><input name="notes" value="{safe(form_values.get('notes'))}"></div>
            </div>
            <div class="form-actions">
                <button class="btn green" type="submit">{"Update Company" if safe_int(form_values.get('id')) > 0 else "Save Company"}</button>
                <a class="btn gray" href="/ui/operations/companies">Clear</a>
            </div>
        </form>
    </div>
    <div class="card">
        <h3 class="sub-title">Company List</h3>
        <table>
            <tr><th>Code</th><th>Name</th><th>Contact</th><th>Phone</th><th>Email</th><th>Status</th><th>Action</th></tr>
            {body}
        </table>
    </div>
    """
    return render_ops_page(request, "Contract Companies", html)


@router.post("/ui/operations/companies/save")
def save_company(
    row_id: int = Form(0),
    code: str = Form(""),
    name: str = Form(""),
    contact_person: str = Form(""),
    phone: str = Form(""),
    email: str = Form(""),
    notes: str = Form(""),
    is_active: str = Form(""),
):
    conn = get_conn()
    active_flag = 1 if safe(is_active) == "1" else 0
    if row_id > 0:
        conn.execute("""
            UPDATE ops_contract_companies
            SET code = ?, name = ?, contact_person = ?, phone = ?, email = ?, notes = ?, is_active = ?
            WHERE id = ?
        """, (safe(code), safe(name), safe(contact_person), safe(phone), safe(email), safe(notes), active_flag, row_id))
        notice = "updated"
    else:
        conn.execute("""
            INSERT INTO ops_contract_companies (code, name, contact_person, phone, email, notes, is_active)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (safe(code) or next_code("ops_contract_companies", "COMP"), safe(name), safe(contact_person), safe(phone), safe(email), safe(notes), active_flag))
        notice = "saved"
    conn.commit()
    conn.close()
    return RedirectResponse(f"/ui/operations/companies?notice={notice}", status_code=303)


@router.get("/ui/operations/fault-types", response_class=HTMLResponse)
def fault_types_page(request: Request, edit_id: int = 0, notice: str = ""):
    conn = get_conn()
    rows = conn.execute("SELECT * FROM ops_fault_types ORDER BY id DESC").fetchall()
    edit_row = conn.execute("SELECT * FROM ops_fault_types WHERE id = ? LIMIT 1", (edit_id,)).fetchone() if edit_id else None
    conn.close()
    form_values = dict(edit_row) if edit_row else {
        "id": 0,
        "code": next_code("ops_fault_types", "FLT"),
        "name": "",
        "service_category": "field_maintenance",
        "notes": "",
        "is_active": 1,
    }
    body = ""
    for row in rows:
        body += f"""
        <tr>
            <td>{safe(row['code'])}</td>
            <td>{safe(row['name'])}</td>
            <td>{safe(row['service_category']).replace('_', ' ').title()}</td>
            <td>{"Active" if int(row['is_active'] or 0) == 1 else "Inactive"}</td>
            <td><a class="btn blue" href="/ui/operations/fault-types?edit_id={row['id']}">Edit</a></td>
        </tr>
        """
    if not body:
        body = "<tr><td colspan='5' style='text-align:center;'>No fault types added yet.</td></tr>"
    html = f"""
    {current_form_notice(notice)}
    <div class="card">
        <div class="toolbar">
            <div>
                <h2 style="margin:0;">Fault Types</h2>
                <div class="section-note">Known fault names used when opening tickets. Pricing and incentive should be configured on the Action side.</div>
            </div>
            <a class="btn gray" href="/ui/operations">Back to Operations</a>
        </div>
        <form method="post" action="/ui/operations/fault-types/save">
            <input type="hidden" name="row_id" value="{safe(form_values.get('id', 0))}">
            <div class="form-grid">
                <div class="form-group"><label>Code</label><input name="code" value="{safe(form_values.get('code'))}" required></div>
                <div class="form-group"><label>Fault Name</label><input name="name" value="{safe(form_values.get('name'))}" required></div>
                <div class="form-group"><label>Service Category</label><select name="service_category">{service_category_options(form_values.get('service_category'))}</select></div>
                <div class="form-group"><label>Active</label><div style="padding-top:10px;"><label><input type="checkbox" name="is_active" value="1" {active_checkbox(int(form_values.get('is_active', 1) or 0) == 1)}> Active Fault</label></div></div>
                <div class="form-group" style="grid-column: span 2;"><label>Notes</label><input name="notes" value="{safe(form_values.get('notes'))}"></div>
            </div>
            <div class="form-actions">
                <button class="btn green" type="submit">{"Update Fault Type" if safe_int(form_values.get('id')) > 0 else "Save Fault Type"}</button>
                <a class="btn gray" href="/ui/operations/fault-types">Clear</a>
            </div>
        </form>
    </div>
    <div class="card">
        <h3 class="sub-title">Fault Type List</h3>
        <table>
            <tr><th>Code</th><th>Fault</th><th>Category</th><th>Status</th><th>Action</th></tr>
            {body}
        </table>
    </div>
    """
    return render_ops_page(request, "Fault Types", html)


@router.post("/ui/operations/fault-types/save")
def save_fault_type(
    row_id: int = Form(0),
    code: str = Form(""),
    name: str = Form(""),
    service_category: str = Form("field_maintenance"),
    notes: str = Form(""),
    is_active: str = Form(""),
):
    conn = get_conn()
    active_flag = 1 if safe(is_active) == "1" else 0
    existing = conn.execute("SELECT default_service_price, default_incentive FROM ops_fault_types WHERE id = ? LIMIT 1", (row_id,)).fetchone() if row_id > 0 else None
    if row_id > 0:
        conn.execute("""
            UPDATE ops_fault_types
            SET code = ?, name = ?, service_category = ?, default_service_price = ?, default_incentive = ?, notes = ?, is_active = ?
            WHERE id = ?
        """, (
            safe(code),
            safe(name),
            safe(service_category),
            float(q2(existing["default_service_price"] if existing else 0)),
            float(q2(existing["default_incentive"] if existing else 0)),
            safe(notes),
            active_flag,
            row_id,
        ))
        notice = "updated"
    else:
        conn.execute("""
            INSERT INTO ops_fault_types (code, name, service_category, default_service_price, default_incentive, notes, is_active)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (
            safe(code) or next_code("ops_fault_types", "FLT"),
            safe(name),
            safe(service_category),
            0.0,
            0.0,
            safe(notes),
            active_flag,
        ))
        notice = "saved"
    conn.commit()
    conn.close()
    return RedirectResponse(f"/ui/operations/fault-types?notice={notice}", status_code=303)


@router.get("/ui/operations/regions", response_class=HTMLResponse)
def regions_page(request: Request, edit_id: int = 0, notice: str = ""):
    conn = get_conn()
    rows = conn.execute("SELECT * FROM ops_regions ORDER BY id DESC").fetchall()
    edit_row = conn.execute("SELECT * FROM ops_regions WHERE id = ? LIMIT 1", (edit_id,)).fetchone() if edit_id else None
    conn.close()
    form_values = dict(edit_row) if edit_row else {
        "id": 0,
        "code": next_code("ops_regions", "REG"),
        "name": "",
        "zone_level": "zone_1",
        "allowance_amount": "0.00",
        "notes": "",
        "is_active": 1,
    }
    body = ""
    for row in rows:
        body += f"""
        <tr>
            <td>{safe(row['code'])}</td>
            <td>{safe(row['name'])}</td>
            <td>{safe(row['zone_level']).replace('_', ' ').title()}</td>
            <td>{money(row['allowance_amount'])}</td>
            <td>{"Active" if int(row['is_active'] or 0) == 1 else "Inactive"}</td>
            <td><a class="btn blue" href="/ui/operations/regions?edit_id={row['id']}">Edit</a></td>
        </tr>
        """
    if not body:
        body = "<tr><td colspan='6' style='text-align:center;'>No regions added yet.</td></tr>"
    html = f"""
    {current_form_notice(notice)}
    <div class="card">
        <div class="toolbar">
            <div>
                <h2 style="margin:0;">Regions</h2>
                <div class="section-note">Area bands used for technician allowance and trip costing by destination.</div>
            </div>
            <a class="btn gray" href="/ui/operations">Back to Operations</a>
        </div>
        <form method="post" action="/ui/operations/regions/save">
            <input type="hidden" name="row_id" value="{safe(form_values.get('id', 0))}">
            <div class="form-grid">
                <div class="form-group"><label>Code</label><input name="code" value="{safe(form_values.get('code'))}" required></div>
                <div class="form-group"><label>Region Name</label><input name="name" value="{safe(form_values.get('name'))}" required></div>
                <div class="form-group"><label>Zone Level</label><select name="zone_level">{zone_level_options(form_values.get('zone_level'))}</select></div>
                <div class="form-group"><label>Allowance Amount</label><input name="allowance_amount" value="{safe(form_values.get('allowance_amount'))}"></div>
                <div class="form-group"><label>Active</label><div style="padding-top:10px;"><label><input type="checkbox" name="is_active" value="1" {active_checkbox(int(form_values.get('is_active', 1) or 0) == 1)}> Active Region</label></div></div>
                <div class="form-group" style="grid-column: span 2;"><label>Notes</label><input name="notes" value="{safe(form_values.get('notes'))}"></div>
            </div>
            <div class="form-actions">
                <button class="btn green" type="submit">{"Update Region" if safe_int(form_values.get('id')) > 0 else "Save Region"}</button>
                <a class="btn gray" href="/ui/operations/regions">Clear</a>
            </div>
        </form>
    </div>
    <div class="card">
        <h3 class="sub-title">Region List</h3>
        <table>
            <tr><th>Code</th><th>Region</th><th>Zone Level</th><th>Allowance</th><th>Status</th><th>Action</th></tr>
            {body}
        </table>
    </div>
    """
    return render_ops_page(request, "Regions", html)


@router.post("/ui/operations/regions/save")
def save_region(
    row_id: int = Form(0),
    code: str = Form(""),
    name: str = Form(""),
    zone_level: str = Form("zone_1"),
    allowance_amount: str = Form("0"),
    notes: str = Form(""),
    is_active: str = Form(""),
):
    conn = get_conn()
    active_flag = 1 if safe(is_active) == "1" else 0
    if row_id > 0:
        conn.execute("""
            UPDATE ops_regions
            SET code = ?, name = ?, zone_level = ?, allowance_amount = ?, notes = ?, is_active = ?
            WHERE id = ?
        """, (safe(code), safe(name), safe(zone_level), float(q2(allowance_amount)), safe(notes), active_flag, row_id))
        notice = "updated"
    else:
        conn.execute("""
            INSERT INTO ops_regions (code, name, zone_level, allowance_amount, notes, is_active)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (safe(code) or next_code("ops_regions", "REG"), safe(name), safe(zone_level), float(q2(allowance_amount)), safe(notes), active_flag))
        notice = "saved"
    conn.commit()
    conn.close()
    return RedirectResponse(f"/ui/operations/regions?notice={notice}", status_code=303)


@router.get("/ui/operations/service-catalog")
def service_catalog_redirect():
    return RedirectResponse("/ui/operations/action-catalog", status_code=302)


@router.get("/ui/operations/action-catalog", response_class=HTMLResponse)
def service_catalog_page(request: Request, edit_id: int = 0, notice: str = ""):
    conn = get_conn()
    rows = conn.execute("SELECT * FROM ops_service_catalog ORDER BY id DESC").fetchall()
    edit_row = conn.execute("SELECT * FROM ops_service_catalog WHERE id = ? LIMIT 1", (edit_id,)).fetchone() if edit_id else None
    form_values = dict(edit_row) if edit_row else {
        "id": 0,
        "code": next_action_code(),
        "name": "",
        "service_category": "field_maintenance",
        "unit_price": "0.00",
        "technician_incentive": "0.00",
        "default_region_level": "zone_1",
        "default_duration_hours": "0.00",
        "notes": "",
        "is_active": 1,
    }
    body = ""
    for row in rows:
        materials_count = conn.execute("SELECT COUNT(*) AS cnt FROM ops_action_materials WHERE action_id = ?", (row["id"],)).fetchone()["cnt"]
        body += f"""
        <tr>
            <td>{safe(row['code'])}</td>
            <td>{safe(row['name'])}</td>
            <td>{safe(row['service_category']).replace('_', ' ').title()}</td>
            <td>{money(row['unit_price'])}</td>
            <td>{money(row['technician_incentive'])}</td>
            <td>{safe(row['default_region_level']).replace('_', ' ').title()}</td>
            <td>{safe(row['default_duration_hours'])}</td>
            <td>{materials_count}</td>
            <td>{"Active" if int(row['is_active'] or 0) == 1 else "Inactive"}</td>
            <td>
                <a class="btn blue" href="/ui/operations/action-catalog?edit_id={row['id']}">Edit</a>
                <a class="btn gray" href="/ui/operations/action-catalog/{row['id']}/materials">Materials</a>
            </td>
        </tr>
        """
    if not body:
        body = "<tr><td colspan='10' style='text-align:center;'>No actions added yet.</td></tr>"
    html = f"""
    {current_form_notice(notice)}
    <div class="card">
        <div class="toolbar">
            <div>
                <h2 style="margin:0;">Action Catalog</h2>
                <div class="section-note">Priced actions used in field maintenance, workshop repairs, and cabinet assembly. Each action can have a fixed list of raw materials.</div>
            </div>
            <div style="display:flex;gap:8px;flex-wrap:wrap;">
                <a class="btn blue" href="/ui/operations/action-catalog/template.csv">Template CSV</a>
                <a class="btn blue" href="/ui/operations/action-catalog/template.xlsx">Template Excel</a>
                <a class="btn gray" href="/ui/operations">Back to Operations</a>
            </div>
        </div>
        <form method="post" action="/ui/operations/action-catalog/save">
            <input type="hidden" name="row_id" value="{safe(form_values.get('id', 0))}">
            <div class="form-grid">
                <div class="form-group"><label>Code</label><input name="code" value="{safe(form_values.get('code'))}" required></div>
                <div class="form-group"><label>Action Name</label><input name="name" value="{safe(form_values.get('name'))}" required></div>
                <div class="form-group"><label>Category</label><select name="service_category">{service_category_options(form_values.get('service_category'))}</select></div>
                <div class="form-group"><label>Unit Price</label><input name="unit_price" value="{safe(form_values.get('unit_price'))}"></div>
                <div class="form-group"><label>Technician Incentive</label><input name="technician_incentive" value="{safe(form_values.get('technician_incentive'))}"></div>
                <div class="form-group"><label>Default Region Level</label><select name="default_region_level">{zone_level_options(form_values.get('default_region_level'))}</select></div>
                <div class="form-group"><label>Default Duration (Hours)</label><input name="default_duration_hours" value="{safe(form_values.get('default_duration_hours'))}"></div>
                <div class="form-group"><label>Active</label><div style="padding-top:10px;"><label><input type="checkbox" name="is_active" value="1" {active_checkbox(int(form_values.get('is_active', 1) or 0) == 1)}> Active Service</label></div></div>
                <div class="form-group" style="grid-column: span 2;"><label>Notes</label><input name="notes" value="{safe(form_values.get('notes'))}"></div>
            </div>
            <div class="form-actions">
                <button class="btn green" type="submit">{"Update Action" if safe_int(form_values.get('id')) > 0 else "Save Action"}</button>
                <a class="btn gray" href="/ui/operations/action-catalog">Clear</a>
            </div>
        </form>
    </div>
    <div class="card">
        <form method="post" action="/ui/operations/action-catalog/import" enctype="multipart/form-data">
            <div class="toolbar">
                <div>
                    <h3 class="sub-title" style="margin:0;">Import Actions</h3>
                    <div class="section-note">Use the template to import priced actions in bulk, then open each action and attach its fixed raw materials list.</div>
                </div>
            </div>
            <div class="form-grid" style="margin-top:14px;">
                <div class="form-group" style="grid-column: span 2;">
                    <label>Import File (CSV / XLSX)</label>
                    <input type="file" name="file" accept=".csv,.xlsx" required>
                </div>
            </div>
            <div class="form-actions">
                <button class="btn green" type="submit">Import Actions</button>
            </div>
        </form>
    </div>
    <div class="card">
        <h3 class="sub-title">Action List</h3>
        <table>
            <tr><th>Code</th><th>Action</th><th>Category</th><th>Price</th><th>Incentive</th><th>Region Level</th><th>Hours</th><th>Materials</th><th>Status</th><th>Action</th></tr>
            {body}
        </table>
    </div>
    """
    conn.close()
    return render_ops_page(request, "Action Catalog", html)


@router.post("/ui/operations/action-catalog/save")
def save_service(
    row_id: int = Form(0),
    code: str = Form(""),
    name: str = Form(""),
    service_category: str = Form("field_maintenance"),
    unit_price: str = Form("0"),
    technician_incentive: str = Form("0"),
    default_region_level: str = Form("zone_1"),
    default_duration_hours: str = Form("0"),
    notes: str = Form(""),
    is_active: str = Form(""),
):
    conn = get_conn()
    active_flag = 1 if safe(is_active) == "1" else 0
    if row_id > 0:
        conn.execute("""
            UPDATE ops_service_catalog
            SET code = ?, name = ?, service_category = ?, unit_price = ?, technician_incentive = ?, default_region_level = ?, default_duration_hours = ?, notes = ?, is_active = ?
            WHERE id = ?
        """, (
            safe(code),
            safe(name),
            safe(service_category),
            float(q2(unit_price)),
            float(q2(technician_incentive)),
            safe(default_region_level),
            float(q2(default_duration_hours)),
            safe(notes),
            active_flag,
            row_id,
        ))
        notice = "updated"
    else:
        conn.execute("""
            INSERT INTO ops_service_catalog (
                code, name, service_category, unit_price, technician_incentive,
                default_region_level, default_duration_hours, notes, is_active
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            safe(code) or next_action_code(),
            safe(name),
            safe(service_category),
            float(q2(unit_price)),
            float(q2(technician_incentive)),
            safe(default_region_level),
            float(q2(default_duration_hours)),
            safe(notes),
            active_flag,
        ))
        notice = "saved"
    conn.commit()
    conn.close()
    return RedirectResponse(f"/ui/operations/action-catalog?notice={notice}", status_code=303)


@router.get("/ui/operations/action-catalog/template.csv")
def actions_template_csv():
    headers = [
        "code",
        "name",
        "service_category",
        "unit_price",
        "technician_incentive",
        "default_region_level",
        "default_duration_hours",
        "notes",
        "is_active",
    ]
    rows = [
        ["ACT-0001", "Replace Rectifier Fan", "workshop_repair", "1500", "250", "zone_1", "2", "Workshop repair action", "1"],
        ["ACT-0002", "Cabinet Assembly Visit", "cabinet_assembly", "3200", "450", "zone_2", "6", "Assembly and field finishing", "1"],
        ["ACT-0003", "BTS Power Fault Fix", "field_maintenance", "2800", "400", "zone_3", "4", "Standard field maintenance action", "1"],
    ]
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    data = output.getvalue().encode("utf-8-sig")
    stream = io.BytesIO(data)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": 'attachment; filename="action_catalog_template.csv"'},
    )


@router.get("/ui/operations/action-catalog/template.xlsx")
def actions_template_xlsx():
    if Workbook is None:
        return RedirectResponse("/ui/operations/action-catalog?notice=" + quote("Excel template is not available. Use CSV template."), status_code=302)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Action Catalog"
    sheet.append([
        "code",
        "name",
        "service_category",
        "unit_price",
        "technician_incentive",
        "default_region_level",
        "default_duration_hours",
        "notes",
        "is_active",
    ])
    sheet.append(["ACT-0001", "Replace Rectifier Fan", "workshop_repair", "1500", "250", "zone_1", "2", "Workshop repair action", "1"])
    sheet.append(["ACT-0002", "Cabinet Assembly Visit", "cabinet_assembly", "3200", "450", "zone_2", "6", "Assembly and field finishing", "1"])
    sheet.append(["ACT-0003", "BTS Power Fault Fix", "field_maintenance", "2800", "400", "zone_3", "4", "Standard field maintenance action", "1"])
    lookups = workbook.create_sheet("Lookups")
    lookups.append(["service_category", "region_level"])
    for idx in range(max(len(SERVICE_CATEGORIES), len(ZONE_LEVELS))):
        cat = SERVICE_CATEGORIES[idx][0] if idx < len(SERVICE_CATEGORIES) else ""
        zone = ZONE_LEVELS[idx][0] if idx < len(ZONE_LEVELS) else ""
        lookups.append([cat, zone])
    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="action_catalog_template.xlsx"'},
    )


@router.post("/ui/operations/action-catalog/import")
async def import_actions(file: UploadFile = File(...)):
    filename = safe(file.filename).lower()
    file_bytes = await file.read()
    if not filename:
        return RedirectResponse("/ui/operations/action-catalog?notice=" + quote("Please choose a file to import."), status_code=302)
    try:
        if filename.endswith(".csv"):
            raw_rows = parse_csv_rows(file_bytes)
        elif filename.endswith(".xlsx"):
            raw_rows = parse_xlsx_rows(file_bytes)
        else:
            return RedirectResponse("/ui/operations/action-catalog?notice=" + quote("Only CSV or XLSX files are supported."), status_code=302)
    except Exception as ex:
        return RedirectResponse("/ui/operations/action-catalog?notice=" + quote(f"Import failed: {safe(ex)}"), status_code=302)
    conn = get_conn()
    imported = 0
    updated = 0
    skipped = 0
    try:
        for raw_row in raw_rows:
            row = normalize_action_import_row(raw_row)
            if not safe(row["name"]):
                skipped += 1
                continue
            code = safe(row["code"]) or next_action_code()
            existing = conn.execute("SELECT id FROM ops_service_catalog WHERE UPPER(COALESCE(code,'')) = UPPER(?) LIMIT 1", (code,)).fetchone()
            payload = (
                code,
                safe(row["name"]),
                safe(row["service_category"]) or "field_maintenance",
                float(q2(row["unit_price"])),
                float(q2(row["technician_incentive"])),
                safe(row["default_region_level"]) or "zone_1",
                float(q2(row["default_duration_hours"])),
                safe(row["notes"]),
                1 if safe(row["is_active"]) not in ["0", "false", "False"] else 0,
            )
            if existing:
                conn.execute("""
                    UPDATE ops_service_catalog
                    SET code = ?, name = ?, service_category = ?, unit_price = ?, technician_incentive = ?,
                        default_region_level = ?, default_duration_hours = ?, notes = ?, is_active = ?
                    WHERE id = ?
                """, payload + (existing["id"],))
                updated += 1
            else:
                conn.execute("""
                    INSERT INTO ops_service_catalog (
                        code, name, service_category, unit_price, technician_incentive,
                        default_region_level, default_duration_hours, notes, is_active
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, payload)
                imported += 1
        conn.commit()
    except Exception as ex:
        conn.rollback()
        conn.close()
        return RedirectResponse("/ui/operations/action-catalog?notice=" + quote(f"Import failed: {safe(ex)}"), status_code=302)
    conn.close()
    msg = f"Import completed. Added: {imported}, Updated: {updated}, Skipped: {skipped}."
    return RedirectResponse("/ui/operations/action-catalog?notice=" + quote(msg), status_code=302)


@router.get("/ui/operations/action-catalog/{action_id}/materials", response_class=HTMLResponse)
def action_materials_page(request: Request, action_id: int):
    conn = get_conn()
    action = conn.execute("SELECT * FROM ops_service_catalog WHERE id = ? LIMIT 1", (action_id,)).fetchone()
    if not action:
        conn.close()
        return HTMLResponse("Action not found.", status_code=404)
    materials = get_action_materials(conn, action_id)
    conn.close()
    rows_html = ""
    total_cost = Decimal("0.00")
    for row in materials:
        line_total = q2(row["qty"]) * q2(row["unit_cost"])
        total_cost += line_total
        rows_html += f"""
        <tr>
            <td>{safe(row['line_no'])}</td>
            <td>{safe(row['item_code'])}</td>
            <td>{safe(row['item_name'])}</td>
            <td>{safe(row['uom'])}</td>
            <td>{money(row['qty'])}</td>
            <td>{money(row['unit_cost'])}</td>
            <td>{money(line_total)}</td>
            <td>{safe(row['notes'])}</td>
            <td>
                <form method="post" action="/ui/operations/action-catalog/{action_id}/materials/{row['id']}/delete" style="display:inline;">
                    <button class="btn red" type="submit">Delete</button>
                </form>
            </td>
        </tr>
        """
    if not rows_html:
        rows_html = "<tr><td colspan='9' style='text-align:center;'>No fixed raw materials added for this action.</td></tr>"
    html = f"""
    <div class="card">
        <div class="toolbar">
            <div>
                <h2 style="margin:0;">Action Materials</h2>
                <div class="section-note"><b>{safe(action['code'])}</b> - {safe(action['name'])}. These are the fixed raw materials used to cost the work order automatically.</div>
            </div>
            <div style="display:flex;gap:8px;flex-wrap:wrap;">
                <a class="btn blue" href="/ui/operations/action-catalog?edit_id={action_id}">Back to Action</a>
                <a class="btn gray" href="/ui/operations/action-catalog">Back to Catalog</a>
            </div>
        </div>
        <div class="table-summary">
            <span class="summary-pill">Action Price: {money(action['unit_price'])}</span>
            <span class="summary-pill">Incentive: {money(action['technician_incentive'])}</span>
            <span class="summary-pill">Fixed Material Cost: {money(total_cost)}</span>
        </div>
        <form method="post" action="/ui/operations/action-catalog/{action_id}/materials/add" style="margin-top:16px;">
            <div class="form-grid">
                <div class="form-group" style="grid-column: span 2;"><label>Item</label><select name="item_id" required>{item_options()}</select></div>
                <div class="form-group"><label>Fixed Qty</label><input name="qty" value="1" required></div>
                <div class="form-group"><label>Standard Unit Cost</label><input name="unit_cost" value="0"></div>
                <div class="form-group" style="grid-column: span 2;"><label>Notes</label><input name="notes"></div>
            </div>
            <div class="form-actions">
                <button class="btn green" type="submit">Add Fixed Material</button>
            </div>
        </form>
    </div>
    <div class="card">
        <h3 class="sub-title">Fixed Raw Materials</h3>
        <table>
            <tr><th>#</th><th>Item Code</th><th>Item Name</th><th>UOM</th><th>Qty</th><th>Unit Cost</th><th>Line Cost</th><th>Notes</th><th>Action</th></tr>
            {rows_html}
        </table>
    </div>
    """
    return render_ops_page(request, "Action Materials", html)


@router.post("/ui/operations/action-catalog/{action_id}/materials/add")
def add_action_material(
    action_id: int,
    item_id: int = Form(...),
    qty: str = Form("1"),
    unit_cost: str = Form("0"),
    notes: str = Form(""),
):
    item = get_item_by_id(item_id)
    if not item:
        return HTMLResponse("Item not found.", status_code=404)
    resolved_cost = q2(unit_cost)
    if resolved_cost <= Decimal("0.00"):
        resolved_cost = get_item_last_cost(item_id)
    conn = get_conn()
    next_line = conn.execute("SELECT COALESCE(MAX(line_no), 0) + 1 AS next_line FROM ops_action_materials WHERE action_id = ?", (action_id,)).fetchone()["next_line"]
    conn.execute("""
        INSERT INTO ops_action_materials (
            action_id, line_no, item_id, item_code, item_name, uom, qty, unit_cost, notes
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        action_id,
        int(next_line or 1),
        item_id,
        safe(item["code"]),
        safe(item["name"]),
        safe(item["uom"]),
        float(q2(qty)),
        float(resolved_cost),
        safe(notes),
    ))
    conn.commit()
    conn.close()
    return RedirectResponse(f"/ui/operations/action-catalog/{action_id}/materials", status_code=303)


@router.post("/ui/operations/action-catalog/{action_id}/materials/{material_id}/delete")
def delete_action_material(action_id: int, material_id: int):
    conn = get_conn()
    conn.execute("DELETE FROM ops_action_materials WHERE id = ? AND action_id = ?", (material_id, action_id))
    conn.commit()
    conn.close()
    return RedirectResponse(f"/ui/operations/action-catalog/{action_id}/materials", status_code=303)


@router.get("/ui/operations/vehicle-rates", response_class=HTMLResponse)
def vehicle_rates_page(request: Request, edit_id: int = 0, notice: str = ""):
    conn = get_conn()
    rows = conn.execute("""
        SELECT vr.*, rs.name AS supplier_name
        FROM ops_vehicle_rates vr
        LEFT JOIN ops_rental_suppliers rs ON rs.id = vr.rental_supplier_id
        ORDER BY vr.id DESC
    """).fetchall()
    edit_row = conn.execute("SELECT * FROM ops_vehicle_rates WHERE id = ? LIMIT 1", (edit_id,)).fetchone() if edit_id else None
    conn.close()
    form_values = dict(edit_row) if edit_row else {
        "id": 0,
        "code": next_code("ops_vehicle_rates", "VEH"),
        "rental_supplier_id": "",
        "slab_name": "standard",
        "vehicle_type": "",
        "ticket_open_price": "0.00",
        "second_slab_to_km": "300.00",
        "km_rate_101_300": "0.00",
        "km_rate_over_300": "0.00",
        "waiting_hour_rate": "0.00",
        "notes": "",
        "is_active": 1,
    }
    body = ""
    for row in rows:
        body += f"""
        <tr>
            <td>{safe(row['code'])}</td>
            <td>{safe(row['supplier_name']) or '-'}</td>
            <td>{safe(row['vehicle_type'])}</td>
            <td>{safe(row['slab_name']) or 'standard'}</td>
            <td>{money(row['ticket_open_price'])}</td>
            <td>{money(row['second_slab_to_km'])}</td>
            <td>{money(row['km_rate_101_300'])}</td>
            <td>{money(row['km_rate_over_300'])}</td>
            <td>{money(row['waiting_hour_rate'])}</td>
            <td>{"Active" if int(row['is_active'] or 0) == 1 else "Inactive"}</td>
            <td><a class="btn blue" href="/ui/operations/vehicle-rates?edit_id={row['id']}">Edit</a></td>
        </tr>
        """
    if not body:
        body = "<tr><td colspan='11' style='text-align:center;'>No vehicle rates added yet.</td></tr>"
    html = f"""
    {current_form_notice(notice)}
    <div class="card">
        <div class="toolbar">
            <div>
                <h2 style="margin:0;">Vehicle Rates</h2>
                <div class="section-note">Trip pricing is now flexible: fixed ticket up to 100 KM, then a second slab up to the limit you define here, then a final over-limit per-KM rate, plus waiting hours.</div>
            </div>
            <a class="btn gray" href="/ui/operations">Back to Operations</a>
        </div>
        <form method="post" action="/ui/operations/vehicle-rates/save">
            <input type="hidden" name="row_id" value="{safe(form_values.get('id', 0))}">
            <div class="form-grid">
                <div class="form-group"><label>Code</label><input name="code" value="{safe(form_values.get('code'))}" required></div>
                <div class="form-group"><label>Rental Office</label><select name="rental_supplier_id" required>{rental_supplier_options(form_values.get('rental_supplier_id'))}</select></div>
                <div class="form-group"><label>Vehicle Type</label><input name="vehicle_type" value="{safe(form_values.get('vehicle_type'))}" required></div>
                <div class="form-group"><label>Pricing Slab</label><input name="slab_name" value="{safe(form_values.get('slab_name') or 'standard')}" required></div>
                <div class="form-group"><label>0 - 100 KM Ticket Price</label><input name="ticket_open_price" value="{safe(form_values.get('ticket_open_price'))}"></div>
                <div class="form-group"><label>Second Slab Upper KM</label><input name="second_slab_to_km" value="{safe(form_values.get('second_slab_to_km', 300))}"></div>
                <div class="form-group"><label>101 - Upper Slab Rate</label><input name="km_rate_101_300" value="{safe(form_values.get('km_rate_101_300'))}"></div>
                <div class="form-group"><label>Over Upper Slab Rate</label><input name="km_rate_over_300" value="{safe(form_values.get('km_rate_over_300'))}"></div>
                <div class="form-group"><label>Waiting Hour Rate</label><input name="waiting_hour_rate" value="{safe(form_values.get('waiting_hour_rate'))}"></div>
                <div class="form-group"><label>Active</label><div style="padding-top:10px;"><label><input type="checkbox" name="is_active" value="1" {active_checkbox(int(form_values.get('is_active', 1) or 0) == 1)}> Active Vehicle Rate</label></div></div>
                <div class="form-group" style="grid-column: span 2;"><label>Notes</label><input name="notes" value="{safe(form_values.get('notes'))}"></div>
            </div>
            <div class="card" style="margin-top:14px;background:#f8fbff;">
                <div class="section-note">
                    Costing concept:
                    1. If trip KM <= 100: total = ticket price + waiting.
                    2. If trip KM is from 101 to the upper slab KM: total = ticket price for first 100 + ((KM - 100) x slab rate) + waiting.
                    3. If trip KM is above the upper slab KM: total = KM x over-slab rate + waiting.
                </div>
            </div>
            <div class="form-actions">
                <button class="btn green" type="submit">{"Update Vehicle Rate" if safe_int(form_values.get('id')) > 0 else "Save Vehicle Rate"}</button>
                <a class="btn gray" href="/ui/operations/vehicle-rates">Clear</a>
            </div>
        </form>
    </div>
    <div class="card">
        <h3 class="sub-title">Vehicle Rate List</h3>
        <table>
            <tr><th>Code</th><th>Rental Office</th><th>Vehicle Type</th><th>Pricing Slab</th><th>0-100 Ticket</th><th>Upper Slab KM</th><th>101-Upper Rate</th><th>Over Upper Rate</th><th>Waiting Hour</th><th>Status</th><th>Action</th></tr>
            {body}
        </table>
    </div>
    """
    return render_ops_page(request, "Vehicle Rates", html)


@router.post("/ui/operations/vehicle-rates/save")
def save_vehicle_rate(
    row_id: int = Form(0),
    code: str = Form(""),
    rental_supplier_id: int = Form(0),
    slab_name: str = Form("standard"),
    vehicle_type: str = Form(""),
    ticket_open_price: str = Form("0"),
    second_slab_to_km: str = Form("300"),
    km_rate_101_300: str = Form("0"),
    km_rate_over_300: str = Form("0"),
    waiting_hour_rate: str = Form("0"),
    notes: str = Form(""),
    is_active: str = Form(""),
):
    if safe_int(rental_supplier_id) <= 0:
        return RedirectResponse("/ui/operations/vehicle-rates?notice=" + quote("Please select the rental office for this pricing slab."), status_code=302)
    conn = get_conn()
    active_flag = 1 if safe(is_active) == "1" else 0
    if row_id > 0:
        conn.execute("""
            UPDATE ops_vehicle_rates
            SET code = ?, rental_supplier_id = ?, slab_name = ?, vehicle_type = ?, ticket_open_price = ?, second_slab_to_km = ?, km_rate_101_300 = ?, km_rate_over_300 = ?, km_rate = ?, waiting_hour_rate = ?, notes = ?, is_active = ?
            WHERE id = ?
        """, (
            safe(code),
            safe_int(rental_supplier_id),
            safe(slab_name) or "standard",
            safe(vehicle_type),
            float(q2(ticket_open_price)),
            float(q2(second_slab_to_km)),
            float(q2(km_rate_101_300)),
            float(q2(km_rate_over_300)),
            float(q2(km_rate_over_300)),
            float(q2(waiting_hour_rate)),
            safe(notes),
            active_flag,
            row_id,
        ))
        notice = "updated"
    else:
        conn.execute("""
            INSERT INTO ops_vehicle_rates (
                code, rental_supplier_id, slab_name, vehicle_type, ticket_open_price, second_slab_to_km, km_rate_101_300, km_rate_over_300, km_rate, waiting_hour_rate, notes, is_active
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            safe(code) or next_code("ops_vehicle_rates", "VEH"),
            safe_int(rental_supplier_id),
            safe(slab_name) or "standard",
            safe(vehicle_type),
            float(q2(ticket_open_price)),
            float(q2(second_slab_to_km)),
            float(q2(km_rate_101_300)),
            float(q2(km_rate_over_300)),
            float(q2(km_rate_over_300)),
            float(q2(waiting_hour_rate)),
            safe(notes),
            active_flag,
        ))
        notice = "saved"
    conn.commit()
    conn.close()
    return RedirectResponse(f"/ui/operations/vehicle-rates?notice={notice}", status_code=303)


@router.get("/ui/operations/rental-suppliers", response_class=HTMLResponse)
def rental_suppliers_page(request: Request, edit_id: int = 0, notice: str = ""):
    conn = get_conn()
    rows = conn.execute("SELECT * FROM ops_rental_suppliers ORDER BY id DESC").fetchall()
    edit_row = conn.execute("SELECT * FROM ops_rental_suppliers WHERE id = ? LIMIT 1", (edit_id,)).fetchone() if edit_id else None
    conn.close()
    form_values = dict(edit_row) if edit_row else {
        "id": 0,
        "code": next_rental_supplier_code(),
        "name": "",
        "contact_person": "",
        "phone": "",
        "notes": "",
        "is_active": 1,
    }
    body = ""
    for row in rows:
        body += f"""
        <tr>
            <td>{safe(row['code'])}</td>
            <td>{safe(row['name'])}</td>
            <td>{safe(row['contact_person'])}</td>
            <td>{safe(row['phone'])}</td>
            <td>{"Active" if int(row['is_active'] or 0) == 1 else "Inactive"}</td>
            <td><a class="btn blue" href="/ui/operations/rental-suppliers?edit_id={row['id']}">Edit</a></td>
        </tr>
        """
    if not body:
        body = "<tr><td colspan='6' style='text-align:center;'>No rental offices added yet.</td></tr>"
    html = f"""
    {current_form_notice(notice)}
    <div class="card">
        <div class="toolbar">
            <div>
                <h2 style="margin:0;">Rental Offices</h2>
                <div class="section-note">Car rental suppliers or offices used for trip tickets. The same vehicle stays tied to one office.</div>
            </div>
            <div style="display:flex;gap:8px;flex-wrap:wrap;">
                <a class="btn blue" href="/ui/operations/vehicles">Vehicles</a>
                <a class="btn gray" href="/ui/operations">Back to Operations</a>
            </div>
        </div>
        <form method="post" action="/ui/operations/rental-suppliers/save">
            <input type="hidden" name="row_id" value="{safe(form_values.get('id', 0))}">
            <div class="form-grid">
                <div class="form-group"><label>Code</label><input name="code" value="{safe(form_values.get('code'))}" required></div>
                <div class="form-group"><label>Office Name</label><input name="name" value="{safe(form_values.get('name'))}" required></div>
                <div class="form-group"><label>Contact Person</label><input name="contact_person" value="{safe(form_values.get('contact_person'))}"></div>
                <div class="form-group"><label>Phone</label><input name="phone" value="{safe(form_values.get('phone'))}"></div>
                <div class="form-group"><label>Active</label><div style="padding-top:10px;"><label><input type="checkbox" name="is_active" value="1" {active_checkbox(int(form_values.get('is_active', 1) or 0) == 1)}> Active Rental Office</label></div></div>
                <div class="form-group" style="grid-column: span 2;"><label>Notes</label><input name="notes" value="{safe(form_values.get('notes'))}"></div>
            </div>
            <div class="form-actions">
                <button class="btn green" type="submit">{"Update Office" if safe_int(form_values.get('id')) > 0 else "Save Office"}</button>
                <a class="btn gray" href="/ui/operations/rental-suppliers">Clear</a>
            </div>
        </form>
    </div>
    <div class="card">
        <h3 class="sub-title">Rental Office List</h3>
        <table>
            <tr><th>Code</th><th>Name</th><th>Contact</th><th>Phone</th><th>Status</th><th>Action</th></tr>
            {body}
        </table>
    </div>
    """
    return render_ops_page(request, "Rental Offices", html)


@router.post("/ui/operations/rental-suppliers/save")
def save_rental_supplier(
    row_id: int = Form(0),
    code: str = Form(""),
    name: str = Form(""),
    contact_person: str = Form(""),
    phone: str = Form(""),
    notes: str = Form(""),
    is_active: str = Form(""),
):
    conn = get_conn()
    active_flag = 1 if safe(is_active) == "1" else 0
    if row_id > 0:
        conn.execute("""
            UPDATE ops_rental_suppliers
            SET code = ?, name = ?, contact_person = ?, phone = ?, notes = ?, is_active = ?
            WHERE id = ?
        """, (safe(code), safe(name), safe(contact_person), safe(phone), safe(notes), active_flag, row_id))
        notice = "updated"
    else:
        conn.execute("""
            INSERT INTO ops_rental_suppliers (code, name, contact_person, phone, notes, is_active)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (safe(code) or next_rental_supplier_code(), safe(name), safe(contact_person), safe(phone), safe(notes), active_flag))
        notice = "saved"
    conn.commit()
    conn.close()
    return RedirectResponse(f"/ui/operations/rental-suppliers?notice={notice}", status_code=303)


@router.get("/ui/operations/vehicles", response_class=HTMLResponse)
def vehicles_page(request: Request, edit_id: int = 0, notice: str = ""):
    conn = get_conn()
    rows = conn.execute("""
        SELECT
            v.*,
            rs.name AS supplier_name,
            vr.code AS rate_code,
            vr.vehicle_type AS rate_vehicle_type,
            vr.slab_name AS rate_slab_name
        FROM ops_vehicles v
        LEFT JOIN ops_rental_suppliers rs ON rs.id = v.rental_supplier_id
        LEFT JOIN ops_vehicle_rates vr ON vr.id = v.vehicle_rate_id
        ORDER BY v.id DESC
    """).fetchall()
    edit_row = conn.execute("SELECT * FROM ops_vehicles WHERE id = ? LIMIT 1", (edit_id,)).fetchone() if edit_id else None
    conn.close()
    form_values = dict(edit_row) if edit_row else {
        "id": 0,
        "code": next_vehicle_code(),
        "vehicle_name": "",
        "vehicle_type": "",
        "rental_supplier_id": "",
        "vehicle_rate_id": "",
        "pricing_slab_name": "standard",
        "plate_no": "",
        "driver_source": "company_driver",
        "supplier_driver_name": "",
        "notes": "",
        "is_active": 1,
    }
    body = ""
    for row in rows:
        rate_label = safe(row["pricing_slab_name"]) or safe(row["rate_slab_name"]) or "standard"
        driver_label = "Company Driver" if safe(row["driver_source"]) == "company_driver" else "Supplier Driver"
        body += f"""
        <tr>
            <td>{safe(row['code'])}</td>
            <td>{safe(row['vehicle_name'])}</td>
            <td>{safe(row['vehicle_type'])}</td>
            <td>{safe(row['supplier_name'])}</td>
            <td>{safe(row['plate_no'])}</td>
            <td>{driver_label}</td>
            <td>{rate_label}</td>
            <td>{"Active" if int(row['is_active'] or 0) == 1 else "Inactive"}</td>
            <td><a class="btn blue" href="/ui/operations/vehicles?edit_id={row['id']}">Edit</a></td>
        </tr>
        """
    if not body:
        body = "<tr><td colspan='9' style='text-align:center;'>No vehicles added yet.</td></tr>"
    html = f"""
    {current_form_notice(notice)}
    <div class="card">
        <div class="toolbar">
            <div>
                <h2 style="margin:0;">Vehicles</h2>
                <div class="section-note">Each vehicle belongs to one rental office and one pricing slab. New trips automatically pick the latest active pricing slab for that office, type, and slab.</div>
            </div>
            <div style="display:flex;gap:8px;flex-wrap:wrap;">
                <a class="btn blue" href="/ui/operations/rental-suppliers">Rental Offices</a>
                <a class="btn blue" href="/ui/operations/vehicle-rates">Vehicle Rates</a>
                <a class="btn gray" href="/ui/operations">Back to Operations</a>
            </div>
        </div>
        <form method="post" action="/ui/operations/vehicles/save">
            <input type="hidden" name="row_id" value="{safe(form_values.get('id', 0))}">
            <div class="form-grid">
                <div class="form-group"><label>Vehicle Code</label><input name="code" value="{safe(form_values.get('code'))}" required></div>
                <div class="form-group"><label>Vehicle Name</label><input name="vehicle_name" value="{safe(form_values.get('vehicle_name'))}" placeholder="Optional display name"></div>
                <div class="form-group"><label>Vehicle Type</label><input name="vehicle_type" value="{safe(form_values.get('vehicle_type'))}" required></div>
                <div class="form-group"><label>Rental Office</label><select name="rental_supplier_id" required>{rental_supplier_options(form_values.get('rental_supplier_id'))}</select></div>
                <div class="form-group"><label>Pricing Slab</label><input name="pricing_slab_name" value="{safe(form_values.get('pricing_slab_name') or 'standard')}" required></div>
                <div class="form-group"><label>Reference Rate Card (Optional)</label><select name="vehicle_rate_id">{vehicle_rate_options(form_values.get('vehicle_rate_id'))}</select></div>
                <div class="form-group"><label>Plate No</label><input name="plate_no" value="{safe(form_values.get('plate_no'))}"></div>
                <div class="form-group"><label>Driver Source</label><select name="driver_source">{driver_source_options(form_values.get('driver_source'))}</select></div>
                <div class="form-group"><label>Default Supplier Driver Name</label><input name="supplier_driver_name" value="{safe(form_values.get('supplier_driver_name'))}"></div>
                <div class="form-group"><label>Active</label><div style="padding-top:10px;"><label><input type="checkbox" name="is_active" value="1" {active_checkbox(int(form_values.get('is_active', 1) or 0) == 1)}> Active Vehicle</label></div></div>
                <div class="form-group" style="grid-column: span 2;"><label>Notes</label><input name="notes" value="{safe(form_values.get('notes'))}"></div>
            </div>
            <div class="form-actions">
                <button class="btn green" type="submit">{"Update Vehicle" if safe_int(form_values.get('id')) > 0 else "Save Vehicle"}</button>
                <a class="btn gray" href="/ui/operations/vehicles">Clear</a>
            </div>
        </form>
    </div>
    <div class="card">
        <h3 class="sub-title">Vehicle List</h3>
        <table>
            <tr><th>Code</th><th>Name</th><th>Type</th><th>Rental Office</th><th>Plate</th><th>Driver Source</th><th>Pricing Slab</th><th>Status</th><th>Action</th></tr>
            {body}
        </table>
    </div>
    """
    return render_ops_page(request, "Vehicles", html)


@router.post("/ui/operations/vehicles/save")
def save_vehicle(
    row_id: int = Form(0),
    code: str = Form(""),
    vehicle_name: str = Form(""),
    vehicle_type: str = Form(""),
    rental_supplier_id: int = Form(0),
    vehicle_rate_id: int = Form(0),
    pricing_slab_name: str = Form("standard"),
    plate_no: str = Form(""),
    driver_source: str = Form("company_driver"),
    supplier_driver_name: str = Form(""),
    notes: str = Form(""),
    is_active: str = Form(""),
):
    if safe_int(rental_supplier_id) <= 0:
        return RedirectResponse("/ui/operations/vehicles?notice=" + quote("Please select a rental office."), status_code=302)
    conn = get_conn()
    active_flag = 1 if safe(is_active) == "1" else 0
    values = (
        safe(code),
        safe(vehicle_name),
        safe(vehicle_type),
        safe_int(rental_supplier_id),
        safe_int(vehicle_rate_id),
        safe(pricing_slab_name) or "standard",
        safe(plate_no),
        safe(driver_source) or "company_driver",
        safe(supplier_driver_name),
        safe(notes),
        active_flag,
    )
    if row_id > 0:
        conn.execute("""
            UPDATE ops_vehicles
            SET code = ?, vehicle_name = ?, vehicle_type = ?, rental_supplier_id = ?, vehicle_rate_id = ?,
                pricing_slab_name = ?, plate_no = ?, driver_source = ?, supplier_driver_name = ?, notes = ?, is_active = ?
            WHERE id = ?
        """, values + (row_id,))
        notice = "updated"
    else:
        conn.execute("""
            INSERT INTO ops_vehicles (
                code, vehicle_name, vehicle_type, rental_supplier_id, vehicle_rate_id,
                pricing_slab_name, plate_no, driver_source, supplier_driver_name, notes, is_active
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            safe(code) or next_vehicle_code(),
            safe(vehicle_name),
            safe(vehicle_type),
            safe_int(rental_supplier_id),
            safe_int(vehicle_rate_id),
            safe(pricing_slab_name) or "standard",
            safe(plate_no),
            safe(driver_source) or "company_driver",
            safe(supplier_driver_name),
            safe(notes),
            active_flag,
        ))
        notice = "saved"
    conn.commit()
    conn.close()
    return RedirectResponse(f"/ui/operations/vehicles?notice={notice}", status_code=303)


@router.get("/ui/operations/contracts", response_class=HTMLResponse)
def contracts_page(request: Request, edit_id: int = 0, notice: str = ""):
    conn = get_conn()
    rows = conn.execute("""
        SELECT c.*, co.code AS company_code, co.name AS company_name
        FROM ops_contracts c
        LEFT JOIN ops_contract_companies co ON co.id = c.company_id
        ORDER BY c.id DESC
    """).fetchall()
    edit_row = conn.execute("SELECT * FROM ops_contracts WHERE id = ? LIMIT 1", (edit_id,)).fetchone() if edit_id else None
    conn.close()
    form_values = dict(edit_row) if edit_row else {
        "id": 0,
        "contract_no": next_contract_no(),
        "company_id": "",
        "contract_name": "",
        "pricing_method": "price_list",
        "start_date": "",
        "end_date": "",
        "status": "active",
        "notes": "",
    }
    body = ""
    for row in rows:
        body += f"""
        <tr>
            <td>{safe(row['contract_no'])}</td>
            <td>{safe(row['company_code'])} - {safe(row['company_name'])}</td>
            <td>{safe(row['contract_name'])}</td>
            <td>{safe(row['pricing_method']).replace('_', ' ').title()}</td>
            <td>{safe(row['start_date'])}</td>
            <td>{safe(row['end_date'])}</td>
            <td>{status_chip(row['status'])}</td>
            <td><a class="btn blue" href="/ui/operations/contracts?edit_id={row['id']}">Edit</a></td>
        </tr>
        """
    if not body:
        body = "<tr><td colspan='8' style='text-align:center;'>No contracts added yet.</td></tr>"
    html = f"""
    {current_form_notice(notice)}
    <div class="card">
        <div class="toolbar">
            <div><h2 style="margin:0;">Contracts</h2></div>
            <a class="btn gray" href="/ui/operations">Back to Operations</a>
        </div>
        <form method="post" action="/ui/operations/contracts/save">
            <input type="hidden" name="row_id" value="{safe(form_values.get('id', 0))}">
            <div class="form-grid">
                <div class="form-group"><label>Contract No</label><input name="contract_no" value="{safe(form_values.get('contract_no'))}" required></div>
                <div class="form-group"><label>Company</label><select name="company_id" required>{company_options(form_values.get('company_id'))}</select></div>
                <div class="form-group"><label>Contract Name</label><input name="contract_name" value="{safe(form_values.get('contract_name'))}" required></div>
                <div class="form-group"><label>Pricing Method</label><select name="pricing_method">{simple_options([('price_list','Price List'),('manual','Manual'),('contract_version','Contract Version')], form_values.get('pricing_method'))}</select></div>
                <div class="form-group"><label>Start Date</label><input type="date" name="start_date" value="{safe(form_values.get('start_date'))}"></div>
                <div class="form-group"><label>End Date</label><input type="date" name="end_date" value="{safe(form_values.get('end_date'))}"></div>
                <div class="form-group"><label>Status</label><select name="status">{simple_options([('active','Active'),('inactive','Inactive'),('expired','Expired')], form_values.get('status'))}</select></div>
                <div class="form-group" style="grid-column: span 2;"><label>Notes</label><input name="notes" value="{safe(form_values.get('notes'))}"></div>
            </div>
            <div class="form-actions">
                <button class="btn green" type="submit">{"Update Contract" if safe_int(form_values.get('id')) > 0 else "Save Contract"}</button>
                <a class="btn gray" href="/ui/operations/contracts">Clear</a>
            </div>
        </form>
    </div>
    <div class="card">
        <h3 class="sub-title">Contract List</h3>
        <table>
            <tr><th>No</th><th>Company</th><th>Name</th><th>Pricing</th><th>Start</th><th>End</th><th>Status</th><th>Action</th></tr>
            {body}
        </table>
    </div>
    """
    return render_ops_page(request, "Contracts", html)


@router.post("/ui/operations/contracts/save")
def save_contract(
    request: Request,
    row_id: int = Form(0),
    contract_no: str = Form(""),
    company_id: int = Form(0),
    contract_name: str = Form(""),
    pricing_method: str = Form("price_list"),
    start_date: str = Form(""),
    end_date: str = Form(""),
    status: str = Form("active"),
    notes: str = Form(""),
):
    conn = get_conn()
    if row_id > 0:
        conn.execute("""
            UPDATE ops_contracts
            SET contract_no = ?, company_id = ?, contract_name = ?, pricing_method = ?,
                start_date = ?, end_date = ?, status = ?, notes = ?
            WHERE id = ?
        """, (safe(contract_no), safe_int(company_id), safe(contract_name), safe(pricing_method), safe(start_date), safe(end_date), safe(status), safe(notes), row_id))
        log_ops_event(request, "ops_contract", row_id, "Updated", f"Contract {safe(contract_no)}", conn=conn)
        notice = "updated"
    else:
        cur = conn.execute("""
            INSERT INTO ops_contracts (contract_no, company_id, contract_name, pricing_method, start_date, end_date, status, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (safe(contract_no) or next_contract_no(), safe_int(company_id), safe(contract_name), safe(pricing_method), safe(start_date), safe(end_date), safe(status), safe(notes)))
        log_ops_event(request, "ops_contract", cur.lastrowid, "Created", f"Contract {safe(contract_no)}", conn=conn)
        notice = "saved"
    conn.commit()
    conn.close()
    return RedirectResponse(f"/ui/operations/contracts?notice={notice}", status_code=303)


@router.get("/ui/operations/tickets", response_class=HTMLResponse)
def tickets_page(request: Request, edit_id: int = 0, notice: str = ""):
    conn = get_conn()
    rows = conn.execute("""
        SELECT t.*, co.name AS company_name, c.contract_no, f.name AS fault_name
        FROM ops_tickets t
        LEFT JOIN ops_contract_companies co ON co.id = t.company_id
        LEFT JOIN ops_contracts c ON c.id = t.contract_id
        LEFT JOIN ops_fault_types f ON f.id = t.fault_type_id
        ORDER BY t.id DESC
    """).fetchall()
    edit_row = conn.execute("SELECT * FROM ops_tickets WHERE id = ? LIMIT 1", (edit_id,)).fetchone() if edit_id else None
    conn.close()
    form_values = dict(edit_row) if edit_row else {
        "id": 0,
        "ticket_no": next_ticket_no(),
        "company_id": "",
        "contract_id": "",
        "ticket_date": "",
        "fault_type_id": "",
        "site_code": "",
        "site_name": "",
        "priority": "normal",
        "request_channel": "",
        "complaint_details": "",
        "status": "open",
    }
    body = ""
    for row in rows:
        body += f"""
        <tr>
            <td>{safe(row['ticket_no'])}</td>
            <td>{safe(row['ticket_date'])}</td>
            <td>{safe(row['company_name'])}</td>
            <td>{safe(row['contract_no'])}</td>
            <td>{safe(row['fault_name'])}</td>
            <td>{safe(row['site_code'])}</td>
            <td>{status_chip(row['status'])}</td>
            <td><a class="btn blue" href="/ui/operations/tickets?edit_id={row['id']}">Edit</a></td>
        </tr>
        """
    if not body:
        body = "<tr><td colspan='8' style='text-align:center;'>No tickets opened yet.</td></tr>"
    html = f"""
    {current_form_notice(notice)}
    <div class="card">
        <div class="toolbar">
            <div><h2 style="margin:0;">Tickets</h2></div>
            <a class="btn gray" href="/ui/operations">Back to Operations</a>
        </div>
        <form method="post" action="/ui/operations/tickets/save">
            <input type="hidden" name="row_id" value="{safe(form_values.get('id', 0))}">
            <div class="form-grid">
                <div class="form-group"><label>Ticket No</label><input name="ticket_no" value="{safe(form_values.get('ticket_no'))}" required></div>
                <div class="form-group"><label>Ticket Date</label><input type="date" name="ticket_date" value="{safe(form_values.get('ticket_date'))}" required></div>
                <div class="form-group"><label>Company</label><select name="company_id" required>{company_options(form_values.get('company_id'))}</select></div>
                <div class="form-group"><label>Contract</label><select name="contract_id">{contract_options(form_values.get('contract_id'))}</select></div>
                <div class="form-group"><label>Fault Type</label><select name="fault_type_id">{fault_options(form_values.get('fault_type_id'))}</select></div>
                <div class="form-group"><label>Priority</label><select name="priority">{simple_options([('low','Low'),('normal','Normal'),('high','High'),('urgent','Urgent')], form_values.get('priority'))}</select></div>
                <div class="form-group"><label>Site Code</label><input name="site_code" value="{safe(form_values.get('site_code'))}"></div>
                <div class="form-group"><label>Site Name</label><input name="site_name" value="{safe(form_values.get('site_name'))}"></div>
                <div class="form-group"><label>Request Channel</label><input name="request_channel" value="{safe(form_values.get('request_channel'))}"></div>
                <div class="form-group"><label>Status</label><select name="status">{simple_options([('open','Open'),('in_progress','In Progress'),('closed','Closed'),('cancelled','Cancelled')], form_values.get('status'))}</select></div>
                <div class="form-group" style="grid-column: span 2;"><label>Complaint Details</label><input name="complaint_details" value="{safe(form_values.get('complaint_details'))}"></div>
            </div>
            <div class="form-actions">
                <button class="btn green" type="submit">{"Update Ticket" if safe_int(form_values.get('id')) > 0 else "Save Ticket"}</button>
                <a class="btn gray" href="/ui/operations/tickets">Clear</a>
            </div>
        </form>
    </div>
    <div class="card">
        <h3 class="sub-title">Ticket List</h3>
        <table>
            <tr><th>No</th><th>Date</th><th>Company</th><th>Contract</th><th>Fault</th><th>Site</th><th>Status</th><th>Action</th></tr>
            {body}
        </table>
    </div>
    """
    if safe_int(form_values.get("id")) > 0:
        html += render_audit_log_card("ops_ticket", safe_int(form_values.get("id")))
    return render_ops_page(request, "Tickets", html)


@router.post("/ui/operations/tickets/save")
def save_ticket(
    request: Request,
    row_id: int = Form(0),
    ticket_no: str = Form(""),
    company_id: int = Form(0),
    contract_id: int = Form(0),
    ticket_date: str = Form(""),
    fault_type_id: int = Form(0),
    site_code: str = Form(""),
    site_name: str = Form(""),
    priority: str = Form("normal"),
    request_channel: str = Form(""),
    complaint_details: str = Form(""),
    status: str = Form("open"),
):
    conn = get_conn()
    values = (safe(ticket_no), safe_int(company_id), safe_int(contract_id), safe(ticket_date), safe_int(fault_type_id), safe(site_code), safe(site_name), safe(priority), safe(request_channel), safe(complaint_details), safe(status))
    if row_id > 0:
        conn.execute("""
            UPDATE ops_tickets
            SET ticket_no = ?, company_id = ?, contract_id = ?, ticket_date = ?, fault_type_id = ?,
                site_code = ?, site_name = ?, priority = ?, request_channel = ?, complaint_details = ?, status = ?
            WHERE id = ?
        """, values + (row_id,))
        log_ops_event(request, "ops_ticket", row_id, "Updated", f"Ticket {safe(ticket_no)}", conn=conn)
        notice = "updated"
    else:
        cur = conn.execute("""
            INSERT INTO ops_tickets (
                ticket_no, company_id, contract_id, ticket_date, fault_type_id, site_code, site_name,
                priority, request_channel, complaint_details, status, created_by
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, values + (actor_name_from_request(request),))
        log_ops_event(request, "ops_ticket", cur.lastrowid, "Created", f"Ticket {safe(ticket_no)}", conn=conn)
        notice = "saved"
    conn.commit()
    conn.close()
    return RedirectResponse(f"/ui/operations/tickets?notice={notice}", status_code=303)


def work_order_options(selected_id=0):
    conn = get_conn()
    rows = conn.execute("""
        SELECT id, work_order_no, site_name, status
        FROM ops_work_orders
        ORDER BY id DESC
    """).fetchall()
    conn.close()
    html = "<option value=''>-- Select Work Order --</option>"
    for row in rows:
        sel = "selected" if str(selected_id or "") == str(row["id"]) else ""
        label = f"{safe(row['work_order_no'])} - {safe(row['site_name'])} ({safe(row['status'])})"
        html += f"<option value='{row['id']}' {sel}>{label}</option>"
    return html


@router.get("/ui/operations/work-orders", response_class=HTMLResponse)
def work_orders_page(request: Request, edit_id: int = 0, notice: str = ""):
    conn = get_conn()
    rows = conn.execute("""
        SELECT
            wo.*,
            co.name AS company_name,
            f.name AS fault_name,
            a.name AS action_name,
            r.name AS region_name,
            COALESCE(te.employee_name, te.name, te.full_name) AS technician_name,
            COALESCE(me.employee_name, me.name, me.full_name) AS manager_name
        FROM ops_work_orders wo
        LEFT JOIN ops_contract_companies co ON co.id = wo.company_id
        LEFT JOIN ops_fault_types f ON f.id = wo.fault_type_id
        LEFT JOIN ops_service_catalog a ON a.id = wo.service_id
        LEFT JOIN ops_regions r ON r.id = wo.region_id
        LEFT JOIN employees te ON te.id = wo.technician_id
        LEFT JOIN employees me ON me.id = wo.manager_id
        ORDER BY wo.id DESC
    """).fetchall()
    edit_row = conn.execute("SELECT * FROM ops_work_orders WHERE id = ? LIMIT 1", (edit_id,)).fetchone() if edit_id else None
    conn.close()
    form_values = dict(edit_row) if edit_row else {
        "id": 0,
        "work_order_no": next_number("ops_work_orders", "work_order_no", "WO"),
        "company_id": "",
        "fault_type_id": "",
        "service_id": "",
        "region_id": "",
        "ticket_id": "",
        "department": "operation",
        "workflow_type": "field_service",
        "action_type": "repair",
        "customer_warehouse_id": "",
        "requested_qty": "0.00",
        "issued_qty": "0.00",
        "completed_qty": "0.00",
        "returned_qty": "0.00",
        "rollout_status": "not_required",
        "rollout_notes": "",
        "actual_actions": "",
        "request_date": "",
        "site_code": "",
        "site_name": "",
        "complaint_details": "",
        "required_materials": "",
        "technician_id": "",
        "manager_id": "",
        "priority": "normal",
        "trip_required": 1,
        "service_price": "0.00",
        "technician_incentive": "0.00",
        "region_allowance": "0.00",
        "status": "new",
        "closure_notes": "",
    }
    body = ""
    for row in rows:
        requested_qty = float(row["requested_qty"] or 0)
        completed_qty = float(row["completed_qty"] or 0)
        remaining_qty = max(requested_qty - completed_qty, 0)
        progress = f"{money(completed_qty)} / {money(requested_qty)}"
        body += f"""
        <tr>
            <td>{safe(row['work_order_no'])}</td>
            <td>{safe(row['request_date'])}</td>
            <td>{safe(row['company_name'])}</td>
            <td>{safe(row['department']).title()}</td>
            <td>{safe(row['workflow_type']).replace('_', ' ').title()}</td>
            <td>{safe(row['fault_name'])}</td>
            <td>{safe(row['action_name'])}</td>
            <td>{safe(row['action_type']).replace('_', ' ').title()}</td>
            <td>{safe(row['region_name'])}</td>
            <td>{safe(row['technician_name'])}</td>
            <td>{progress}<br><span class="section-note">Remaining {money(remaining_qty)}</span></td>
            <td>{status_chip(row['status'])}</td>
            <td>
                <a class="btn blue" href="/ui/operations/work-orders/{row['id']}">Open</a>
                <a class="btn gray" href="/ui/operations/work-orders?edit_id={row['id']}">Edit</a>
            </td>
        </tr>
        """
    if not body:
        body = "<tr><td colspan='13' style='text-align:center;'>No work orders created yet.</td></tr>"
    trip_checked = "checked" if int(form_values.get("trip_required") or 0) == 1 else ""
    html = f"""
    {current_form_notice(notice)}
    <div class="card">
        <div class="toolbar">
            <div><h2 style="margin:0;">Work Orders</h2></div>
            <a class="btn gray" href="/ui/operations">Back to Operations</a>
        </div>
        <form method="post" action="/ui/operations/work-orders/save">
            <input type="hidden" name="row_id" value="{safe(form_values.get('id', 0))}">
            <div class="form-grid">
                <div class="form-group"><label>Work Order No</label><input name="work_order_no" value="{safe(form_values.get('work_order_no'))}" required></div>
                <div class="form-group"><label>Request Date</label><input type="date" name="request_date" value="{safe(form_values.get('request_date'))}" required></div>
                <div class="form-group"><label>Company</label><select name="company_id" required>{company_options(form_values.get('company_id'))}</select></div>
                <div class="form-group"><label>Ticket</label><select name="ticket_id">{ticket_options(form_values.get('ticket_id'))}</select></div>
                <div class="form-group"><label>Department</label><select name="department" required>{simple_options(OPS_DEPARTMENTS, form_values.get('department') or 'operation')}</select></div>
                <div class="form-group"><label>Workflow</label><select name="workflow_type" required>{simple_options(WORKFLOW_TYPES, form_values.get('workflow_type') or 'field_service')}</select></div>
                <div class="form-group"><label>Fault Type</label><select name="fault_type_id">{fault_options(form_values.get('fault_type_id'))}</select></div>
                <div class="form-group"><label>Action Type</label><select name="service_id">{action_options(form_values.get('service_id'))}</select></div>
                <div class="form-group"><label>Execution Action</label><select name="action_type">{simple_options(FIELD_ACTION_TYPES, form_values.get('action_type') or 'repair')}</select></div>
                <div class="form-group"><label>Customer Warehouse</label><select name="customer_warehouse_id">{custody_warehouse_options(form_values.get('customer_warehouse_id'), form_values.get('company_id'), form_values.get('department'))}</select></div>
                <div class="form-group"><label>Region</label><select name="region_id">{region_options(form_values.get('region_id'))}</select></div>
                <div class="form-group"><label>Technician</label><select name="technician_id">{employee_options(form_values.get('technician_id'))}</select></div>
                <div class="form-group"><label>Supervisor</label><select name="manager_id">{employee_options(form_values.get('manager_id'))}</select></div>
                <div class="form-group"><label>Priority</label><select name="priority">{simple_options([('low','Low'),('normal','Normal'),('high','High'),('urgent','Urgent')], form_values.get('priority'))}</select></div>
                <div class="form-group"><label>Status</label><select name="status">{status_options(form_values.get('status'))}</select></div>
                <div class="form-group"><label>Site Code</label><input name="site_code" value="{safe(form_values.get('site_code'))}"></div>
                <div class="form-group"><label>Site Name</label><input name="site_name" value="{safe(form_values.get('site_name'))}"></div>
                <div class="form-group"><label>Billing Price</label><input name="service_price" value="{safe(form_values.get('service_price'))}"></div>
                <div class="form-group"><label>Technician Incentive</label><input name="technician_incentive" value="{safe(form_values.get('technician_incentive'))}"></div>
                <div class="form-group"><label>Region Allowance</label><input name="region_allowance" value="{safe(form_values.get('region_allowance'))}"></div>
                <div class="form-group"><label>Requested Qty</label><input type="number" step="0.01" name="requested_qty" value="{safe(form_values.get('requested_qty'))}"></div>
                <div class="form-group"><label>Completed Qty</label><input type="number" step="0.01" name="completed_qty" value="{safe(form_values.get('completed_qty'))}"></div>
                <div class="form-group"><label>Rollout Status</label><select name="rollout_status">{simple_options(ROLLOUT_STATUSES, form_values.get('rollout_status') or 'not_required')}</select></div>
                <div class="form-group"><label>Trip Required</label><div style="padding-top:10px;"><label><input type="checkbox" name="trip_required" value="1" {trip_checked}> Needs Trip Ticket</label></div></div>
                <div class="form-group" style="grid-column: span 2;"><label>Complaint Details</label><input name="complaint_details" value="{safe(form_values.get('complaint_details'))}"></div>
                <div class="form-group" style="grid-column: span 2;"><label>Required Materials / Notes</label><input name="required_materials" value="{safe(form_values.get('required_materials'))}"></div>
                <div class="form-group" style="grid-column: span 2;"><label>Actual Actions</label><input name="actual_actions" value="{safe(form_values.get('actual_actions'))}"></div>
                <div class="form-group" style="grid-column: span 2;"><label>Rollout Notes</label><input name="rollout_notes" value="{safe(form_values.get('rollout_notes'))}"></div>
                <div class="form-group" style="grid-column: span 2;"><label>Closure Notes</label><input name="closure_notes" value="{safe(form_values.get('closure_notes'))}"></div>
            </div>
            <div class="form-actions">
                <button class="btn green" type="submit">{"Update Work Order" if safe_int(form_values.get('id')) > 0 else "Save Work Order"}</button>
                <a class="btn gray" href="/ui/operations/work-orders">Clear</a>
            </div>
        </form>
    </div>
    <div class="card">
        <h3 class="sub-title">Work Order List</h3>
        <table>
            <tr><th>No</th><th>Date</th><th>Company</th><th>Department</th><th>Workflow</th><th>Fault</th><th>Action</th><th>Execution</th><th>Region</th><th>Technician</th><th>Progress</th><th>Status</th><th>Action</th></tr>
            {body}
        </table>
    </div>
    """
    if safe_int(form_values.get("id")) > 0:
        html += render_audit_log_card("ops_work_order", safe_int(form_values.get("id")))
    return render_ops_page(request, "Work Orders", html)


@router.post("/ui/operations/work-orders/save")
def save_work_order(
    request: Request,
    row_id: int = Form(0),
    work_order_no: str = Form(""),
    company_id: int = Form(0),
    fault_type_id: int = Form(0),
    service_id: int = Form(0),
    region_id: int = Form(0),
    ticket_id: int = Form(0),
    department: str = Form("operation"),
    workflow_type: str = Form("field_service"),
    action_type: str = Form("repair"),
    customer_warehouse_id: int = Form(0),
    requested_qty: str = Form("0"),
    completed_qty: str = Form("0"),
    rollout_status: str = Form("not_required"),
    rollout_notes: str = Form(""),
    actual_actions: str = Form(""),
    request_date: str = Form(""),
    site_code: str = Form(""),
    site_name: str = Form(""),
    complaint_details: str = Form(""),
    required_materials: str = Form(""),
    technician_id: int = Form(0),
    manager_id: int = Form(0),
    priority: str = Form("normal"),
    trip_required: str = Form(""),
    service_price: str = Form("0"),
    technician_incentive: str = Form("0"),
    region_allowance: str = Form("0"),
    status: str = Form("new"),
    closure_notes: str = Form(""),
):
    conn = get_conn()
    trip_flag = 1 if safe(trip_required) == "1" else 0
    values = (
        safe(work_order_no),
        safe_int(company_id),
        safe_int(fault_type_id),
        safe_int(service_id),
        safe_int(region_id),
        safe_int(ticket_id),
        safe(department) or "operation",
        safe(workflow_type) or "field_service",
        safe(action_type),
        safe_int(customer_warehouse_id),
        float(q2(requested_qty)),
        float(q2(completed_qty)),
        safe(rollout_status) or "not_required",
        safe(rollout_notes),
        safe(actual_actions),
        safe(request_date),
        safe(site_code),
        safe(site_name),
        safe(complaint_details),
        safe(required_materials),
        safe_int(technician_id),
        safe_int(manager_id),
        safe(priority),
        trip_flag,
        float(q2(service_price)),
        float(q2(technician_incentive)),
        float(q2(region_allowance)),
        safe(status),
        safe(closure_notes),
    )
    if row_id > 0:
        conn.execute("""
            UPDATE ops_work_orders
            SET work_order_no = ?, company_id = ?, fault_type_id = ?, service_id = ?, region_id = ?,
                ticket_id = ?, department = ?, workflow_type = ?, action_type = ?, customer_warehouse_id = ?,
                requested_qty = ?, completed_qty = ?, rollout_status = ?, rollout_notes = ?, actual_actions = ?,
                request_date = ?, site_code = ?, site_name = ?, complaint_details = ?, required_materials = ?,
                technician_id = ?, manager_id = ?, priority = ?, trip_required = ?, service_price = ?,
                technician_incentive = ?, region_allowance = ?, status = ?, closure_notes = ?, updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
        """, values + (row_id,))
        log_ops_event(request, "ops_work_order", row_id, "Updated", f"Work order {safe(work_order_no)}", conn=conn)
        notice = "updated"
    else:
        cur = conn.execute("""
            INSERT INTO ops_work_orders (
                work_order_no, company_id, fault_type_id, service_id, region_id,
                ticket_id, department, workflow_type, action_type, customer_warehouse_id,
                requested_qty, completed_qty, rollout_status, rollout_notes, actual_actions,
                request_date,
                site_code, site_name, complaint_details, required_materials, technician_id, manager_id,
                priority, trip_required, service_price, technician_incentive, region_allowance,
                status, closure_notes, created_by
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, values + (actor_name_from_request(request),))
        log_ops_event(request, "ops_work_order", cur.lastrowid, "Created", f"Work order {safe(work_order_no)}", conn=conn)
        notice = "saved"
    conn.commit()
    conn.close()
    return RedirectResponse(f"/ui/operations/work-orders?notice={notice}", status_code=303)


def work_order_movement_link(row, label, movement_type, from_status, to_status, qty=None, notes=""):
    params = {
        "company_id": row["company_id"],
        "department": safe(row["department"]) or "operation",
        "warehouse_id": row["customer_warehouse_id"],
        "ticket_id": row["ticket_id"],
        "work_order_id": row["id"],
        "movement_type": movement_type,
        "module_code": "",
        "module_name": "",
        "from_status": from_status,
        "to_status": to_status,
        "qty": qty if qty is not None else "",
        "uom": "PCS",
        "notes": notes or f"{label} for {safe(row['work_order_no'])}",
    }
    query = "&".join(f"{key}={quote(safe(value))}" for key, value in params.items() if safe(value) != "")
    return f"/ui/operations/customer-custody-transactions?{query}"


@router.get("/ui/operations/work-orders/{work_order_id}", response_class=HTMLResponse)
def work_order_detail_page(request: Request, work_order_id: int, notice: str = ""):
    conn = get_conn()
    row = conn.execute("""
        SELECT
            wo.*,
            co.code AS company_code,
            co.name AS company_name,
            f.name AS fault_name,
            s.name AS service_name,
            r.name AS region_name,
            cw.code AS warehouse_code,
            cw.name AS warehouse_name,
            tk.ticket_no,
            COALESCE(te.employee_name, te.name, te.full_name) AS technician_name,
            COALESCE(me.employee_name, me.name, me.full_name) AS manager_name
        FROM ops_work_orders wo
        LEFT JOIN ops_contract_companies co ON co.id = wo.company_id
        LEFT JOIN ops_fault_types f ON f.id = wo.fault_type_id
        LEFT JOIN ops_service_catalog s ON s.id = wo.service_id
        LEFT JOIN ops_regions r ON r.id = wo.region_id
        LEFT JOIN ops_customer_custody_warehouses cw ON cw.id = wo.customer_warehouse_id
        LEFT JOIN ops_tickets tk ON tk.id = wo.ticket_id
        LEFT JOIN employees te ON te.id = wo.technician_id
        LEFT JOIN employees me ON me.id = wo.manager_id
        WHERE wo.id = ?
        LIMIT 1
    """, (work_order_id,)).fetchone()
    if not row:
        conn.close()
        return render_ops_page(request, "Work Order", "<div class='card'>Work order not found.</div>")

    movement_rows = conn.execute("""
        SELECT *
        FROM ops_customer_custody_transactions
        WHERE work_order_id = ?
        ORDER BY transaction_date, id
    """, (work_order_id,)).fetchall()
    conn.close()

    requested_qty = float(row["requested_qty"] or 0)
    issued_qty = float(row["issued_qty"] or 0)
    completed_qty = float(row["completed_qty"] or 0)
    returned_qty = float(row["returned_qty"] or 0)
    remaining_qty = max(requested_qty - completed_qty, 0)
    warehouse_label = ""
    if safe(row["warehouse_code"]) or safe(row["warehouse_name"]):
        warehouse_label = f"{safe(row['warehouse_code'])} - {safe(row['warehouse_name'])}"

    quick_actions = [
        ("Receive Faulty", "receipt", "", "faulty", requested_qty or "", f"Receive faulty modules for {safe(row['work_order_no'])}"),
        ("Issue To Repair", "issue_to_repair", "faulty", "under_repair", remaining_qty or "", f"Issue modules to repair for {safe(row['work_order_no'])}"),
        ("Return From Repair", "return_from_repair", "under_repair", "working", remaining_qty or "", f"Return repaired modules for {safe(row['work_order_no'])}"),
        ("Issue To Site", "site_issue", "working", "installed", remaining_qty or "", f"Issue working modules to site for {safe(row['work_order_no'])}"),
        ("Return From Site", "site_return", "installed", "working", "", f"Return unused modules from site for {safe(row['work_order_no'])}"),
        ("Swap Removed Faulty", "swap_removed", "installed", "faulty", "", f"Removed faulty module from swap for {safe(row['work_order_no'])}"),
    ]
    action_buttons = ""
    if safe_int(row["customer_warehouse_id"]) > 0:
        for label, movement_type, from_status, to_status, qty, notes in quick_actions:
            href = work_order_movement_link(row, label, movement_type, from_status, to_status, qty, notes)
            action_buttons += f"<a class='btn blue' href='{href}'>{label}</a>"
    else:
        action_buttons = "<span class='section-note'>Select customer warehouse on the work order to enable custody movements.</span>"

    movement_body = ""
    for mov in movement_rows:
        movement_body += f"""
        <tr>
            <td>{safe(mov['transaction_no'])}</td>
            <td>{safe(mov['transaction_date'])}</td>
            <td>{safe(mov['movement_type']).replace('_', ' ').title()}</td>
            <td>{safe(mov['module_code'])} {safe(mov['module_name'])}</td>
            <td>{safe(mov['from_status']).replace('_', ' ').title()} -> {safe(mov['to_status']).replace('_', ' ').title()}</td>
            <td>{money(mov['qty'])}</td>
            <td>{safe(mov['notes'])}</td>
        </tr>
        """
    if not movement_body:
        movement_body = "<tr><td colspan='7' style='text-align:center;'>No custody movements linked yet.</td></tr>"

    rollout_form = f"""
    <div class="card">
        <div class="toolbar">
            <h3 class="sub-title">Rollout Approval</h3>
            {rollout_status_chip(row['rollout_status'])}
        </div>
        <form method="post" action="/ui/operations/work-orders/{row['id']}/rollout-confirm">
            <div class="form-grid">
                <div class="form-group" style="grid-column: span 2;">
                    <label>Rollout Notes</label>
                    <input name="rollout_notes" value="{safe(row['rollout_notes'])}" required>
                </div>
            </div>
            <div class="form-actions">
                <button class="btn green" type="submit">Rollout Confirm</button>
            </div>
        </form>
    </div>
    """

    html = f"""
    {current_form_notice(notice)}
    <div class="card">
        <div class="toolbar">
            <div>
                <h2 style="margin:0;">Work Order {safe(row['work_order_no'])}</h2>
                <div class="section-note">{safe(row['company_code'])} - {safe(row['company_name'])}</div>
            </div>
            <div class="form-actions" style="margin:0;">
                <a class="btn gray" href="/ui/operations/work-orders">Work Orders</a>
                <a class="btn blue" href="/ui/operations/work-orders?edit_id={row['id']}">Edit</a>
            </div>
        </div>
        <div class="info-grid">
            <div><b>Status:</b> {status_chip(row['status'])}</div>
            <div><b>Workflow:</b> {safe(row['workflow_type']).replace('_', ' ').title()}</div>
            <div><b>Department:</b> {safe(row['department']).title()}</div>
            <div><b>Execution:</b> {safe(row['action_type']).replace('_', ' ').title()}</div>
            <div><b>Ticket:</b> {safe(row['ticket_no'])}</div>
            <div><b>Fault:</b> {safe(row['fault_name'])}</div>
            <div><b>Action:</b> {safe(row['service_name'])}</div>
            <div><b>Region:</b> {safe(row['region_name'])}</div>
            <div><b>Site:</b> {safe(row['site_code'])} {safe(row['site_name'])}</div>
            <div><b>Technician:</b> {safe(row['technician_name'])}</div>
            <div><b>Supervisor:</b> {safe(row['manager_name'])}</div>
            <div><b>Customer Warehouse:</b> {warehouse_label}</div>
        </div>
    </div>
    <div class="stats-grid">
        <div class="stat-card"><div class="stat-title">Requested</div><div class="stat-value">{money(requested_qty)}</div></div>
        <div class="stat-card"><div class="stat-title">Issued</div><div class="stat-value">{money(issued_qty)}</div></div>
        <div class="stat-card"><div class="stat-title">Completed</div><div class="stat-value">{money(completed_qty)}</div></div>
        <div class="stat-card"><div class="stat-title">Remaining</div><div class="stat-value">{money(remaining_qty)}</div></div>
        <div class="stat-card"><div class="stat-title">Returned</div><div class="stat-value">{money(returned_qty)}</div></div>
    </div>
    <div class="card">
        <h3 class="sub-title">Custody Movements</h3>
        <div class="form-actions">{action_buttons}</div>
        <table>
            <tr><th>No</th><th>Date</th><th>Movement</th><th>Module</th><th>Status Flow</th><th>Qty</th><th>Notes</th></tr>
            {movement_body}
        </table>
    </div>
    {rollout_form}
    <div class="card">
        <h3 class="sub-title">Engineer Closure</h3>
        <form method="post" action="/ui/operations/work-orders/{row['id']}/complete">
            <div class="form-grid">
                <div class="form-group"><label>Completed Qty</label><input type="number" step="0.01" name="completed_qty" value="{money(completed_qty)}"></div>
                <div class="form-group"><label>Actual Actions</label><input name="actual_actions" value="{safe(row['actual_actions'])}"></div>
                <div class="form-group" style="grid-column: span 2;"><label>Closure Notes</label><input name="closure_notes" value="{safe(row['closure_notes'])}"></div>
            </div>
            <div class="form-actions">
                <button class="btn green" type="submit">Save Progress / Close</button>
            </div>
        </form>
    </div>
    {render_audit_log_card("ops_work_order", row['id'])}
    """
    return render_ops_page(request, f"Work Order {safe(row['work_order_no'])}", html)


@router.post("/ui/operations/work-orders/{work_order_id}/rollout-confirm")
def confirm_work_order_rollout(request: Request, work_order_id: int, rollout_notes: str = Form("")):
    if not safe(rollout_notes):
        return RedirectResponse(f"/ui/operations/work-orders/{work_order_id}?notice={quote('Rollout notes are required')}", status_code=303)
    conn = get_conn()
    row = conn.execute("SELECT work_order_no FROM ops_work_orders WHERE id = ? LIMIT 1", (work_order_id,)).fetchone()
    if not row:
        conn.close()
        return RedirectResponse("/ui/operations/work-orders?notice=not_found", status_code=303)
    actor = actor_name_from_request(request)
    conn.execute("""
        UPDATE ops_work_orders
        SET rollout_status = 'approved',
            rollout_notes = ?,
            rollout_by = ?,
            rollout_at = CURRENT_TIMESTAMP,
            status = CASE WHEN status IN ('new', 'assigned') THEN 'approved' ELSE status END,
            updated_at = CURRENT_TIMESTAMP
        WHERE id = ?
    """, (safe(rollout_notes), actor, work_order_id))
    log_ops_event(request, "ops_work_order", work_order_id, "Rollout Approved", f"Rollout approved for {safe(row['work_order_no'])}: {safe(rollout_notes)}", conn=conn)
    conn.commit()
    conn.close()
    return RedirectResponse(f"/ui/operations/work-orders/{work_order_id}?notice=rollout_approved", status_code=303)


@router.post("/ui/operations/work-orders/{work_order_id}/complete")
def complete_work_order(request: Request, work_order_id: int, completed_qty: str = Form("0"), actual_actions: str = Form(""), closure_notes: str = Form("")):
    conn = get_conn()
    row = conn.execute("SELECT * FROM ops_work_orders WHERE id = ? LIMIT 1", (work_order_id,)).fetchone()
    if not row:
        conn.close()
        return RedirectResponse("/ui/operations/work-orders?notice=not_found", status_code=303)
    requested_qty = float(row["requested_qty"] or 0)
    final_completed_qty = float(q2(completed_qty))
    next_status = "closed" if requested_qty > 0 and final_completed_qty >= requested_qty else "in_progress"
    conn.execute("""
        UPDATE ops_work_orders
        SET completed_qty = ?,
            actual_actions = ?,
            closure_notes = ?,
            status = ?,
            completed_at = CASE WHEN ? = 'closed' THEN CURRENT_TIMESTAMP ELSE completed_at END,
            updated_at = CURRENT_TIMESTAMP
        WHERE id = ?
    """, (final_completed_qty, safe(actual_actions), safe(closure_notes), next_status, next_status, work_order_id))
    log_ops_event(
        request,
        "ops_work_order",
        work_order_id,
        "Closure Updated",
        f"Completed {money(final_completed_qty)} of {money(requested_qty)}. {safe(actual_actions)} {safe(closure_notes)}",
        conn=conn,
    )
    conn.commit()
    conn.close()
    return RedirectResponse(f"/ui/operations/work-orders/{work_order_id}?notice=progress_saved", status_code=303)


def region_options(selected_id=0):
    conn = get_conn()
    rows = conn.execute("""
        SELECT id, code, name, zone_level
        FROM ops_regions
        WHERE COALESCE(is_active, 1) = 1
        ORDER BY name
    """).fetchall()
    conn.close()
    html = "<option value=''>-- Select Region --</option>"
    for row in rows:
        sel = "selected" if str(selected_id or "") == str(row["id"]) else ""
        label = f"{safe(row['code'])} - {safe(row['name'])} ({safe(row['zone_level'])})"
        html += f"<option value='{row['id']}' {sel}>{label}</option>"
    return html


@router.get("/ui/operations/technician-reports", response_class=HTMLResponse)
def technician_reports_page(request: Request, edit_id: int = 0, notice: str = ""):
    conn = get_conn()
    rows = conn.execute("""
        SELECT tr.*, wo.work_order_no, wo.site_name
        FROM ops_technician_reports tr
        LEFT JOIN ops_work_orders wo ON wo.id = tr.work_order_id
        ORDER BY tr.id DESC
    """).fetchall()
    edit_row = conn.execute("SELECT * FROM ops_technician_reports WHERE id = ? LIMIT 1", (edit_id,)).fetchone() if edit_id else None
    conn.close()
    form_values = dict(edit_row) if edit_row else {
        "id": 0,
        "work_order_id": "",
        "report_date": "",
        "arrival_time": "",
        "completion_time": "",
        "issue_found": "",
        "action_taken": "",
        "materials_used": "",
        "technician_notes": "",
        "customer_notes": "",
        "report_status": "submitted",
        "service_price": "0.00",
        "technician_incentive": "0.00",
        "region_allowance": "0.00",
        "review_notes": "",
        "status": "submitted",
    }
    body = ""
    for row in rows:
        body += f"""
        <tr>
            <td>{safe(row['report_date'])}</td>
            <td>{safe(row['work_order_no'])}</td>
            <td>{safe(row['site_name'])}</td>
            <td>{safe(row['issue_found'])}</td>
            <td>{safe(row['action_taken'])}</td>
            <td>{status_chip(row['status'])}</td>
            <td><a class="btn blue" href="/ui/operations/technician-reports?edit_id={row['id']}">Edit</a></td>
        </tr>
        """
    if not body:
        body = "<tr><td colspan='7' style='text-align:center;'>No technician reports submitted yet.</td></tr>"
    html = f"""
    {current_form_notice(notice)}
    <div class="card">
        <div class="toolbar">
            <div><h2 style="margin:0;">Technician Reports</h2></div>
            <a class="btn gray" href="/ui/operations">Back to Operations</a>
        </div>
        <form method="post" action="/ui/operations/technician-reports/save">
            <input type="hidden" name="row_id" value="{safe(form_values.get('id', 0))}">
            <div class="form-grid">
                <div class="form-group"><label>Work Order</label><select name="work_order_id" required>{work_order_options(form_values.get('work_order_id'))}</select></div>
                <div class="form-group"><label>Report Date</label><input type="date" name="report_date" value="{safe(form_values.get('report_date'))}" required></div>
                <div class="form-group"><label>Arrival Time</label><input type="time" name="arrival_time" value="{safe(form_values.get('arrival_time'))}"></div>
                <div class="form-group"><label>Completion Time</label><input type="time" name="completion_time" value="{safe(form_values.get('completion_time'))}"></div>
                <div class="form-group"><label>Report Status</label><select name="report_status">{simple_options([('submitted','Submitted'),('reviewed','Reviewed'),('approved','Approved'),('rejected','Rejected')], form_values.get('report_status'))}</select></div>
                <div class="form-group"><label>Status</label><select name="status">{simple_options([('submitted','Submitted'),('approved','Approved'),('rejected','Rejected')], form_values.get('status'))}</select></div>
                <div class="form-group" style="grid-column: span 2;"><label>Issue Found</label><input name="issue_found" value="{safe(form_values.get('issue_found'))}"></div>
                <div class="form-group" style="grid-column: span 2;"><label>Action Taken</label><input name="action_taken" value="{safe(form_values.get('action_taken'))}"></div>
                <div class="form-group" style="grid-column: span 2;"><label>Materials Used</label><input name="materials_used" value="{safe(form_values.get('materials_used'))}"></div>
                <div class="form-group"><label>Service Price</label><input name="service_price" value="{safe(form_values.get('service_price'))}"></div>
                <div class="form-group"><label>Technician Incentive</label><input name="technician_incentive" value="{safe(form_values.get('technician_incentive'))}"></div>
                <div class="form-group"><label>Region Allowance</label><input name="region_allowance" value="{safe(form_values.get('region_allowance'))}"></div>
                <div class="form-group" style="grid-column: span 2;"><label>Technician Notes</label><input name="technician_notes" value="{safe(form_values.get('technician_notes'))}"></div>
                <div class="form-group" style="grid-column: span 2;"><label>Customer Notes</label><input name="customer_notes" value="{safe(form_values.get('customer_notes'))}"></div>
                <div class="form-group" style="grid-column: span 2;"><label>Review Notes</label><input name="review_notes" value="{safe(form_values.get('review_notes'))}"></div>
            </div>
            <div class="form-actions">
                <button class="btn green" type="submit">{"Update Report" if safe_int(form_values.get('id')) > 0 else "Save Report"}</button>
                <a class="btn gray" href="/ui/operations/technician-reports">Clear</a>
            </div>
        </form>
    </div>
    <div class="card">
        <h3 class="sub-title">Report List</h3>
        <table>
            <tr><th>Date</th><th>Work Order</th><th>Site</th><th>Issue</th><th>Action Taken</th><th>Status</th><th>Action</th></tr>
            {body}
        </table>
    </div>
    """
    if safe_int(form_values.get("id")) > 0:
        html += render_audit_log_card("ops_technician_report", safe_int(form_values.get("id")))
    return render_ops_page(request, "Technician Reports", html)


@router.post("/ui/operations/technician-reports/save")
def save_technician_report(
    request: Request,
    row_id: int = Form(0),
    work_order_id: int = Form(0),
    report_date: str = Form(""),
    arrival_time: str = Form(""),
    completion_time: str = Form(""),
    issue_found: str = Form(""),
    action_taken: str = Form(""),
    materials_used: str = Form(""),
    technician_notes: str = Form(""),
    customer_notes: str = Form(""),
    report_status: str = Form("submitted"),
    service_price: str = Form("0"),
    technician_incentive: str = Form("0"),
    region_allowance: str = Form("0"),
    review_notes: str = Form(""),
    status: str = Form("submitted"),
):
    conn = get_conn()
    values = (
        safe_int(work_order_id),
        safe(report_date),
        safe(arrival_time),
        safe(completion_time),
        safe(issue_found),
        safe(action_taken),
        safe(materials_used),
        safe(technician_notes),
        safe(customer_notes),
        safe(report_status),
        float(q2(service_price)),
        float(q2(technician_incentive)),
        float(q2(region_allowance)),
        safe(review_notes),
        safe(status),
    )
    if row_id > 0:
        conn.execute("""
            UPDATE ops_technician_reports
            SET work_order_id = ?, report_date = ?, arrival_time = ?, completion_time = ?,
                issue_found = ?, action_taken = ?, materials_used = ?, technician_notes = ?,
                customer_notes = ?, report_status = ?, service_price = ?, technician_incentive = ?,
                region_allowance = ?, review_notes = ?, status = ?
            WHERE id = ?
        """, values + (row_id,))
        log_ops_event(request, "ops_technician_report", row_id, "Updated", f"Report for work order id {safe_int(work_order_id)}", conn=conn)
        notice = "updated"
    else:
        cur = conn.execute("""
            INSERT INTO ops_technician_reports (
                work_order_id, report_date, arrival_time, completion_time, issue_found, action_taken,
                materials_used, technician_notes, customer_notes, report_status, service_price,
                technician_incentive, region_allowance, review_notes, status, submitted_by, submitted_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
        """, values + (actor_name_from_request(request),))
        log_ops_event(request, "ops_technician_report", cur.lastrowid, "Created", f"Report for work order id {safe_int(work_order_id)}", conn=conn)
        notice = "saved"
    conn.commit()
    conn.close()
    return RedirectResponse(f"/ui/operations/technician-reports?notice={notice}", status_code=303)


@router.get("/ui/operations/pricing-versions", response_class=HTMLResponse)
def pricing_versions_page(request: Request, edit_action_id: int = 0, edit_vehicle_id: int = 0, notice: str = ""):
    conn = get_conn()
    action_rows = conn.execute("""
        SELECT pv.*, c.contract_no, a.code AS action_code, a.name AS action_name
        FROM ops_action_price_versions pv
        LEFT JOIN ops_contracts c ON c.id = pv.contract_id
        LEFT JOIN ops_service_catalog a ON a.id = pv.action_id
        ORDER BY pv.id DESC
    """).fetchall()
    vehicle_rows = conn.execute("""
        SELECT pv.*, c.contract_no, vr.code AS rate_code, vr.vehicle_type, vr.slab_name
        FROM ops_vehicle_price_versions pv
        LEFT JOIN ops_contracts c ON c.id = pv.contract_id
        LEFT JOIN ops_vehicle_rates vr ON vr.id = pv.vehicle_rate_id
        ORDER BY pv.id DESC
    """).fetchall()
    action_edit = conn.execute("SELECT * FROM ops_action_price_versions WHERE id = ?", (edit_action_id,)).fetchone() if edit_action_id else None
    vehicle_edit = conn.execute("SELECT * FROM ops_vehicle_price_versions WHERE id = ?", (edit_vehicle_id,)).fetchone() if edit_vehicle_id else None
    conn.close()
    action_form = dict(action_edit) if action_edit else {"id": 0, "contract_id": "", "action_id": "", "version_name": "", "effective_from": "", "effective_to": "", "fuel_reference": "", "action_price": "0.00", "technician_incentive": "0.00", "region_allowance": "0.00", "is_active": 1, "notes": ""}
    vehicle_form = dict(vehicle_edit) if vehicle_edit else {"id": 0, "contract_id": "", "vehicle_rate_id": "", "version_name": "", "effective_from": "", "effective_to": "", "fuel_reference": "", "ticket_open_price": "0.00", "second_slab_to_km": "300.00", "km_rate_101_300": "0.00", "km_rate_over_300": "0.00", "waiting_hour_rate": "0.00", "is_active": 1, "notes": ""}
    action_body = ""
    for row in action_rows:
        action_body += f"""
        <tr>
            <td>{safe(row['version_name'])}</td>
            <td>{safe(row['contract_no'])}</td>
            <td>{safe(row['action_code'])} - {safe(row['action_name'])}</td>
            <td>{safe(row['effective_from'])}</td>
            <td>{safe(row['effective_to'])}</td>
            <td>{money(row['action_price'])}</td>
            <td>{money(row['technician_incentive'])}</td>
            <td>{money(row['region_allowance'])}</td>
            <td>{"Active" if int(row['is_active'] or 0) == 1 else "Inactive"}</td>
            <td><a class="btn blue" href="/ui/operations/pricing-versions?edit_action_id={row['id']}">Edit</a></td>
        </tr>
        """
    if not action_body:
        action_body = "<tr><td colspan='10' style='text-align:center;'>No action pricing versions yet.</td></tr>"
    vehicle_body = ""
    for row in vehicle_rows:
        vehicle_body += f"""
        <tr>
            <td>{safe(row['version_name'])}</td>
            <td>{safe(row['contract_no'])}</td>
            <td>{safe(row['rate_code'])} - {safe(row['vehicle_type'])} / {safe(row['slab_name'])}</td>
            <td>{safe(row['effective_from'])}</td>
            <td>{safe(row['effective_to'])}</td>
            <td>{money(row['ticket_open_price'])}</td>
            <td>{money(row['km_rate_101_300'])}</td>
            <td>{money(row['km_rate_over_300'])}</td>
            <td>{money(row['waiting_hour_rate'])}</td>
            <td>{"Active" if int(row['is_active'] or 0) == 1 else "Inactive"}</td>
            <td><a class="btn blue" href="/ui/operations/pricing-versions?edit_vehicle_id={row['id']}">Edit</a></td>
        </tr>
        """
    if not vehicle_body:
        vehicle_body = "<tr><td colspan='11' style='text-align:center;'>No vehicle pricing versions yet.</td></tr>"
    html = f"""
    {current_form_notice(notice)}
    <div class="card">
        <div class="toolbar">
            <div><h2 style="margin:0;">Pricing Versions</h2></div>
            <a class="btn gray" href="/ui/operations">Back to Operations</a>
        </div>
        <form method="post" action="/ui/operations/pricing-versions/action/save">
            <input type="hidden" name="row_id" value="{safe(action_form.get('id', 0))}">
            <h3 class="sub-title">Action Price Version</h3>
            <div class="form-grid">
                <div class="form-group"><label>Contract</label><select name="contract_id" required>{contract_options(action_form.get('contract_id'))}</select></div>
                <div class="form-group"><label>Action</label><select name="action_id" required>{action_options(action_form.get('action_id'))}</select></div>
                <div class="form-group"><label>Version Name</label><input name="version_name" value="{safe(action_form.get('version_name'))}" required></div>
                <div class="form-group"><label>Effective From</label><input type="date" name="effective_from" value="{safe(action_form.get('effective_from'))}" required></div>
                <div class="form-group"><label>Effective To</label><input type="date" name="effective_to" value="{safe(action_form.get('effective_to'))}"></div>
                <div class="form-group"><label>Fuel Reference</label><input name="fuel_reference" value="{safe(action_form.get('fuel_reference'))}"></div>
                <div class="form-group"><label>Customer Billing Price</label><input name="action_price" value="{safe(action_form.get('action_price'))}"></div>
                <div class="form-group"><label>Technician Incentive</label><input name="technician_incentive" value="{safe(action_form.get('technician_incentive'))}"></div>
                <div class="form-group"><label>Region Allowance</label><input name="region_allowance" value="{safe(action_form.get('region_allowance'))}"></div>
                <div class="form-group"><label>Active</label><div style="padding-top:10px;"><label><input type="checkbox" name="is_active" value="1" {active_checkbox(int(action_form.get('is_active', 1) or 0) == 1)}> Active Version</label></div></div>
                <div class="form-group" style="grid-column: span 2;"><label>Notes</label><input name="notes" value="{safe(action_form.get('notes'))}"></div>
            </div>
            <div class="form-actions"><button class="btn green" type="submit">Save Action Price</button></div>
        </form>
    </div>
    <div class="card">
        <form method="post" action="/ui/operations/pricing-versions/vehicle/save">
            <input type="hidden" name="row_id" value="{safe(vehicle_form.get('id', 0))}">
            <h3 class="sub-title">Vehicle Price Version</h3>
            <div class="form-grid">
                <div class="form-group"><label>Contract</label><select name="contract_id" required>{contract_options(vehicle_form.get('contract_id'))}</select></div>
                <div class="form-group"><label>Vehicle Rate</label><select name="vehicle_rate_id" required>{vehicle_rate_options(vehicle_form.get('vehicle_rate_id'))}</select></div>
                <div class="form-group"><label>Version Name</label><input name="version_name" value="{safe(vehicle_form.get('version_name'))}" required></div>
                <div class="form-group"><label>Effective From</label><input type="date" name="effective_from" value="{safe(vehicle_form.get('effective_from'))}" required></div>
                <div class="form-group"><label>Effective To</label><input type="date" name="effective_to" value="{safe(vehicle_form.get('effective_to'))}"></div>
                <div class="form-group"><label>Fuel Reference</label><input name="fuel_reference" value="{safe(vehicle_form.get('fuel_reference'))}"></div>
                <div class="form-group"><label>Ticket Open Price</label><input name="ticket_open_price" value="{safe(vehicle_form.get('ticket_open_price'))}"></div>
                <div class="form-group"><label>Second Slab To KM</label><input name="second_slab_to_km" value="{safe(vehicle_form.get('second_slab_to_km'))}"></div>
                <div class="form-group"><label>KM Rate 101-300</label><input name="km_rate_101_300" value="{safe(vehicle_form.get('km_rate_101_300'))}"></div>
                <div class="form-group"><label>KM Rate Over 300</label><input name="km_rate_over_300" value="{safe(vehicle_form.get('km_rate_over_300'))}"></div>
                <div class="form-group"><label>Waiting Hour Rate</label><input name="waiting_hour_rate" value="{safe(vehicle_form.get('waiting_hour_rate'))}"></div>
                <div class="form-group"><label>Active</label><div style="padding-top:10px;"><label><input type="checkbox" name="is_active" value="1" {active_checkbox(int(vehicle_form.get('is_active', 1) or 0) == 1)}> Active Version</label></div></div>
                <div class="form-group" style="grid-column: span 2;"><label>Notes</label><input name="notes" value="{safe(vehicle_form.get('notes'))}"></div>
            </div>
            <div class="form-actions"><button class="btn green" type="submit">Save Vehicle Price</button></div>
        </form>
    </div>
    <div class="card">
        <h3 class="sub-title">Action Price Versions</h3>
        <table><tr><th>Version</th><th>Contract</th><th>Action</th><th>From</th><th>To</th><th>Price</th><th>Incentive</th><th>Allowance</th><th>Status</th><th>Action</th></tr>{action_body}</table>
    </div>
    <div class="card">
        <h3 class="sub-title">Vehicle Price Versions</h3>
        <table><tr><th>Version</th><th>Contract</th><th>Rate</th><th>From</th><th>To</th><th>Open</th><th>101-300</th><th>Over 300</th><th>Waiting</th><th>Status</th><th>Action</th></tr>{vehicle_body}</table>
    </div>
    """
    return render_ops_page(request, "Pricing Versions", html)


@router.post("/ui/operations/pricing-versions/action/save")
def save_action_price_version(
    request: Request,
    row_id: int = Form(0),
    contract_id: int = Form(0),
    action_id: int = Form(0),
    version_name: str = Form(""),
    effective_from: str = Form(""),
    effective_to: str = Form(""),
    fuel_reference: str = Form(""),
    action_price: str = Form("0"),
    technician_incentive: str = Form("0"),
    region_allowance: str = Form("0"),
    is_active: str = Form(""),
    notes: str = Form(""),
):
    conn = get_conn()
    active_flag = 1 if safe(is_active) == "1" else 0
    values = (safe_int(contract_id), safe_int(action_id), safe(version_name), safe(effective_from), safe(effective_to), safe(fuel_reference), float(q2(action_price)), float(q2(technician_incentive)), float(q2(region_allowance)), active_flag, safe(notes))
    if row_id > 0:
        conn.execute("""
            UPDATE ops_action_price_versions
            SET contract_id = ?, action_id = ?, version_name = ?, effective_from = ?, effective_to = ?,
                fuel_reference = ?, action_price = ?, technician_incentive = ?, region_allowance = ?, is_active = ?, notes = ?
            WHERE id = ?
        """, values + (row_id,))
        log_ops_event(request, "ops_action_price_version", row_id, "Updated", f"Action price {safe(version_name)}", conn=conn)
        notice = "updated"
    else:
        cur = conn.execute("""
            INSERT INTO ops_action_price_versions (
                contract_id, action_id, version_name, effective_from, effective_to, fuel_reference,
                action_price, technician_incentive, region_allowance, is_active, notes
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, values)
        log_ops_event(request, "ops_action_price_version", cur.lastrowid, "Created", f"Action price {safe(version_name)}", conn=conn)
        notice = "saved"
    conn.commit()
    conn.close()
    return RedirectResponse(f"/ui/operations/pricing-versions?notice={notice}", status_code=303)


@router.post("/ui/operations/pricing-versions/vehicle/save")
def save_vehicle_price_version(
    request: Request,
    row_id: int = Form(0),
    contract_id: int = Form(0),
    vehicle_rate_id: int = Form(0),
    version_name: str = Form(""),
    effective_from: str = Form(""),
    effective_to: str = Form(""),
    fuel_reference: str = Form(""),
    ticket_open_price: str = Form("0"),
    second_slab_to_km: str = Form("300"),
    km_rate_101_300: str = Form("0"),
    km_rate_over_300: str = Form("0"),
    waiting_hour_rate: str = Form("0"),
    is_active: str = Form(""),
    notes: str = Form(""),
):
    conn = get_conn()
    active_flag = 1 if safe(is_active) == "1" else 0
    values = (safe_int(contract_id), safe_int(vehicle_rate_id), safe(version_name), safe(effective_from), safe(effective_to), safe(fuel_reference), float(q2(ticket_open_price)), float(q2(second_slab_to_km)), float(q2(km_rate_101_300)), float(q2(km_rate_over_300)), float(q2(waiting_hour_rate)), active_flag, safe(notes))
    if row_id > 0:
        conn.execute("""
            UPDATE ops_vehicle_price_versions
            SET contract_id = ?, vehicle_rate_id = ?, version_name = ?, effective_from = ?, effective_to = ?,
                fuel_reference = ?, ticket_open_price = ?, second_slab_to_km = ?, km_rate_101_300 = ?,
                km_rate_over_300 = ?, waiting_hour_rate = ?, is_active = ?, notes = ?
            WHERE id = ?
        """, values + (row_id,))
        log_ops_event(request, "ops_vehicle_price_version", row_id, "Updated", f"Vehicle price {safe(version_name)}", conn=conn)
        notice = "updated"
    else:
        cur = conn.execute("""
            INSERT INTO ops_vehicle_price_versions (
                contract_id, vehicle_rate_id, version_name, effective_from, effective_to, fuel_reference,
                ticket_open_price, second_slab_to_km, km_rate_101_300, km_rate_over_300, waiting_hour_rate, is_active, notes
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, values)
        log_ops_event(request, "ops_vehicle_price_version", cur.lastrowid, "Created", f"Vehicle price {safe(version_name)}", conn=conn)
        notice = "saved"
    conn.commit()
    conn.close()
    return RedirectResponse(f"/ui/operations/pricing-versions?notice={notice}", status_code=303)


@router.get("/ui/operations/trips", response_class=HTMLResponse)
def trips_page(request: Request, notice: str = ""):
    conn = get_conn()
    rows = conn.execute("""
        SELECT
            t.*,
            v.code AS vehicle_code,
            v.plate_no,
            rs.name AS supplier_name,
            COALESCE(e.employee_name, e.name, e.full_name) AS driver_name,
            (SELECT COUNT(*) FROM ops_trip_work_orders l WHERE l.trip_id = t.id) AS work_order_count
        FROM ops_trip_tickets t
        LEFT JOIN ops_vehicles v ON v.id = t.vehicle_id
        LEFT JOIN ops_rental_suppliers rs ON rs.id = t.rental_supplier_id
        LEFT JOIN employees e ON e.id = t.driver_employee_id
        ORDER BY t.id DESC
    """).fetchall()
    conn.close()
    body = ""
    for row in rows:
        driver_label = safe(row["driver_name"]) or safe(row["supplier_driver_name"]) or "-"
        vehicle_label = safe(row["vehicle_code"]) or safe(row["vehicle_type"])
        if safe(row["supplier_name"]):
            vehicle_label = f"{vehicle_label} / {safe(row['supplier_name'])}"
        cost_label = "Pending Accounting Approval" if safe(row["status"]).lower() != "approved" else money(row["total_cost"])
        alloc_label = "Pending" if safe(row["status"]).lower() != "approved" else money(row["allocated_cost_per_work_order"])
        body += f"""
        <tr>
            <td>{safe(row['trip_no'])}</td>
            <td>{safe(row['trip_date'])}</td>
            <td>{vehicle_label}</td>
            <td>{driver_label}</td>
            <td>{safe(row['work_order_count'])}</td>
            <td>{money(row['total_km']) if q2(row['total_km']) > 0 else '-'}</td>
            <td>{cost_label}</td>
            <td>{alloc_label}</td>
            <td>{trip_status_chip(row['status'])}</td>
            <td><a class="btn blue" href="/ui/operations/trips/{row['id']}">Open</a></td>
        </tr>
        """
    if not body:
        body = "<tr><td colspan='10' style='text-align:center;'>No trip tickets yet.</td></tr>"
    html = f"""
    {current_form_notice(notice)}
    <div class="card">
        <div class="toolbar">
            <div>
                <h2 style="margin:0;">Trip Tickets</h2>
                <div class="section-note">Draft by technical manager, completion by movement manager, and accounting approval before transport cost hits work orders.</div>
            </div>
            <div style="display:flex;gap:8px;flex-wrap:wrap;">
                <a class="btn green" href="/ui/operations/trips/new">+ New Draft Trip</a>
                <a class="btn blue" href="/ui/operations/vehicles">Vehicles</a>
                <a class="btn blue" href="/ui/operations/rental-suppliers">Rental Offices</a>
                <a class="btn gray" href="/ui/operations">Back to Operations</a>
            </div>
        </div>
        <div class="table-summary">
            <span class="summary-pill">Draft: {sum(1 for r in rows if safe(r['status']).lower() == 'draft')}</span>
            <span class="summary-pill">Dispatched: {sum(1 for r in rows if safe(r['status']).lower() == 'dispatched')}</span>
            <span class="summary-pill">Completed: {sum(1 for r in rows if safe(r['status']).lower() == 'completed')}</span>
            <span class="summary-pill">Approved: {sum(1 for r in rows if safe(r['status']).lower() == 'approved')}</span>
        </div>
    </div>
    <div class="card">
        <h3 class="sub-title">Trip Register</h3>
        <table>
            <tr><th>Trip No</th><th>Date</th><th>Vehicle</th><th>Driver</th><th>Work Orders</th><th>KM</th><th>Total Cost</th><th>Cost / WO</th><th>Status</th><th>Action</th></tr>
            {body}
        </table>
    </div>
    """
    return render_ops_page(request, "Trip Tickets", html)


@router.get("/ui/operations/trips/new", response_class=HTMLResponse)
def new_trip_page(request: Request, notice: str = ""):
    return trip_detail_page(request, 0, notice)


@router.get("/ui/operations/trips/{trip_id}", response_class=HTMLResponse)
def trip_detail_page(request: Request, trip_id: int, notice: str = ""):
    conn = get_conn()
    trip = None
    lines = []
    if trip_id > 0:
        trip = conn.execute("""
            SELECT
                t.*,
                v.code AS vehicle_code,
                v.plate_no,
                rs.name AS supplier_name,
                COALESCE(e.employee_name, e.name, e.full_name) AS driver_name
            FROM ops_trip_tickets t
            LEFT JOIN ops_vehicles v ON v.id = t.vehicle_id
            LEFT JOIN ops_rental_suppliers rs ON rs.id = t.rental_supplier_id
            LEFT JOIN employees e ON e.id = t.driver_employee_id
            WHERE t.id = ?
            LIMIT 1
        """, (trip_id,)).fetchone()
        if not trip:
            conn.close()
            return HTMLResponse("Trip not found.", status_code=404)
        lines = conn.execute("""
            SELECT
                l.*,
                wo.work_order_no,
                wo.site_name,
                wo.status,
                c.name AS company_name
            FROM ops_trip_work_orders l
            LEFT JOIN ops_work_orders wo ON wo.id = l.work_order_id
            LEFT JOIN ops_contract_companies c ON c.id = wo.company_id
            WHERE l.trip_id = ?
            ORDER BY l.line_no, l.id
        """, (trip_id,)).fetchall()
    conn.close()

    selected_work_orders = [row["work_order_id"] for row in lines]
    form_values = dict(trip) if trip else {
        "id": 0,
        "trip_no": next_trip_no(),
        "trip_date": "",
        "vehicle_id": "",
        "driver_employee_id": "",
        "supplier_driver_name": "",
        "notes": "",
        "status": "draft",
        "start_odometer": "",
        "end_odometer": "",
        "waiting_hours": "0",
        "movement_notes": "",
        "accounting_notes": "",
        "driver_source": "company_driver",
    }
    if trip and not form_values.get("supplier_driver_name") and safe(form_values.get("driver_source")) == "supplier_driver":
        form_values["supplier_driver_name"] = safe(form_values.get("driver_name"))

    line_rows = ""
    for idx, row in enumerate(lines, start=1):
        alloc = "Pending approval" if safe(form_values.get("status")).lower() != "approved" else money(row["allocated_cost"])
        line_rows += f"""
        <tr>
            <td>{idx}</td>
            <td>{safe(row['work_order_no'])}</td>
            <td>{safe(row['company_name'])}</td>
            <td>{safe(row['site_name'])}</td>
            <td>{safe(row['status']).title()}</td>
            <td>{alloc}</td>
        </tr>
        """
    if not line_rows:
        line_rows = "<tr><td colspan='6' style='text-align:center;'>No work orders linked yet.</td></tr>"

    status_value = safe(form_values.get("status") or "draft").lower()
    start_photo_html = f"<a class='btn gray' target='_blank' href='{safe(form_values.get('start_photo_path'))}'>View Start Meter</a>" if safe(form_values.get("start_photo_path")) else "<span class='section-note'>No start photo yet.</span>"
    end_photo_html = f"<a class='btn gray' target='_blank' href='{safe(form_values.get('end_photo_path'))}'>View End Meter</a>" if safe(form_values.get("end_photo_path")) else "<span class='section-note'>No end photo yet.</span>"
    driver_source_value = safe(form_values.get("driver_source") or "company_driver")
    trip_title = "New Draft Trip" if safe_int(form_values.get("id")) == 0 else f"Trip {safe(form_values.get('trip_no'))}"

    html = f"""
    <style>
        .checkbox-grid {{ display:grid; grid-template-columns: repeat(auto-fit, minmax(260px, 1fr)); gap:10px; margin-top:10px; }}
        .checkbox-card {{ display:flex; gap:10px; align-items:flex-start; padding:12px; border:1px solid #d6e1f4; border-radius:14px; background:#fff; }}
        .checkbox-card input {{ margin-top:4px; }}
        .empty-note {{ padding:14px; border:1px dashed #c9d8f0; border-radius:14px; color:#61779b; background:#f8fbff; }}
    </style>
    {current_form_notice(notice)}
    <div class="card">
        <div class="toolbar">
            <div>
                <h2 style="margin:0;">{trip_title}</h2>
                <div class="section-note">Draft trip is prepared by the technical manager. Vehicle can still be changed while the trip is in draft only.</div>
            </div>
            <div style="display:flex;gap:8px;flex-wrap:wrap;">
                <a class="btn blue" href="/ui/operations/trips">Back to Trips</a>
                <a class="btn gray" href="/ui/operations/vehicles">Vehicles</a>
            </div>
        </div>
        <div class="table-summary">
            <span class="summary-pill">Status: {safe(form_values.get('status')).replace('_', ' ').title()}</span>
            <span class="summary-pill">Trip No: {safe(form_values.get('trip_no'))}</span>
            <span class="summary-pill">Linked Work Orders: {len(lines) if lines else len(selected_work_orders)}</span>
            <span class="summary-pill">Total Cost: {"Pending Accounting Approval" if status_value != "approved" else money(form_values.get('total_cost'))}</span>
        </div>
    </div>
    <div class="card">
        <h3 class="sub-title">Draft Setup</h3>
        <form method="post" action="/ui/operations/trips/save-draft">
            <input type="hidden" name="trip_id" value="{safe(form_values.get('id', 0))}">
            <div class="form-grid">
                <div class="form-group"><label>Trip No</label><input name="trip_no" value="{safe(form_values.get('trip_no'))}" required></div>
                <div class="form-group"><label>Trip Date</label><input type="date" name="trip_date" value="{safe(form_values.get('trip_date'))}" required></div>
                <div class="form-group"><label>Vehicle</label><select name="vehicle_id" required {'' if status_value == 'draft' or safe_int(form_values.get('id')) == 0 else 'disabled'}>{vehicle_options(form_values.get('vehicle_id'))}</select></div>
                <div class="form-group"><label>Driver</label><select name="driver_employee_id">{employee_options(form_values.get('driver_employee_id'), driver_only=True)}</select></div>
                <div class="form-group"><label>Supplier Driver Name</label><input name="supplier_driver_name" value="{safe(form_values.get('supplier_driver_name'))}" placeholder="Only if the rental office provides the driver"></div>
                <div class="form-group" style="grid-column: span 2;"><label>Draft Notes</label><input name="notes" value="{safe(form_values.get('notes'))}"></div>
            </div>
            <div class="form-group">
                <label>Linked Work Orders</label>
                <div class="section-note">One trip can serve more than one work order, and later the trip cost will be divided equally between them after accounting approval.</div>
                <div class="checkbox-grid">
                    {work_order_choices(selected_work_orders)}
                </div>
            </div>
            <div class="form-actions">
                <button class="btn green" type="submit">{"Update Draft Trip" if safe_int(form_values.get('id')) > 0 else "Save Draft Trip"}</button>
                <a class="btn gray" href="/ui/operations/trips/new">Clear</a>
            </div>
        </form>
    </div>
    """

    if safe_int(form_values.get("id")) > 0:
        html += f"""
        <div class="card">
            <h3 class="sub-title">Linked Work Orders</h3>
            <table>
                <tr><th>#</th><th>Work Order</th><th>Company</th><th>Site</th><th>Status</th><th>Allocated Transport Cost</th></tr>
                {line_rows}
            </table>
        </div>
        <div class="card">
            <div class="toolbar">
                <div>
                    <h3 class="sub-title" style="margin:0;">Movement Manager Section</h3>
                    <div class="section-note">Start and end odometer cannot be saved without the meter photo. Waiting hours are entered manually.</div>
                </div>
                <div class="table-summary">
                    <span class="summary-pill">Start Photo: {"Saved" if safe(form_values.get('start_photo_path')) else "Missing"}</span>
                    <span class="summary-pill">End Photo: {"Saved" if safe(form_values.get('end_photo_path')) else "Missing"}</span>
                </div>
            </div>
            <form method="post" action="/ui/operations/trips/{trip_id}/movement" enctype="multipart/form-data">
                <div class="form-grid">
                    <div class="form-group"><label>Start Odometer</label><input name="start_odometer" value="{safe(form_values.get('start_odometer'))}"></div>
                    <div class="form-group"><label>Start Meter Photo</label><input type="file" name="start_photo" accept="image/*"></div>
                    <div class="form-group">{start_photo_html}</div>
                    <div class="form-group"><label>End Odometer</label><input name="end_odometer" value="{safe(form_values.get('end_odometer'))}"></div>
                    <div class="form-group"><label>End Meter Photo</label><input type="file" name="end_photo" accept="image/*"></div>
                    <div class="form-group">{end_photo_html}</div>
                    <div class="form-group"><label>Waiting Hours</label><input name="waiting_hours" value="{safe(form_values.get('waiting_hours'))}"></div>
                    <div class="form-group" style="grid-column: span 2;"><label>Movement Notes</label><input name="movement_notes" value="{safe(form_values.get('movement_notes'))}"></div>
                </div>
                <div class="form-actions">
                    <button class="btn blue" type="submit" name="action" value="dispatch" {"disabled" if status_value in ['completed', 'approved'] else ''}>Save Start / Dispatch</button>
                    <button class="btn green" type="submit" name="action" value="complete" {"disabled" if status_value == 'approved' else ''}>Complete Trip</button>
                </div>
            </form>
            <div class="table-summary">
                <span class="summary-pill">KM: {money(form_values.get('total_km')) if q2(form_values.get('total_km')) > 0 else '-'}</span>
                <span class="summary-pill">Trip Cost: {money(form_values.get('total_cost')) if q2(form_values.get('total_cost')) > 0 else '-'}</span>
                <span class="summary-pill">Driver Commission: {money(form_values.get('driver_commission_amount')) if q2(form_values.get('driver_commission_amount')) > 0 else '-'}</span>
            </div>
        </div>
        <div class="card">
            <div class="toolbar">
                <div>
                    <h3 class="sub-title" style="margin:0;">Accounting Approval</h3>
                    <div class="section-note">Trip cost should only become visible on work orders after accounting approval.</div>
                </div>
                <div class="table-summary">
                    <span class="summary-pill">Per Work Order: {money(form_values.get('allocated_cost_per_work_order')) if status_value == 'approved' else 'Pending'}</span>
                </div>
            </div>
            <form method="post" action="/ui/operations/trips/{trip_id}/approve">
                <div class="form-grid">
                    <div class="form-group" style="grid-column: span 2;"><label>Accounting Notes</label><input name="accounting_notes" value="{safe(form_values.get('accounting_notes'))}"></div>
                </div>
                <div class="form-actions">
                    <button class="btn green" type="submit" {"disabled" if status_value != 'completed' else ''}>Approve Trip</button>
                </div>
            </form>
        </div>
        """
    return render_ops_page(request, "Trip Ticket", html)


@router.post("/ui/operations/trips/save-draft")
async def save_trip_draft(
    request: Request,
    trip_id: int = Form(0),
    trip_no: str = Form(""),
    trip_date: str = Form(""),
    vehicle_id: int = Form(0),
    driver_employee_id: int = Form(0),
    supplier_driver_name: str = Form(""),
    notes: str = Form(""),
    work_order_ids: list[int] = Form([]),
):
    if not safe(trip_date):
        target = "/ui/operations/trips/new" if trip_id <= 0 else f"/ui/operations/trips/{trip_id}"
        return RedirectResponse(target + "?notice=" + quote("Trip date is required."), status_code=302)
    if safe_int(vehicle_id) <= 0:
        target = "/ui/operations/trips/new" if trip_id <= 0 else f"/ui/operations/trips/{trip_id}"
        return RedirectResponse(target + "?notice=" + quote("Please select the actual vehicle."), status_code=302)
    selected_work_orders = [safe_int(x) for x in work_order_ids if safe_int(x) > 0]
    if not selected_work_orders:
        target = "/ui/operations/trips/new" if trip_id <= 0 else f"/ui/operations/trips/{trip_id}"
        return RedirectResponse(target + "?notice=" + quote("Please link at least one work order to the trip."), status_code=302)

    vehicle = get_vehicle_row(vehicle_id)
    if not vehicle:
        target = "/ui/operations/trips/new" if trip_id <= 0 else f"/ui/operations/trips/{trip_id}"
        return RedirectResponse(target + "?notice=" + quote("Selected vehicle was not found."), status_code=302)
    if safe_int(vehicle["vehicle_rate_id_resolved"]) <= 0:
        target = "/ui/operations/trips/new" if trip_id <= 0 else f"/ui/operations/trips/{trip_id}"
        return RedirectResponse(target + "?notice=" + quote("No active pricing slab was found for this vehicle. Please activate a matching vehicle pricing slab first."), status_code=302)
    driver_source = safe(vehicle["driver_source"]) or "company_driver"
    if driver_source == "company_driver" and safe_int(driver_employee_id) <= 0:
        target = "/ui/operations/trips/new" if trip_id <= 0 else f"/ui/operations/trips/{trip_id}"
        return RedirectResponse(target + "?notice=" + quote("Please select the company driver from HR employees."), status_code=302)
    if driver_source == "supplier_driver" and not safe(supplier_driver_name) and not safe(vehicle["supplier_driver_name"]):
        target = "/ui/operations/trips/new" if trip_id <= 0 else f"/ui/operations/trips/{trip_id}"
        return RedirectResponse(target + "?notice=" + quote("Please enter the supplier driver name for this trip."), status_code=302)

    conn = get_conn()
    current_trip = conn.execute("SELECT id, status FROM ops_trip_tickets WHERE id = ? LIMIT 1", (trip_id,)).fetchone() if trip_id > 0 else None
    if current_trip and safe(current_trip["status"]).lower() != "draft":
        conn.close()
        return RedirectResponse(f"/ui/operations/trips/{trip_id}?notice=" + quote("Only draft trips can be edited by the technical manager."), status_code=302)

    base_values = (
        safe(trip_no) or next_trip_no(),
        safe(trip_date),
        safe_int(vehicle_id),
        safe_int(vehicle["rental_supplier_id"]),
        safe(vehicle["vehicle_type"]),
        safe_int(vehicle["vehicle_rate_id_resolved"]),
        driver_source,
        safe_int(driver_employee_id) if driver_source == "company_driver" else 0,
        safe(supplier_driver_name) or safe(vehicle["supplier_driver_name"]),
        float(q2(vehicle["ticket_open_price"] if vehicle else 0)),
        float(q2(vehicle["second_slab_to_km"] if vehicle else 300)),
        float(q2(vehicle["km_rate_101_300"] if vehicle else 0)),
        float(q2(vehicle["km_rate_over_300"] if vehicle else 0)),
        float(q2(vehicle["waiting_hour_rate"] if vehicle else 0)),
        safe(notes),
    )
    if trip_id > 0:
        conn.execute("""
            UPDATE ops_trip_tickets
            SET trip_no = ?, trip_date = ?, vehicle_id = ?, rental_supplier_id = ?, vehicle_type = ?,
                vehicle_rate_id = ?, driver_source = ?, driver_employee_id = ?, supplier_driver_name = ?,
                ticket_open_price = ?, second_slab_to_km = ?, km_rate_101_300 = ?, km_rate_over_300 = ?, waiting_hour_rate = ?, notes = ?
            WHERE id = ?
        """, base_values + (trip_id,))
        target_trip_id = trip_id
    else:
        conn.execute("""
            INSERT INTO ops_trip_tickets (
                trip_no, trip_date, vehicle_id, rental_supplier_id, vehicle_type, vehicle_rate_id,
                driver_source, driver_employee_id, supplier_driver_name,
                ticket_open_price, second_slab_to_km, km_rate_101_300, km_rate_over_300, waiting_hour_rate,
                notes, created_by
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, base_values + (actor_name_from_request(request),))
        target_trip_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]

    conn.execute("DELETE FROM ops_trip_work_orders WHERE trip_id = ?", (target_trip_id,))
    for idx, work_order_id in enumerate(selected_work_orders, start=1):
        conn.execute("""
            INSERT OR IGNORE INTO ops_trip_work_orders (trip_id, work_order_id, line_no, allocated_cost)
            VALUES (?, ?, ?, 0)
        """, (target_trip_id, work_order_id, idx))
    conn.commit()
    conn.close()
    return RedirectResponse(f"/ui/operations/trips/{target_trip_id}?notice=" + quote("Draft trip saved successfully."), status_code=303)


@router.post("/ui/operations/trips/{trip_id}/movement")
async def save_trip_movement(
    request: Request,
    trip_id: int,
    action: str = Form("dispatch"),
    start_odometer: str = Form(""),
    end_odometer: str = Form(""),
    waiting_hours: str = Form("0"),
    movement_notes: str = Form(""),
    start_photo: UploadFile = File(None),
    end_photo: UploadFile = File(None),
):
    conn = get_conn()
    trip = conn.execute("SELECT * FROM ops_trip_tickets WHERE id = ? LIMIT 1", (trip_id,)).fetchone()
    if not trip:
        conn.close()
        return HTMLResponse("Trip not found.", status_code=404)
    if safe(trip["status"]).lower() == "approved":
        conn.close()
        return RedirectResponse(f"/ui/operations/trips/{trip_id}?notice=" + quote("Approved trips cannot be changed."), status_code=302)

    start_value = q2(start_odometer) if safe(start_odometer) else q2(trip["start_odometer"])
    end_value = q2(end_odometer) if safe(end_odometer) else q2(trip["end_odometer"])
    waiting_value = q2(waiting_hours) if safe(waiting_hours) else q2(trip["waiting_hours"])
    start_photo_path = safe(trip["start_photo_path"])
    end_photo_path = safe(trip["end_photo_path"])
    if start_photo and safe(start_photo.filename):
        start_photo_path = await save_trip_photo(start_photo, f"trip_{trip_id}_start")
    if end_photo and safe(end_photo.filename):
        end_photo_path = await save_trip_photo(end_photo, f"trip_{trip_id}_end")

    if safe(action).lower() == "dispatch":
        if start_value <= Decimal("0.00"):
            conn.close()
            return RedirectResponse(f"/ui/operations/trips/{trip_id}?notice=" + quote("Start odometer is required before dispatch."), status_code=302)
        if not safe(start_photo_path):
            conn.close()
            return RedirectResponse(f"/ui/operations/trips/{trip_id}?notice=" + quote("Please upload the start meter photo before dispatch."), status_code=302)
        conn.execute("""
            UPDATE ops_trip_tickets
            SET start_odometer = ?, start_photo_path = ?, movement_notes = ?, status = 'dispatched'
            WHERE id = ?
        """, (float(start_value), safe(start_photo_path), safe(movement_notes), trip_id))
        conn.commit()
        conn.close()
        return RedirectResponse(f"/ui/operations/trips/{trip_id}?notice=" + quote("Trip dispatched successfully."), status_code=303)

    if start_value <= Decimal("0.00"):
        conn.close()
        return RedirectResponse(f"/ui/operations/trips/{trip_id}?notice=" + quote("Start odometer is required before completion."), status_code=302)
    if not safe(start_photo_path):
        conn.close()
        return RedirectResponse(f"/ui/operations/trips/{trip_id}?notice=" + quote("Start meter photo is required before completion."), status_code=302)
    if end_value <= Decimal("0.00"):
        conn.close()
        return RedirectResponse(f"/ui/operations/trips/{trip_id}?notice=" + quote("End odometer is required to complete the trip."), status_code=302)
    if not safe(end_photo_path):
        conn.close()
        return RedirectResponse(f"/ui/operations/trips/{trip_id}?notice=" + quote("Please upload the end meter photo before completing the trip."), status_code=302)
    if end_value < start_value:
        conn.close()
        return RedirectResponse(f"/ui/operations/trips/{trip_id}?notice=" + quote("End odometer cannot be less than start odometer."), status_code=302)

    total_km = q2(end_value - start_value)
    total_cost = calculate_trip_cost(
        total_km,
        waiting_value,
        trip["ticket_open_price"],
        trip["second_slab_to_km"],
        trip["km_rate_101_300"],
        trip["km_rate_over_300"],
        trip["waiting_hour_rate"],
    )
    work_order_count = conn.execute("SELECT COUNT(*) AS cnt FROM ops_trip_work_orders WHERE trip_id = ?", (trip_id,)).fetchone()["cnt"] or 0
    allocated_cost = q2(total_cost / Decimal(str(work_order_count))) if int(work_order_count) > 0 else Decimal("0.00")
    driver_commission_pct = Decimal("0.00")
    driver_commission_amount = Decimal("0.00")
    if safe(trip["driver_source"]) == "company_driver" and safe_int(trip["driver_employee_id"]) > 0:
        driver = conn.execute("SELECT COALESCE(trip_commission_pct, 0) AS trip_commission_pct FROM employees WHERE id = ? LIMIT 1", (trip["driver_employee_id"],)).fetchone()
        driver_commission_pct = q2(driver["trip_commission_pct"] if driver else 0)
        driver_commission_amount = q2(total_cost * driver_commission_pct / Decimal("100.00"))

    conn.execute("""
        UPDATE ops_trip_tickets
        SET start_odometer = ?, start_photo_path = ?, end_odometer = ?, end_photo_path = ?,
            waiting_hours = ?, total_km = ?, total_cost = ?, allocated_cost_per_work_order = ?,
            driver_commission_pct = ?, driver_commission_amount = ?, movement_notes = ?,
            movement_closed_by = ?, completed_at = CURRENT_TIMESTAMP, status = 'completed'
        WHERE id = ?
    """, (
        float(start_value),
        safe(start_photo_path),
        float(end_value),
        safe(end_photo_path),
        float(waiting_value),
        float(total_km),
        float(total_cost),
        float(allocated_cost),
        float(driver_commission_pct),
        float(driver_commission_amount),
        safe(movement_notes),
        actor_name_from_request(request),
        trip_id,
    ))
    conn.commit()
    conn.close()
    return RedirectResponse(f"/ui/operations/trips/{trip_id}?notice=" + quote("Trip completed and ready for accounting approval."), status_code=303)


@router.post("/ui/operations/trips/{trip_id}/approve")
def approve_trip(
    request: Request,
    trip_id: int,
    accounting_notes: str = Form(""),
):
    conn = get_conn()
    trip = conn.execute("SELECT * FROM ops_trip_tickets WHERE id = ? LIMIT 1", (trip_id,)).fetchone()
    if not trip:
        conn.close()
        return HTMLResponse("Trip not found.", status_code=404)
    if safe(trip["status"]).lower() != "completed":
        conn.close()
        return RedirectResponse(f"/ui/operations/trips/{trip_id}?notice=" + quote("Only completed trips can be approved by accounting."), status_code=302)
    lines = conn.execute("SELECT id FROM ops_trip_work_orders WHERE trip_id = ? ORDER BY id", (trip_id,)).fetchall()
    if not lines:
        conn.close()
        return RedirectResponse(f"/ui/operations/trips/{trip_id}?notice=" + quote("No work orders linked to this trip."), status_code=302)
    allocated_cost = q2(trip["allocated_cost_per_work_order"])
    for row in lines:
        conn.execute("UPDATE ops_trip_work_orders SET allocated_cost = ? WHERE id = ?", (float(allocated_cost), row["id"]))
    conn.execute("""
        UPDATE ops_trip_tickets
        SET status = 'approved', accounting_notes = ?, approved_by = ?, approved_at = CURRENT_TIMESTAMP
        WHERE id = ?
    """, (safe(accounting_notes), actor_name_from_request(request), trip_id))
    conn.commit()
    conn.close()
    return RedirectResponse(f"/ui/operations/trips/{trip_id}?notice=" + quote("Trip approved successfully. Transport cost is now ready for work orders."), status_code=303)
