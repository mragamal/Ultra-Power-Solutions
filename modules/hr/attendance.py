import csv
import io
from datetime import datetime
from html import escape
from urllib.parse import quote

from fastapi import APIRouter, Request, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse

from db import get_conn
from layout import render_page
from modules.hr.categories import ensure_categories_table
from modules.hr.employees import ensure_employees_table, safe, to_float

try:
    from openpyxl import Workbook, load_workbook
except Exception:
    Workbook = None
    load_workbook = None

router = APIRouter()


def ensure_column(conn, table_name, column_name, alter_sql):
    cols = conn.execute(f"PRAGMA table_info({table_name})").fetchall()
    names = [c["name"] for c in cols]
    if column_name not in names:
        conn.execute(alter_sql)


def ensure_attendance_tables():
    ensure_employees_table()
    ensure_categories_table()
    conn = get_conn()
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS attendance_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            employee_id INTEGER,
            employee_code TEXT,
            biometric_code TEXT,
            employee_name TEXT,
            category_name TEXT,
            attendance_role TEXT,
            attendance_date TEXT,
            check_in TEXT,
            check_out TEXT,
            worked_hours REAL DEFAULT 0,
            overtime_hours REAL DEFAULT 0,
            late_minutes REAL DEFAULT 0,
            early_leave_minutes REAL DEFAULT 0,
            status TEXT DEFAULT 'present',
            source TEXT DEFAULT 'biometric_import',
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
        """
    )

    ensure_column(conn, "attendance_logs", "employee_id", "ALTER TABLE attendance_logs ADD COLUMN employee_id INTEGER")
    ensure_column(conn, "attendance_logs", "employee_code", "ALTER TABLE attendance_logs ADD COLUMN employee_code TEXT")
    ensure_column(conn, "attendance_logs", "biometric_code", "ALTER TABLE attendance_logs ADD COLUMN biometric_code TEXT")
    ensure_column(conn, "attendance_logs", "employee_name", "ALTER TABLE attendance_logs ADD COLUMN employee_name TEXT")
    ensure_column(conn, "attendance_logs", "category_name", "ALTER TABLE attendance_logs ADD COLUMN category_name TEXT")
    ensure_column(conn, "attendance_logs", "attendance_role", "ALTER TABLE attendance_logs ADD COLUMN attendance_role TEXT")
    ensure_column(conn, "attendance_logs", "attendance_date", "ALTER TABLE attendance_logs ADD COLUMN attendance_date TEXT")
    ensure_column(conn, "attendance_logs", "check_in", "ALTER TABLE attendance_logs ADD COLUMN check_in TEXT")
    ensure_column(conn, "attendance_logs", "check_out", "ALTER TABLE attendance_logs ADD COLUMN check_out TEXT")
    ensure_column(conn, "attendance_logs", "worked_hours", "ALTER TABLE attendance_logs ADD COLUMN worked_hours REAL DEFAULT 0")
    ensure_column(conn, "attendance_logs", "overtime_hours", "ALTER TABLE attendance_logs ADD COLUMN overtime_hours REAL DEFAULT 0")
    ensure_column(conn, "attendance_logs", "late_minutes", "ALTER TABLE attendance_logs ADD COLUMN late_minutes REAL DEFAULT 0")
    ensure_column(conn, "attendance_logs", "early_leave_minutes", "ALTER TABLE attendance_logs ADD COLUMN early_leave_minutes REAL DEFAULT 0")
    ensure_column(conn, "attendance_logs", "status", "ALTER TABLE attendance_logs ADD COLUMN status TEXT DEFAULT 'present'")
    ensure_column(conn, "attendance_logs", "source", "ALTER TABLE attendance_logs ADD COLUMN source TEXT DEFAULT 'biometric_import'")
    ensure_column(conn, "attendance_logs", "created_at", "ALTER TABLE attendance_logs ADD COLUMN created_at TEXT DEFAULT CURRENT_TIMESTAMP")
    conn.commit()
    conn.close()


def parse_csv_rows(file_bytes):
    text = file_bytes.decode("utf-8-sig", errors="ignore")
    reader = csv.DictReader(io.StringIO(text))
    return [dict(r) for r in reader]


def parse_xlsx_rows(file_bytes):
    if load_workbook is None:
        raise Exception("Excel import is not available right now.")
    workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [safe(h).lower() for h in rows[0]]
    result = []
    for data_row in rows[1:]:
        item = {}
        for i, header in enumerate(headers):
            if not header:
                continue
            item[header] = "" if i >= len(data_row) or data_row[i] is None else str(data_row[i])
        result.append(item)
    return result


def parse_time_to_hours(value):
    text = safe(value)
    if not text:
        return None
    for fmt in ("%H:%M", "%H:%M:%S"):
        try:
            dt = datetime.strptime(text, fmt)
            return dt.hour + (dt.minute / 60) + (dt.second / 3600)
        except Exception:
            pass
    return None


def parse_minutes_from_time(value):
    hours = parse_time_to_hours(value)
    if hours is None:
        return None
    return hours * 60.0


def derive_worked_hours(check_in, check_out, fallback=0):
    start = parse_time_to_hours(check_in)
    end = parse_time_to_hours(check_out)
    if start is None or end is None:
        return to_float(fallback, 0)
    diff = end - start
    if diff < 0:
        return to_float(fallback, 0)
    return diff


def is_attendance_exempt_employee(employee):
    keys = employee.keys() if hasattr(employee, "keys") else employee
    try:
        if "attendance_exempt" in keys and int(employee["attendance_exempt"] or 0) == 1:
            return True
    except Exception:
        pass
    role = safe(employee["attendance_role"] if "attendance_role" in keys else "")
    role = role.lower().replace("_", " ").replace("-", " ")
    return role in ("manager", "managerial", "managerial level", "no attendance")


def get_header_value(row, aliases):
    for key in aliases:
        if key in row and safe(row.get(key)):
            return safe(row.get(key))
    return ""


def normalize_attendance_row(row):
    return {
        "employee_code": get_header_value(row, ["employee_code", "code", "emp_code"]),
        "biometric_code": get_header_value(row, ["biometric_code", "machine_code", "attendance_code"]),
        "attendance_date": get_header_value(row, ["attendance_date", "date", "work_date"]),
        "check_in": get_header_value(row, ["check_in", "in_time", "time_in"]),
        "check_out": get_header_value(row, ["check_out", "out_time", "time_out"]),
        "worked_hours": get_header_value(row, ["worked_hours", "hours_worked", "work_hours"]),
        "status": (get_header_value(row, ["status", "attendance_status"]) or "present").lower(),
    }


ensure_attendance_tables()


@router.get("/ui/hr/attendance", response_class=HTMLResponse)
def attendance_list(request: Request):
    conn = get_conn()
    rows = conn.execute(
        """
        SELECT *
        FROM attendance_logs
        ORDER BY attendance_date DESC, employee_code, id DESC
        LIMIT 500
        """
    ).fetchall()
    conn.close()

    msg = safe(request.query_params.get("msg"))
    msg_html = f'<div class="msg success">{escape(msg)}</div>' if msg else ""

    body = ""
    for row in rows:
        body += f"""
        <tr>
            <td>{escape(safe(row['attendance_date']))}</td>
            <td>{escape(safe(row['employee_code']))}</td>
            <td>{escape(safe(row['biometric_code']))}</td>
            <td>{escape(safe(row['employee_name']))}</td>
            <td>{escape(safe(row['category_name']))}</td>
            <td>{escape(safe(row['attendance_role']))}</td>
            <td>{escape(safe(row['check_in']))}</td>
            <td>{escape(safe(row['check_out']))}</td>
            <td class="number-cell">{float(row['worked_hours'] or 0):,.2f}</td>
            <td class="number-cell">{float(row['overtime_hours'] or 0):,.2f}</td>
            <td class="number-cell">{float(row['late_minutes'] or 0):,.0f}</td>
            <td class="number-cell">{float(row['early_leave_minutes'] or 0):,.0f}</td>
            <td>{escape(safe(row['status']))}</td>
        </tr>
        """
    if not body:
        body = "<tr><td colspan='13' style='text-align:center;'>No attendance logs found.</td></tr>"

    html = f"""
    <div class="card">
        {msg_html}
        <div style="display:flex;justify-content:space-between;align-items:center;gap:10px;flex-wrap:wrap;">
            <h2>Attendance</h2>
            <div style="display:flex;gap:8px;flex-wrap:wrap;">
                <a class="btn blue" href="/ui/hr/attendance/export/excel">Export Excel</a>
                <a class="btn blue" href="/ui/hr/attendance/template.csv">Template CSV</a>
                <a class="btn blue" href="/ui/hr/attendance/template.xlsx">Template Excel</a>
            </div>
        </div>

        <form method="post" action="/ui/hr/attendance/import" enctype="multipart/form-data" style="margin-top:14px;">
            <div class="row">
                <div class="col">
                    <label>Import Attendance / Biometric File</label>
                    <input type="file" name="file" accept=".csv,.xlsx" required>
                </div>
                <div class="col" style="display:flex;align-items:end;">
                    <button class="btn green" type="submit">Import Attendance</button>
                </div>
            </div>
        </form>
    </div>

    <div class="card">
        <p class="section-note">Import exported attendance from the fingerprint device using employee code or biometric code. Payroll will use these logs to calculate worked days, overtime, and absence deduction.</p>
    </div>

    <div class="card">
        <div class="table-wrap">
            <table>
                <tr>
                    <th>Date</th>
                    <th>Employee Code</th>
                    <th>Biometric Code</th>
                    <th>Name</th>
                    <th>Category</th>
                    <th>Role</th>
                    <th>Check In</th>
                    <th>Check Out</th>
                    <th>Worked Hours</th>
                    <th>Overtime Hours</th>
                    <th>Late Min</th>
                    <th>Early Leave Min</th>
                    <th>Status</th>
                </tr>
                {body}
            </table>
        </div>
    </div>
    """
    return HTMLResponse(render_page("Attendance", html, "en", current_path=request.url.path))


@router.get("/ui/hr/attendance/template.csv")
def attendance_template_csv():
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["employee_code", "biometric_code", "attendance_date", "check_in", "check_out", "worked_hours", "status"])
    writer.writerow(["EMP-0001", "1001", "2026-04-01", "08:00", "17:15", "9.25", "present"])
    writer.writerow(["EMP-0002", "1002", "2026-04-01", "08:30", "16:30", "8.00", "present"])
    data = output.getvalue().encode("utf-8-sig")
    stream = io.BytesIO(data)
    stream.seek(0)
    return StreamingResponse(stream, media_type="text/csv; charset=utf-8", headers={"Content-Disposition": 'attachment; filename="attendance_template.csv"'})


@router.get("/ui/hr/attendance/export/excel")
def export_attendance_excel():
    if Workbook is None:
        return RedirectResponse("/ui/hr/attendance?msg=" + quote("Excel export is not available. Install openpyxl."))
    
    ensure_attendance_tables()
    conn = get_conn()
    
    try:
        rows = conn.execute("""
            SELECT 
                al.id,
                al.employee_id,
                al.employee_code,
                al.biometric_code,
                al.employee_name,
                al.category_name,
                al.attendance_role,
                al.attendance_date,
                al.check_in,
                al.check_out,
                al.worked_hours,
                al.overtime_hours,
                al.late_minutes,
                al.early_leave_minutes,
                al.status,
                al.notes
            FROM attendance_logs al
            ORDER BY al.attendance_date DESC, al.employee_code
        """).fetchall()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Data"
        
        headers = [
            'ID', 'Employee ID', 'Employee Code', 'Biometric Code', 'Employee Name',
            'Category', 'Role', 'Date', 'Check In', 'Check Out', 'Worked Hours', 
            'OT Hours', 'Late Minutes', 'Early Leave Minutes', 'Status', 'Notes'
        ]
        
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)
        
        for row_num, att in enumerate(rows, 2):
            ws.cell(row=row_num, column=1, value=att[0])
            ws.cell(row=row_num, column=2, value=att[1])
            ws.cell(row=row_num, column=3, value=att[2])
            ws.cell(row=row_num, column=4, value=att[3])
            ws.cell(row=row_num, column=5, value=att[4])
            ws.cell(row=row_num, column=6, value=att[5])
            ws.cell(row=row_num, column=7, value=att[6])
            ws.cell(row=row_num, column=8, value=att[7])
            ws.cell(row=row_num, column=9, value=att[8])
            ws.cell(row=row_num, column=10, value=att[9])
            ws.cell(row=row_num, column=11, value=att[10])
            ws.cell(row=row_num, column=12, value=att[11])
            ws.cell(row=row_num, column=13, value=att[12])
            ws.cell(row=row_num, column=14, value=att[13])
            ws.cell(row=row_num, column=15, value=att[14])
            ws.cell(row=row_num, column=16, value=att[15])
            ws.cell(row=row_num, column=17, value=att[16])
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"attendance_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return StreamingResponse(
            io.BytesIO(output.getvalue()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        return RedirectResponse("/ui/hr/attendance?msg=" + quote(f"Export error: {str(e)}"))
    finally:
        conn.close()


@router.get("/ui/hr/attendance/template.xlsx")
def attendance_template_xlsx():
    if Workbook is None:
        return RedirectResponse("/ui/hr/attendance?msg=" + quote("Excel template is not available. Use CSV template."))
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Attendance Template"
    sheet.append(["employee_code", "biometric_code", "attendance_date", "check_in", "check_out", "worked_hours", "status"])
    sheet.append(["EMP-0001", "1001", "2026-04-01", "08:00", "17:15", "9.25", "present"])
    sheet.append(["EMP-0002", "1002", "2026-04-01", "08:30", "16:30", "8.00", "present"])
    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": 'attachment; filename="attendance_template.xlsx"'})


@router.post("/ui/hr/attendance/import")
async def attendance_import(file: UploadFile = File(...)):
    ensure_attendance_tables()
    filename = safe(file.filename).lower()
    file_bytes = await file.read()
    if not filename:
        return RedirectResponse("/ui/hr/attendance?msg=" + quote("Please choose an attendance file."), status_code=302)

    try:
        if filename.endswith(".csv"):
            raw_rows = parse_csv_rows(file_bytes)
        elif filename.endswith(".xlsx"):
            raw_rows = parse_xlsx_rows(file_bytes)
        else:
            return RedirectResponse("/ui/hr/attendance?msg=" + quote("Only CSV or XLSX files are supported."), status_code=302)
    except Exception as ex:
        return RedirectResponse("/ui/hr/attendance?msg=" + quote(f"Attendance import failed: {safe(ex)}"), status_code=302)

    conn = get_conn()
    employees = conn.execute(
        """
        SELECT e.*, c.name AS category_name, c.attendance_role, c.shift_start, c.shift_end,
               COALESCE(c.attendance_exempt, 0) AS attendance_exempt,
               c.expected_daily_hours AS category_daily_hours,
               c.late_grace_minutes, c.early_leave_grace_minutes, c.overtime_after_hours
        FROM employees e
        LEFT JOIN employee_categories c ON c.id = e.category_id
        """
    ).fetchall()
    by_code = {safe(e["code"]).upper(): e for e in employees if safe(e["code"])}
    by_bio = {safe(e["biometric_code"]).upper(): e for e in employees if safe(e["biometric_code"])}

    imported = 0
    skipped = 0
    for raw_row in raw_rows:
        row = normalize_attendance_row(raw_row)
        employee = None
        code_key = safe(row["employee_code"]).upper()
        bio_key = safe(row["biometric_code"]).upper()
        if code_key and code_key in by_code:
            employee = by_code[code_key]
        elif bio_key and bio_key in by_bio:
            employee = by_bio[bio_key]

        if not employee or not safe(row["attendance_date"]):
            skipped += 1
            continue

        worked_hours = derive_worked_hours(row["check_in"], row["check_out"], row["worked_hours"])
        overtime_after_hours = float(employee["overtime_after_hours"] or employee["expected_daily_hours"] or employee["category_daily_hours"] or 8)
        overtime_hours = max(worked_hours - overtime_after_hours, 0.0)

        late_minutes = 0.0
        early_leave_minutes = 0.0
        check_in_minutes = parse_minutes_from_time(row["check_in"])
        check_out_minutes = parse_minutes_from_time(row["check_out"])
        shift_start_minutes = parse_minutes_from_time(employee["shift_start"])
        shift_end_minutes = parse_minutes_from_time(employee["shift_end"])
        late_grace = float(employee["late_grace_minutes"] or 0)
        early_grace = float(employee["early_leave_grace_minutes"] or 0)
        if check_in_minutes is not None and shift_start_minutes is not None:
            late_minutes = max(check_in_minutes - (shift_start_minutes + late_grace), 0.0)
        if check_out_minutes is not None and shift_end_minutes is not None:
            early_leave_minutes = max((shift_end_minutes - early_grace) - check_out_minutes, 0.0)

        status = safe(row["status"]) or "present"
        if late_minutes > 0:
            status = "late"
        if early_leave_minutes > 0 and status == "present":
            status = "early_leave"
        if late_minutes > 0 and early_leave_minutes > 0:
            status = "late_and_early_leave"

        if is_attendance_exempt_employee(employee):
            late_minutes = 0.0
            early_leave_minutes = 0.0
            overtime_hours = 0.0
            status = safe(row["status"]) or "present"

        exists = conn.execute(
            """
            SELECT id
            FROM attendance_logs
            WHERE employee_id = ?
              AND attendance_date = ?
              AND COALESCE(check_in, '') = ?
              AND COALESCE(check_out, '') = ?
            LIMIT 1
            """,
            (employee["id"], safe(row["attendance_date"]), safe(row["check_in"]), safe(row["check_out"])),
        ).fetchone()
        if exists:
            skipped += 1
            continue

        conn.execute(
            """
            INSERT INTO attendance_logs (
                employee_id, employee_code, biometric_code, employee_name, category_name, attendance_role,
                attendance_date, check_in, check_out, worked_hours, overtime_hours, late_minutes, early_leave_minutes, status, source
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'biometric_import')
            """,
            (
                employee["id"],
                safe(employee["code"]),
                safe(employee["biometric_code"]),
                safe(employee["name"]),
                safe(employee["category_name"]),
                safe(employee["attendance_role"]),
                safe(row["attendance_date"]),
                safe(row["check_in"]),
                safe(row["check_out"]),
                worked_hours,
                overtime_hours,
                late_minutes,
                early_leave_minutes,
                status,
            ),
        )
        imported += 1

    conn.commit()
    conn.close()
    return RedirectResponse("/ui/hr/attendance?msg=" + quote(f"Attendance imported: {imported}. Skipped: {skipped}."), status_code=302)
