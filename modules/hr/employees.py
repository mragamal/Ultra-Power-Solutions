import csv
import io
from datetime import datetime
from html import escape
from urllib.parse import quote

from fastapi import APIRouter, Request, Form, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse

from db import get_conn
from layout import render_page
from audit import render_audit_log_card, safe_log_request_action
from modules.hr.categories import category_options_html, ensure_categories_table

try:
    from openpyxl import Workbook, load_workbook
except Exception:
    Workbook = None
    load_workbook = None

router = APIRouter()


def safe(value):
    return "" if value is None else str(value).strip()


def money(value):
    try:
        return f"{float(value or 0):,.2f}"
    except Exception:
        return "0.00"


def to_float(value, default=0.0):
    try:
        return float(safe(value).replace(",", "") or default)
    except Exception:
        return float(default)


def to_int_flag(value, default=1):
    text = safe(value).lower()
    if text in ("1", "true", "yes", "y", "active"):
        return 1
    if text in ("0", "false", "no", "n", "inactive"):
        return 0
    try:
        return 1 if int(value) == 1 else 0
    except Exception:
        return int(default)


def resolve_category_id(conn, category_id="", category_code="", category_name=""):
    if safe(category_id).isdigit():
        return int(safe(category_id))

    code = safe(category_code)
    name = safe(category_name)
    if not code and not name:
        return None

    row = None
    if code:
        row = conn.execute(
            "SELECT id FROM employee_categories WHERE UPPER(COALESCE(code, '')) = UPPER(?) LIMIT 1",
            (code,),
        ).fetchone()
    if not row and name:
        row = conn.execute(
            "SELECT id FROM employee_categories WHERE UPPER(COALESCE(name, '')) = UPPER(?) LIMIT 1",
            (name,),
        ).fetchone()
    return int(row["id"]) if row else None


def ensure_column(conn, table_name, column_name, alter_sql):
    cols = conn.execute(f"PRAGMA table_info({table_name})").fetchall()
    names = [c["name"] for c in cols]
    if column_name not in names:
        conn.execute(alter_sql)


def ensure_employees_table():
    ensure_categories_table()
    conn = get_conn()

    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS employees (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT,
            category_id INTEGER,
            biometric_code TEXT,
            name TEXT NOT NULL,
            phone TEXT,
            email TEXT,
            department TEXT,
            job_title TEXT,
            hire_date TEXT,
            national_id TEXT,
            payment_method TEXT DEFAULT 'bank_transfer',
            bank_name TEXT,
            bank_account TEXT,
            basic_salary REAL DEFAULT 0,
            housing_allowance REAL DEFAULT 0,
            transport_allowance REAL DEFAULT 0,
            other_allowance REAL DEFAULT 0,
            insurance_applicable INTEGER DEFAULT 1,
            insurance_number TEXT,
            insurance_salary REAL DEFAULT 0,
            insurance_employee_rate REAL DEFAULT 11,
            insurance_employer_rate REAL DEFAULT 18.75,
            expected_daily_hours REAL DEFAULT 8,
            is_active INTEGER DEFAULT 1,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
        """
    )

    ensure_column(conn, "employees", "code", "ALTER TABLE employees ADD COLUMN code TEXT")
    ensure_column(conn, "employees", "category_id", "ALTER TABLE employees ADD COLUMN category_id INTEGER")
    ensure_column(conn, "employees", "biometric_code", "ALTER TABLE employees ADD COLUMN biometric_code TEXT")
    ensure_column(conn, "employees", "phone", "ALTER TABLE employees ADD COLUMN phone TEXT")
    ensure_column(conn, "employees", "email", "ALTER TABLE employees ADD COLUMN email TEXT")
    ensure_column(conn, "employees", "department", "ALTER TABLE employees ADD COLUMN department TEXT")
    ensure_column(conn, "employees", "job_title", "ALTER TABLE employees ADD COLUMN job_title TEXT")
    ensure_column(conn, "employees", "hire_date", "ALTER TABLE employees ADD COLUMN hire_date TEXT")
    ensure_column(conn, "employees", "national_id", "ALTER TABLE employees ADD COLUMN national_id TEXT")
    ensure_column(conn, "employees", "payment_method", "ALTER TABLE employees ADD COLUMN payment_method TEXT DEFAULT 'bank_transfer'")
    ensure_column(conn, "employees", "bank_name", "ALTER TABLE employees ADD COLUMN bank_name TEXT")
    ensure_column(conn, "employees", "bank_account", "ALTER TABLE employees ADD COLUMN bank_account TEXT")
    ensure_column(conn, "employees", "basic_salary", "ALTER TABLE employees ADD COLUMN basic_salary REAL DEFAULT 0")
    ensure_column(conn, "employees", "housing_allowance", "ALTER TABLE employees ADD COLUMN housing_allowance REAL DEFAULT 0")
    ensure_column(conn, "employees", "transport_allowance", "ALTER TABLE employees ADD COLUMN transport_allowance REAL DEFAULT 0")
    ensure_column(conn, "employees", "other_allowance", "ALTER TABLE employees ADD COLUMN other_allowance REAL DEFAULT 0")
    ensure_column(conn, "employees", "insurance_applicable", "ALTER TABLE employees ADD COLUMN insurance_applicable INTEGER DEFAULT 1")
    ensure_column(conn, "employees", "insurance_number", "ALTER TABLE employees ADD COLUMN insurance_number TEXT")
    ensure_column(conn, "employees", "insurance_salary", "ALTER TABLE employees ADD COLUMN insurance_salary REAL DEFAULT 0")
    ensure_column(conn, "employees", "insurance_employee_rate", "ALTER TABLE employees ADD COLUMN insurance_employee_rate REAL DEFAULT 11")
    ensure_column(conn, "employees", "insurance_employer_rate", "ALTER TABLE employees ADD COLUMN insurance_employer_rate REAL DEFAULT 18.75")
    ensure_column(conn, "employees", "expected_daily_hours", "ALTER TABLE employees ADD COLUMN expected_daily_hours REAL DEFAULT 8")
    ensure_column(conn, "employees", "is_active", "ALTER TABLE employees ADD COLUMN is_active INTEGER DEFAULT 1")

    conn.commit()
    conn.close()


def next_employee_code():
    ensure_employees_table()
    conn = get_conn()
    row = conn.execute(
        """
        SELECT code
        FROM employees
        WHERE COALESCE(code, '') <> ''
        ORDER BY id DESC
        LIMIT 1
        """
    ).fetchone()
    conn.close()

    last = safe(row["code"]) if row else ""
    if not last:
        return "EMP-0001"
    try:
        num = int(last.split("-")[-1])
    except Exception:
        num = 0
    return f"EMP-{num + 1:04d}"


def q_count(conn, sql, params=()):
    try:
        return int(conn.execute(sql, params).fetchone()[0] or 0)
    except Exception:
        return 0


def q_sum(conn, sql, params=()):
    try:
        return float(conn.execute(sql, params).fetchone()[0] or 0)
    except Exception:
        return 0.0


def employee_package(employee):
    return (
        float(employee["basic_salary"] or 0)
        + float(employee["housing_allowance"] or 0)
        + float(employee["transport_allowance"] or 0)
        + float(employee["other_allowance"] or 0)
    )


def employee_stat_button(label, value, href, accent="#2563eb"):
    return f"""
    <a href="{href}" style="display:flex;align-items:center;gap:10px;min-width:150px;padding:12px 16px;border:1px solid #dbe4f0;border-radius:8px;background:#fff;text-decoration:none;color:#0b2d5b;">
        <span style="width:34px;height:34px;border-radius:8px;background:{accent};color:white;display:flex;align-items:center;justify-content:center;font-weight:900;">#</span>
        <span>
            <span style="display:block;font-size:12px;color:#5f718d;font-weight:700;">{label}</span>
            <span style="display:block;font-size:18px;font-weight:900;">{value}</span>
        </span>
    </a>
    """


def employee_profile_html(conn, employee, request_path):
    employee_id = int(employee["id"])
    full_name = safe(employee["name"])
    initials = escape((full_name or "?")[:1].upper())
    total_package = employee_package(employee)
    payroll_count = q_count(conn, "SELECT COUNT(*) FROM payroll_lines WHERE employee_id = ?", (employee_id,))
    advances_balance = q_sum(conn, """
        SELECT COALESCE(SUM(balance), 0)
        FROM employee_advances
        WHERE employee_id = ?
    """, (employee_id,))
    custody_balance = q_sum(conn, """
        SELECT COALESCE(SUM(l.debit - l.credit), 0)
        FROM journal_lines l
        JOIN journal_entries j ON j.id = l.journal_id
        LEFT JOIN accounts a ON a.code = l.account_code
        WHERE LOWER(COALESCE(j.status,'')) = 'posted'
          AND LOWER(COALESCE(l.partner_type,'')) = 'employee'
          AND COALESCE(l.partner_id,0) = ?
          AND (COALESCE(a.name,'') LIKE '%عهد%' OR LOWER(COALESCE(a.name,'')) LIKE '%custody%')
    """, (employee_id,))
    status = '<span class="status-chip green">Active</span>' if int(employee["is_active"] or 0) == 1 else '<span class="status-chip gray">Inactive</span>'
    smart_buttons = "".join([
        employee_stat_button("Payroll", str(payroll_count), f"/ui/hr/payroll?employee_id={employee_id}", "#2563eb"),
        employee_stat_button("Salary Package", money(total_package), f"/ui/hr/employees/{employee_id}/edit", "#0f766e"),
        employee_stat_button("Advances", money(advances_balance), f"/ui/accounting/employee-advances?employee_id={employee_id}", "#dc2626"),
        employee_stat_button("Custody", money(custody_balance), f"/ui/accounting/petty-cash/statement?employee_id={employee_id}", "#7c3aed"),
        employee_stat_button("Ledger", "Open", f"/ui/accounting/partner-ledger?partner_type=employee&partner_id={employee_id}", "#f59e0b"),
    ])

    html = f"""
    <div class="card">
        <div class="toolbar">
            <div>
                <h2 style="margin:0;">Employee {escape(safe(employee['code']))}</h2>
                <div style="color:#6f819d;margin-top:6px;">{escape(full_name)}</div>
            </div>
            <div style="display:flex;gap:8px;flex-wrap:wrap;">
                <a class="btn blue" href="/ui/hr/employees/{employee_id}/edit">Edit</a>
                <a class="btn gray" href="/ui/hr/employees">Back</a>
            </div>
        </div>
    </div>

    <div class="card" style="display:flex;gap:10px;flex-wrap:wrap;align-items:stretch;">
        {smart_buttons}
    </div>

    <div style="display:grid;grid-template-columns:minmax(0,2fr) minmax(300px,1fr);gap:18px;align-items:start;">
        <div class="card">
            <div style="display:flex;gap:18px;align-items:flex-start;flex-wrap:wrap;">
                <div style="width:118px;height:118px;border-radius:8px;background:#2f5fb8;color:white;display:flex;align-items:center;justify-content:center;font-size:58px;font-weight:900;">{initials}</div>
                <div style="flex:1;min-width:260px;">
                    <div style="font-size:30px;font-weight:900;color:#0b2d5b;">{escape(full_name)}</div>
                    <div style="margin-top:10px;color:#102a4c;line-height:1.8;">
                        <div><b>Code:</b> {escape(safe(employee['code']))}</div>
                        <div><b>Biometric:</b> {escape(safe(employee['biometric_code']))}</div>
                        <div><b>Phone:</b> {escape(safe(employee['phone']))}</div>
                        <div><b>Email:</b> {escape(safe(employee['email']))}</div>
                        <div><b>Department:</b> {escape(safe(employee['department']))}</div>
                        <div><b>Job Title:</b> {escape(safe(employee['job_title']))}</div>
                        <div><b>Hire Date:</b> {escape(safe(employee['hire_date']))}</div>
                        <div><b>National ID:</b> {escape(safe(employee['national_id']))}</div>
                        <div><b>Payment Method:</b> {escape(safe(employee['payment_method']))}</div>
                        <div><b>Bank:</b> {escape(safe(employee['bank_name']))} {escape(safe(employee['bank_account']))}</div>
                        <div><b>Status:</b> {status}</div>
                    </div>
                </div>
            </div>

            <div style="margin-top:24px;border-top:1px solid #e5edf7;padding-top:18px;">
                <h3 style="margin-top:0;">Payroll Setup</h3>
                <div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:10px;color:#102a4c;">
                    <div><b>Basic:</b> {money(employee['basic_salary'])}</div>
                    <div><b>Housing:</b> {money(employee['housing_allowance'])}</div>
                    <div><b>Transport:</b> {money(employee['transport_allowance'])}</div>
                    <div><b>Other:</b> {money(employee['other_allowance'])}</div>
                    <div><b>Insurance Salary:</b> {money(employee['insurance_salary'])}</div>
                    <div><b>Daily Hours:</b> {money(employee['expected_daily_hours'])}</div>
                </div>
            </div>
        </div>

        <div>
            {render_audit_log_card('employee', employee_id, 'Activity Log')}
        </div>
    </div>
    """
    return render_page(f"Employee {safe(employee['code'])}", html, "en", current_path=request_path)


def employee_form_html(action, data=None):
    data = data or {}
    active_yes = "selected" if str(data.get("is_active", "1")) == "1" else ""
    active_no = "selected" if str(data.get("is_active", "1")) != "1" else ""

    payment_method = safe(data.get("payment_method") or "bank_transfer")
    selected_bank = "selected" if payment_method == "bank_transfer" else ""
    selected_cash = "selected" if payment_method == "cash" else ""
    insurance_yes = "selected" if str(data.get("insurance_applicable", "1")) == "1" else ""
    insurance_no = "selected" if str(data.get("insurance_applicable", "1")) != "1" else ""

    return f"""
    <div class="card">
        <h2>{'Edit Employee' if '/edit' in action else 'New Employee'}</h2>

        <form method="post" action="{action}">
            <div class="row">
                <div class="col">
                    <label>Code</label>
                    <input name="code" value="{escape(safe(data.get('code') or next_employee_code()))}" readonly>
                </div>
                <div class="col">
                    <label>Employee Name</label>
                    <input name="name" value="{escape(safe(data.get('name')))}" required>
                </div>
            </div>

            <div class="row" style="margin-top:14px;">
                <div class="col">
                    <label>Phone</label>
                    <input name="phone" value="{escape(safe(data.get('phone')))}">
                </div>
                <div class="col">
                    <label>Email</label>
                    <input name="email" value="{escape(safe(data.get('email')))}">
                </div>
            </div>

            <div class="row" style="margin-top:14px;">
                <div class="col">
                    <label>Biometric Code</label>
                    <input name="biometric_code" value="{escape(safe(data.get('biometric_code')))}">
                </div>
                <div class="col">
                    <label>Employee Category</label>
                    <select name="category_id">
                        {category_options_html(data.get('category_id') or '')}
                    </select>
                </div>
            </div>

            <div class="row" style="margin-top:14px;">
                <div class="col">
                    <label>Department</label>
                    <input name="department" value="{escape(safe(data.get('department')))}">
                </div>
                <div class="col">
                    <label>Job Title</label>
                    <input name="job_title" value="{escape(safe(data.get('job_title')))}">
                </div>
            </div>

            <div class="row" style="margin-top:14px;">
                <div class="col">
                    <label>Hire Date</label>
                    <input type="date" name="hire_date" value="{escape(safe(data.get('hire_date')))}">
                </div>
                <div class="col">
                    <label>National ID</label>
                    <input name="national_id" value="{escape(safe(data.get('national_id')))}">
                </div>
            </div>

            <div class="card" style="margin-top:18px;">
                <h3>Payroll Setup</h3>
                <div class="row">
                    <div class="col">
                        <label>Basic Salary</label>
                        <input type="number" step="0.01" name="basic_salary" value="{safe(data.get('basic_salary') or '0')}">
                    </div>
                    <div class="col">
                        <label>Housing Allowance</label>
                        <input type="number" step="0.01" name="housing_allowance" value="{safe(data.get('housing_allowance') or '0')}">
                    </div>
                </div>

                <div class="row" style="margin-top:14px;">
                    <div class="col">
                        <label>Expected Daily Hours</label>
                        <input type="number" step="0.01" name="expected_daily_hours" value="{safe(data.get('expected_daily_hours') or '8')}">
                    </div>
                    <div class="col">
                        <label>Transport Allowance</label>
                        <input type="number" step="0.01" name="transport_allowance" value="{safe(data.get('transport_allowance') or '0')}">
                    </div>
                </div>

                <div class="row" style="margin-top:14px;">
                    <div class="col">
                        <label>Other Allowance</label>
                        <input type="number" step="0.01" name="other_allowance" value="{safe(data.get('other_allowance') or '0')}">
                    </div>
                    <div class="col"></div>
                </div>

                <div class="row" style="margin-top:14px;">
                    <div class="col">
                        <label>Insurance Applicable</label>
                        <select name="insurance_applicable">
                            <option value="1" {insurance_yes}>Yes</option>
                            <option value="0" {insurance_no}>No</option>
                        </select>
                    </div>
                    <div class="col">
                        <label>Insurance Number</label>
                        <input name="insurance_number" value="{escape(safe(data.get('insurance_number')))}">
                    </div>
                </div>

                <div class="row" style="margin-top:14px;">
                    <div class="col">
                        <label>Insurance Salary</label>
                        <input type="number" step="0.01" name="insurance_salary" value="{safe(data.get('insurance_salary') or '0')}">
                    </div>
                    <div class="col">
                        <label>Employee Insurance Rate %</label>
                        <input type="number" step="0.01" name="insurance_employee_rate" value="{safe(data.get('insurance_employee_rate') or '11')}">
                    </div>
                </div>

                <div class="row" style="margin-top:14px;">
                    <div class="col">
                        <label>Employer Insurance Rate %</label>
                        <input type="number" step="0.01" name="insurance_employer_rate" value="{safe(data.get('insurance_employer_rate') or '18.75')}">
                    </div>
                    <div class="col"></div>
                </div>

                <div class="row" style="margin-top:14px;">
                    <div class="col">
                        <label>Payment Method</label>
                        <select name="payment_method">
                            <option value="bank_transfer" {selected_bank}>Bank Transfer</option>
                            <option value="cash" {selected_cash}>Cash</option>
                        </select>
                    </div>
                    <div class="col">
                        <label>Bank Name</label>
                        <input name="bank_name" value="{escape(safe(data.get('bank_name')))}">
                    </div>
                </div>

                <div class="row" style="margin-top:14px;">
                    <div class="col">
                        <label>Bank Account</label>
                        <input name="bank_account" value="{escape(safe(data.get('bank_account')))}">
                    </div>
                    <div class="col">
                        <label>Status</label>
                        <select name="is_active">
                            <option value="1" {active_yes}>Active</option>
                            <option value="0" {active_no}>Inactive</option>
                        </select>
                    </div>
                </div>
            </div>

            <div style="margin-top:20px;">
                <button class="btn green" type="submit">Save</button>
                <a class="btn gray" href="/ui/hr/employees">Back</a>
            </div>
        </form>
    </div>
    """


def get_header_value(row, aliases):
    for key in aliases:
        if key in row and safe(row.get(key)):
            return safe(row.get(key))
    return ""


def normalize_import_row(row):
    return {
        "code": get_header_value(row, ["code", "employee_code", "emp_code"]),
        "category_id": get_header_value(row, ["category_id"]),
        "category_code": get_header_value(row, ["category_code", "employee_category_code"]),
        "category_name": get_header_value(row, ["category_name", "employee_category"]),
        "biometric_code": get_header_value(row, ["biometric_code", "machine_code", "attendance_code"]),
        "name": get_header_value(row, ["name", "employee_name", "full_name"]),
        "phone": get_header_value(row, ["phone", "mobile"]),
        "email": get_header_value(row, ["email", "mail"]),
        "department": get_header_value(row, ["department", "dept"]),
        "job_title": get_header_value(row, ["job_title", "title", "position"]),
        "hire_date": get_header_value(row, ["hire_date", "joining_date", "join_date"]),
        "national_id": get_header_value(row, ["national_id", "id_no", "id_number"]),
        "payment_method": get_header_value(row, ["payment_method", "salary_method"]) or "bank_transfer",
        "bank_name": get_header_value(row, ["bank_name", "bank"]),
        "bank_account": get_header_value(row, ["bank_account", "account_no", "account_number"]),
        "basic_salary": get_header_value(row, ["basic_salary", "salary_basic", "basic"]) or "0",
        "housing_allowance": get_header_value(row, ["housing_allowance", "housing"]) or "0",
        "transport_allowance": get_header_value(row, ["transport_allowance", "transport"]) or "0",
        "other_allowance": get_header_value(row, ["other_allowance", "other"]) or "0",
        "insurance_applicable": get_header_value(row, ["insurance_applicable", "insured"]) or "1",
        "insurance_number": get_header_value(row, ["insurance_number", "social_insurance_no"]),
        "insurance_salary": get_header_value(row, ["insurance_salary", "social_insurance_salary"]) or "0",
        "insurance_employee_rate": get_header_value(row, ["insurance_employee_rate", "employee_insurance_rate"]) or "11",
        "insurance_employer_rate": get_header_value(row, ["insurance_employer_rate", "employer_insurance_rate"]) or "18.75",
        "expected_daily_hours": get_header_value(row, ["expected_daily_hours", "daily_hours"]) or "8",
        "is_active": get_header_value(row, ["is_active", "active", "status"]) or "1",
    }


def parse_csv_rows(file_bytes):
    text = file_bytes.decode("utf-8-sig", errors="ignore")
    reader = csv.DictReader(io.StringIO(text))
    return [dict(r) for r in reader]


def parse_xlsx_rows(file_bytes):
    if load_workbook is None:
        raise Exception("Excel import is not available right now.")

    workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [safe(h).lower() for h in rows[0]]
    result = []
    for data_row in rows[1:]:
        item = {}
        for i, header in enumerate(headers):
            if not header:
                continue
            item[header] = "" if i >= len(data_row) or data_row[i] is None else str(data_row[i])
        result.append(item)
    return result


ensure_employees_table()


@router.get("/ui/hr/employees", response_class=HTMLResponse)
def employees_list(request: Request):
    ensure_employees_table()
    conn = get_conn()
    rows = conn.execute(
        """
        SELECT e.*, c.name AS category_name
        FROM employees e
        LEFT JOIN employee_categories c ON c.id = e.category_id
        ORDER BY id DESC
        """
    ).fetchall()
    conn.close()

    msg = safe(request.query_params.get("msg"))
    msg_html = f'<div class="msg success">{escape(msg)}</div>' if msg else ""

    body = ""
    for r in rows:
        total_package = (
            float(r["basic_salary"] or 0)
            + float(r["housing_allowance"] or 0)
            + float(r["transport_allowance"] or 0)
            + float(r["other_allowance"] or 0)
        )
        status_html = '<span class="status-chip green">Active</span>' if int(r["is_active"] or 0) == 1 else '<span class="status-chip gray">Inactive</span>'
        body += f"""
        <tr>
            <td>{escape(safe(r['code']))}</td>
            <td>{escape(safe(r['biometric_code']))}</td>
            <td>{escape(safe(r['name']))}</td>
            <td>{escape(safe(r['category_name']))}</td>
            <td>{escape(safe(r['department']))}</td>
            <td>{escape(safe(r['job_title']))}</td>
            <td>{escape(safe(r['hire_date']))}</td>
            <td class="number-cell">{money(r['basic_salary'])}</td>
            <td class="number-cell">{money(r['insurance_salary'])}</td>
            <td class="number-cell">{money(total_package)}</td>
            <td>{status_html}</td>
            <td>
                <a class="btn gray" href="/ui/hr/employees/{r['id']}">Open</a>
                <a class="btn blue" href="/ui/hr/employees/{r['id']}/edit">Edit</a>
                <a class="btn red" href="/ui/hr/employees/{r['id']}/delete" onclick="return confirm('Are you sure you want to delete this employee? This action cannot be undone.')">Delete</a>
            </td>
        </tr>
        """

    if not body:
        body = "<tr><td colspan='12' style='text-align:center;'>No employees found.</td></tr>"

    html = f"""
    <div class="card">
        {msg_html}
        <div style="display:flex;justify-content:space-between;align-items:center;gap:10px;flex-wrap:wrap;">
            <h2>Employees</h2>
            <div style="display:flex;gap:8px;flex-wrap:wrap;">
                <a class="btn gray" href="/ui/hr/categories">Employee Categories</a>
                <a class="btn blue" href="/ui/hr/employees/export/excel">Export Excel</a>
                <a class="btn blue" href="/ui/hr/employees/template.csv">Template CSV</a>
                <a class="btn blue" href="/ui/hr/employees/template.xlsx">Template Excel</a>
                <a class="btn green" href="/ui/hr/employees/new">+ New Employee</a>
            </div>
        </div>

        <form method="post" action="/ui/hr/employees/import" enctype="multipart/form-data" style="margin-top:14px;">
            <div class="row">
                <div class="col">
                    <label>Import Employees (CSV / XLSX)</label>
                    <input type="file" name="file" accept=".csv,.xlsx" required>
                </div>
                <div class="col" style="display:flex;align-items:end;gap:8px;">
                    <select name="import_option" style="padding:8px;border:1px solid #ddd;border-radius:4px;">
                        <option value="append">Append to existing employees</option>
                        <option value="replace">Replace all employees</option>
                    </select>
                    <button class="btn green" type="submit" onclick="return confirm('Are you sure you want to import employees?')">Import</button>
                </div>
            </div>
        </form>
    </div>

    <div class="card">
        <p class="section-note">Link each employee to a category so attendance role, shift timing, grace minutes, and overtime policy can be applied automatically during attendance import and payroll processing.</p>
    </div>

    <div class="card">
        <table>
            <tr>
                <th>Code</th>
                <th>Biometric</th>
                <th>Name</th>
                <th>Category</th>
                <th>Department</th>
                <th>Job Title</th>
                <th>Hire Date</th>
                <th>Basic Salary</th>
                <th>Insurance Salary</th>
                <th>Total Package</th>
                <th>Status</th>
                <th>Actions</th>
            </tr>
            {body}
        </table>
    </div>
    """
    return HTMLResponse(render_page("Employees", html, "en", current_path=request.url.path))


@router.get("/ui/hr/employees/new", response_class=HTMLResponse)
def new_employee(request: Request):
    html = employee_form_html("/ui/hr/employees/new")
    return HTMLResponse(render_page("New Employee", html, "en", current_path=request.url.path))


@router.post("/ui/hr/employees/new")
def create_employee(
    request: Request,
    code: str = Form(""),
    category_id: str = Form(""),
    biometric_code: str = Form(""),
    name: str = Form(...),
    phone: str = Form(""),
    email: str = Form(""),
    department: str = Form(""),
    job_title: str = Form(""),
    hire_date: str = Form(""),
    national_id: str = Form(""),
    payment_method: str = Form("bank_transfer"),
    bank_name: str = Form(""),
    bank_account: str = Form(""),
    basic_salary: str = Form("0"),
    housing_allowance: str = Form("0"),
    transport_allowance: str = Form("0"),
    other_allowance: str = Form("0"),
    insurance_applicable: int = Form(1),
    insurance_number: str = Form(""),
    insurance_salary: str = Form("0"),
    insurance_employee_rate: str = Form("11"),
    insurance_employer_rate: str = Form("18.75"),
    expected_daily_hours: str = Form("8"),
    is_active: int = Form(1),
):
    ensure_employees_table()
    conn = get_conn()
    employee_code = safe(code) or next_employee_code()

    cur = conn.execute(
        """
        INSERT INTO employees (
            code, category_id, biometric_code, name, phone, email, department, job_title, hire_date, national_id,
            payment_method, bank_name, bank_account,
            basic_salary, housing_allowance, transport_allowance, other_allowance,
            insurance_applicable, insurance_number, insurance_salary, insurance_employee_rate, insurance_employer_rate,
            expected_daily_hours, is_active
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            employee_code,
            int(category_id) if safe(category_id).isdigit() else None,
            safe(biometric_code),
            safe(name),
            safe(phone),
            safe(email),
            safe(department),
            safe(job_title),
            safe(hire_date),
            safe(national_id),
            safe(payment_method) or "bank_transfer",
            safe(bank_name),
            safe(bank_account),
            to_float(basic_salary),
            to_float(housing_allowance),
            to_float(transport_allowance),
            to_float(other_allowance),
            int(insurance_applicable or 0),
            safe(insurance_number),
            to_float(insurance_salary),
            to_float(insurance_employee_rate),
            to_float(insurance_employer_rate),
            to_float(expected_daily_hours, 8),
            int(is_active or 0),
        ),
    )
    safe_log_request_action(
        request,
        "employee",
        int(cur.lastrowid),
        "Created",
        f"Employee {employee_code} - {safe(name)} created.",
        conn=conn,
        module="hr",
    )
    conn.commit()
    conn.close()
    return RedirectResponse("/ui/hr/employees", status_code=302)


@router.get("/ui/hr/employees/{employee_id}", response_class=HTMLResponse)
def open_employee(request: Request, employee_id: int):
    ensure_employees_table()
    conn = get_conn()
    row = conn.execute("SELECT * FROM employees WHERE id = ? LIMIT 1", (employee_id,)).fetchone()
    if not row:
        conn.close()
        return HTMLResponse("Employee not found", status_code=404)
    html = employee_profile_html(conn, row, str(request.url.path))
    conn.close()
    return HTMLResponse(html)


@router.get("/ui/hr/employees/{employee_id}/delete", response_class=HTMLResponse)
def delete_employee(request: Request, employee_id: int):
    ensure_employees_table()
    conn = get_conn()
    
    try:
        # Check if employee exists
        employee = conn.execute("SELECT * FROM employees WHERE id = ? LIMIT 1", (employee_id,)).fetchone()
        if not employee:
            return RedirectResponse("/ui/hr/employees?msg=" + quote("Employee not found"), status_code=302)
        
        # Check if employee has related data (payroll, attendance, advances)
        payroll_count = conn.execute("SELECT COUNT(*) FROM payroll_lines WHERE employee_id = ?", (employee_id,)).fetchone()[0]
        attendance_count = conn.execute("SELECT COUNT(*) FROM attendance_logs WHERE employee_id = ?", (employee_id,)).fetchone()[0]
        advances_count = conn.execute("SELECT COUNT(*) FROM employee_advances WHERE employee_id = ?", (employee_id,)).fetchone()[0]
        
        # If employee has related data, show warning
        if payroll_count > 0 or attendance_count > 0 or advances_count > 0:
            msg = f"Cannot delete employee {employee['code']} - {employee['name']}. Related data found: "
            if payroll_count > 0:
                msg += f"{payroll_count} payroll records, "
            if attendance_count > 0:
                msg += f"{attendance_count} attendance records, "
            if advances_count > 0:
                msg += f"{advances_count} advance records"
            
            return RedirectResponse("/ui/hr/employees?msg=" + quote(msg), status_code=302)
        
        # Delete the employee
        conn.execute("DELETE FROM employees WHERE id = ?", (employee_id,))
        conn.commit()
        
        return RedirectResponse("/ui/hr/employees?msg=" + quote(f"Employee {employee['code']} - {employee['name']} deleted successfully"), status_code=302)
        
    except Exception as e:
        return RedirectResponse("/ui/hr/employees?msg=" + quote(f"Error deleting employee: {str(e)}"), status_code=302)
    finally:
        conn.close()


@router.get("/ui/hr/employees/{employee_id}/edit", response_class=HTMLResponse)
def edit_employee(request: Request, employee_id: int):
    ensure_employees_table()
    conn = get_conn()
    row = conn.execute("SELECT * FROM employees WHERE id = ? LIMIT 1", (employee_id,)).fetchone()
    conn.close()

    if not row:
        return HTMLResponse("Employee not found", status_code=404)

    html = employee_form_html(f"/ui/hr/employees/{employee_id}/edit", dict(row))
    return HTMLResponse(render_page("Edit Employee", html, "en", current_path=request.url.path))


@router.post("/ui/hr/employees/{employee_id}/edit")
def update_employee(
    request: Request,
    employee_id: int,
    code: str = Form(""),
    category_id: str = Form(""),
    biometric_code: str = Form(""),
    name: str = Form(...),
    phone: str = Form(""),
    email: str = Form(""),
    department: str = Form(""),
    job_title: str = Form(""),
    hire_date: str = Form(""),
    national_id: str = Form(""),
    payment_method: str = Form("bank_transfer"),
    bank_name: str = Form(""),
    bank_account: str = Form(""),
    basic_salary: str = Form("0"),
    housing_allowance: str = Form("0"),
    transport_allowance: str = Form("0"),
    other_allowance: str = Form("0"),
    insurance_applicable: int = Form(1),
    insurance_number: str = Form(""),
    insurance_salary: str = Form("0"),
    insurance_employee_rate: str = Form("11"),
    insurance_employer_rate: str = Form("18.75"),
    expected_daily_hours: str = Form("8"),
    is_active: int = Form(1),
):
    ensure_employees_table()
    conn = get_conn()
    current_row = conn.execute("SELECT code FROM employees WHERE id = ? LIMIT 1", (employee_id,)).fetchone()
    if not current_row:
        conn.close()
        return HTMLResponse("Employee not found", status_code=404)

    employee_code = safe(code) or safe(current_row["code"]) or next_employee_code()

    conn.execute(
        """
        UPDATE employees
        SET code = ?,
            category_id = ?,
            biometric_code = ?,
            name = ?,
            phone = ?,
            email = ?,
            department = ?,
            job_title = ?,
            hire_date = ?,
            national_id = ?,
            payment_method = ?,
            bank_name = ?,
            bank_account = ?,
            basic_salary = ?,
            housing_allowance = ?,
            transport_allowance = ?,
            other_allowance = ?,
            insurance_applicable = ?,
            insurance_number = ?,
            insurance_salary = ?,
            insurance_employee_rate = ?,
            insurance_employer_rate = ?,
            expected_daily_hours = ?,
            is_active = ?
        WHERE id = ?
        """,
        (
            employee_code,
            int(category_id) if safe(category_id).isdigit() else None,
            safe(biometric_code),
            safe(name),
            safe(phone),
            safe(email),
            safe(department),
            safe(job_title),
            safe(hire_date),
            safe(national_id),
            safe(payment_method) or "bank_transfer",
            safe(bank_name),
            safe(bank_account),
            to_float(basic_salary),
            to_float(housing_allowance),
            to_float(transport_allowance),
            to_float(other_allowance),
            int(insurance_applicable or 0),
            safe(insurance_number),
            to_float(insurance_salary),
            to_float(insurance_employee_rate),
            to_float(insurance_employer_rate),
            to_float(expected_daily_hours, 8),
            int(is_active or 0),
            employee_id,
        ),
    )
    safe_log_request_action(
        request,
        "employee",
        employee_id,
        "Updated",
        f"Employee {employee_code} - {safe(name)} updated.",
        conn=conn,
        module="hr",
    )
    conn.commit()
    conn.close()

    return RedirectResponse(f"/ui/hr/employees/{employee_id}", status_code=302)


@router.get("/ui/hr/employees/template.csv")
def employees_template_csv():
    headers = [
        "code",
        "category_id",
        "category_code",
        "category_name",
        "biometric_code",
        "name",
        "phone",
        "email",
        "department",
        "job_title",
        "hire_date",
        "national_id",
        "payment_method",
        "bank_name",
        "bank_account",
        "basic_salary",
        "housing_allowance",
        "transport_allowance",
        "other_allowance",
        "insurance_applicable",
        "insurance_number",
        "insurance_salary",
        "insurance_employee_rate",
        "insurance_employer_rate",
        "expected_daily_hours",
        "is_active",
    ]
    rows = [
        ["EMP-0001", "", "CAT-OFF", "Office Staff", "1001", "Ahmed Ali", "01000000000", "ahmed@company.com", "Finance", "Accountant", "2026-01-01", "29801011234567", "bank_transfer", "CIB", "00123456789", "12000", "2500", "1500", "500", "1", "INS-1001", "12000", "11", "18.75", "8", "1"],
        ["EMP-0002", "", "CAT-MGT", "Management", "1002", "Mona Salah", "01000000001", "mona@company.com", "HR", "HR Specialist", "2026-02-01", "29802021234567", "cash", "", "", "9000", "1500", "1000", "0", "1", "INS-1002", "9000", "11", "18.75", "8", "1"],
    ]
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)

    data = output.getvalue().encode("utf-8-sig")
    stream = io.BytesIO(data)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": 'attachment; filename="employees_template.csv"'},
    )


@router.get("/ui/hr/employees/export/excel")
def export_employees_excel():
    if Workbook is None:
        return RedirectResponse("/ui/hr/employees?msg=" + quote("Excel export is not available. Install openpyxl."))
    
    ensure_employees_table()
    conn = get_conn()
    
    try:
        rows = conn.execute("""
            SELECT 
                e.id,
                e.code as employee_code,
                e.name as employee_name,
                e.department,
                e.job_title,
                e.basic_salary,
                e.housing_allowance,
                e.transport_allowance,
                e.other_allowance,
                e.insurance_salary,
                e.insurance_applicable,
                e.insurance_employee_rate,
                e.insurance_employer_rate,
                e.expected_daily_hours,
                e.is_active,
                COALESCE(e.hire_date, ''),
                e.phone,
                e.email,
                e.biometric_code,
                e.status
            FROM employees e
            ORDER BY e.code
        """).fetchall()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Employees Data"
        
        headers = [
            'ID', 'Employee Code', 'Employee Name', 'Department', 'Job Title',
            'Basic Salary', 'Housing Allowance', 'Transport Allowance', 'Other Allowance',
            'Insurance Salary', 'Insurance Applicable', 'Employee Rate %', 'Employer Rate %',
            'Daily Hours', 'Active', 'Hire Date', 'Phone', 'Email', 'Biometric Code', 'Status'
        ]
        
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)
        
        for row_num, emp in enumerate(rows, 2):
            ws.cell(row=row_num, column=1, value=emp[0])
            ws.cell(row=row_num, column=2, value=emp[1])
            ws.cell(row=row_num, column=3, value=emp[2])
            ws.cell(row=row_num, column=4, value=emp[3])
            ws.cell(row=row_num, column=5, value=emp[4])
            ws.cell(row=row_num, column=6, value=emp[5])
            ws.cell(row=row_num, column=7, value=emp[6])
            ws.cell(row=row_num, column=8, value=emp[7])
            ws.cell(row=row_num, column=9, value=emp[8])
            ws.cell(row=row_num, column=10, value=emp[9])
            ws.cell(row=row_num, column=11, value=emp[10])
            ws.cell(row=row_num, column=12, value=emp[11])
            ws.cell(row=row_num, column=13, value=emp[12])
            ws.cell(row=row_num, column=14, value=emp[13])
            ws.cell(row=row_num, column=15, value=emp[14])
            ws.cell(row=row_num, column=16, value=emp[15])
            ws.cell(row=row_num, column=17, value=emp[16])
            ws.cell(row=row_num, column=18, value=emp[17])
            ws.cell(row=row_num, column=19, value=emp[18])
            ws.cell(row=row_num, column=20, value=emp[19])
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"employees_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return StreamingResponse(
            io.BytesIO(output.getvalue()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        return RedirectResponse("/ui/hr/employees?msg=" + quote(f"Export error: {str(e)}"))
    finally:
        conn.close()


@router.get("/ui/hr/employees/template.xlsx")
def employees_template_xlsx():
    if Workbook is None:
        return RedirectResponse("/ui/hr/employees?msg=" + quote("Excel template is not available. Use CSV template."))

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Employees Template"
    headers = [
        "code",
        "category_id",
        "category_code",
        "category_name",
        "biometric_code",
        "name",
        "phone",
        "email",
        "department",
        "job_title",
        "hire_date",
        "national_id",
        "payment_method",
        "bank_name",
        "bank_account",
        "basic_salary",
        "housing_allowance",
        "transport_allowance",
        "other_allowance",
        "insurance_applicable",
        "insurance_number",
        "insurance_salary",
        "insurance_employee_rate",
        "insurance_employer_rate",
        "expected_daily_hours",
        "is_active",
    ]
    sheet.append(headers)
    sheet.append(["EMP-0001", "", "CAT-OFF", "Office Staff", "1001", "Ahmed Ali", "01000000000", "ahmed@company.com", "Finance", "Accountant", "2026-01-01", "29801011234567", "bank_transfer", "CIB", "00123456789", "12000", "2500", "1500", "500", "1", "INS-1001", "12000", "11", "18.75", "8", "1"])
    sheet.append(["EMP-0002", "", "CAT-MGT", "Management", "1002", "Mona Salah", "01000000001", "mona@company.com", "HR", "HR Specialist", "2026-02-01", "29802021234567", "cash", "", "", "9000", "1500", "1000", "0", "1", "INS-1002", "9000", "11", "18.75", "8", "1"])

    lookup_sheet = workbook.create_sheet("Lookups")
    lookup_sheet.append(["payment_method", "is_active"])
    lookup_sheet.append(["bank_transfer", "1"])
    lookup_sheet.append(["cash", "0"])

    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="employees_template.xlsx"'},
    )


@router.post("/ui/hr/employees/import")
async def employees_import(file: UploadFile = File(...), import_option: str = Form("append")):
    ensure_employees_table()
    filename = safe(file.filename).lower()
    file_bytes = await file.read()

    if not filename:
        return RedirectResponse("/ui/hr/employees?msg=" + quote("Please choose a file to import."), status_code=302)

    try:
        if filename.endswith(".csv"):
            raw_rows = parse_csv_rows(file_bytes)
        elif filename.endswith(".xlsx"):
            raw_rows = parse_xlsx_rows(file_bytes)
        else:
            return RedirectResponse("/ui/hr/employees?msg=" + quote("Only CSV or XLSX files are supported."), status_code=302)
        
        # Debug: Show how many rows were parsed
        print(f"DEBUG: Parsed {len(raw_rows)} rows from file")
        
    except Exception as ex:
        print(f"DEBUG: Parse error: {ex}")
        return RedirectResponse("/ui/hr/employees?msg=" + quote(f"Import failed: {safe(ex)}"), status_code=302)

    conn = get_conn()
    imported = 0
    skipped = 0
    next_seq = 1
    try:
        # If replace option, delete all existing employees
        if import_option == "replace":
            conn.execute("DELETE FROM employees")
            conn.commit()
            next_seq = 1
        else:
            # For append option, get existing codes
            existing = conn.execute(
                """
                SELECT code
                FROM employees
                WHERE COALESCE(code, '') <> ''
                ORDER BY id DESC
                LIMIT 1
                """
            ).fetchone()
            if existing:
                try:
                    next_seq = int(safe(existing["code"]).split("-")[-1]) + 1
                except Exception:
                    next_seq = 1

        print(f"DEBUG: Processing {len(raw_rows)} rows")
        
        for i, raw_row in enumerate(raw_rows):
            print(f"DEBUG: Processing row {i+1}: {raw_row}")
            
            row = normalize_import_row(raw_row)
            print(f"DEBUG: Normalized row: {row}")
            
            if not safe(row["name"]):
                print(f"DEBUG: Skipping row {i+1} - no name")
                skipped += 1
                continue

            code = safe(row["code"])
            if not code:
                code = f"EMP-{next_seq:04d}"
                next_seq += 1

            exists = conn.execute("SELECT id FROM employees WHERE code = ? LIMIT 1", (code,)).fetchone()
            if exists:
                print(f"DEBUG: Skipping row {i+1} - code {code} already exists")
                skipped += 1
                continue

            print(f"DEBUG: Inserting employee: {code} - {safe(row['name'])}")
            
            conn.execute(
                """
                INSERT INTO employees (
                    code, category_id, biometric_code, name, phone, email, department, job_title, hire_date, national_id,
                    payment_method, bank_name, bank_account,
                    basic_salary, housing_allowance, transport_allowance, other_allowance,
                    insurance_applicable, insurance_number, insurance_salary, insurance_employee_rate, insurance_employer_rate,
                    expected_daily_hours, is_active
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    code,
                    resolve_category_id(conn, row["category_id"], row["category_code"], row["category_name"]),
                    safe(row["biometric_code"]),
                    safe(row["name"]),
                    safe(row["phone"]),
                    safe(row["email"]),
                    safe(row["department"]),
                    safe(row["job_title"]),
                    safe(row["hire_date"]),
                    safe(row["national_id"]),
                    safe(row["payment_method"]) or "bank_transfer",
                    safe(row["bank_name"]),
                    safe(row["bank_account"]),
                    to_float(row["basic_salary"]),
                    to_float(row["housing_allowance"]),
                    to_float(row["transport_allowance"]),
                    to_float(row["other_allowance"]),
                    to_int_flag(row["insurance_applicable"], 1),
                    safe(row["insurance_number"]),
                    to_float(row["insurance_salary"]),
                    to_float(row["insurance_employee_rate"], 11),
                    to_float(row["insurance_employer_rate"], 18.75),
                    to_float(row["expected_daily_hours"], 8),
                    to_int_flag(row["is_active"], 1),
                ),
            )
            imported += 1
            print(f"DEBUG: Successfully imported employee {code}")

        conn.commit()
    finally:
        conn.close()

    print(f"DEBUG: Import completed - Imported: {imported}, Skipped: {skipped}")
    return RedirectResponse(
        "/ui/hr/employees?msg=" + quote(f"Employees imported: {imported}. Skipped: {skipped}."),
        status_code=302,
    )
